"""
Вспомогательные функции для работы с табличными файлами (XLSX, XLS, CSV)
и отображения прогресса.
"""
from __future__ import annotations

import csv
import os
import re
import time
import zipfile
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

from openpyxl import load_workbook

from app_logger import get_logger

_log = get_logger(__name__)


# ---------------------------------------------------------------------------
# Пределы размера входных файлов (защита от OOM / ZIP-bomb)
# ---------------------------------------------------------------------------
# Дефолт: 200 МБ. Переопределяется через env MAX_XLSX_BYTES / MAX_CSV_BYTES.
_DEFAULT_MAX_XLSX_BYTES = 200 * 1024 * 1024
_DEFAULT_MAX_CSV_BYTES = 500 * 1024 * 1024
# Жёсткий потолок для распакованного содержимого XLSX (ZIP-bomb guard).
# XLSX обычно сжимается 5–20×; 2 ГБ распакованного — уже явно подозрительно.
_DEFAULT_MAX_XLSX_UNCOMPRESSED_BYTES = 2 * 1024 * 1024 * 1024


def _env_int(name: str, default: int) -> int:
    raw = os.environ.get(name)
    if not raw:
        return default
    try:
        value = int(raw)
        return value if value > 0 else default
    except ValueError:
        _log.warning("invalid %s=%r, using default=%d", name, raw, default)
        return default


def max_xlsx_bytes() -> int:
    return _env_int("MAX_XLSX_BYTES", _DEFAULT_MAX_XLSX_BYTES)


def max_csv_bytes() -> int:
    return _env_int("MAX_CSV_BYTES", _DEFAULT_MAX_CSV_BYTES)


def max_xlsx_uncompressed_bytes() -> int:
    return _env_int("MAX_XLSX_UNCOMPRESSED_BYTES", _DEFAULT_MAX_XLSX_UNCOMPRESSED_BYTES)


class TabularFileTooLargeError(ValueError):
    """Файл превышает допустимый размер (защита от OOM / ZIP-bomb)."""


def _check_size_limits(path: Path) -> None:
    """Проверяет, что файл и (для XLSX) распакованное содержимое укладываются в лимиты.

    Raises:
        TabularFileTooLargeError — если файл превышает лимит.
    """
    try:
        compressed_size = path.stat().st_size
    except OSError:
        return

    if _is_csv(path):
        limit = max_csv_bytes()
        if compressed_size > limit:
            raise TabularFileTooLargeError(
                f"CSV {path.name}: {compressed_size} байт > лимит {limit}. "
                f"Увеличьте MAX_CSV_BYTES или разделите файл."
            )
        return

    if _is_excel(path):
        limit = max_xlsx_bytes()
        if compressed_size > limit:
            raise TabularFileTooLargeError(
                f"XLSX {path.name}: {compressed_size} байт > лимит {limit}. "
                f"Увеличьте MAX_XLSX_BYTES или разделите файл."
            )
        uncompressed_limit = max_xlsx_uncompressed_bytes()
        total_uncompressed = 0
        try:
            with zipfile.ZipFile(path, "r") as zf:
                for info in zf.infolist():
                    total_uncompressed += info.file_size
                    if total_uncompressed > uncompressed_limit:
                        raise TabularFileTooLargeError(
                            f"XLSX {path.name}: распакованный размер "
                            f"> {uncompressed_limit} байт — возможна ZIP-bomb."
                        )
        except zipfile.BadZipFile as exc:
            raise TabularFileTooLargeError(
                f"XLSX {path.name}: не удалось открыть как ZIP — {exc}"
            ) from exc


# ---------------------------------------------------------------------------
# Определение типа файла
# ---------------------------------------------------------------------------

def _is_csv(path: Path) -> bool:
    return path.suffix.lower() == ".csv"


def _is_excel(path: Path) -> bool:
    return path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm")


# ---------------------------------------------------------------------------
# Универсальный итератор строк
# ---------------------------------------------------------------------------

@contextmanager
def open_tabular(path: Path) -> Iterator[Iterator[tuple]]:
    """Контекстный менеджер — отдаёт итератор строк для CSV или XLSX/XLS.

    Каждая строка — кортеж значений (str | None).
    Первая строка — заголовок (как обычно).

    Использование::

        with open_tabular(path) as rows:
            header = next(rows)
            for row in rows:
                ...
    """
    _check_size_limits(path)

    if _is_csv(path):
        # Пробуем несколько кодировок, типичных для CSV из Excel
        encodings = ["utf-8-sig", "utf-8", "cp1251", "latin-1"]
        fobj = None
        for enc in encodings:
            try:
                fobj = open(path, newline="", encoding=enc)
                fobj.read(1024)          # тестовое чтение
                fobj.seek(0)
                break
            except Exception:
                if fobj:
                    fobj.close()
                fobj = None
        if fobj is None:
            fobj = open(path, newline="", encoding="latin-1", errors="replace")

        # Определяем разделитель (,  ;  \t)
        sample = fobj.read(4096)
        fobj.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel  # по умолчанию запятая

        reader = csv.reader(fobj, dialect)

        def _csv_iter():
            for raw_row in reader:
                # Приводим к единому виду: None вместо пустых строк
                yield tuple(v if v.strip() != "" else None for v in raw_row)

        try:
            yield _csv_iter()
        finally:
            fobj.close()

    else:
        # XLSX / XLSM и прочие поддерживаемые openpyxl форматы
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        try:
            yield ws.iter_rows(values_only=True)
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# Чтение заголовков
# ---------------------------------------------------------------------------

def read_headers(path: Path) -> list[str]:
    """Читает первую строку файла как список заголовков."""
    with open_tabular(path) as rows:
        header = next(rows)
    return ["" if h is None else str(h).strip() for h in header]


# ---------------------------------------------------------------------------
# Индекс колонки по имени
# ---------------------------------------------------------------------------

def idx_of(headers: list[str], col_name: str) -> int | None:
    """Возвращает индекс колонки по имени или None если не найдена."""
    if not col_name:
        return None
    try:
        return headers.index(col_name)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Оценка числа строк
# ---------------------------------------------------------------------------

def _count_csv_rows_fast(path: Path) -> int:
    """Подсчёт строк CSV на уровне байтов (~10× быстрее парсинга).

    Читает файл блоками по 1 МБ и считает символы '\\n'.
    Вычитает 1 строку заголовка. Результат — оценка, достаточная для прогресс-бара.
    """
    try:
        newlines = 0
        last_byte = b""
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1 << 20), b""):
                newlines += chunk.count(b"\n")
                last_byte = chunk[-1:]
        # Если файл не оканчивается на \n — последняя строка без newline
        if last_byte and last_byte != b"\n":
            newlines += 1
        return max(newlines - 1, 0)  # минус заголовок
    except OSError as _e:
        _log.warning("_count_csv_rows_fast(%s): %s; using fallback=1000", path, _e)
        return 1000  # fallback


_EXACT_COUNT_THRESHOLD_BYTES = 5 * 1024 * 1024  # 5 МБ


def _count_xlsx_rows_exact(p: Path) -> int:
    """Точный подсчёт непустых строк данных (без заголовка) для небольших XLSX.

    ws.max_row опирается на метаданные файла и может включать строки с форматированием
    но без данных («призрачные строки»). Итерация по строкам возвращает только реальные.
    """
    _check_size_limits(p)
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        count = 0
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue  # пропускаем заголовок
            if any(v is not None for v in row):
                count += 1
        return count  # может быть 0 — вызывающий код решает, как обрабатывать пустой файл
    finally:
        wb.close()


def estimate_total_rows(paths: list[Path]) -> int:
    """Оценивает суммарное количество строк данных (без заголовка) по всем файлам.

    Для XLSX ≤ 5 МБ использует точный подсчёт итерацией (избегает «призрачных строк»
    с форматированием, которые завышают ws.max_row). Для больших файлов — быстрый режим.
    """
    total = 0
    for p in paths:
        if _is_csv(p):
            total += _count_csv_rows_fast(p)
        elif p.stat().st_size <= _EXACT_COUNT_THRESHOLD_BYTES:
            _exact = _count_xlsx_rows_exact(p)
            if _exact == 0:
                import warnings as _wempty
                _wempty.warn(
                    f"[excel_utils] {p.name}: файл не содержит строк данных (только заголовок или пустой).",
                    UserWarning, stacklevel=2,
                )
            total += _exact
        else:
            _check_size_limits(p)
            wb = load_workbook(p, read_only=True, data_only=True)
            try:
                ws = wb.active
                total += max((ws.max_row or 1) - 1, 0)
            finally:
                wb.close()
    return max(total, 1)


# ---------------------------------------------------------------------------
# Форматирование прогресса
# ---------------------------------------------------------------------------

def fmt_eta(start_ts: float, done: int, total: int) -> str:
    """Форматирует оставшееся время в виде строки ETA."""
    if done <= 0 or total <= 0:
        return ""
    elapsed = max(0.001, time.time() - start_ts)
    speed = done / elapsed
    if speed <= 0:
        return ""
    remain = max(0, total - done)
    eta = remain / speed
    if eta < 60:
        return f"ETA {eta:.0f}s"
    if eta < 3600:
        return f"ETA {eta/60:.1f}m"
    return f"ETA {eta/3600:.2f}h"


def fmt_speed(start_ts: float, done: int) -> str:
    """Форматирует текущую скорость обработки в строках/сек."""
    elapsed = max(0.001, time.time() - start_ts)
    sp = done / elapsed
    return f"{sp:.1f} rows/s" if sp >= 1 else f"{sp:.2f} rows/s"


# ---------------------------------------------------------------------------
# Post-save Excel formatting (column widths) via ZIP/XML patching
# ---------------------------------------------------------------------------

def patch_xlsx_col_widths(path: Path, patches: list[tuple[int, list[float]]]) -> None:
    """Быстро добавляет ширины колонок в сохранённый xlsx-файл.

    Работает напрямую с ZIP-архивом без загрузки данных ячеек в openpyxl.
    patches: список (sheet_index_0based, [width_col1, width_col2, ...])
    """
    if not patches:
        return

    # Находим XML-файлы листов внутри архива
    with zipfile.ZipFile(path, "r") as zin:
        all_names = zin.namelist()
        sheet_files = sorted(
            [n for n in all_names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)],
            key=lambda n: int(re.search(r"(\d+)\.xml$", n).group(1)),
        )

        # Строим карту: имя файла → новые ширины
        patch_map: dict[str, list[float]] = {}
        for sheet_idx, widths in patches:
            if sheet_idx < len(sheet_files):
                patch_map[sheet_files[sheet_idx]] = widths

        if not patch_map:
            return

        # Читаем все данные пока файл ещё открыт
        file_contents: dict[str, bytes] = {name: zin.read(name) for name in all_names}

    # zin закрыт — теперь можно безопасно заменить файл (особенно на Windows)
    tmp = path.with_suffix(".tmp.xlsx")
    try:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in all_names:
                data = file_contents[name]
                if name in patch_map:
                    widths = patch_map[name]
                    cols_xml = "<cols>" + "".join(
                        f'<col min="{i}" max="{i}" width="{w:.2f}" customWidth="1"/>'
                        for i, w in enumerate(widths, 1)
                    ) + "</cols>"
                    xml_str = data.decode("utf-8")
                    # Убираем существующий <cols> если есть (openpyxl иногда добавляет пустой)
                    xml_str = re.sub(
                        r"<cols\b[^/]*/?>(?:.*?</cols>)?", "", xml_str, flags=re.DOTALL
                    )
                    # Инжектируем перед <sheetData
                    xml_str = xml_str.replace("<sheetData", cols_xml + "<sheetData", 1)
                    data = xml_str.encode("utf-8")
                zout.writestr(name, data)
        tmp.replace(path)
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except OSError as _e:
            _log.debug("patch_xlsx_col_widths: tmp cleanup failed: %s", _e)
        raise
