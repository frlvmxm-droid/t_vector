# -*- coding: utf-8 -*-
"""
Comprehensive test suite for the LinearSVC-Calibrated codebase.
Each test prints "PASS: <name>" or "FAIL: <name> - <error>".
Summary at the end: "Results: X passed, Y failed"
"""
from __future__ import annotations

import csv
import math
import os
import sys
import tempfile
import traceback
from pathlib import Path

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

passed = 0
failed = 0


def run_test(name: str, fn):
    global passed, failed
    try:
        fn()
        print(f"PASS: {name}")
        passed += 1
    except Exception as e:
        tb = traceback.format_exc()
        print(f"FAIL: {name} - {e}")
        print("  Traceback (most recent call last):")
        for line in tb.splitlines()[1:]:
            print("  " + line)
        failed += 1


# ===========================================================================
# text_utils tests
# ===========================================================================

def test_strip_html_basic():
    from text_utils import strip_html
    result = strip_html("<b>Hello</b> <i>World</i>")
    assert "Hello" in result and "World" in result
    assert "<b>" not in result


def test_strip_html_none():
    from text_utils import strip_html
    assert strip_html(None) == ""


def test_strip_html_nbsp():
    from text_utils import strip_html
    result = strip_html("foo&nbsp;bar")
    assert "foo" in result and "bar" in result
    assert "&nbsp;" not in result


def test_normalize_text_basic():
    from text_utils import normalize_text
    result = normalize_text("Hello\r\nWorld")
    assert "Hello" in result
    assert "World" in result
    assert "\r\n" not in result


def test_normalize_text_none():
    from text_utils import normalize_text
    assert normalize_text(None) == ""


def test_normalize_text_masks():
    from text_utils import normalize_text
    result = normalize_text("Звонил !fio! по вопросу карты")
    assert "!fio!" not in result
    # text should remain meaningful
    assert "карт" in result


def test_normalize_text_drops_noise_lines():
    from text_utils import normalize_text
    # DROP_LINES: sdk_request
    result = normalize_text("sdk_request\nнормальная строка")
    assert "sdk_request" not in result
    assert "нормальная строка" in result


def test_clean_masks():
    from text_utils import clean_masks
    result = clean_masks("Имя !fio! и номер !num! важны")
    assert "!fio!" not in result
    assert "!num!" not in result


def test_clean_masks_none():
    from text_utils import clean_masks
    result = clean_masks(None)
    assert result == ""


def test_clean_answer_text_html():
    from text_utils import clean_answer_text
    result = clean_answer_text("<p>Ваш ответ готов</p>")
    assert "<p>" not in result
    assert "ответ" in result or "готов" in result


def test_clean_answer_text_none():
    from text_utils import clean_answer_text
    assert clean_answer_text(None) == ""


def test_clean_answer_text_strips_header():
    from text_utils import clean_answer_text
    # Should strip the template answer header
    result = clean_answer_text(
        "Мы рассмотрели ваше обращение №123 22.01.2026. Ваш вопрос решён."
    )
    assert "рассмотрели" not in result.lower()
    assert "решён" in result or "вопрос" in result


def test_parse_dialog_roles_basic():
    from text_utils import parse_dialog_roles
    text = (
        "CLIENT: Здравствуйте, у меня проблема с картой, не могу снять деньги\n"
        "OPERATOR: Добрый день, сейчас помогу вам разобраться с этой ситуацией"
    )
    dialog_clean, client_text, operator_text, roles_found = parse_dialog_roles(text)
    assert roles_found is True
    assert "проблема" in client_text or "карт" in client_text


def test_parse_dialog_roles_no_roles():
    from text_utils import parse_dialog_roles
    text = "Это просто текст без ролей и структуры диалога"
    dialog_clean, client_text, operator_text, roles_found = parse_dialog_roles(text)
    assert roles_found is False
    assert client_text == ""
    assert operator_text == ""
    assert "текст" in dialog_clean


def test_parse_dialog_roles_empty():
    from text_utils import parse_dialog_roles
    dialog_clean, client_text, operator_text, roles_found = parse_dialog_roles("")
    assert roles_found is False
    assert dialog_clean == ""


def test_parse_dialog_roles_filters_noise():
    from text_utils import parse_dialog_roles
    # Operator noise: "здравствуйте слушаю" — short + noise pattern → filtered
    text = (
        "CLIENT: Мне нужна помощь с переводом средств на счёт другого банка\n"
        "OPERATOR: Здравствуйте слушаю вас\n"
        "OPERATOR: Сейчас посмотрю ваш вопрос по переводу\n"
    )
    _, client_text, operator_text, roles_found = parse_dialog_roles(text)
    assert roles_found is True
    # The substantive CLIENT line should survive
    assert "переводом" in client_text or "счёт" in client_text


def test_parse_dialog_roles_chatbot_ignored():
    from text_utils import parse_dialog_roles
    text = (
        "CHATBOT: Привет, выберите тему обращения\n"
        "CLIENT: Проблема с кредитной картой, не работает оплата\n"
    )
    dialog_clean, client_text, operator_text, roles_found = parse_dialog_roles(
        text, ignore_chatbot=True
    )
    assert roles_found is True
    # chatbot lines should not appear in clean dialog
    assert "CHATBOT" not in dialog_clean


def test_parse_dialog_roles_continuation():
    from text_utils import parse_dialog_roles
    # Continuation line (no role prefix) should be appended to previous
    text = (
        "CLIENT: Здравствуйте, у меня вопрос по вкладу\n"
        "продолжение реплики без метки роли клиента\n"
        "OPERATOR: Добрый день, слушаю вас внимательно и готов помочь\n"
    )
    _, client_text, operator_text, roles_found = parse_dialog_roles(text)
    assert roles_found is True


# ===========================================================================
# feature_builder tests
# ===========================================================================

def test_build_feature_text_basic():
    from feature_builder import build_feature_text
    weights = {
        "w_desc": 2, "w_client": 3, "w_operator": 1,
        "w_summary": 2, "w_answer_short": 1, "w_answer_full": 0
    }
    result = build_feature_text(
        channel="call",
        desc="Вопрос по кредиту",
        client_text="Клиент: Хочу узнать условия кредита",
        operator_text="Оператор: Расскажу об условиях",
        summary="Клиент интересовался кредитом",
        ans_short="Условия кредита описаны",
        ans_full="",
        weights=weights,
    )
    assert "[CHANNEL]" in result
    assert "call" in result
    assert "[DESC]" in result
    assert "кредит" in result.lower()


def test_build_feature_text_weight_zero_skips_section():
    from feature_builder import build_feature_text
    weights = {
        "w_desc": 0, "w_client": 2, "w_operator": 0,
        "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0
    }
    result = build_feature_text(
        channel="chat",
        desc="Это описание не должно попасть",
        client_text="Клиент написал о проблеме с переводом денег",
        operator_text="Оператор дал ответ",
        summary="Суммаризация",
        ans_short="Ответ",
        ans_full="",
        weights=weights,
    )
    assert "[DESC]" not in result
    assert "[OPERATOR]" not in result
    assert "[CLIENT]" in result


def test_build_feature_text_empty_fields():
    from feature_builder import build_feature_text
    weights = {
        "w_desc": 2, "w_client": 3, "w_operator": 1,
        "w_summary": 2, "w_answer_short": 1, "w_answer_full": 1
    }
    result = build_feature_text(
        channel="call",
        desc="",
        client_text="",
        operator_text="",
        summary="",
        ans_short="",
        ans_full="",
        weights=weights,
    )
    # Should still have CHANNEL section at minimum
    assert "[CHANNEL]" in result


def test_head_tail_compress_short_text():
    """_head_tail_lines: when total lines <= n_head + n_tail, returns all lines."""
    from feature_builder import _head_tail_lines
    text = "line1\nline2\nline3"
    result = _head_tail_lines(text, n_head=5, n_tail=4)
    assert "line1" in result
    assert "line2" in result
    assert "line3" in result
    assert "…" not in result


def test_head_tail_compress_long_text():
    """_head_tail_lines: long dialog keeps head + selects tail, inserts separator."""
    from feature_builder import _head_tail_lines
    lines = [f"реплика номер {i} с уникальным словом {i}" for i in range(20)]
    text = "\n".join(lines)
    result = _head_tail_lines(text, n_head=3, n_tail=3)
    # Head lines should be present
    assert "реплика номер 0" in result
    assert "реплика номер 1" in result
    assert "реплика номер 2" in result
    # Separator should be inserted
    assert "…" in result


def test_head_tail_compress_single_line():
    from feature_builder import _head_tail_lines
    text = "одна строка"
    result = _head_tail_lines(text, n_head=5, n_tail=4)
    assert "одна строка" in result


def test_head_tail_compress_empty():
    from feature_builder import _head_tail_lines
    result = _head_tail_lines("", n_head=5, n_tail=4)
    assert result == ""


def test_choose_row_profile_weights_off():
    from feature_builder import choose_row_profile_weights
    base = {"w_desc": 2, "w_client": 3, "w_operator": 1,
            "w_summary": 2, "w_answer_short": 1, "w_answer_full": 1}
    result = choose_row_profile_weights(
        base=base, auto_profile="off",
        has_desc=True, has_dialog=True, roles_found=True,
        has_summary=True, has_ans_s=True, has_ans_f=True,
    )
    assert result == base


def test_choose_row_profile_weights_zeroes_empty_fields():
    from feature_builder import choose_row_profile_weights
    base = {"w_desc": 2, "w_client": 3, "w_operator": 1,
            "w_summary": 2, "w_answer_short": 1, "w_answer_full": 1}
    result = choose_row_profile_weights(
        base=base, auto_profile="off",
        has_desc=False, has_dialog=True, roles_found=True,
        has_summary=False, has_ans_s=False, has_ans_f=False,
    )
    assert result["w_desc"] == 0
    assert result["w_summary"] == 0
    assert result["w_answer_short"] == 0
    assert result["w_answer_full"] == 0
    assert result["w_client"] == 3


def test_choose_row_profile_weights_smart():
    from feature_builder import choose_row_profile_weights
    from config import PRESET_WEIGHTS
    base = {"w_desc": 1, "w_client": 1, "w_operator": 0,
            "w_summary": 1, "w_answer_short": 0, "w_answer_full": 0}
    result = choose_row_profile_weights(
        base=base, auto_profile="smart",
        has_desc=True, has_dialog=True, roles_found=True,
        has_summary=True, has_ans_s=False, has_ans_f=False,
    )
    # smart with roles+summary+no answers → "no_answers" preset; result = max(base, preset) for each key
    # preset no_answers: w_desc=2, w_client=3, w_operator=2, w_summary=4
    preset = PRESET_WEIGHTS["no_answers"]
    assert result["w_desc"] == max(base["w_desc"], preset["w_desc"])       # max(1,2)=2
    assert result["w_client"] == max(base["w_client"], preset["w_client"])  # max(1,3)=3
    assert result["w_summary"] == max(base["w_summary"], preset["w_summary"])  # max(1,4)=4; has_summary=True
    assert result["w_answer_short"] == 0  # has_ans_s=False → zeroed regardless of preset
    assert result["w_answer_full"] == 0   # has_ans_f=False → zeroed


def test_choose_row_profile_weights_strict():
    from feature_builder import choose_row_profile_weights
    base = {"w_desc": 1, "w_client": 1, "w_operator": 0,
            "w_summary": 1, "w_answer_short": 0, "w_answer_full": 0}
    result = choose_row_profile_weights(
        base=base, auto_profile="strict",
        has_desc=True, has_dialog=False, roles_found=False,
        has_summary=True, has_ans_s=True, has_ans_f=False,
    )
    # strict with has_summary only → "summary" preset
    from config import PRESET_WEIGHTS
    preset = PRESET_WEIGHTS["summary"]
    assert result["w_desc"] == preset["w_desc"]
    assert result["w_answer_full"] == 0  # has_ans_f=False


# ===========================================================================
# excel_utils tests
# ===========================================================================

def _make_csv(content: str) -> Path:
    """Create a temp CSV file with given content and return its path."""
    f = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8", newline=""
    )
    f.write(content)
    f.close()
    return Path(f.name)


def test_open_tabular_csv_basic():
    from excel_utils import open_tabular
    p = _make_csv("col1,col2,col3\nval1,val2,val3\nfoo,bar,baz\n")
    try:
        with open_tabular(p) as rows:
            header = next(rows)
            data = list(rows)
        assert len(header) == 3
        assert header[0] == "col1"
        assert len(data) == 2
        assert data[0][0] == "val1"
    finally:
        p.unlink()


def test_open_tabular_csv_empty_fields():
    from excel_utils import open_tabular
    p = _make_csv("a,b,c\n1,,3\n")
    try:
        with open_tabular(p) as rows:
            next(rows)  # header
            row = next(rows)
        # Empty field should be None
        assert row[1] is None
        assert row[0] == "1"
        assert row[2] == "3"
    finally:
        p.unlink()


def test_open_tabular_csv_semicolon_delimiter():
    from excel_utils import open_tabular
    p = _make_csv("a;b;c\n1;2;3\n")
    try:
        with open_tabular(p) as rows:
            header = next(rows)
            data_row = next(rows)
        assert len(header) == 3
        assert header[0] == "a"
        assert data_row[1] == "2"
    finally:
        p.unlink()


def test_count_csv_rows_fast():
    from excel_utils import _count_csv_rows_fast
    # 3 data rows + 1 header = 4 lines
    p = _make_csv("h1,h2\nv1,v2\nv3,v4\nv5,v6\n")
    try:
        count = _count_csv_rows_fast(p)
        assert count == 3, f"Expected 3, got {count}"
    finally:
        p.unlink()


def test_count_csv_rows_fast_no_trailing_newline():
    from excel_utils import _count_csv_rows_fast
    # File without trailing newline
    f = tempfile.NamedTemporaryFile(
        mode="wb", suffix=".csv", delete=False
    )
    f.write(b"h1,h2\nv1,v2\nv3,v4")  # no trailing \n
    f.close()
    p = Path(f.name)
    try:
        count = _count_csv_rows_fast(p)
        assert count == 2, f"Expected 2, got {count}"
    finally:
        p.unlink()


def test_read_headers():
    from excel_utils import read_headers
    p = _make_csv("Name,Age,City\nAlice,30,Moscow\n")
    try:
        headers = read_headers(p)
        assert headers == ["Name", "Age", "City"]
    finally:
        p.unlink()


def test_idx_of_found():
    from excel_utils import idx_of
    headers = ["col_a", "col_b", "col_c"]
    assert idx_of(headers, "col_b") == 1


def test_idx_of_not_found():
    from excel_utils import idx_of
    headers = ["col_a", "col_b"]
    assert idx_of(headers, "nonexistent") is None


def test_idx_of_empty_name():
    from excel_utils import idx_of
    headers = ["col_a", "col_b"]
    assert idx_of(headers, "") is None


def test_estimate_total_rows():
    from excel_utils import estimate_total_rows
    p = _make_csv("h1,h2\nv1,v2\nv3,v4\n")
    try:
        total = estimate_total_rows([p])
        assert total == 2
    finally:
        p.unlink()


def test_fmt_eta_zero_done():
    from excel_utils import fmt_eta
    import time
    # done=0 should return empty string
    result = fmt_eta(time.time(), 0, 100)
    assert result == ""


def test_fmt_eta_non_zero():
    from excel_utils import fmt_eta
    import time
    start = time.time() - 10  # 10 seconds ago
    result = fmt_eta(start, 50, 100)
    assert len(result) > 0
    assert "ETA" in result


def test_fmt_speed():
    from excel_utils import fmt_speed
    import time
    start = time.time() - 10
    result = fmt_speed(start, 100)
    assert "rows/s" in result


# ===========================================================================
# ml_core tests — PerFieldVectorizer
# ===========================================================================

def _make_feature_texts(n=20):
    """Generate synthetic feature texts with proper section tags."""
    texts = []
    for i in range(n):
        label = "ClassA" if i % 2 == 0 else "ClassB"
        texts.append(
            f"[CHANNEL]\ncall\n"
            f"[DESC]\nОписание обращения номер {i} про {label.lower()}\n"
            f"[CLIENT]\nКлиент сообщил что у него проблема с продуктом {i}\n"
            f"[OPERATOR]\nОператор помог решить вопрос клиента номер {i}\n"
        )
    return texts


def test_per_field_vectorizer_fit_transform():
    from ml_core import PerFieldVectorizer
    texts = _make_feature_texts(20)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vect = PerFieldVectorizer(base_weights=weights, min_df=1, max_features=1000)
    X = vect.fit_transform(texts)
    assert X.shape[0] == 20
    assert X.shape[1] > 0


def test_per_field_vectorizer_fit_then_transform():
    from ml_core import PerFieldVectorizer
    texts = _make_feature_texts(20)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vect = PerFieldVectorizer(base_weights=weights, min_df=1, max_features=1000)
    vect.fit(texts)
    X = vect.transform(texts)
    assert X.shape[0] == 20
    assert X.shape[1] > 0


def test_per_field_vectorizer_empty_texts():
    """Edge case: empty texts should not crash."""
    from ml_core import PerFieldVectorizer
    texts = ["", "", ""]
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vect = PerFieldVectorizer(base_weights=weights, min_df=1, max_features=1000)
    X = vect.fit_transform(texts)
    assert X.shape[0] == 3


def test_per_field_vectorizer_single_row():
    """Edge case: single row should work."""
    from ml_core import PerFieldVectorizer
    texts = ["[CHANNEL]\ncall\n[DESC]\nОдин документ для теста"]
    weights = {"w_desc": 2, "w_client": 1, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vect = PerFieldVectorizer(base_weights=weights, min_df=1, max_features=1000)
    X = vect.fit_transform(texts)
    assert X.shape[0] == 1


def test_per_field_vectorizer_all_stopwords():
    """Edge case: texts consisting entirely of stop-words should not crash."""
    from ml_core import PerFieldVectorizer
    from constants import RUSSIAN_STOP_WORDS
    sw_text = " ".join(list(RUSSIAN_STOP_WORDS)[:10])
    texts = [
        f"[CHANNEL]\ncall\n[DESC]\n{sw_text}\n[CLIENT]\n{sw_text}\n",
        f"[CHANNEL]\nchat\n[DESC]\n{sw_text}\n[CLIENT]\n{sw_text}\n",
    ]
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    stop_words = list(RUSSIAN_STOP_WORDS)
    vect = PerFieldVectorizer(
        base_weights=weights, min_df=1, max_features=1000,
        stop_words=stop_words
    )
    X = vect.fit_transform(texts)
    # Should not crash; might produce zero columns (handled with fallback)
    assert X.shape[0] == 2


def test_per_field_vectorizer_sparse_0_columns_filtered():
    """Bug fix test: sparse matrices with 0 columns are filtered before hstack."""
    import scipy.sparse as sp
    from ml_core import PerFieldVectorizer
    # All texts have only CLIENT section, DESC is empty → DESC vectorizer may have 0 cols
    texts = [
        "[CHANNEL]\ncall\n[CLIENT]\nклиент написал о проблеме с платежом",
        "[CHANNEL]\ncall\n[CLIENT]\nвопрос про кредит и условия займа",
        "[CHANNEL]\nchat\n[CLIENT]\nпроблема со входом в личный кабинет",
    ]
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vect = PerFieldVectorizer(base_weights=weights, min_df=1, max_features=500)
    # This should not raise an error even when some fields are completely empty
    X = vect.fit_transform(texts)
    assert X.shape[0] == 3
    assert X.shape[1] > 0


def test_per_field_vectorizer_parse_fields():
    """Unit test for _parse_fields internal method."""
    from ml_core import PerFieldVectorizer
    vect = PerFieldVectorizer(base_weights={"w_desc": 1})
    text = (
        "[CHANNEL]\ncall\n"
        "[DESC]\nОписание проблемы\n"
        "[CLIENT]\nТекст клиента\n"
        "[OPERATOR]\nТекст оператора\n"
    )
    fields = vect._parse_fields(text)
    assert "DESC" in fields
    assert "CLIENT" in fields
    assert "OPERATOR" in fields
    assert "описание" in fields["DESC"].lower() or "описан" in fields["DESC"].lower()


def test_per_field_vectorizer_repeated_tags_deduplicated():
    """_parse_fields should take only FIRST occurrence of repeated tags (weight repetition)."""
    from ml_core import PerFieldVectorizer
    vect = PerFieldVectorizer(base_weights={"w_desc": 2})
    text = (
        "[DESC]\nпервое описание\n"
        "[DESC]\nвторое описание (повтор для веса)\n"
    )
    fields = vect._parse_fields(text)
    # Only first occurrence should be recorded
    assert "DESC" in fields
    assert "первое" in fields["DESC"]
    assert "второе" not in fields["DESC"]


# ===========================================================================
# ml_core tests — MetaFeatureExtractor
# ===========================================================================

def test_meta_feature_extractor_basic():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    texts = [
        "[CHANNEL]\ncall\n[DESC]\nОписание\n[CLIENT]\nКлиент написал\n",
        "[CHANNEL]\nchat\n[SUMMARY]\nСуммаризация диалога\n",
    ]
    mfe.fit(texts)
    X = mfe.transform(texts)
    assert X.shape == (2, 16)
    # First sample: is_call=1, is_chat=0
    assert X[0, 6] == 1.0  # is_call
    assert X[0, 7] == 0.0  # is_chat
    # Second sample: is_chat=1
    assert X[1, 7] == 1.0  # is_chat


def test_meta_feature_extractor_has_flags():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    text = "[CHANNEL]\ncall\n[DESC]\nТекст\n[CLIENT]\nКлиент\n[OPERATOR]\nОператор\n"
    X = mfe.transform([text])
    assert X[0, 0] == 1.0  # has_desc
    assert X[0, 1] == 1.0  # has_client
    assert X[0, 2] == 1.0  # has_operator
    assert X[0, 3] == 0.0  # has_summary
    assert X[0, 4] == 0.0  # has_ans_short
    assert X[0, 5] == 0.0  # has_ans_full


def test_meta_feature_extractor_lengths():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    # DESC with 3 words
    text = "[CHANNEL]\ncall\n[DESC]\nодно два три\n"
    X = mfe.transform([text])
    # desc_len = log1p(3)
    assert abs(X[0, 8] - math.log1p(3)) < 1e-4


def test_meta_feature_extractor_empty_text():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    X = mfe.transform([""])
    assert X.shape == (1, 16)
    # All zeros (no sections)
    assert X[0, 0] == 0.0  # has_desc
    assert X[0, 6] == 0.0  # is_call
    assert X[0, 14] == 0.0  # total_len


def test_meta_feature_extractor_client_share():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    # CLIENT has 4 words, OPERATOR has 4 words → share = 0.5
    text = (
        "[CHANNEL]\ncall\n"
        "[CLIENT]\nодно два три четыре\n"
        "[OPERATOR]\nпять шесть семь восемь\n"
    )
    X = mfe.transform([text])
    client_share = X[0, 13]
    assert abs(client_share - 0.5) < 1e-4


def test_meta_feature_extractor_no_dialog():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    # No CLIENT or OPERATOR → client_share defaults to 0.5
    text = "[CHANNEL]\ncall\n[DESC]\nТолько описание без диалога\n"
    X = mfe.transform([text])
    assert X[0, 13] == 0.5  # client_share default


def test_meta_feature_extractor_various_dialogs():
    from ml_core import MetaFeatureExtractor
    mfe = MetaFeatureExtractor()
    texts = [
        # call with lots of client speech
        "[CHANNEL]\ncall\n[CLIENT]\nмного слов от клиента здесь и там\n[OPERATOR]\nкоротко\n",
        # chat only
        "[CHANNEL]\nchat\n[SUMMARY]\nсуммаризация длинная\n[ANSWER_SHORT]\nответ\n",
        # empty
        "",
    ]
    mfe.fit(texts)
    X = mfe.transform(texts)
    assert X.shape == (3, 16)
    # sample 0: has_client=1, has_operator=1
    assert X[0, 1] == 1.0
    assert X[0, 2] == 1.0
    # sample 1: has_summary=1, has_ans_short=1
    assert X[1, 3] == 1.0
    assert X[1, 4] == 1.0
    # sample 2: all zeros
    assert X[2, 0] == 0.0


# ===========================================================================
# ml_core tests — make_hybrid_vectorizer
# ===========================================================================

def test_make_hybrid_vectorizer_per_field():
    from ml_core import make_hybrid_vectorizer
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vec = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=1000,
    )
    texts = _make_feature_texts(10)
    X = vec.fit_transform(texts)
    assert X.shape[0] == 10
    assert X.shape[1] > 0


def test_make_hybrid_vectorizer_legacy():
    from ml_core import make_hybrid_vectorizer
    vec = make_hybrid_vectorizer(
        use_per_field=False,
        min_df=1, max_features=1000,
    )
    texts = _make_feature_texts(10)
    X = vec.fit_transform(texts)
    assert X.shape[0] == 10
    assert X.shape[1] > 0


def test_make_hybrid_vectorizer_with_meta():
    from ml_core import make_hybrid_vectorizer
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vec = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=1000, use_meta=True,
    )
    texts = _make_feature_texts(10)
    X = vec.fit_transform(texts)
    assert X.shape[0] == 10
    # Should have more columns than without meta (15 extra)
    assert X.shape[1] > 15


def test_make_hybrid_vectorizer_with_stop_words():
    from ml_core import make_hybrid_vectorizer
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vec = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=1000,
        use_stop_words=True, use_noise_tokens=True, use_noise_phrases=True,
    )
    texts = _make_feature_texts(10)
    X = vec.fit_transform(texts)
    assert X.shape[0] == 10


def test_make_hybrid_vectorizer_no_base_weights():
    """use_per_field=True but no base_weights → falls back to legacy mode."""
    from ml_core import make_hybrid_vectorizer
    vec = make_hybrid_vectorizer(
        use_per_field=True, base_weights=None,
        min_df=1, max_features=1000,
    )
    texts = _make_feature_texts(10)
    X = vec.fit_transform(texts)
    assert X.shape[0] == 10


# ===========================================================================
# ml_core tests — make_classifier
# ===========================================================================

def test_make_classifier_large_dataset():
    """With enough samples per class, should return CalibratedClassifierCV."""
    from ml_core import make_classifier
    from sklearn.calibration import CalibratedClassifierCV
    y = ["A"] * 20 + ["B"] * 20
    clf, clf_type = make_classifier(y, C=1.0, max_iter=100, balanced=True)
    assert isinstance(clf, CalibratedClassifierCV)
    assert "LinearSVC" in clf_type


def test_make_classifier_small_dataset():
    """With too few samples per class, should fall back to LogisticRegression."""
    from ml_core import make_classifier
    from sklearn.linear_model import LogisticRegression
    y = ["A"] * 2 + ["B"] * 2
    clf, clf_type = make_classifier(y, C=1.0, max_iter=100, balanced=True)
    assert isinstance(clf, LogisticRegression)
    assert "LogReg" in clf_type


def test_make_classifier_single_class():
    """Single class → LogisticRegression fallback."""
    from ml_core import make_classifier
    from sklearn.linear_model import LogisticRegression
    y = ["A"] * 10
    clf, clf_type = make_classifier(y, C=1.0, max_iter=100, balanced=False)
    assert isinstance(clf, LogisticRegression)


def test_make_classifier_calib_method_isotonic():
    from ml_core import make_classifier
    from sklearn.calibration import CalibratedClassifierCV
    y = ["A"] * 10 + ["B"] * 10
    clf, clf_type = make_classifier(
        y, C=1.0, max_iter=100, balanced=True, calib_method="isotonic"
    )
    assert isinstance(clf, CalibratedClassifierCV)
    assert "isotonic" in clf_type


def test_make_classifier_calib_method_auto():
    from ml_core import make_classifier
    y = ["A"] * 10 + ["B"] * 10
    clf, clf_type = make_classifier(
        y, C=1.0, max_iter=100, balanced=True, calib_method="auto"
    )
    # auto chooses based on avg_per_class — with 10 samples/class → sigmoid
    assert clf is not None


def test_make_classifier_dynamic_cv_small():
    """min_class=3 → _CV=3 → LinearSVC+Calibrated(cv=3) вместо LogReg."""
    from ml_core import make_classifier
    from sklearn.calibration import CalibratedClassifierCV
    # 3 примера на класс: раньше уходило в LogReg, теперь LinearSVC с CV=3
    y = ["A"] * 3 + ["B"] * 3
    clf, clf_type = make_classifier(y, C=1.0, max_iter=100, balanced=True)
    assert isinstance(clf, CalibratedClassifierCV)
    assert "cv=3" in clf_type


def test_make_classifier_dynamic_cv_medium():
    """min_class=6 → _CV=3 (6//2=3)."""
    from ml_core import make_classifier
    from sklearn.calibration import CalibratedClassifierCV
    y = ["A"] * 6 + ["B"] * 6
    clf, clf_type = make_classifier(y, C=1.0, max_iter=100, balanced=True)
    assert isinstance(clf, CalibratedClassifierCV)
    assert "cv=3" in clf_type


def test_make_classifier_dynamic_cv_large():
    """min_class=20 → _CV=5 (стандарт)."""
    from ml_core import make_classifier
    from sklearn.calibration import CalibratedClassifierCV
    y = ["A"] * 20 + ["B"] * 20
    clf, clf_type = make_classifier(y, C=1.0, max_iter=100, balanced=True)
    assert isinstance(clf, CalibratedClassifierCV)
    assert "cv=5" in clf_type


def test_make_hybrid_vectorizer_meta_has_scaler():
    """use_meta=True → meta-ветка содержит StandardScaler."""
    from ml_core import make_hybrid_vectorizer
    from sklearn.pipeline import Pipeline, FeatureUnion
    from sklearn.preprocessing import StandardScaler
    weights = {"w_desc": 2, "w_client": 2, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vec = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=500, use_meta=True,
    )
    assert isinstance(vec, FeatureUnion)
    # Находим meta-ветку
    meta_step = dict(vec.transformer_list).get("meta")
    assert meta_step is not None
    assert isinstance(meta_step, Pipeline)
    # Последний шаг должен быть StandardScaler
    _, scaler = meta_step.steps[-1]
    assert isinstance(scaler, StandardScaler)


def test_upgrade_config_dict_lemma_default_true():
    """Новый дефолт use_lemma=True попадает в upgrade_config_dict."""
    from constants import upgrade_config_dict
    result = upgrade_config_dict({})
    assert result["use_lemma"] is True


def test_ctfidf_bigrams_in_keywords():
    """c-TF-IDF с ngram_range=(1,3) может выдавать многословные ключевые слова."""
    from ml_core import extract_cluster_keywords_ctfidf
    import numpy as np

    docs = [
        "досрочное погашение кредита банк",
        "досрочное погашение займа досрочно",
        "технический сбой мобильного приложения",
        "ошибка авторизации мобильного приложения",
    ] * 5  # увеличиваем выборку чтобы биграммы набрали min_df
    labels = np.array([0, 0, 1, 1] * 5)

    kws = extract_cluster_keywords_ctfidf(docs, labels, n_clusters=2, top_n=10)
    assert len(kws) == 2
    all_kws = " ".join(kws)
    # При ngram_range=(1,3) должны появляться многословные фразы
    has_bigram = any(" " in kw for kw in all_kws.split(", "))
    assert has_bigram, f"Биграммы не найдены в ключевых словах: {kws}"


# ===========================================================================
# ml_core tests — train_model (build_pipeline)
# ===========================================================================

def _make_synthetic_dataset(n_per_class=20):
    """Create simple synthetic texts for training."""
    texts = []
    labels = []
    for i in range(n_per_class):
        texts.append(
            f"[CHANNEL]\ncall\n"
            f"[DESC]\nпроблема с картой и снятием наличных {i}\n"
            f"[CLIENT]\nклиент жалуется на блокировку счёта {i}\n"
        )
        labels.append("Карта")
    for i in range(n_per_class):
        texts.append(
            f"[CHANNEL]\ncall\n"
            f"[DESC]\nвопрос по кредиту и ставке процентов {i}\n"
            f"[CLIENT]\nклиент интересуется условиями кредита {i}\n"
        )
        labels.append("Кредит")
    return texts, labels


def test_train_model_basic():
    from ml_core import make_hybrid_vectorizer, train_model
    texts, labels = _make_synthetic_dataset(n_per_class=15)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=500,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0.2,
        random_state=42,
    )
    assert pipe is not None
    assert clf_type != ""
    assert report != ""


def test_train_model_empty_validation_set_no_crash():
    """Bug fix: empty validation set should not crash on np.percentile."""
    from ml_core import make_hybrid_vectorizer, train_model
    # Use test_size=0 to force no validation
    texts, labels = _make_synthetic_dataset(n_per_class=5)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=500,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0,   # force no validation
        random_state=42,
    )
    assert pipe is not None
    # With test_size=0, validation is skipped
    assert labels_out is None
    assert cm is None
    assert "ВАЛИДАЦИЯ ПРОПУЩЕНА" in report


def test_train_model_too_few_samples_no_crash():
    """train_model with < 30 total samples falls back gracefully."""
    from ml_core import make_hybrid_vectorizer, train_model
    texts = [
        "[CHANNEL]\ncall\n[DESC]\nпроблема с картой\n",
        "[CHANNEL]\ncall\n[DESC]\nвопрос по кредиту\n",
        "[CHANNEL]\ncall\n[DESC]\nблокировка счёта\n",
    ]
    labels = ["Карта", "Кредит", "Карта"]
    weights = {"w_desc": 2, "w_client": 0, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=200,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=100,
        balanced=False,
        test_size=0.2,
        random_state=42,
    )
    assert pipe is not None
    # Should skip validation due to too few samples
    assert labels_out is None


def test_train_model_with_legacy_vectorizer():
    from ml_core import make_hybrid_vectorizer, train_model
    texts, labels = _make_synthetic_dataset(n_per_class=15)
    features = make_hybrid_vectorizer(
        use_per_field=False,
        min_df=1, max_features=500,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0.2,
        random_state=42,
    )
    assert pipe is not None
    assert report != ""


def test_train_model_predict_after_training():
    """Pipeline trained by train_model should be able to predict."""
    from ml_core import make_hybrid_vectorizer, train_model
    texts, labels = _make_synthetic_dataset(n_per_class=20)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=500,
    )
    pipe, _, _, _, _, _ = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0.2,
        random_state=42,
    )
    new_texts = [
        "[CHANNEL]\ncall\n[DESC]\nпроблема с картой снятие\n[CLIENT]\nКлиент жалуется\n",
        "[CHANNEL]\ncall\n[DESC]\nвопрос кредит процент\n[CLIENT]\nинтересуется кредитом\n",
    ]
    preds = pipe.predict(new_texts)
    assert len(preds) == 2
    assert all(p in ["Карта", "Кредит"] for p in preds)


def test_train_model_predict_proba():
    """train_model with enough data should produce a model with predict_proba."""
    from ml_core import make_hybrid_vectorizer, train_model
    texts, labels = _make_synthetic_dataset(n_per_class=20)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=500,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0.2,
        random_state=42,
    )
    if hasattr(pipe, "predict_proba"):
        new_texts = ["[CHANNEL]\ncall\n[DESC]\nпроблема с картой\n"]
        proba = pipe.predict_proba(new_texts)
        assert proba.shape[0] == 1
        assert abs(proba[0].sum() - 1.0) < 1e-5


def test_train_model_extras_thresholds():
    """Extras dict should contain thresh_90, thresh_75 when validation runs."""
    from ml_core import make_hybrid_vectorizer, train_model
    texts, labels = _make_synthetic_dataset(n_per_class=20)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 0,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=weights,
        min_df=1, max_features=500,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=texts, y=labels,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0.2,
        random_state=42,
    )
    if labels_out is not None and hasattr(pipe, "predict_proba"):
        assert "thresh_90" in extras
        assert "thresh_75" in extras
        assert 0.0 <= extras["thresh_90"] <= 1.0
        assert 0.0 <= extras["thresh_75"] <= 1.0


# ===========================================================================
# ml_core tests — PhraseRemover
# ===========================================================================

def test_phrase_remover_basic():
    from ml_core import PhraseRemover
    pr = PhraseRemover(["мы рассмотрели ваше обращение", "спасибо за ожидание"])
    result = pr.transform(["Мы рассмотрели ваше обращение и дадим ответ"])
    assert "рассмотрели" not in result[0].lower()


def test_phrase_remover_empty_phrases():
    from ml_core import PhraseRemover
    pr = PhraseRemover([])
    result = pr.transform(["текст без изменений"])
    assert result[0] == "текст без изменений"


def test_phrase_remover_none_input():
    from ml_core import PhraseRemover
    pr = PhraseRemover(["шаблон"])
    result = pr.transform([None])
    assert result[0] == ""


def test_phrase_remover_fit_transform():
    from ml_core import PhraseRemover
    pr = PhraseRemover(["чем могу помочь"])
    result = pr.fit_transform(["Здравствуйте чем могу помочь сегодня"])
    assert "чем могу помочь" not in result[0].lower()


# ===========================================================================
# ml_core tests — dataset_health_checks
# ===========================================================================

def test_dataset_health_checks_ok():
    from ml_core import dataset_health_checks
    y = ["A"] * 50 + ["B"] * 50
    stats = {"rows_used": 100}
    fatal, warn = dataset_health_checks(stats, y)
    assert len(fatal) == 0


def test_dataset_health_checks_no_rows():
    from ml_core import dataset_health_checks
    fatal, warn = dataset_health_checks({"rows_used": 0}, [])
    assert any("строк" in f.lower() or "фильтрац" in f.lower() for f in fatal)


def test_dataset_health_checks_single_class():
    from ml_core import dataset_health_checks
    y = ["A"] * 10
    fatal, warn = dataset_health_checks({"rows_used": 10}, y)
    assert any("класс" in f.lower() or "причин" in f.lower() for f in fatal)


# ===========================================================================
# ml_core tests — extract_cluster_keywords
# ===========================================================================

def test_extract_cluster_keywords_per_field():
    from ml_core import PerFieldVectorizer, extract_cluster_keywords
    import numpy as np
    texts = _make_feature_texts(20)
    weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
               "w_summary": 0, "w_answer_short": 0, "w_answer_full": 0}
    vect = PerFieldVectorizer(base_weights=weights, min_df=1, max_features=500)
    X = vect.fit_transform(texts)
    # Fake 2 cluster centroids
    centers = np.array([
        np.asarray(X[:10].mean(axis=0)).ravel(),
        np.asarray(X[10:].mean(axis=0)).ravel(),
    ])
    kws = extract_cluster_keywords(vect, centers, top_n=5)
    assert len(kws) == 2
    assert isinstance(kws[0], str)


def test_extract_cluster_keywords_tfidf():
    """extract_cluster_keywords работает с обычным TfidfVectorizer."""
    from ml_core import extract_cluster_keywords
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = [
        "клиент позвонил с вопросом об оплате",
        "оператор помог решить проблему",
        "заявка принята в работу",
        "технический сбой устранён",
        "клиент доволен результатом",
    ]
    vec = TfidfVectorizer(min_df=1)
    X = vec.fit_transform(texts)
    centers = np.array([
        np.asarray(X[:3].mean(axis=0)).ravel(),
        np.asarray(X[3:].mean(axis=0)).ravel(),
    ])
    kws = extract_cluster_keywords(vec, centers, top_n=5)
    assert len(kws) == 2
    assert all(isinstance(k, str) for k in kws)
    # Хотя бы один кластер должен содержать слова
    assert any(k.strip() for k in kws)


def test_extract_cluster_keywords_dim_mismatch():
    """При несовпадении размерности центроидов и словаря возвращаются пустые строки."""
    from ml_core import extract_cluster_keywords
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = ["кошка сидела на крыше", "собака бежала по полю", "птица летела над лесом"]
    vec = TfidfVectorizer(min_df=1)
    vec.fit_transform(texts)
    # Центроиды другой размерности (не совпадает с vocab)
    n_feat = len(vec.get_feature_names_out())
    centers = np.ones((2, max(1, n_feat - 3)))  # намеренно меньший размер
    kws = extract_cluster_keywords(vec, centers, top_n=5)
    assert len(kws) == 2
    assert all(k == "" for k in kws)


def test_extract_cluster_keywords_no_positive_values():
    """Центроиды из нулей/отрицательных значений → пустые ключевые слова."""
    from ml_core import extract_cluster_keywords
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = ["тест один два", "тест три четыре"]
    vec = TfidfVectorizer(min_df=1)
    vec.fit_transform(texts)
    n_feat = len(vec.get_feature_names_out())
    centers = np.zeros((2, n_feat))  # нули — нет положительных компонент
    kws = extract_cluster_keywords(vec, centers, top_n=5)
    assert len(kws) == 2
    assert all(k == "" for k in kws)


def test_extract_cluster_keywords_from_labels_basic():
    """extract_cluster_keywords_from_labels возвращает ключевые слова для каждого кластера."""
    from ml_core import extract_cluster_keywords_from_labels
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = [
        "оплата кредита задержка",
        "закрыть счёт банк",
        "техническая проблема приложение",
        "ошибка входа пароль",
        "кредит досрочное погашение",
    ]
    labels = np.array([0, 0, 1, 1, 0])
    vec = TfidfVectorizer(min_df=1)
    Xv = vec.fit_transform(texts)

    kws = extract_cluster_keywords_from_labels(vec, Xv, labels, n_clusters=2, top_n=5)
    assert len(kws) == 2
    assert all(isinstance(k, str) for k in kws)
    assert any(k.strip() for k in kws)


def test_extract_cluster_keywords_from_labels_empty_cluster():
    """Кластер без строк получает пустую строку ключевых слов."""
    from ml_core import extract_cluster_keywords_from_labels
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = ["текст один два три", "текст четыре пять шесть"]
    labels = np.array([0, 0])  # кластер 1 пустой
    vec = TfidfVectorizer(min_df=1)
    Xv = vec.fit_transform(texts)

    kws = extract_cluster_keywords_from_labels(vec, Xv, labels, n_clusters=2, top_n=3)
    assert len(kws) == 2
    assert kws[1] == ""  # кластер 1 пустой


def test_extract_cluster_keywords_from_labels_dense():
    """extract_cluster_keywords_from_labels работает с плотной матрицей (np.ndarray)."""
    from ml_core import extract_cluster_keywords_from_labels
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = ["банк перевод карта", "кредит долг погашение", "перевод счёт банк"]
    labels = np.array([0, 1, 0])
    vec = TfidfVectorizer(min_df=1)
    Xv = vec.fit_transform(texts).toarray()  # dense

    kws = extract_cluster_keywords_from_labels(vec, Xv, labels, n_clusters=2, top_n=3)
    assert len(kws) == 2
    assert all(isinstance(k, str) for k in kws)


def test_extract_cluster_keywords_ctfidf_basic():
    """c-TF-IDF возвращает специфичные для кластеров ключевые слова."""
    from ml_core import extract_cluster_keywords_ctfidf
    import numpy as np

    docs = [
        "оплата счёт кредит банк",
        "кредит погашение долг",
        "техническая ошибка приложение",
        "сбой вход пароль система",
        "банк перевод карта оплата",
    ]
    labels = np.array([0, 0, 1, 1, 0])

    kws = extract_cluster_keywords_ctfidf(docs, labels, n_clusters=2, top_n=5)
    assert len(kws) == 2
    assert all(isinstance(k, str) for k in kws)
    # Кластер 0 должен содержать финансовые термины
    assert any(k.strip() for k in kws)
    # Ключевые слова двух кластеров должны различаться
    assert kws[0] != kws[1]


def test_extract_cluster_keywords_ctfidf_empty_cluster():
    """c-TF-IDF: кластер без документов получает пустую строку."""
    from ml_core import extract_cluster_keywords_ctfidf
    import numpy as np

    docs = ["один два три", "четыре пять шесть"]
    labels = np.array([0, 0])  # кластер 1 пустой

    kws = extract_cluster_keywords_ctfidf(docs, labels, n_clusters=2, top_n=3)
    assert len(kws) == 2
    assert kws[1] == ""


def test_extract_cluster_keywords_ctfidf_with_stop_words():
    """c-TF-IDF с list-стоп-словами не падает и фильтрует стоп-слова."""
    from ml_core import extract_cluster_keywords_ctfidf
    import numpy as np

    docs = [
        "и клиент обратился в банк",
        "и оператор принял звонок",
        "технический сбой в системе",
        "ошибка при входе в систему",
    ]
    labels = np.array([0, 0, 1, 1])
    stop_words = ["и", "в", "при"]

    kws = extract_cluster_keywords_ctfidf(docs, labels, n_clusters=2,
                                          stop_words=stop_words, top_n=5)
    assert len(kws) == 2
    # Стоп-слова не должны попасть в ключевые слова
    for kw_str in kws:
        for word in kw_str.split(", "):
            assert word not in stop_words, f"Стоп-слово '{word}' попало в ключевые слова"


def test_elbow_k_fallback_logic():
    """Fallback-алгоритм второй разности (метод локтя) выбирает правильный K."""
    # Реплицируем fallback логику из app_cluster.pick_elbow_k
    def pick_elbow_k_fallback(inertias, ks):
        if len(ks) < 3:
            return ks[0]
        y = inertias[:]
        y0, y1 = min(y), max(y)
        yn = [(v - y0) / (y1 - y0 + 1e-9) for v in y]
        best_i = 1
        best_val = -1.0
        for i in range(1, len(yn) - 1):
            val = abs(yn[i + 1] - 2 * yn[i] + yn[i - 1])
            if val > best_val:
                best_val = val
                best_i = i
        return ks[best_i]

    # Инерция резко падает от K=2 до K=5, потом выравнивается — локоть должен быть в K=5
    ks = [2, 3, 4, 5, 6, 7, 8]
    inertias = [1000.0, 700.0, 500.0, 200.0, 185.0, 175.0, 170.0]
    k = pick_elbow_k_fallback(inertias, ks)
    assert k in ks, f"Returned K={k} not in {ks}"

    # Монотонно убывающая — всё равно не падает
    ks2 = [2, 3, 4]
    inertias2 = [300.0, 200.0, 100.0]
    k2 = pick_elbow_k_fallback(inertias2, ks2)
    assert k2 in ks2

    # Меньше 3 элементов — возвращает первый K
    assert pick_elbow_k_fallback([500.0, 300.0], [2, 3]) == 2
    assert pick_elbow_k_fallback([500.0], [5]) == 5


def test_kmeans_cluster_pipeline():
    """MiniBatchKMeans + extract_cluster_keywords — полный mini-pipeline кластеризации."""
    from ml_core import (extract_cluster_keywords, extract_cluster_keywords_from_labels,
                          extract_cluster_keywords_ctfidf)
    from sklearn.cluster import MiniBatchKMeans
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np

    texts = [
        "кредит оплата долг банк",
        "кредит погашение платёж",
        "техническая ошибка приложение",
        "сбой система вход пароль",
        "банк перевод карта счёт",
        "ошибка авторизация мобильный",
    ] * 3  # 18 строк, 2 кластера

    labels_true = np.array([0, 0, 1, 1, 0, 1] * 3)

    vec = TfidfVectorizer(min_df=1)
    Xv = vec.fit_transform(texts)

    km = MiniBatchKMeans(n_clusters=2, random_state=42, n_init=3)
    km.fit(Xv)

    # Ключевые слова по центроидам KMeans
    kws_centers = extract_cluster_keywords(vec, km.cluster_centers_, top_n=5)
    assert len(kws_centers) == 2
    assert all(isinstance(k, str) for k in kws_centers)

    # Ключевые слова по меткам (для SBERT-режима)
    kws_labels = extract_cluster_keywords_from_labels(
        vec, Xv, km.labels_, n_clusters=2, top_n=5
    )
    assert len(kws_labels) == 2

    # c-TF-IDF
    kws_ctfidf = extract_cluster_keywords_ctfidf(texts, km.labels_, n_clusters=2, top_n=5)
    assert len(kws_ctfidf) == 2


# ===========================================================================
# Clustering pipeline: full_cluster mapping & output correctness
# ===========================================================================

def test_cluster_full_cluster_single_file():
    """full_cluster корректно строится для одного файла с пустыми и непустыми строками."""
    # Симулируем: 5 строк в файле, строки 0,2,3 - с текстом; 1,4 - пустые
    row_has_text = [True, False, True, True, False]
    # labels для 3 ok-строк:
    labels = [0, 1, 0]
    label_offset = 0
    full_cluster = []
    label_cursor = label_offset
    for ok in row_has_text:
        if ok:
            full_cluster.append(int(labels[label_cursor]))
            label_cursor += 1
        else:
            full_cluster.append(-1)
    assert full_cluster == [0, -1, 1, 0, -1], f"Unexpected full_cluster: {full_cluster}"
    assert label_cursor == 3


def test_cluster_full_cluster_multi_file():
    """full_cluster корректно переносит label_offset между файлами."""
    # Файл 1: 3 строки, 2 ok
    file1_has_text = [True, False, True]
    # Файл 2: 2 строки, 2 ok
    file2_has_text = [True, True]
    # Общий labels для 4 ok-строк
    labels = [0, 1, 2, 0]

    label_offset = 0
    results = []
    for file_has_text in [file1_has_text, file2_has_text]:
        full_cluster = []
        label_cursor = label_offset
        for ok in file_has_text:
            if ok:
                full_cluster.append(int(labels[label_cursor]))
                label_cursor += 1
            else:
                full_cluster.append(-1)
        label_offset = label_cursor
        results.append(full_cluster)

    assert results[0] == [0, -1, 1], f"File1 full_cluster: {results[0]}"
    assert results[1] == [2, 0], f"File2 full_cluster: {results[1]}"
    assert label_offset == 4


def test_cluster_kw_dict_from_list():
    """kw_dict корректно строится из List[str] и поддерживает поиск по cluster_id."""
    from ml_core import extract_cluster_keywords_ctfidf

    docs = [
        "снятие наличных карта банкомат",
        "снять деньги карта",
        "перевод счёт получатель",
        "перевести сумма счёт",
    ]
    labels = [0, 0, 1, 1]
    K = 2

    kw = extract_cluster_keywords_ctfidf(docs, labels, K, top_n=5)

    # kw - это List[str], не dict
    assert isinstance(kw, list), f"kw should be list, got {type(kw)}"
    assert len(kw) == K, f"len(kw)={len(kw)}, expected K={K}"
    assert all(isinstance(s, str) for s in kw), "All kw elements must be str"

    # Проверяем логику построения kw_dict как в app_cluster.py
    kw_dict = {i: kw[i] for i in range(len(kw))}
    assert kw_dict[0] != "", "Cluster 0 should have keywords"
    assert kw_dict[1] != "", "Cluster 1 should have keywords"

    # Проверяем логику лога (исправленный вариант): kw - List, не dict
    _kw_n = len(kw)
    _kw_avg = sum(len(v) for v in kw) // max(1, _kw_n)
    assert _kw_n == K, f"_kw_n={_kw_n}, expected {K}"
    assert _kw_avg > 0, "Expected non-zero avg keyword length"


def test_cluster_output_columns():
    """Выходная строка файла содержит правильный cluster_id и cluster_keywords."""
    from ml_core import extract_cluster_keywords_ctfidf
    import numpy as np

    docs = [
        "кредит погашение платёж банк",
        "кредит долг ставка",
        "перевод карта счёт операция",
        "счёт перевести получатель",
        "ошибка приложение вход пароль",
        "сбой система авторизация",
    ]
    labels_arr = np.array([0, 0, 1, 1, 2, 2])
    K = 3

    kw = extract_cluster_keywords_ctfidf(docs, labels_arr, K, top_n=5)
    kw_dict = {i: kw[i] for i in range(len(kw))}

    # Симулируем row_has_text: 8 строк, 2 пустые
    row_has_text = [True, True, False, True, True, False, True, True]
    full_cluster = []
    label_cursor = 0
    for ok in row_has_text:
        if ok:
            full_cluster.append(int(labels_arr[label_cursor]))
            label_cursor += 1
        else:
            full_cluster.append(-1)

    assert full_cluster == [0, 0, -1, 1, 1, -1, 2, 2]

    # Проверяем значения, которые будут записаны в файл
    for i, cid in enumerate(full_cluster):
        kws = "" if cid < 0 else kw_dict.get(cid, "")
        if cid == -1:
            assert kws == "", f"Empty row i={i} должна иметь kws=''"
        else:
            assert isinstance(kws, str), f"kws[{i}] должна быть строкой"
            assert len(kws) > 0, f"Строка i={i} (cid={cid}) должна иметь непустые ключевые слова"


def test_cluster_end_to_end_with_csv():
    """Полный pipeline: CSV → кластеризация → запись → проверка выходного файла."""
    import csv
    import tempfile
    import os
    import numpy as np
    from pathlib import Path
    from ml_core import extract_cluster_keywords_ctfidf
    from sklearn.cluster import MiniBatchKMeans
    from sklearn.feature_extraction.text import TfidfVectorizer
    from openpyxl import Workbook, load_workbook

    # Создаём временный CSV
    rows_data = [
        ("кредит погашение платёж",),
        ("кредит долг ставка банк",),
        ("",),                                 # пустая строка
        ("перевод карта счёт операция",),
        ("счёт перевести получатель",),
        ("ошибка приложение вход",),
        ("сбой система авторизация пароль",),
        ("",),                                 # пустая строка
    ]
    cid_col = "cluster_id"
    ckw_col = "cluster_keywords"
    text_col = "description"

    with tempfile.TemporaryDirectory() as tmp_dir:
        csv_path = Path(tmp_dir) / "test_input.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([text_col])
            for row in rows_data:
                w.writerow(row)

        # === Первый проход: читаем тексты ===
        row_has_text = []
        X_all = []
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for raw_row in reader:
                text = raw_row[0].strip() if raw_row else ""
                ok = bool(text)
                row_has_text.append(ok)
                if ok:
                    X_all.append(text)

        n_ok = len(X_all)
        assert n_ok == 6, f"Expected 6 non-empty rows, got {n_ok}"

        # === Кластеризация ===
        K = 3
        vec = TfidfVectorizer(min_df=1)
        Xv = vec.fit_transform(X_all)
        km = MiniBatchKMeans(n_clusters=K, random_state=42, n_init=5)
        labels = km.fit_predict(Xv)

        assert len(labels) == n_ok

        # === Извлечение ключевых слов ===
        kw = extract_cluster_keywords_ctfidf(X_all, labels, K, top_n=5)
        assert isinstance(kw, list) and len(kw) == K

        # kw_dict как в app_cluster.py
        kw_dict = {i: kw[i] for i in range(len(kw))}

        # Проверяем исправленный расчёт для лога
        _kw_n = len(kw)
        _kw_avg = sum(len(v) for v in kw) // max(1, _kw_n)
        assert _kw_n == K
        assert _kw_avg > 0

        # === Строим full_cluster ===
        full_cluster = []
        label_cursor = 0
        for ok in row_has_text:
            if ok:
                full_cluster.append(int(labels[label_cursor]))
                label_cursor += 1
            else:
                full_cluster.append(-1)

        assert len(full_cluster) == len(rows_data)
        # Пустые строки должны быть -1
        assert full_cluster[2] == -1, "Строка 3 (пустая) должна быть cluster_id=-1"
        assert full_cluster[7] == -1, "Строка 8 (пустая) должна быть cluster_id=-1"
        # Непустые строки должны иметь корректный cluster_id
        for i, (ok, cid) in enumerate(zip(row_has_text, full_cluster)):
            if ok:
                assert 0 <= cid < K, f"Row {i}: cluster_id={cid} вне диапазона [0, {K})"
            else:
                assert cid == -1, f"Row {i}: пустая строка должна иметь cluster_id=-1"

        # === Записываем выходной Excel ===
        out_path = Path(tmp_dir) / "test_output.xlsx"
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Sheet1")
        header_out = [text_col, cid_col, ckw_col]
        ws.append(header_out)

        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for i, raw_row in enumerate(reader):
                text = raw_row[0] if raw_row else ""
                cid = full_cluster[i]
                kws = "" if cid < 0 else kw_dict.get(cid, "")
                ws.append([text, cid, kws])
        wb.save(out_path)

        # === Проверяем выходной файл ===
        wb2 = load_workbook(out_path, read_only=True, data_only=True)
        ws2 = wb2.active
        out_rows = list(ws2.iter_rows(values_only=True))
        wb2.close()

        assert out_rows[0] == (text_col, cid_col, ckw_col), \
            f"Заголовок не совпадает: {out_rows[0]}"
        assert len(out_rows) == len(rows_data) + 1, \
            f"Ожидалось {len(rows_data)+1} строк, получено {len(out_rows)}"

        for i, (orig, out_row) in enumerate(zip(rows_data, out_rows[1:])):
            text_val, cid_val, kws_val = out_row
            ok = bool(orig[0].strip())
            if ok:
                assert isinstance(cid_val, int) and 0 <= cid_val < K, \
                    f"Строка {i}: cid_val={cid_val!r} некорректен"
                assert isinstance(kws_val, str) and len(kws_val) > 0, \
                    f"Строка {i}: кластер {cid_val} должен иметь непустые ключевые слова, получено: {kws_val!r}"
            else:
                assert cid_val == -1, \
                    f"Строка {i}: пустая строка должна иметь cid=-1, получено {cid_val!r}"
                # openpyxl читает пустые строки как None в data_only режиме
                assert kws_val is None or kws_val == "", \
                    f"Строка {i}: пустая строка должна иметь kws='' или None, получено {kws_val!r}"


# ===========================================================================
# ml_core tests — constants and config
# ===========================================================================

def test_russian_stop_words_non_empty():
    from constants import RUSSIAN_STOP_WORDS
    assert len(RUSSIAN_STOP_WORDS) > 0
    assert "и" in RUSSIAN_STOP_WORDS


def test_noise_tokens_non_empty():
    from constants import NOISE_TOKENS
    assert len(NOISE_TOKENS) > 0
    assert "fio" in NOISE_TOKENS


def test_noise_phrases_non_empty():
    from constants import NOISE_PHRASES
    assert len(NOISE_PHRASES) > 0


def test_preset_weights_valid():
    from config import PRESET_WEIGHTS
    required_keys = {"w_desc", "w_client", "w_operator",
                     "w_summary", "w_answer_short", "w_answer_full"}
    for preset_name, pw in PRESET_WEIGHTS.items():
        assert required_keys == set(pw.keys()), f"Preset {preset_name} missing keys"
        for k, v in pw.items():
            assert isinstance(v, int), f"Preset {preset_name}[{k}] is not int"
            assert 0 <= v <= 5, f"Preset {preset_name}[{k}] out of range"


def test_section_prefix_valid():
    from constants import SECTION_PREFIX
    expected_tags = {"CHANNEL", "DESC", "CLIENT", "OPERATOR",
                     "SUMMARY", "ANSWER_SHORT", "ANSWER_FULL"}
    assert expected_tags == set(SECTION_PREFIX.keys())
    for tag, prefix in SECTION_PREFIX.items():
        assert prefix.startswith("[") and prefix.endswith("]")


def test_upgrade_config_dict():
    from constants import upgrade_config_dict
    partial = {"C": 5.0, "max_iter": 100}
    result = upgrade_config_dict(partial)
    # Should have all default fields
    assert "use_noise_tokens" in result
    assert "use_per_field" in result
    # Custom values should override defaults
    assert result["C"] == 5.0
    assert result["max_iter"] == 100


# ===========================================================================
# Integration test: end-to-end pipeline
# ===========================================================================

def test_end_to_end_pipeline():
    """Full pipeline: build texts → vectorize → train → predict."""
    from feature_builder import build_feature_text, choose_row_profile_weights
    from text_utils import parse_dialog_roles, clean_answer_text
    from ml_core import make_hybrid_vectorizer, train_model

    raw_data = [
        {
            "channel": "call",
            "desc": "Проблема с картой, не работает банкомат",
            "dialog": "CLIENT: Не могу снять деньги в банкомате\nOPERATOR: Уточню информацию по вашему вопросу",
            "summary": "Клиент не может снять наличные",
            "ans_short": "Банкомат временно не работает",
            "label": "Карта",
        },
        {
            "channel": "call",
            "desc": "Вопрос по условиям кредита",
            "dialog": "CLIENT: Хочу узнать ставку по кредиту наличными\nOPERATOR: Расскажу об условиях кредитования",
            "summary": "Клиент интересуется кредитом",
            "ans_short": "Условия кредита описаны в тарифах",
            "label": "Кредит",
        },
    ] * 15  # 30 samples total

    base_weights = {"w_desc": 2, "w_client": 3, "w_operator": 1,
                    "w_summary": 2, "w_answer_short": 1, "w_answer_full": 0}

    feature_texts = []
    y = []
    for row in raw_data:
        _, client_text, operator_text, roles_found = parse_dialog_roles(row["dialog"])
        ans_short = clean_answer_text(row["ans_short"])
        weights = choose_row_profile_weights(
            base=base_weights, auto_profile="off",
            has_desc=bool(row["desc"]), has_dialog=bool(row["dialog"]),
            roles_found=roles_found, has_summary=bool(row["summary"]),
            has_ans_s=bool(ans_short), has_ans_f=False,
        )
        ft = build_feature_text(
            channel=row["channel"],
            desc=row["desc"],
            client_text=client_text,
            operator_text=operator_text,
            summary=row["summary"],
            ans_short=ans_short,
            ans_full="",
            weights=weights,
        )
        feature_texts.append(ft)
        y.append(row["label"])

    features = make_hybrid_vectorizer(
        use_per_field=True, base_weights=base_weights,
        min_df=1, max_features=500,
        use_stop_words=True, use_noise_tokens=True,
    )
    pipe, clf_type, report, labels_out, cm, extras = train_model(
        X=feature_texts, y=y,
        features=features,
        C=1.0, max_iter=200,
        balanced=True,
        test_size=0.2,
        random_state=42,
    )
    assert pipe is not None
    preds = pipe.predict(feature_texts[:2])
    assert len(preds) == 2


# ===========================================================================
# Run all tests
# ===========================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("Running test suite for LinearSVC-Calibrated codebase")
    print("=" * 70)

    # --- text_utils ---
    print("\n--- text_utils ---")
    run_test("strip_html basic", test_strip_html_basic)
    run_test("strip_html None", test_strip_html_none)
    run_test("strip_html nbsp", test_strip_html_nbsp)
    run_test("normalize_text basic", test_normalize_text_basic)
    run_test("normalize_text None", test_normalize_text_none)
    run_test("normalize_text masks", test_normalize_text_masks)
    run_test("normalize_text drops noise lines", test_normalize_text_drops_noise_lines)
    run_test("clean_masks", test_clean_masks)
    run_test("clean_masks None", test_clean_masks_none)
    run_test("clean_answer_text html", test_clean_answer_text_html)
    run_test("clean_answer_text None", test_clean_answer_text_none)
    run_test("clean_answer_text strips header", test_clean_answer_text_strips_header)
    run_test("parse_dialog_roles basic", test_parse_dialog_roles_basic)
    run_test("parse_dialog_roles no roles", test_parse_dialog_roles_no_roles)
    run_test("parse_dialog_roles empty", test_parse_dialog_roles_empty)
    run_test("parse_dialog_roles filters noise", test_parse_dialog_roles_filters_noise)
    run_test("parse_dialog_roles chatbot ignored", test_parse_dialog_roles_chatbot_ignored)
    run_test("parse_dialog_roles continuation", test_parse_dialog_roles_continuation)

    # --- feature_builder ---
    print("\n--- feature_builder ---")
    run_test("build_feature_text basic", test_build_feature_text_basic)
    run_test("build_feature_text weight zero skips section", test_build_feature_text_weight_zero_skips_section)
    run_test("build_feature_text empty fields", test_build_feature_text_empty_fields)
    run_test("head_tail_compress short text", test_head_tail_compress_short_text)
    run_test("head_tail_compress long text", test_head_tail_compress_long_text)
    run_test("head_tail_compress single line", test_head_tail_compress_single_line)
    run_test("head_tail_compress empty", test_head_tail_compress_empty)
    run_test("choose_row_profile_weights off", test_choose_row_profile_weights_off)
    run_test("choose_row_profile_weights zeroes empty fields", test_choose_row_profile_weights_zeroes_empty_fields)
    run_test("choose_row_profile_weights smart", test_choose_row_profile_weights_smart)
    run_test("choose_row_profile_weights strict", test_choose_row_profile_weights_strict)

    # --- excel_utils ---
    print("\n--- excel_utils ---")
    run_test("open_tabular CSV basic", test_open_tabular_csv_basic)
    run_test("open_tabular CSV empty fields → None", test_open_tabular_csv_empty_fields)
    run_test("open_tabular CSV semicolon delimiter", test_open_tabular_csv_semicolon_delimiter)
    run_test("count_csv_rows_fast", test_count_csv_rows_fast)
    run_test("count_csv_rows_fast no trailing newline", test_count_csv_rows_fast_no_trailing_newline)
    run_test("read_headers", test_read_headers)
    run_test("idx_of found", test_idx_of_found)
    run_test("idx_of not found", test_idx_of_not_found)
    run_test("idx_of empty name", test_idx_of_empty_name)
    run_test("estimate_total_rows", test_estimate_total_rows)
    run_test("fmt_eta zero done", test_fmt_eta_zero_done)
    run_test("fmt_eta non zero", test_fmt_eta_non_zero)
    run_test("fmt_speed", test_fmt_speed)

    # --- ml_core: PerFieldVectorizer ---
    print("\n--- ml_core: PerFieldVectorizer ---")
    run_test("PerFieldVectorizer fit_transform", test_per_field_vectorizer_fit_transform)
    run_test("PerFieldVectorizer fit then transform", test_per_field_vectorizer_fit_then_transform)
    run_test("PerFieldVectorizer empty texts", test_per_field_vectorizer_empty_texts)
    run_test("PerFieldVectorizer single row", test_per_field_vectorizer_single_row)
    run_test("PerFieldVectorizer all stop-words", test_per_field_vectorizer_all_stopwords)
    run_test("PerFieldVectorizer sparse 0-col matrices filtered (bug fix)", test_per_field_vectorizer_sparse_0_columns_filtered)
    run_test("PerFieldVectorizer _parse_fields", test_per_field_vectorizer_parse_fields)
    run_test("PerFieldVectorizer repeated tags deduplicated", test_per_field_vectorizer_repeated_tags_deduplicated)

    # --- ml_core: MetaFeatureExtractor ---
    print("\n--- ml_core: MetaFeatureExtractor ---")
    run_test("MetaFeatureExtractor basic shape", test_meta_feature_extractor_basic)
    run_test("MetaFeatureExtractor has_* flags", test_meta_feature_extractor_has_flags)
    run_test("MetaFeatureExtractor desc lengths log1p", test_meta_feature_extractor_lengths)
    run_test("MetaFeatureExtractor empty text", test_meta_feature_extractor_empty_text)
    run_test("MetaFeatureExtractor client_share 50/50", test_meta_feature_extractor_client_share)
    run_test("MetaFeatureExtractor no dialog → client_share=0.5", test_meta_feature_extractor_no_dialog)
    run_test("MetaFeatureExtractor various dialogs", test_meta_feature_extractor_various_dialogs)

    # --- ml_core: make_hybrid_vectorizer ---
    print("\n--- ml_core: make_hybrid_vectorizer ---")
    run_test("make_hybrid_vectorizer per_field", test_make_hybrid_vectorizer_per_field)
    run_test("make_hybrid_vectorizer legacy", test_make_hybrid_vectorizer_legacy)
    run_test("make_hybrid_vectorizer with meta", test_make_hybrid_vectorizer_with_meta)
    run_test("make_hybrid_vectorizer meta has StandardScaler", test_make_hybrid_vectorizer_meta_has_scaler)
    run_test("make_hybrid_vectorizer with stop_words", test_make_hybrid_vectorizer_with_stop_words)
    run_test("make_hybrid_vectorizer no base_weights fallback", test_make_hybrid_vectorizer_no_base_weights)

    # --- ml_core: make_classifier ---
    print("\n--- ml_core: make_classifier ---")
    run_test("make_classifier large dataset → CalibratedClassifierCV", test_make_classifier_large_dataset)
    run_test("make_classifier small dataset → LogisticRegression", test_make_classifier_small_dataset)
    run_test("make_classifier single class → fallback", test_make_classifier_single_class)
    run_test("make_classifier calib_method=isotonic", test_make_classifier_calib_method_isotonic)
    run_test("make_classifier calib_method=auto", test_make_classifier_calib_method_auto)
    run_test("make_classifier dynamic CV small (min_class=3 → cv=3)", test_make_classifier_dynamic_cv_small)
    run_test("make_classifier dynamic CV medium (min_class=6 → cv=3)", test_make_classifier_dynamic_cv_medium)
    run_test("make_classifier dynamic CV large (min_class=20 → cv=5)", test_make_classifier_dynamic_cv_large)

    # --- ml_core: train_model ---
    print("\n--- ml_core: train_model ---")
    run_test("train_model basic", test_train_model_basic)
    run_test("train_model empty validation set → no crash (bug fix)", test_train_model_empty_validation_set_no_crash)
    run_test("train_model too few samples → no crash", test_train_model_too_few_samples_no_crash)
    run_test("train_model with legacy vectorizer", test_train_model_with_legacy_vectorizer)
    run_test("train_model predict after training", test_train_model_predict_after_training)
    run_test("train_model predict_proba", test_train_model_predict_proba)
    run_test("train_model extras thresholds", test_train_model_extras_thresholds)

    # --- ml_core: PhraseRemover ---
    print("\n--- ml_core: PhraseRemover ---")
    run_test("PhraseRemover basic", test_phrase_remover_basic)
    run_test("PhraseRemover empty phrases", test_phrase_remover_empty_phrases)
    run_test("PhraseRemover None input", test_phrase_remover_none_input)
    run_test("PhraseRemover fit_transform", test_phrase_remover_fit_transform)

    # --- ml_core: dataset_health_checks ---
    print("\n--- ml_core: dataset_health_checks ---")
    run_test("dataset_health_checks OK", test_dataset_health_checks_ok)
    run_test("dataset_health_checks no rows → fatal", test_dataset_health_checks_no_rows)
    run_test("dataset_health_checks single class → fatal", test_dataset_health_checks_single_class)

    # --- ml_core: extract_cluster_keywords ---
    print("\n--- ml_core: extract_cluster_keywords ---")
    run_test("extract_cluster_keywords PerFieldVectorizer", test_extract_cluster_keywords_per_field)
    run_test("extract_cluster_keywords TfidfVectorizer", test_extract_cluster_keywords_tfidf)
    run_test("extract_cluster_keywords dim mismatch → empty strings", test_extract_cluster_keywords_dim_mismatch)
    run_test("extract_cluster_keywords no positive values → empty strings", test_extract_cluster_keywords_no_positive_values)

    # --- ml_core: extract_cluster_keywords_from_labels ---
    print("\n--- ml_core: extract_cluster_keywords_from_labels ---")
    run_test("extract_cluster_keywords_from_labels basic", test_extract_cluster_keywords_from_labels_basic)
    run_test("extract_cluster_keywords_from_labels empty cluster", test_extract_cluster_keywords_from_labels_empty_cluster)
    run_test("extract_cluster_keywords_from_labels dense matrix", test_extract_cluster_keywords_from_labels_dense)

    # --- ml_core: extract_cluster_keywords_ctfidf ---
    print("\n--- ml_core: extract_cluster_keywords_ctfidf ---")
    run_test("extract_cluster_keywords_ctfidf basic", test_extract_cluster_keywords_ctfidf_basic)
    run_test("extract_cluster_keywords_ctfidf empty cluster", test_extract_cluster_keywords_ctfidf_empty_cluster)
    run_test("extract_cluster_keywords_ctfidf with stop_words", test_extract_cluster_keywords_ctfidf_with_stop_words)
    run_test("extract_cluster_keywords_ctfidf bigrams in keywords", test_ctfidf_bigrams_in_keywords)

    # --- Elbow / KMeans pipeline ---
    print("\n--- Clustering: elbow & pipeline ---")
    run_test("pick_elbow_k fallback logic", test_elbow_k_fallback_logic)
    run_test("KMeans + keywords full pipeline", test_kmeans_cluster_pipeline)

    # --- Clustering: full pipeline & output correctness ---
    print("\n--- Clustering: full_cluster mapping & output correctness ---")
    run_test("full_cluster single file mapping", test_cluster_full_cluster_single_file)
    run_test("full_cluster multi-file label_offset", test_cluster_full_cluster_multi_file)
    run_test("kw_dict built from List[str]", test_cluster_kw_dict_from_list)
    run_test("output columns: cluster_id & keywords", test_cluster_output_columns)
    run_test("end-to-end clustering with CSV file", test_cluster_end_to_end_with_csv)

    # --- constants / config ---
    print("\n--- constants / config ---")
    run_test("RUSSIAN_STOP_WORDS non-empty", test_russian_stop_words_non_empty)
    run_test("NOISE_TOKENS non-empty", test_noise_tokens_non_empty)
    run_test("NOISE_PHRASES non-empty", test_noise_phrases_non_empty)
    run_test("PRESET_WEIGHTS valid structure", test_preset_weights_valid)
    run_test("SECTION_PREFIX valid structure", test_section_prefix_valid)
    run_test("upgrade_config_dict fills defaults", test_upgrade_config_dict)
    run_test("upgrade_config_dict use_lemma default=True", test_upgrade_config_dict_lemma_default_true)

    # --- Integration ---
    print("\n--- Integration ---")
    run_test("end-to-end pipeline", test_end_to_end_pipeline)

    # --- Summary ---
    print("\n" + "=" * 70)
    print(f"Results: {passed} passed, {failed} failed")
    print("=" * 70)
    sys.exit(0 if failed == 0 else 1)
