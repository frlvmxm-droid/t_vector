# -*- coding: utf-8 -*-
"""
app_apply.py — ApplyTabMixin: вкладка «Классификация» и логика применения модели.

Содержит:
  • _build_apply_tab()  — построение UI вкладки
  • run_apply()         — фоновый поток классификации (с flush_chunk)
  • Вспомогательные методы (sort/copy таблицы, пикеры файлов)
"""
from __future__ import annotations

import heapq
import sys
import threading
import time
import traceback as _traceback
from pathlib import Path
from typing import Any, Dict, List, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from constants import CLASS_DIR, MODEL_DIR, now_stamp, upgrade_config_dict
from config import DEFAULT_COLS, DEFAULT_OTHER_LABEL
from excel_utils import (
    read_headers, idx_of, estimate_total_rows,
    fmt_eta, fmt_speed, open_tabular, patch_xlsx_col_widths,
)
from feature_builder import build_feature_text, choose_row_profile_weights
from ml_core import SBERTVectorizer, find_sbert_in_pipeline, find_setfit_classifier, SBERT_LOCAL_DIR
from hw_profile import tune_runtime_by_input_size
from app_logger import get_logger
from exceptions import ModelLoadError, FeatureBuildError, PredictPipelineError
from model_loader import ensure_trusted_model_path, load_model_artifact, get_trusted_model_hash, make_tkinter_confirm_fn
from app_apply_service import EnsemblePredictor
from apply_prediction_service import ApplyRunState
from app_apply_view import build_apply_files_card
from app_apply_workflow import validate_apply_preconditions, build_validated_apply_snapshot
from task_runner import ErrorEnvelope, OperationLifecycle, begin_long_task, prepare_long_task_ui
from ui_theme import BG, FG, PANEL2, ENTRY_BG, ACCENT, ACCENT3, MUTED, MUTED2, BORDER, SUCCESS
from artifact_contracts import TRAIN_MODEL_ARTIFACT_TYPE
from ui.tabs.apply_presenter import format_apply_autoprofile_log

_log = get_logger(__name__)
from ui_widgets_tk import Tooltip, CollapsibleSection


class ApplyTabMixin:
    """Методы вкладки «Классификация» и логика применения модели к данным."""

    # ------------------------------------------------------------------
    # Единый безопасный загрузчик моделей
    # ------------------------------------------------------------------

    def _ensure_trusted_path(self, path: str, label: str = "Модель") -> bool:
        """Проверяет что path в доверенном списке сессии.

        Если нет — показывает consent-диалог из фонового потока через self.after().
        Ожидает ответа максимум 30 секунд; по истечении timeout автоматически
        отказывает (возвращает False и пишет RuntimeWarning).

        Returns:
            True  — путь доверенный (уже в списке или пользователь подтвердил).
            False — пользователь отказался или timeout.
        """
        return ensure_trusted_model_path(
            self._trust_store, path, label,
            confirm_fn=make_tkinter_confirm_fn(self),
            logger=_log,
        )

    # Текущая поддерживаемая schema_version
    _SUPPORTED_SCHEMA_VERSION = 1

    def _load_model_pkg(self, path: str, log_fn=None) -> dict:
        """Загружает model-bundle через единый безопасный loader и проверяет schema_version.

        Важно: чтение артефакта инкапсулировано в `model_loader.load_model_artifact`
        (включая формат-политику, SHA/checks и валидацию), а этот метод задаёт
        apply-callsite policy и унифицирует диагностику версии.

        Политика schema_version:
          None               → мягкое предупреждение (старый формат, обратная совместимость).
          int == 1           → OK.
          int > 1            → ModelLoadError (формат из будущей версии приложения).
          не-int (str, dict…) → ModelLoadError (повреждённый / незнакомый артефакт).

        Args:
            path:   путь к .joblib файлу (должен быть в trust_store).
            log_fn: функция логирования, принимающая строку (или None).

        Returns:
            Словарь-пакет модели (ключи: pipeline, config, schema_version, …).

        Raises:
            ModelLoadError: если schema_version несовместима (неверный тип или > supported).
        """
        return load_model_artifact(
            path,
            precomputed_sha256=self._trust_store.get_hash(path),
            supported_schema_version=self._SUPPORTED_SCHEMA_VERSION,
            expected_artifact_types=(TRAIN_MODEL_ARTIFACT_TYPE,),
            required_keys=("pipeline",),
            allowed_extensions=(".joblib",),
            require_trusted=True,
            trusted_paths=self._trust_store.trusted_canonical_paths(),
            allow_missing_schema=True,
            log_fn=log_fn,
            logger=_log,
        )

    def _draw_result_chart(self, distribution: "Dict[str, int]", total: int) -> None:
        """Рисует горизонтальный bar-chart распределения меток на self._chart_canvas."""
        from ui_theme_ctk import COLORS as _CTK
        _FG    = _CTK["fg"]
        _MUTED = _CTK["muted"]
        _BAR1  = _CTK["accent"]   # #f97316 — orange, top bar
        _BAR_N = _CTK["accent3"]  # #c2410c — darker orange, rest

        cv = self._chart_canvas
        if not cv.winfo_exists():
            return
        cv.delete("all")
        if not distribution or total <= 0:
            cv.configure(height=0)
            return

        sorted_items = sorted(distribution.items(), key=lambda x: x[1], reverse=True)[:10]
        n = len(sorted_items)
        if n == 0:
            cv.configure(height=0)
            return

        row_h    = 28
        pad_left = 200
        pad_right = 90
        pad_top  = 10
        bar_gap  = 4   # vertical gap inside row
        chart_h  = n * row_h + pad_top * 2

        cv.configure(height=chart_h)
        cv.update_idletasks()
        w = cv.winfo_width()
        if w < 100:
            w = 700

        bar_w = max(1, w - pad_left - pad_right)
        max_count = sorted_items[0][1] if sorted_items else 1

        from ui_theme import _best_font
        _fnt  = (_best_font(), 9)
        _fnt_b = (_best_font(), 9, "bold")

        for i, (label, count) in enumerate(sorted_items):
            y = pad_top + i * row_h + row_h // 2
            pct = count / total * 100

            # label text (right-aligned before bar)
            short_lbl = label if len(label) <= 28 else label[:26] + "…"
            cv.create_text(
                pad_left - 10, y, text=short_lbl,
                anchor="e",
                fill=_FG if i == 0 else _MUTED,
                font=_fnt_b if i == 0 else _fnt,
            )

            # bar
            fill_w = int(bar_w * count / max_count)
            if fill_w > 0:
                shade = _BAR1 if i == 0 else _BAR_N
                cv.create_rectangle(
                    pad_left, y - row_h // 2 + bar_gap,
                    pad_left + fill_w, y + row_h // 2 - bar_gap,
                    fill=shade, outline="", width=0,
                )

            # count + pct
            cv.create_text(
                pad_left + fill_w + 10, y,
                text=f"{count:,}  ({pct:.1f}%)",
                anchor="w", fill=_MUTED, font=_fnt,
            )

    def _build_apply_ensemble_card(self, parent: "ttk.Frame") -> None:
        # ── Ансамбль моделей ────────────────────────────────────────────────
        ens_lf = ttk.LabelFrame(
            parent,
            text="Ансамбль моделей",
            padding=10,
        )
        ens_lf.pack(fill="x", pady=(0, 10))

        _ens_top = ttk.Frame(ens_lf)
        _ens_top.pack(fill="x")
        cb_ens = ttk.Checkbutton(
            _ens_top, text="Режим ансамбля", variable=self.use_ensemble,
            command=self._toggle_ensemble_ui,
        )
        cb_ens.pack(side="left")
        self.attach_help(
            cb_ens,
            "Ансамбль моделей",
            "Объединяет предсказания двух моделей через средневзвешенное predict_proba.\n\n"
            "Как использовать:\n"
            "  1. Обучите модель A (например, с SBERT ruBERT-tiny2).\n"
            "  2. Обучите модель B (например, с USER-base / USER2-base от deepvk).\n"
            "  3. Выберите модель A как основную (Шаг 1), модель B — здесь.\n"
            "  4. Настройте веса (обычно 50/50 или 60/40 в пользу лучшей модели).\n\n"
            "Ансамбль разных архитектур (TF-IDF + DeBERTa, ruBERT + DeBERTa)\n"
            "даёт +2–5% macro F1 по сравнению с лучшей одиночной моделью.\n\n"
            "Оба .joblib файла должны быть обучены на одних и тех же классах.",
            "Режим ансамбля: усреднение predict_proba двух моделей",
        )

        self._ens_inner = ttk.Frame(ens_lf)
        self._ens_inner.pack(fill="x", pady=(6, 0))

        ttk.Label(self._ens_inner, text="Модель 2 (.joblib):", style="Card.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Entry(self._ens_inner, textvariable=self.ensemble_model2, width=75).grid(
            row=0, column=1, sticky="we", padx=10
        )
        btn_ens2 = ttk.Button(self._ens_inner, text="Выбрать…", command=self.pick_ensemble_model)
        btn_ens2.grid(row=0, column=2)
        Tooltip(btn_ens2, "Выбрать второй .joblib файл модели для ансамбля.\nОба файла должны содержать одинаковый набор классов.")
        self._ens_inner.columnconfigure(1, weight=1)

        _w_row = ttk.Frame(self._ens_inner)
        _w_row.grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 0))
        ttk.Label(_w_row, text="Вес модели 1:").pack(side="left")
        _w1_sv = tk.StringVar(value="50%")
        def _upd_w1(*_):
            _w1_sv.set(f"{self.ensemble_w1.get() * 100:.0f}%")
        self.ensemble_w1.trace_add("write", _upd_w1)
        ttk.Scale(_w_row, from_=0.1, to=0.9, orient="horizontal",
                  variable=self.ensemble_w1, length=160).pack(side="left", padx=8)
        ttk.Label(_w_row, textvariable=_w1_sv, width=5).pack(side="left")
        _w2_sv = tk.StringVar(value="(модель 2: 50%)")
        def _upd_w2(*_):
            _w2_sv.set(f"(модель 2: {(1 - self.ensemble_w1.get()) * 100:.0f}%)")
        self.ensemble_w1.trace_add("write", _upd_w2)
        ttk.Label(_w_row, textvariable=_w2_sv, style="Card.Muted.TLabel").pack(side="left", padx=(8, 0))

        self._toggle_ensemble_ui()   # скрыть inner пока чекбокс выключен


    def _build_apply_adv_card(self, parent: "ttk.Frame") -> None:
        # ── Дополнительные колонки и авто-замена (свёрнуты по умолчанию) ────
        _adv_apply = CollapsibleSection(
            parent,
            title="Доп. колонки и авто-замена (суммаризация, ответы, замена «Другое»)",
            collapsed=True,
        )
        _adv_apply.pack(fill="x", pady=(0, 10))

        _adv_cols = ttk.LabelFrame(_adv_apply.body, text="Дополнительные колонки", padding=12)
        _adv_cols.pack(fill="x", pady=(4, 6))
        self._combobox(_adv_cols, 0, "Суммаризация:", self.summary_col, "Выжимка диалога (если есть)")
        self._combobox(_adv_cols, 1, "Ответ краткий:", self.ans_short_col, "Короткий ответ оператора")
        self._combobox(_adv_cols, 2, "Ответ полный:", self.ans_full_col, "Развёрнутый ответ / HTML")
        _adv_cols.columnconfigure(1, weight=1)

        # --- авто-замена на «Другое» ---
        row_other = ttk.Frame(_adv_apply.body)
        row_other.pack(fill="x", padx=10, pady=(0, 6))
        cb_other = ttk.Checkbutton(row_other, text="Заменять на:", variable=self.use_other_label)
        cb_other.pack(side="left")
        self.attach_help(cb_other, "Авто-замена при низкой уверенности",
                         "Если включено, строки с уверенностью модели (pred_proba) НИЖЕ\n"
                         "указанного порога получат выбранную метку вместо предсказания.\n\n"
                         "pred_proba и pred_top3 сохраняют оригинальные значения —\n"
                         "виден исходный результат модели даже у замененных строк.\n\n"
                         "Типичный сценарий: всё, в чём модель неуверена → «Другое»,\n"
                         "чтобы в датасете не было «мусорных» предсказаний.\n\n"
                         "Работает независимо от порога ревью (pred_needs_review).",
                         "Авто-замена на «Другое» при низкой уверенности")
        ent_other = ttk.Entry(row_other, textvariable=self.other_label_text, width=14)
        ent_other.pack(side="left", padx=(4, 0))
        self.attach_help(ent_other, "Метка для замены",
                         "Текст, который будет записан в pred_col для строк\n"
                         "с уверенностью ниже порога.\n\n"
                         "По умолчанию: «Другое»\n"
                         "Можно задать любой текст, например «Нераспознано».",
                         "Метка-замена для неуверенных предсказаний")
        ttk.Label(row_other, text="при уверенности ниже").pack(side="left", padx=(10, 0))
        _oth_sv = tk.StringVar(value=f"{self.other_label_threshold.get() * 100:.0f}%")
        def _upd_oth(*_):
            _oth_sv.set(f"{self.other_label_threshold.get() * 100:.0f}%")
        self.other_label_threshold.trace_add("write", _upd_oth)
        ttk.Scale(row_other, from_=0.10, to=0.90, orient="horizontal",
                  variable=self.other_label_threshold, length=150).pack(side="left", padx=(8, 0))
        ttk.Label(row_other, textvariable=_oth_sv, width=5).pack(side="left")
        cb_pc_other = ttk.Checkbutton(row_other, text="per-class пороги из модели",
                                      variable=self.use_per_class_other_threshold)
        cb_pc_other.pack(side="left", padx=(10, 0))
        self.attach_help(cb_pc_other, "Использовать per-class пороги обученной модели",
                         "Если включено и модель сохранила per_class_thresholds\n"
                         "(выбирается автоматически по PR-кривой для каждого класса\n"
                         "на валидации) — используются они. Порог выше = «Другое» применяется\n"
                         "к классу-кандидату, только если уверенность ниже его индивидуального\n"
                         "порога. Глобальный слайдер выступает fallback для отсутствующих классов.\n\n"
                         "Эффект: меньше ложно-положительных на сложных классах, больше —\n"
                         "правильно попавших в «Другое».",
                         "per-class пороги для «Другое»")

        # --- детектор неоднозначных предсказаний ---
        row_ambig = ttk.Frame(_adv_apply.body)
        row_ambig.pack(fill="x", padx=10, pady=(0, 8))
        cb_ambig = ttk.Checkbutton(row_ambig, text="Детектор неоднозначных предсказаний",
                                   variable=self.use_ambiguity_detector)
        cb_ambig.pack(side="left")
        self.attach_help(cb_ambig, "Детектор неоднозначных предсказаний",
                         "Если разрыв между топ-1 и топ-2 уверенностью меньше ε,\n"
                         "строка получает pred_needs_review=1 даже при высокой уверенности.\n\n"
                         "Помогает поймать случаи, когда модель «колеблется» между двумя классами.\n\n"
                         "Рекомендуемое ε: 0.05–0.10")
        ttk.Label(row_ambig, text="  ε =").pack(side="left")
        ttk.Spinbox(row_ambig, from_=0.01, to=0.30, increment=0.01,
                    textvariable=self.ambiguity_epsilon, width=5, format="%.2f").pack(side="left")


    def _build_apply_tab(self):
        # ── Sub-tab frames ─────────────────────────────────────────────────
        _s0 = ttk.Frame(self.tab_apply)   # Настройки
        _s1 = ttk.Frame(self.tab_apply)   # Запуск & Лог

        build_apply_files_card(self, _s0)

        # ── Кнопки пресетов ──────────────────────────────────────────────────
        _preset_row = ttk.Frame(_s0)
        _preset_row.pack(fill="x", pady=(0, 6))
        _btn_save_preset_a = ttk.Button(_preset_row, text="💾 Пресет",
                                        command=self._save_preset)
        _btn_save_preset_a.pack(side="left", padx=(0, 4))
        Tooltip(_btn_save_preset_a, "Сохранить текущие настройки классификации как пресет.")
        _btn_load_preset_a = ttk.Button(_preset_row, text="📂 Пресет ▾",
                                        command=lambda: self._open_preset_menu(_btn_load_preset_a))
        _btn_load_preset_a.pack(side="left")
        Tooltip(_btn_load_preset_a, "Загрузить ранее сохранённый пресет настроек.")

        self._build_apply_ensemble_card(_s0)
        # ── Основные колонки (всегда видны) ─────────────────────────────────
        cols = ttk.LabelFrame(_s0, text="Колонки и настройки классификации", padding=12)
        cols.pack(fill="x", pady=(0, 6))
        self._combobox(cols, 0, "Описание:", self.desc_col, "Описание обращения")
        self._combobox(cols, 1, "Текст звонка:", self.call_col, "Транскрипт звонка")
        self._combobox(cols, 2, "Текст чата:", self.chat_col, "Транскрипт чата")

        ttk.Label(cols, text="Колонка для предсказания:", style="Card.TLabel").grid(row=3, column=0, sticky="w", pady=(6, 0))
        self.cb_pred = ttk.Combobox(cols, textvariable=self.pred_col, state="normal", width=64, values=[])
        self.cb_pred.grid(row=3, column=1, sticky="w", padx=10, pady=(6, 0))
        self.attach_help(self.cb_pred, "Колонка pred", "В эту колонку будет записан итоговый label (причина).", "Напр.: pred_marker1")

        row_thr = ttk.Frame(cols)
        row_thr.grid(row=4, column=1, sticky="w", padx=10, pady=(8, 6))
        lbl_thr = ttk.Label(row_thr, text="Порог уверенности (ревью):")
        lbl_thr.pack(side="left")
        Tooltip(lbl_thr,
                "Строки, где уверенность модели (pred_proba) НИЖЕ этого порога,\n"
                "получат pred_needs_review=1 в выходном Excel.\n\n"
                "Это сигнал для ручной проверки сомнительных предсказаний.\n\n"
                "Используйте кнопку «↓ Из обучения» для авто-выбора порога\n"
                "по результатам последнего обучения (75-й перцентиль).\n\n"
                "Рекомендуется: 0.55–0.75")
        _thr_sv = tk.StringVar(value=f"{self.review_threshold.get():.2f}")
        def _upd_thr(*_):
            _thr_sv.set(f"{self.review_threshold.get():.2f}")
        self.review_threshold.trace_add("write", _upd_thr)
        ttk.Scale(row_thr, from_=0.30, to=0.95, orient="horizontal", variable=self.review_threshold, length=190).pack(side="left", padx=10)
        ttk.Label(row_thr, textvariable=_thr_sv, width=5).pack(side="left")

        def _apply_rec_thr():
            v = self._rec_threshold_75.get()
            if v > 0:
                self.review_threshold.set(round(v, 3))
        btn_rec_thr = ttk.Button(row_thr, text="↓ Из обучения", command=_apply_rec_thr)
        btn_rec_thr.pack(side="left", padx=(10, 0))
        self.attach_help(btn_rec_thr, "Применить порог из обучения",
                         "Применяет рекомендованный порог (75-й перцентиль уверенности),\n"
                         "рассчитанный по валидационной выборке последнего обучения.\n\n"
                         "Это значит: 75% предсказаний были уверены выше этого порога.\n"
                         "Строки ниже попадут в pred_needs_review=1 (нуждаются в проверке).\n\n"
                         "Кнопка активна только после успешного обучения с test_size > 0.",
                         "Применить рекомендованный порог из последнего обучения")
        ttk.Label(row_thr, textvariable=self._rec_thr_label,
                  style="Card.Muted.TLabel").pack(side="left", padx=(6, 0))

        cols.columnconfigure(1, weight=1)

        self._build_apply_adv_card(_s0)
        # ── Per-class пороги (тюнинг без переобучения) ──────────────────────
        _thresh_section = CollapsibleSection(
            _s0,
            title="Per-class пороги (тюнинг без переобучения)",
            collapsed=True,
        )
        _thresh_section.pack(fill="x", pady=(0, 10))
        self._build_class_threshold_editor(_thresh_section.body)

        self._build_readiness_bar(_s1, [
            lambda: (
                bool(self.model_file.get()),
                f"🤖 Модель: {Path(self.model_file.get()).name}" if self.model_file.get()
                else "🤖 Модель: не выбрана",
            ),
            lambda: (
                bool(self.apply_file.get()),
                f"📄 Файл: {Path(self.apply_file.get()).name}" if self.apply_file.get()
                else "📄 Файл: не выбран",
            ),
        ])
        self.btn_apply, self.btn_apply_stop = self._build_action_block(
            _s1,
            btn_text="▶  Классифицировать",
            btn_cmd=self.run_apply,
            progress_var=self.apply_progress,
            pct_var=self.apply_pct,
            phase_var=self.apply_phase,
            speed_var=self.apply_speed,
            eta_var=self.apply_eta,
            label="Прогресс классификации",
        )
        summ = ttk.Frame(_s1, style="Card.TFrame", padding=(12, 6))
        summ.pack(fill="x", pady=(0, 6))
        ttk.Label(summ, text="Результат:", style="Card.TLabel").pack(side="left")
        ttk.Label(summ, textvariable=self.last_apply_summary, style="Card.Muted.TLabel").pack(side="left", padx=10)

        btn_open_dir = ttk.Button(summ, text="📂 Открыть папку результатов",
                                  command=lambda: self._open_directory(CLASS_DIR))
        btn_open_dir.pack(side="right")
        Tooltip(btn_open_dir, f"Открывает папку с файлами классификации:\n{CLASS_DIR}")
        # ── Bar-chart распределения меток (заполняется после run_apply) ──────────
        self._chart_canvas = tk.Canvas(
            _s1, bg=PANEL2, highlightthickness=0, bd=0, height=0,
        )
        self._chart_canvas.pack(fill="x", pady=(6, 0))

        ttk.Label(_s1, text="Сводная таблица по результатам классификации  (ПКМ = копировать):",
                  style="Card.TLabel").pack(anchor="w", padx=4, pady=(4, 0))

        _tree_frame = ttk.Frame(_s1)
        _tree_frame.pack(fill="both", expand=True)

        self.apply_tree = ttk.Treeview(
            _tree_frame,
            columns=("label", "count", "pct", "avg_proba", "examples"),
            show="headings",
            height=14,
        )
        col_defs = [
            ("label",     "Описание проблемы",    280),
            ("count",     "Кол-во обращений",     130),
            ("pct",       "% попадания",            90),
            ("avg_proba", "Ср. уверенность",        90),
            ("examples",  "Примеры текстов",       700),
        ]
        self._apply_sort_reverse: Dict[str, bool] = {}
        for cid, heading, width in col_defs:
            self.apply_tree.heading(
                cid, text=heading,
                anchor="w",
                command=lambda c=cid: self._sort_apply_tree(c),
            )
            self.apply_tree.column(cid, width=width, anchor="w")

        _vsb = ttk.Scrollbar(_tree_frame, orient="vertical", command=self.apply_tree.yview)
        self.apply_tree.configure(yscrollcommand=_vsb.set)
        self.apply_tree.pack(side="left", fill="both", expand=True)
        _vsb.pack(side="right", fill="y")

        # Контекстное меню: копирование строки / всей таблицы
        self._apply_ctx_menu = tk.Menu(self.apply_tree, tearoff=0)
        self._apply_ctx_menu.add_command(
            label="Копировать строку", command=self._copy_apply_row,
        )
        self._apply_ctx_menu.add_command(
            label="Копировать всю таблицу", command=self._copy_apply_all,
        )
        self.apply_tree.bind("<Button-3>", self._on_apply_tree_rclick)

        # ── Register bottom sub-tab strip for Classification tab ──────────────
        self._register_sub_tabs(
            1,
            ["Настройки", "Запуск & Лог"],
            [_s0, _s1],
        )

    # ------------------------------------------------------------------
    # Per-class threshold editor
    # ------------------------------------------------------------------

    def _build_class_threshold_editor(self, parent: ttk.Frame) -> None:
        """Строит редактор per-class порогов (вызывается из _build_apply_tab)."""
        ctrl_row = ttk.Frame(parent)
        ctrl_row.pack(fill="x", padx=8, pady=(4, 4))

        cb_use = ttk.Checkbutton(
            ctrl_row,
            text="Использовать кастомные per-class пороги",
            variable=self.use_custom_class_thresholds,
        )
        cb_use.pack(side="left")
        self.attach_help(
            cb_use,
            "Кастомные per-class пороги",
            "Если включено, для каждого класса применяется порог из таблицы ниже\n"
            "вместо глобального порога ревью.\n\n"
            "Позволяет тонко настраивать чувствительность модели по каждому классу\n"
            "без переобучения.\n\n"
            "Нажмите «↓ Из модели», чтобы заполнить пороги значениями из обученной модели.",
        )

        btn_from_model = ttk.Button(
            ctrl_row, text="↓ Из модели",
            command=self._load_thresholds_from_model,
        )
        btn_from_model.pack(side="left", padx=(12, 4))
        self.attach_help(
            btn_from_model,
            "Загрузить пороги из модели",
            "Считывает per_class_thresholds из загруженной модели и заполняет таблицу.\n\n"
            "Модель должна быть выбрана в поле «Файл модели».",
        )

        btn_reset = ttk.Button(
            ctrl_row, text="Сброс",
            command=lambda: self._reset_class_thresholds(0.60),
        )
        btn_reset.pack(side="left", padx=(0, 4))
        self.attach_help(btn_reset, "Сброс порогов", "Сбрасывает все пороги к глобальному значению 0.60.")

        ttk.Label(
            ctrl_row,
            text="  Пороги применяются к pred_needs_review (и к «Другое» если включено)",
            style="Card.Muted.TLabel",
        ).pack(side="left", padx=(8, 0))

        # Scrollable area for class rows
        _outer = ttk.Frame(parent)
        _outer.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        _canvas = tk.Canvas(_outer, height=160, highlightthickness=0)
        _vsb = ttk.Scrollbar(_outer, orient="vertical", command=_canvas.yview)
        _canvas.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side="right", fill="y")
        _canvas.pack(side="left", fill="both", expand=True)
        self._thresh_inner = ttk.Frame(_canvas)
        self._thresh_canvas_win = _canvas.create_window((0, 0), window=self._thresh_inner, anchor="nw")
        self._thresh_inner.bind(
            "<Configure>",
            lambda e: _canvas.configure(scrollregion=_canvas.bbox("all")),
        )
        _canvas.bind(
            "<Configure>",
            lambda e: _canvas.itemconfig(self._thresh_canvas_win, width=e.width),
        )
        self._thresh_canvas = _canvas

        # Header row
        hdr = ttk.Frame(self._thresh_inner)
        hdr.pack(fill="x", pady=(2, 0))
        ttk.Label(hdr, text="Класс", width=36, style="Card.TLabel").pack(side="left")
        ttk.Label(hdr, text="Порог", width=8, style="Card.TLabel").pack(side="left")
        ttk.Label(hdr, text="Слайдер", width=22, style="Card.TLabel").pack(side="left")
        ttk.Label(hdr, text="↓ из модели", width=12, style="Card.Muted.TLabel").pack(side="left")

        self._thresh_class_rows: List[Dict] = []

    def populate_class_thresholds(
        self,
        class_thresholds: Dict[str, float],
        *,
        model_thresholds: Optional[Dict[str, float]] = None,
    ) -> None:
        """Заполняет редактор строками классов с порогами.

        Args:
            class_thresholds: {class_label: threshold} — текущие значения для отображения.
            model_thresholds: {class_label: threshold} — оригиналы из модели (для кнопки «↓»).
        """
        # Clear existing rows
        for w in self._thresh_inner.winfo_children():
            if getattr(w, "_is_thresh_hdr", False):
                continue
            w.destroy()
        self._thresh_class_rows.clear()
        self._custom_class_thresholds.clear()

        # Rebuild header
        hdr = ttk.Frame(self._thresh_inner)
        hdr._is_thresh_hdr = True  # type: ignore[attr-defined]
        hdr.pack(fill="x", pady=(2, 0))
        ttk.Label(hdr, text="Класс", width=36, style="Card.TLabel").pack(side="left")
        ttk.Label(hdr, text="Порог", width=8, style="Card.TLabel").pack(side="left")
        ttk.Label(hdr, text="Слайдер", width=22, style="Card.TLabel").pack(side="left")
        ttk.Label(hdr, text="↓ из модели", width=12, style="Card.Muted.TLabel").pack(side="left")

        _orig = dict(model_thresholds or class_thresholds)

        for cls_label, thr_val in sorted(class_thresholds.items()):
            var = tk.DoubleVar(value=round(float(thr_val), 3))
            self._custom_class_thresholds[cls_label] = var

            row = ttk.Frame(self._thresh_inner)
            row.pack(fill="x", pady=1)

            ttk.Label(row, text=cls_label, width=36, style="Card.TLabel").pack(side="left")

            sv = tk.StringVar(value=f"{var.get():.2f}")
            def _upd_sv(*_, v=var, s=sv):
                s.set(f"{v.get():.2f}")
            var.trace_add("write", _upd_sv)

            ttk.Label(row, textvariable=sv, width=6).pack(side="left")
            ttk.Scale(
                row, from_=0.10, to=0.99, orient="horizontal",
                variable=var, length=150,
            ).pack(side="left", padx=(4, 8))

            orig_val = _orig.get(cls_label, thr_val)
            ttk.Button(
                row, text="↓",
                command=lambda v=var, o=orig_val: v.set(round(o, 3)),
                width=3,
            ).pack(side="left")

            self._thresh_class_rows.append({"label": cls_label, "var": var})

        # Reset canvas scroll
        try:
            self._thresh_canvas.yview_moveto(0)
        except Exception as _scroll_exc:
            _log.debug("thresh canvas scroll reset failed: %s", _scroll_exc)

    def _load_thresholds_from_model(self) -> None:
        """Считывает per_class_thresholds из текущей модели и заполняет редактор."""
        model_path = self.model_file.get().strip()
        if not model_path:
            self.log_apply("⚠ Сначала выберите файл модели.")
            return
        try:
            pkg = self._load_model_pkg(model_path, log_fn=self.log_apply)
            thr = pkg.get("per_class_thresholds") or {}
            if not thr:
                self.log_apply("⚠ Модель не содержит per_class_thresholds (обучена без test_size?).")
                return
            self.populate_class_thresholds(thr, model_thresholds=thr)
            self.use_custom_class_thresholds.set(True)
            self.log_apply(f"[Пороги] Загружено {len(thr)} классов из модели.")
        except Exception as e:
            self.log_apply(f"⚠ Не удалось загрузить пороги из модели: {e}")

    def _reset_class_thresholds(self, value: float = 0.60) -> None:
        """Сбрасывает все кастомные пороги к заданному значению."""
        for row in self._thresh_class_rows:
            row["var"].set(round(value, 3))

    def _sort_apply_tree(self, col: str) -> None:
        """Сортирует строки apply_tree по клику на заголовок колонки (toggle asc/desc)."""
        reverse = not self._apply_sort_reverse.get(col, True)
        self._apply_sort_reverse[col] = reverse

        raw = [(self.apply_tree.set(iid, col), iid) for iid in self.apply_tree.get_children()]
        try:
            if col == "count":
                raw.sort(key=lambda x: int(x[0]) if x[0] else 0, reverse=reverse)
            elif col in ("pct", "avg_proba"):
                raw.sort(
                    key=lambda x: float(x[0].rstrip("%")) if x[0] and x[0] not in ("—", "") else 0.0,
                    reverse=reverse,
                )
            else:
                raw.sort(key=lambda x: x[0].lower(), reverse=reverse)
        except Exception as _e:
            _log.debug(
                "_sort_apply_tree col=%r failed (%s); falling back to str sort. "
                "First bad value: %r",
                col, _e,
                next((v for v, _ in raw if v not in ("", "—")), None),
            )
            raw.sort(key=lambda x: x[0], reverse=reverse)

        for i, (_, iid) in enumerate(raw):
            self.apply_tree.move(iid, "", i)

        # Обновляем заголовок — показываем стрелку направления
        for cid in self.apply_tree["columns"]:
            cur = self.apply_tree.heading(cid)["text"].rstrip(" ▲▼")
            arrow = (" ▼" if reverse else " ▲") if cid == col else ""
            self.apply_tree.heading(cid, text=cur + arrow)

    def _on_apply_tree_rclick(self, event: "tk.Event") -> None:
        """Показывает контекстное меню по правой кнопке мыши."""
        iid = self.apply_tree.identify_row(event.y)
        if iid:
            self.apply_tree.selection_set(iid)
        try:
            self._apply_ctx_menu.post(event.x_root, event.y_root)
        finally:
            self._apply_ctx_menu.grab_release()

    def _copy_apply_row(self) -> None:
        """Копирует выделенные строки в буфер обмена (TSV)."""
        sel = self.apply_tree.selection()
        if not sel:
            return
        cols = self.apply_tree["columns"]
        lines = ["\t".join(self.apply_tree.set(iid, c) for c in cols) for iid in sel]
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))

    def _copy_apply_all(self) -> None:
        """Копирует всю таблицу в буфер обмена (TSV с заголовком)."""
        cols = self.apply_tree["columns"]
        header_line = "\t".join(
            self.apply_tree.heading(c)["text"].rstrip(" ▲▼") for c in cols
        )
        rows = [
            "\t".join(self.apply_tree.set(iid, c) for c in cols)
            for iid in self.apply_tree.get_children()
        ]
        self.clipboard_clear()
        self.clipboard_append(header_line + "\n" + "\n".join(rows))

    def _toggle_ensemble_ui(self):
        """Показывает/скрывает поля ансамбля в зависимости от чекбокса."""
        if self.use_ensemble.get():
            self._ens_inner.pack(fill="x", pady=(6, 0))
        else:
            self._ens_inner.pack_forget()

    def pick_ensemble_model(self):
        p = filedialog.askopenfilename(
            title="Модель 2 для ансамбля (.joblib)",
            filetypes=[("Joblib", "*.joblib"), ("All files", "*.*")],
        )
        if not p:
            return
        if not messagebox.askyesno(
            "Безопасность",
            "Файлы моделей (.joblib) содержат сериализованный Python-код.\n\n"
            "⚠ Загружай ТОЛЬКО модели из доверенных источников!\n\n"
            "Продолжить загрузку?",
        ):
            return
        self.ensemble_model2.set(p)
        self._trust_store.add_trusted(p)
        self.log_apply(f"[Ансамбль] Модель 2 выбрана: {p}")

    def log_apply(self, msg: str): self._log_to(self.apply_log, msg)

    def export_predictions(self):
        """Копирует последний файл с предсказаниями в выбранное место."""
        src = getattr(self, "_last_apply_out_path", None)
        if src is None or not Path(src).exists():
            messagebox.showinfo(
                "Экспорт предсказаний",
                "Сначала выполните классификацию — затем можно экспортировать результат."
            )
            return
        dst = filedialog.asksaveasfilename(
            title="Сохранить предсказания как…",
            defaultextension=".xlsx",
            initialfile=Path(src).name,
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not dst:
            return
        import shutil
        try:
            shutil.copyfile(src, dst)
            self.log_apply(f"Экспорт: {dst}")
        except Exception as e:
            messagebox.showerror("Ошибка экспорта", str(e))

    def pick_apply_file(self):
        p = filedialog.askopenfilename(
            title="Файл для классификации",
            filetypes=[
                ("Таблицы (xlsx, csv)", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Все файлы", "*.*"),
            ],
        )
        if not p:
            return
        self.apply_file.set(p)
        self._add_recent("apply_file", p)
        try:
            headers = read_headers(Path(p))
            self._refresh_combobox_values(headers)
            self.log_apply(f"Заголовки: {len(headers)} колонок.")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _pick_apply_from_path(self, p: str) -> None:
        """Выбирает файл классификации из недавних без диалога."""
        if not Path(p).exists():
            messagebox.showwarning("Файл не найден", f"Файл не найден:\n{p}")
            return
        self.apply_file.set(p)
        try:
            headers = read_headers(Path(p))
            self._refresh_combobox_values(headers)
            self.log_apply(f"Заголовки: {len(headers)} колонок.")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def pick_model(self):
        _init = str(MODEL_DIR) if MODEL_DIR.exists() else "."
        p = filedialog.askopenfilename(
            title="Модель .joblib",
            initialdir=_init,
            filetypes=[("Joblib", "*.joblib"), ("All files", "*.*")],
        )
        if not p:
            return
        # Предупреждение: joblib использует pickle — загружай только доверенные файлы
        if not messagebox.askyesno(
            "Безопасность",
            "Файлы моделей (.joblib) содержат сериализованный Python-код.\n\n"
            "⚠ Загружай ТОЛЬКО модели из доверенных источников!\n"
            "Вредоносный файл может выполнить произвольный код на твоём компьютере.\n\n"
            "Продолжить загрузку?",
        ):
            return
        self.model_file.set(p)
        self._trust_store.add_trusted(p)
        self._add_recent("model_file", p)
        try:
            pkg = self._load_model_pkg(p, log_fn=self.log_apply)
            cfg = upgrade_config_dict(pkg.get("config", {}))
            if cfg:
                self._apply_config_to_ui(cfg)
                self.log_apply("Модель загружена: config применён.")
            else:
                self.log_apply("Модель загружена, но config не найден (старый формат?).")
            _m = pkg.get("eval_metrics") or {}
            if _m.get("macro_f1") is not None:
                self.log_apply(
                    f"  Метрики: F1={_m['macro_f1']} | "
                    f"Accuracy={_m.get('accuracy', '—')} | "
                    f"n_train={_m.get('n_train', '—')} | "
                    f"Обучена: {str(_m.get('trained_at', '—'))[:10]}"
                )
            _thr = pkg.get("per_class_thresholds") or {}
            if _thr:
                self.populate_class_thresholds(_thr, model_thresholds=_thr)
                self.log_apply(f"  [Пороги] {len(_thr)} классов загружены в редактор.")
        except Exception as e:
            messagebox.showwarning("Модель", f"Не удалось прочитать config модели:\n{e}")

    def _pick_model_from_path(self, p: str) -> None:
        """Загружает модель из недавних без диалога (без consent-диалога — доверие уже было дано)."""
        if not Path(p).exists():
            messagebox.showwarning("Файл не найден", f"Файл не найден:\n{p}")
            return
        self.model_file.set(p)
        self._trust_store.add_trusted(p)
        try:
            pkg = self._load_model_pkg(p, log_fn=self.log_apply)
            cfg = upgrade_config_dict(pkg.get("config", {}))
            if cfg:
                self._apply_config_to_ui(cfg)
                self.log_apply("Модель загружена из недавних: config применён.")
            else:
                self.log_apply("Модель загружена из недавних, config не найден.")
            _m2 = pkg.get("eval_metrics") or {}
            if _m2.get("macro_f1") is not None:
                self.log_apply(
                    f"  Метрики: F1={_m2['macro_f1']} | "
                    f"Accuracy={_m2.get('accuracy', '—')} | "
                    f"n_train={_m2.get('n_train', '—')} | "
                    f"Обучена: {str(_m2.get('trained_at', '—'))[:10]}"
                )
            _thr2 = pkg.get("per_class_thresholds") or {}
            if _thr2:
                self.populate_class_thresholds(_thr2, model_thresholds=_thr2)
                self.log_apply(f"  [Пороги] {len(_thr2)} классов загружены в редактор.")
        except Exception as e:
            messagebox.showwarning("Модель", f"Не удалось прочитать config модели:\n{e}")


    def _apply_postprocess(
        self,
        snap: dict,
        _ars: "ApplyRunState",
    ):
        """Write summary, histogram, AL sheets; save workbook; return out_path."""
        # --- лист «Сводная таблица» ---
        self.after(0, lambda n=len(_ars.summary_dict): self.log_apply(
            f"Формирую сводную таблицу: {n} меток, {_ars.total_classified} строк классифицировано"
        ))
        ws_summ = _ars.wb_out.create_sheet(title="Сводная таблица")
        ws_summ.freeze_panes = "A2"
        ws_summ.append([
            "Описание проблемы",
            "Кол-во обращений",
            "% попадания",
            "Ср. уверенность",
            "Нуждается в ревью",
            "Пример 1",
            "Пример 2",
            "Пример 3",
        ])
        _data_sheet_name = _ars.ws_out.title
        # В Excel-гиперссылках одинарная кавычка в имени листа экранируется удвоением
        _safe_sheet_name = _data_sheet_name.replace("'", "''")
        _hyperlink_font = Font(color="0563C1", underline="single")
        _sorted_labels = sorted(_ars.summary_dict.items(), key=lambda x: x[1]["count"], reverse=True)
        for _lbl, _data in _sorted_labels:
            _cnt = _data["count"]
            _pct = round(_cnt / max(1, _ars.total_classified) * 100, 1)
            _avg_p = (
                round(_data["prob_sum"] / _data["prob_count"], 3)
                if _data["prob_count"] > 0 else ""
            )
            _rev = _data.get("review_count", 0)

            # Ячейка с меткой — гиперссылка на первую строку этой метки в листе данных
            cell_lbl = WriteOnlyCell(ws_summ, value=_lbl)
            _first_row = _data.get("first_row")
            if _first_row:
                cell_lbl.hyperlink = f"#'{_safe_sheet_name}'!A{_first_row}"
                cell_lbl.font = _hyperlink_font

            # Ячейки примеров — каждая гиперссылка на соответствующую строку в листе данных
            example_cells = []
            for _ex_text, _ex_row in _data["examples"]:
                cell_ex = WriteOnlyCell(ws_summ, value=_ex_text)
                cell_ex.hyperlink = f"#'{_safe_sheet_name}'!A{_ex_row}"
                cell_ex.font = _hyperlink_font
                example_cells.append(cell_ex)
            while len(example_cells) < 3:
                example_cells.append(WriteOnlyCell(ws_summ, value=""))

            ws_summ.append([cell_lbl, _cnt, _pct, _avg_p, _rev] + example_cells)

        # --- лист «Гистограмма уверенности» ---
        self.after(0, lambda: self.log_apply("Формирую гистограмму уверенности…"))
        ws_hist = _ars.wb_out.create_sheet(title="Гистограмма уверенности")
        ws_hist.freeze_panes = "A2"
        ws_hist.append(["Диапазон уверенности", "Кол-во предсказаний", "% от всех"])
        for _b in range(20):
            _bin_lbl = f"{_b / 20:.2f}–{(_b + 1) / 20:.2f}"
            _bin_cnt = _ars.conf_bins[_b]
            _bin_pct = round(_bin_cnt / max(1, _ars.total_classified) * 100, 1)
            ws_hist.append([_bin_lbl, _bin_cnt, _bin_pct])

        # ── Drift-диагностика: классы с высокой долей needs_review ────
        # Классы, где review_count/count >> среднего, сигнализируют о возможном
        # дрейфе данных (модель неуверена именно в этих категориях).
        _total_rev = sum(d["review_count"] for d in _ars.summary_dict.values())
        if _total_rev > 0 and _ars.total_classified > 0:
            _global_rev_rate = _total_rev / _ars.total_classified
            _high_rev = [
                (lbl, d["review_count"], d["count"],
                 d["review_count"] / d["count"])
                for lbl, d in _ars.summary_dict.items()
                if d["count"] >= 5
                   and d["review_count"] / d["count"] > _global_rev_rate * 1.5
            ]
            _high_rev.sort(key=lambda x: x[3], reverse=True)
            if _high_rev:
                _rev_lines = "\n".join(
                    f"    «{lbl}»  review={rv}/{cnt}  ({rate:.0%})"
                    for lbl, rv, cnt, rate in _high_rev[:10]
                )
                self.after(0, lambda lines=_rev_lines, gr=_global_rev_rate: self.log_apply(
                    f"[⚠ Drift] Классы с повышенной долей ревью "
                    f"(глобальная: {gr:.0%}):\n{lines}"
                ))
            else:
                self.after(0, lambda gr=_global_rev_rate: self.log_apply(
                    f"[✓ Drift] Доля ревью равномерна по классам "
                    f"(глобальная: {gr:.0%})"
                ))

        # --- лист «На разметку» (active learning) ---
        _ars.al_rows.sort(key=lambda _r: _r[0])
        _al_top = _ars.al_rows[:500]
        if _al_top:
            self.after(0, lambda n=len(_al_top): self.log_apply(
                f"[Активное обучение] Отобрано {n} строк для ручной разметки → лист «На разметку»"
            ))
            ws_al = _ars.wb_out.create_sheet(title="На разметку")
            ws_al.freeze_panes = "A2"
            _al_header = ["Текст", "Предсказание", "Уверенность", "Топ-3 классы"] + _ars.header_in
            ws_al.append(_al_header)
            for _al_prob, _al_pred, _al_top3, _al_text, _al_row_vals in _al_top:
                ws_al.append([_al_text, _al_pred, round(_al_prob, 4), _al_top3] + _al_row_vals)

        out_path = CLASS_DIR / f"{_ars.in_path.stem}_classified_{now_stamp()}.xlsx"
        self.after(0, lambda p=out_path: self.log_apply(f"Сохраняю файл: {p.name}…"))
        _ars.wb_out.save(out_path)

        # Ширины колонок — инжектируем в ZIP после сохранения (write_only не поддерживает column_dimensions)
        def _col_w(col: str) -> float:
            if col == _ars.pred_col:       return 35.0
            if col == "pred_proba":   return 15.0
            if col == "pred_top3":    return 50.0
            if col == "pred_needs_review": return 18.0
            if col == "pred_ambiguous": return 16.0
            return 22.0
        data_widths = [_col_w(c) for c in _ars.header_out]
        summ_widths = [45.0, 18.0, 12.0, 18.0, 18.0, 60.0, 60.0, 60.0]  # 8 колонок: метка + 5 числовых + 3 примера
        hist_widths = [22.0, 22.0, 18.0]
        _width_specs = [(0, data_widths), (1, summ_widths), (2, hist_widths)]
        if _al_top:
            _al_widths = [55.0, 35.0, 16.0, 55.0] + [22.0] * len(_ars.header_in)
            _width_specs.append((3, _al_widths))
        patch_xlsx_col_widths(out_path, _width_specs)
        return out_path

    def run_apply(self):
        self._right_tabs.select(self._log_tab_indices["apply"])
        with self._proc_lock:
            if self._processing:
                return
            self._processing = True

        if not validate_apply_preconditions(self):
            return

        # Снимаем все UI-параметры ДО запуска потока
        snap = build_validated_apply_snapshot(self)
        if snap is None:
            return

        # Пути читаем из snap (сделан до запуска потока, потокобезопасно)
        model_path = snap["model_file"]
        xlsx = snap["apply_file"]

        # Предупреждение о несогласованных порогах
        if snap.get("use_other_label"):
            _oth = snap.get("other_label_threshold", 0.50)
            _rev = snap.get("review_threshold", 0.60)
            if _oth > _rev:
                messagebox.showwarning(
                    "Несогласованные пороги",
                    f"Порог авто-замены на «Другое» ({_oth:.0%}) выше порога ревью ({_rev:.0%}).\n\n"
                    f"Это значит: строки, которые будут заменены на «Другое», "
                    f"не попадут в pred_needs_review=1 (ревью уже не нужно, раз метка заменена).\n\n"
                    f"Обычно порог замены ≤ порога ревью. Проверьте настройки.",
                )

        begin_long_task(
            cancel_event=self._cancel_event,
            run_button=self.btn_apply,
            run_button_busy_text="⏳ Классификация…",
            stop_button=self.btn_apply_stop,
            progress_var=self.apply_progress,
            status_var=self.apply_status,
            pct_var=self.apply_pct,
            phase_var=self.apply_phase,
            speed_var=self.apply_speed,
            eta_var=self.apply_eta,
            clear_summary_var=self.last_apply_summary,
            start_phase="Загрузка модели…",
            start_log=self.log_apply,
        )
        self.log_apply("==== APPLY START ====")

        t0 = time.time()

        _task_ui = prepare_long_task_ui(
            owner=self,
            progress_var=self.apply_progress,
            status_var=self.apply_status,
            pct_var=self.apply_pct,
            phase_var=self.apply_phase,
            speed_var=self.apply_speed,
            eta_var=self.apply_eta,
            run_button=self.btn_apply,
            run_button_idle_text="▶  Классифицировать",
            stop_button=self.btn_apply_stop,
            log_fn=self.log_apply,
        )
        _ctrl = _task_ui.controller
        ui_prog = _task_ui.ui_prog
        _lifecycle = _task_ui.lifecycle

        def worker():
            try:
                _ars = ApplyRunState()
                # Проверка: путь должен быть подтверждён пользователем в этой сессии.
                # Защищает от случаев, когда model_path установлен программно (конфиг/снапшот)
                # без явного прохождения через pick_model() с предупреждением безопасности.
                if not self._ensure_trusted_path(model_path, label="Модель 1"):
                    self.after(0, lambda: ui_prog(0.0, "Отменено"))
                    return

                self.after(0, lambda: ui_prog(3.0, "Загрузка модели…"))
                _log_apply_main = lambda msg: self.after(0, lambda m=msg: self.log_apply(m))
                pkg = self._load_model_pkg(model_path, log_fn=_log_apply_main)
                pipe = pkg["pipeline"]
                model_cfg: dict = upgrade_config_dict(pkg.get("config", {}))

                _clf_step = pipe.steps[-1][1] if hasattr(pipe, "steps") else pipe
                _clf_name = type(_clf_step).__name__
                _n_cls_loaded = len(pipe.classes_) if hasattr(pipe, "classes_") else "?"
                self.after(0, lambda n=_n_cls_loaded, c=_clf_name: self.log_apply(
                    f"Модель: {c} | {n} классов"
                ))

                # --- настроить SetFit device + callbacks если pipeline является SetFitClassifier ---
                _setfit_clf = find_setfit_classifier(pipe)
                if _setfit_clf is not None:
                    _setfit_clf.device = snap.get("sbert_device", "auto")
                    _setfit_clf.cache_dir = str(SBERT_LOCAL_DIR)
                    _setfit_clf.log_cb = lambda msg: self.after(0, lambda m=msg: self.log_apply(m))
                    _setfit_clf.progress_cb = None
                    self.after(0, lambda: self.log_apply(
                        f"[SetFit] classifier — device={snap.get('sbert_device','auto')} "
                        f"(прогресс модели — в логе)"
                    ))

                # --- настроить SBERT device + batch_size + callbacks если pipeline его содержит ---
                _sbert_in_pipe = find_sbert_in_pipeline(pipe)
                if _sbert_in_pipe is not None:
                    _sbert_in_pipe.device = snap.get("sbert_device", "auto")
                    # batch_size: пикленная модель могла быть обучена на другом железе;
                    # используем snap-значение (инициализировано из hw_profile на текущей машине).
                    _sbert_in_pipe.batch_size = snap.get("sbert_batch", self._hw.sbert_batch)
                    _sbert_in_pipe.cache_dir = str(SBERT_LOCAL_DIR)
                    _sbert_in_pipe.log_cb = lambda msg: self.after(0, lambda m=msg: self.log_apply(m))
                    # progress_cb = None при classify: прогресс-бар управляется счётчиком чанков,
                    # иначе SBERT шлёт значения 78-90% и бар прыгает вперёд-назад на каждом чанке.
                    # Статус SBERT виден в apply_log через log_cb.
                    _sbert_in_pipe.progress_cb = None
                    self.after(0, lambda: self.log_apply(
                        f"[SBERT] pipeline — device={snap.get('sbert_device','auto')} "
                        f"batch={snap.get('sbert_batch', self._hw.sbert_batch)}"
                    ))

                # --- загрузить модель 2 для ансамбля ---
                pipe2 = None
                has_proba2: bool = False
                classes_map2: Optional[Dict[int, int]] = None
                ens_w1: float = 1.0
                ens_w2: float = 0.0
                if snap.get("use_ensemble") and snap.get("ensemble_model2"):
                    ens_w1 = float(snap.get("ensemble_w1", 0.5))
                    ens_w2 = 1.0 - ens_w1
                    _ens2_path = snap["ensemble_model2"]
                    if not self._ensure_trusted_path(_ens2_path, label="Модель 2 (ансамбль)"):
                        self.after(0, lambda: ui_prog(0.0, "Отменено"))
                        return
                    self.after(0, lambda: ui_prog(4.0, "Загрузка модели 2 (ансамбль)…"))
                    _log_ens2 = lambda msg: self.after(0, lambda m=f"[Ансамбль] {msg}": self.log_apply(m))
                    pkg2 = self._load_model_pkg(_ens2_path, log_fn=_log_ens2)
                    pipe2 = pkg2["pipeline"]
                    has_proba2 = hasattr(pipe2, "predict_proba")
                    _sbert2 = find_sbert_in_pipeline(pipe2)
                    if _sbert2 is not None:
                        _sbert2.device = snap.get("sbert_device", "auto")
                        _sbert2.batch_size = snap.get("sbert_batch", self._hw.sbert_batch)
                        _sbert2.cache_dir = str(SBERT_LOCAL_DIR)
                        _sbert2.log_cb = lambda msg: self.after(0, lambda m=msg: self.log_apply(m))
                        _sbert2.progress_cb = None
                    self.after(0, lambda p=snap["ensemble_model2"], w1=ens_w1, w2=ens_w2: self.log_apply(
                        f"[Ансамбль] Модель 2 загружена: {p} | вес1={w1:.0%} вес2={w2:.0%}"
                    ))

                pred_col = snap["pred_col"]
                thr = snap["review_threshold"]
                # per-class пороги: кастомные (пользовательские) или из артефакта модели
                _per_cls_thr: dict = pkg.get("per_class_thresholds", {})
                if snap.get("use_ensemble") and snap.get("ensemble_model2"):
                    # При ансамбле объединяем пороги: берём из pkg2 только там, где нет в pkg1
                    _per_cls_thr = {**pkg2.get("per_class_thresholds", {}), **_per_cls_thr}
                # Кастомные пороги (редактируемые без переобучения) перекрывают модельные
                if snap.get("use_custom_class_thresholds"):
                    _custom = {
                        cls: float(var.get())
                        for cls, var in self._custom_class_thresholds.items()
                    }
                    _per_cls_thr = {**_per_cls_thr, **_custom}

                in_path = Path(xlsx)
                total_rows = estimate_total_rows([in_path])
                _tuned = tune_runtime_by_input_size(
                    input_bytes=in_path.stat().st_size if in_path.exists() else 0,
                    chunk=int(snap.get("chunk", 4000)),
                    sbert_batch=int(snap.get("sbert_batch", self._hw.sbert_batch)),
                    kmeans_batch=int(snap.get("kmeans_batch", self._hw.kmeans_batch)),
                )
                snap["chunk"] = _tuned["chunk"]
                snap["sbert_batch"] = _tuned["sbert_batch"]
                snap["kmeans_batch"] = _tuned["kmeans_batch"]
                self.after(0, lambda t=_tuned: self.log_apply(format_apply_autoprofile_log(t)))
                # RAM-aware chunk cap: prevent OOM on huge files
                try:
                    import psutil as _psutil
                    _mb_free = _psutil.virtual_memory().available // 1024 // 1024
                    # Rough estimate: each row ~2KB in memory (text + features + output)
                    _ram_cap = max(200, min(int(_mb_free // 2), 8000))
                    if snap["chunk"] > _ram_cap:
                        self.after(0, lambda old=snap["chunk"], cap=_ram_cap: self.log_apply(
                            f"[AutoChunk] RAM={_mb_free}МБ свободно → chunk: {old} → {cap}"
                        ))
                        snap["chunk"] = _ram_cap
                except Exception as _psutil_exc:
                    _log.debug("psutil RAM check failed, keeping chunk as-is: %s", _psutil_exc)

                wb_out = Workbook(write_only=True)
                ws_out = wb_out.create_sheet(title=in_path.stem[:31])

                start_ts = time.time()
                done = 0
                CHUNK = snap.get("chunk", 4000)   # авто по RAM + входному объёму
                buf_rows: List[List[Any]] = []
                buf_X: List[str] = []
                buf_ok: List[bool] = []

                # --- сводная таблица ---
                summary_dict: Dict[str, Dict] = {}   # label → {count, examples, prob_sum, prob_count, review_count, first_row}
                total_classified = 0
                conf_bins: List[int] = [0] * 20      # гистограмма: бины 0.00–0.05, 0.05–0.10, …, 0.95–1.00
                excel_row_num = 2  # строка 1 — заголовок; данные начинаются со строки 2

                # --- активное обучение: буфер неуверенных строк (сортируется и пишется в конце) ---
                _al_rows: List = []   # (prob, pred, top3, text, row_vals)

                has_proba = hasattr(pipe, "predict_proba")
                classes = list(pipe.classes_) if hasattr(pipe, "classes_") else None

                # --- температурное масштабирование вероятностей ---
                _temperature = float(pkg.get("temperature", 1.0))
                if abs(_temperature - 1.0) > 0.02 and has_proba:
                    self.after(0, lambda T=_temperature: self.log_apply(
                        f"[Калибровка] Температурное масштабирование T={T:.3f} активно"
                    ))

                # --- вычислить маппинг классов для ансамбля ---
                if pipe2 is not None and has_proba2 and classes is not None:
                    classes2 = list(pipe2.classes_) if hasattr(pipe2, "classes_") else None
                    if classes2 is not None:
                        if set(classes) == set(classes2):
                            # Полное совпадение: строим индексный маппинг classes2 → classes1
                            classes_map2 = {j2: classes.index(c) for j2, c in enumerate(classes2)}
                            self.after(0, lambda: self.log_apply(
                                f"[Ансамбль] Классы совпадают ({len(classes)} шт.) — ансамбль активен ✅"
                            ))
                        else:
                            common = set(classes) & set(classes2)
                            extra2 = set(classes2) - set(classes)
                            missing2 = set(classes) - set(classes2)
                            self.after(0, lambda e=extra2, m=missing2, c=common: self.log_apply(
                                f"⚠️  [Ансамбль] Классы не совпадают!\n"
                                f"    Лишние в модели 2: {sorted(e)}\n"
                                f"    Отсутствуют в модели 2: {sorted(m)}\n"
                                f"    Ансамбль применяется только для {len(c)} общих классов."
                            ))
                            classes_map2 = {
                                j2: classes.index(c)
                                for j2, c in enumerate(classes2)
                                if c in common
                            }

                header_in: List[str] = []
                header_out: List[str] = []
                i_desc_in: Optional[int] = None   # индекс колонки desc в header_in
                # idx_map вычисляется ОДИН РАЗ после чтения заголовка (не на каждый чанк)
                idx_map: Dict[str, Optional[int]] = {}

                def flush_chunk():
                    nonlocal buf_rows, buf_X, buf_ok, total_classified, excel_row_num, _al_rows

                    if not buf_rows:
                        return

                    _chunk_size = len(buf_rows)

                    preds = [""] * len(buf_rows)
                    prob_best: List[Optional[float]] = [None] * len(buf_rows)
                    prob_second: List[Optional[float]] = [None] * len(buf_rows)
                    top3: List[str] = [""] * len(buf_rows)

                    idxs = [i for i, ok in enumerate(buf_ok) if ok]
                    if idxs:
                        Xp = [buf_X[i] for i in idxs]
                        if has_proba:
                            proba = pipe.predict_proba(Xp)
                            # --- Температурное масштабирование ---
                            if abs(_temperature - 1.0) > 0.02:
                                _p_s = np.power(np.clip(proba, 1e-10, 1.0), 1.0 / _temperature)
                                _p_s /= _p_s.sum(axis=1, keepdims=True).clip(1e-10)
                                proba = _p_s
                            # --- Ансамбль: усреднение или мета-классификатор ---
                            if pipe2 is not None and has_proba2 and classes_map2 is not None:
                                proba2 = pipe2.predict_proba(Xp)
                                _meta_clf_ap = pkg.get("meta_learner")
                                if _meta_clf_ap is not None:
                                    # Мета-классификатор: align обоих к meta_classes, stack, predict
                                    import numpy as _np_ap
                                    _meta_cls_ap = pkg.get("meta_learner_classes", classes)
                                    _p1_ap = EnsemblePredictor.align_probabilities(
                                        proba, _meta_cls_ap, classes
                                    )
                                    _p2_ap = EnsemblePredictor.align_probabilities(
                                        proba2, _meta_cls_ap, classes2
                                    )
                                    _stacked_ap = _np_ap.hstack([_p1_ap, _p2_ap])
                                    _meta_proba = _meta_clf_ap.predict_proba(_stacked_ap)
                                    # Выровнять к classes (порядок pipe1)
                                    _meta_out_cls = list(_meta_clf_ap.classes_)
                                    proba = EnsemblePredictor.align_probabilities(
                                        _meta_proba, classes, _meta_out_cls
                                    )
                                else:
                                    proba2_aligned = EnsemblePredictor.align_probabilities(
                                        proba2, classes, classes2
                                    )
                                    proba = EnsemblePredictor.blend(proba, proba2_aligned, ens_w1)
                            # --- K-fold ансамбль: усредняем K моделей ---
                            _kfold_models_ap = pkg.get("kfold_models")
                            _kfold_classes_ap = pkg.get("kfold_classes")
                            if _kfold_models_ap and _kfold_classes_ap and classes is not None:
                                try:
                                    _kf_sum = np.zeros_like(proba)
                                    _kf_n = 0
                                    for _kfm in _kfold_models_ap:
                                        _kfp = _kfm.predict_proba(Xp)
                                        _kfp_aligned = EnsemblePredictor.align_probabilities(
                                            _kfp, classes, list(_kfm.classes_)
                                        )
                                        _kf_sum += _kfp_aligned
                                        _kf_n += 1
                                    if _kf_n > 0:
                                        # Взвешенное среднее: 50% основная модель + 50% K-fold
                                        proba = 0.5 * proba + 0.5 * (_kf_sum / _kf_n)
                                except Exception as _kf_exc:
                                    _log.debug("kfold ensemble averaging failed, using main model: %s", _kf_exc)
                            # --- Иерархическое усиление (group_model) ---
                            _group_model_ap = pkg.get("group_model")
                            _c2g_ap = pkg.get("class_to_group", {})
                            if _group_model_ap is not None and _c2g_ap and classes is not None:
                                try:
                                    _gp = _group_model_ap.predict_proba(Xp)
                                    _g_classes = [str(g) for g in _group_model_ap.classes_]
                                    for _ki in range(len(Xp)):
                                        for _ci, _cls in enumerate(classes):
                                            _grp = str(_c2g_ap.get(_cls, -1))
                                            if _grp in _g_classes:
                                                _gi = _g_classes.index(_grp)
                                                proba[_ki, _ci] *= float(_gp[_ki, _gi])
                                        _row_sum = proba[_ki].sum()
                                        if _row_sum > 0:
                                            proba[_ki] /= _row_sum
                                except Exception as _grp_exc:
                                    _log.debug("group_model hierarchical boost failed, using base proba: %s", _grp_exc)
                            # Сохраняем top-K id для последующего LLM-ре-ранка неуверенных.
                            _rerank_topk_by_i: Dict[int, List[str]] = {}
                            _rerank_topk_k = max(2, min(5, int(snap.get("llm_rerank_top_k", 3))))
                            for k, i in enumerate(idxs):
                                rowp = proba[k]
                                bi = int(rowp.argmax())
                                preds[i] = str(classes[bi]) if classes is not None else ""
                                prob_best[i] = float(rowp[bi])
                                top_idx = heapq.nlargest(3, range(len(rowp)), key=lambda j: rowp[j])
                                top3[i] = "; ".join([f"{classes[j]}:{float(rowp[j]):.3f}" for j in top_idx])
                                if len(top_idx) >= 2:
                                    prob_second[i] = float(rowp[top_idx[1]])
                                if classes is not None:
                                    _topk_ids = heapq.nlargest(_rerank_topk_k, range(len(rowp)),
                                                               key=lambda j: rowp[j])
                                    _rerank_topk_by_i[i] = [str(classes[j]) for j in _topk_ids]

                            # --- LLM ре-ранк неуверенных предсказаний ---
                            if snap.get("use_llm_rerank") and classes is not None:
                                _low = float(snap.get("llm_rerank_low", 0.50))
                                _high = float(snap.get("llm_rerank_high", 0.70))
                                _rerank_idx = [
                                    i for i in idxs
                                    if prob_best[i] is not None
                                    and _low <= float(prob_best[i]) < _high
                                ]
                                if _rerank_idx:
                                    try:
                                        from llm_reranker import rerank_top_k
                                        _class_exs = pkg.get("class_examples", {}) or {}
                                        _texts_r = [str(buf_X[i])[:900] for i in _rerank_idx]
                                        _cands_r = [_rerank_topk_by_i.get(i, [preds[i]])
                                                    for i in _rerank_idx]
                                        _fb_r = [preds[i] for i in _rerank_idx]
                                        _new_labels = rerank_top_k(
                                            _texts_r, _cands_r, _fb_r,
                                            provider=str(snap.get("llm_provider", "openai")),
                                            model=str(snap.get("llm_model", "")),
                                            api_key=str(snap.get("llm_api_key", "")),
                                            class_examples=_class_exs,
                                            timeout_sec=float(snap.get("llm_timeout", 20.0)),
                                            max_retries=1,
                                            log_fn=lambda m: self.after(0, lambda msg=m: self.log_apply(msg)),
                                        )
                                        _n_changed = 0
                                        for _i_row, _new in zip(_rerank_idx, _new_labels):
                                            if _new and _new != preds[_i_row]:
                                                preds[_i_row] = _new
                                                _n_changed += 1
                                        self.after(0, lambda n=_n_changed, tot=len(_rerank_idx):
                                            self.log_apply(
                                                f"[LLM-ре-ранк] изменено {n}/{tot} меток"
                                            ))
                                    except Exception as _rr_exc:
                                        self.after(0, lambda e=_rr_exc: self.log_apply(
                                            f"[LLM-ре-ранк] ошибка: {type(e).__name__}: {e}"
                                        ))
                        else:
                            p_labels = pipe.predict(Xp)
                            for k, i in enumerate(idxs):
                                preds[i] = str(p_labels[k])

                    # --- авто-замена на «Другое» при низкой уверенности ---
                    # Выполняется ДО вычисления needs_review: заменённые строки ревью не требуют.
                    # Если снапшот включает use_per_class_other_threshold И модель хранит
                    # per_class_thresholds — используем per-class порог; иначе глобальный.
                    _replaced_with_other: set = set()
                    if snap.get("use_other_label") and has_proba:
                        _oth_thr = snap.get("other_label_threshold", 0.50)
                        _oth_lbl = snap.get("other_label_text", DEFAULT_OTHER_LABEL)
                        _use_pc_other = bool(snap.get("use_per_class_other_threshold", False)) and bool(_per_cls_thr)
                        for i in idxs:
                            if prob_best[i] is None:
                                continue
                            _thr_i = _per_cls_thr.get(preds[i], _oth_thr) if _use_pc_other else _oth_thr
                            if prob_best[i] < _thr_i:
                                preds[i] = _oth_lbl
                                _replaced_with_other.add(i)

                    # needs_review вычисляется ПОСЛЕ авто-замены:
                    # если строка заменена на «Другое» — ревью не нужно (метка уже определена)
                    # per-class порог используется там, где он есть; иначе — глобальный thr
                    _use_ambig = snap.get("use_ambiguity_detector", False) and has_proba
                    _ambig_eps = float(snap.get("ambiguity_epsilon", 0.07))
                    needs_review = []
                    for i in range(len(buf_rows)):
                        if prob_best[i] is None or i in _replaced_with_other:
                            needs_review.append(0)
                        elif prob_best[i] < _per_cls_thr.get(preds[i], thr):
                            needs_review.append(1)
                        elif _use_ambig and prob_second[i] is not None and (prob_best[i] - prob_second[i]) < _ambig_eps:
                            needs_review.append(1)
                        else:
                            needs_review.append(0)

                    for i, row_vals in enumerate(buf_rows):
                        out_row = []
                        for col in header_out:
                            if col == pred_col:
                                out_row.append(preds[i])
                            elif col == "pred_proba":
                                out_row.append(prob_best[i])
                            elif col == "pred_top3":
                                out_row.append(top3[i])
                            elif col == "pred_needs_review":
                                out_row.append(needs_review[i])
                            elif col == "pred_ambiguous":
                                _amb = (
                                    1
                                    if (
                                        prob_best[i] is not None
                                        and prob_second[i] is not None
                                        and (prob_best[i] - prob_second[i]) < float(snap.get("ambiguity_epsilon", 0.07))
                                    )
                                    else 0
                                )
                                out_row.append(_amb)
                            else:
                                j = idx_map.get(col)
                                out_row.append(row_vals[j] if (j is not None and 0 <= j < len(row_vals)) else None)
                        ws_out.append(out_row)
                        current_excel_row = excel_row_num
                        excel_row_num += 1

                        # --- обновить сводную таблицу ---
                        lbl = preds[i]
                        if lbl:
                            total_classified += 1
                            if lbl not in summary_dict:
                                # prob_sum/prob_count — running average вместо списка
                                # (чтобы не расти O(N) при миллионах строк одного класса)
                                summary_dict[lbl] = {
                                    "count": 0, "examples": [],
                                    "prob_sum": 0.0, "prob_count": 0,
                                    "review_count": 0,
                                    "first_row": current_excel_row,  # номер первой строки метки в листе данных
                                }
                            summary_dict[lbl]["count"] += 1
                            # собираем до 3 примеров: (текст, номер_строки_excel)
                            if len(summary_dict[lbl]["examples"]) < 3:
                                if i_desc_in is not None and i_desc_in < len(row_vals):
                                    ex_raw = str(row_vals[i_desc_in] or "").strip()[:220]
                                    ex_text = f"[Описание] {ex_raw}" if ex_raw else ""
                                else:
                                    ex_raw = (buf_X[i] or "").replace("\n", " ")[:220]
                                    ex_text = f"[текст] {ex_raw}" if ex_raw else ""
                                if ex_text:
                                    summary_dict[lbl]["examples"].append((ex_text, current_excel_row))
                            if prob_best[i] is not None:
                                summary_dict[lbl]["prob_sum"] += float(prob_best[i])
                                summary_dict[lbl]["prob_count"] += 1
                                if needs_review[i]:
                                    summary_dict[lbl]["review_count"] += 1
                                conf_bins[max(0, min(19, int(float(prob_best[i]) * 20)))] += 1

                    _n_preds    = len(idxs)
                    _n_empty    = _chunk_size - _n_preds
                    _n_replaced = len(_replaced_with_other)
                    _n_review   = sum(needs_review)
                    _d_now      = done
                    self.after(0, lambda d=_d_now, tot=total_rows, cs=_chunk_size,
                               np=_n_preds, ne=_n_empty, nr=_n_replaced, rv=_n_review:
                        self.log_apply(
                            f"  Чанк @{d}/{tot}: строк={cs} | "
                            f"предсказано={np} | пустых={ne} | "
                            f"→«Другое»={nr} | на ревью={rv}"
                        )
                    )

                    # --- Активное обучение: собираем строки с низкой уверенностью ---
                    _al_thr = float(snap.get("active_learning_threshold", 0.70))
                    for i in range(len(buf_rows)):
                        if prob_best[i] is not None and float(prob_best[i]) < _al_thr:
                            _al_rows.append((
                                float(prob_best[i]),
                                str(preds[i]),
                                str(top3[i]),
                                str(buf_X[i])[:300],
                                list(buf_rows[i]),
                            ))
                    # Ограничиваем рост буфера в памяти: при >5000 оставляем 2000 наименее уверенных
                    if len(_al_rows) > 5000:
                        _al_rows.sort(key=lambda _r: _r[0])
                        del _al_rows[2000:]

                    buf_rows, buf_X, buf_ok = [], [], []

                with open_tabular(in_path) as it:
                    header_in = ["" if h is None else str(h) for h in next(it)]
                    i_desc_in = idx_of(header_in, snap["desc_col"]) if snap["desc_col"] else None

                    # --- Fix 8: проверка совместимости колонок модели и файла ---
                    _model_text_keys = ("desc_col", "call_col", "chat_col",
                                        "summary_col", "answer_short_col", "answer_full_col")
                    _trained_cols = [model_cfg.get(k, "") for k in _model_text_keys]
                    _trained_cols = [c for c in _trained_cols if c]  # убираем пустые
                    _missing = [c for c in _trained_cols if c not in header_in]
                    if _missing:
                        self.after(0, lambda mc=_missing: self.log_apply(
                            f"⚠️  Колонки из обучения отсутствуют в файле: {mc}\n"
                            f"    Убедись, что в UI выбраны правильные колонки для этого файла."
                        ))

                    extra_cols = [pred_col, "pred_proba", "pred_top3", "pred_needs_review", "pred_ambiguous"]
                    header_out = header_in[:]
                    for c in extra_cols:
                        if c not in header_out:
                            header_out.append(c)

                    has_any = any(
                        col and col in header_in
                        for col in [
                            snap["desc_col"], snap["call_col"], snap["chat_col"],
                            (snap["summary_col"] if snap["use_summary"] else ""),
                            snap["ans_short_col"], snap["ans_full_col"],
                        ]
                    )
                    if not has_any:
                        raise FeatureBuildError("Не выбраны/не найдены колонки текста в файле (проверь заголовки).")

                    # Форматирование: freeze + autofilter (устанавливаются до первого append)
                    ws_out.freeze_panes = "A2"
                    ws_out.auto_filter.ref = f"A1:{get_column_letter(len(header_out))}1"

                    ws_out.append(header_out)
                    # Вычисляем один раз — closure flush_chunk читает из enclosing scope
                    # O(N) вместо O(N²): прямое построение словаря позиций без linear index()
                    idx_map.update({col: i for i, col in enumerate(header_in)})

                    for row in it:
                        row_vals = list(row)
                        xfeat = self._row_to_feature_text(row_vals, header_in, snap, header_index=idx_map)
                        ok = bool(xfeat.strip())

                        buf_rows.append(row_vals)
                        buf_X.append(xfeat)
                        buf_ok.append(ok)

                        done += 1
                        if len(buf_rows) >= CHUNK:
                            flush_chunk()
                            if self._cancel_event.is_set():
                                raise InterruptedError()
                        if done % 500 == 0 or done == 1:
                            pct = 8.0 + 80.0 * done / float(total_rows)
                            msg = f"Классификация: {done}/{total_rows} | {fmt_speed(start_ts, done)} | {fmt_eta(start_ts, done, total_rows)}"
                            self.after(0, lambda pct=pct, msg=msg: ui_prog(pct, msg))

                    flush_chunk()

                # ── Постобработка (сводка / гистограмма / AL / сохранение) ──────────
                _ars.wb_out = wb_out
                _ars.ws_out = ws_out
                _ars.summary_dict = summary_dict
                _ars.total_classified = total_classified
                _ars.conf_bins = conf_bins
                _ars.al_rows = _al_rows
                _ars.header_in = header_in
                _ars.header_out = header_out
                _ars.in_path = in_path
                _ars.pred_col = pred_col
                out_path = self._apply_postprocess(snap, _ars)

                elapsed = time.time() - t0

                self._last_apply_out_path = out_path

                def done_ui():
                    ui_prog(100.0, "Готово ✅")
                    self.apply_speed.set("")
                    self.apply_eta.set(f"Итого: {elapsed:.0f}с")

                    # ── Распределение уверенности ─────────────────────────────
                    _total_conf = sum(conf_bins)
                    if _total_conf > 0:
                        _high = sum(conf_bins[16:]) / _total_conf      # ≥ 0.80
                        _mid  = sum(conf_bins[12:16]) / _total_conf    # 0.60–0.79
                        _low  = sum(conf_bins[:12]) / _total_conf      # < 0.60
                        _avg_c = sum(
                            (i * 0.05 + 0.025) * conf_bins[i] for i in range(20)
                        ) / _total_conf
                        self.log_apply(
                            f"[Уверенность] Среднее: {_avg_c:.3f}"
                            f" | ≥0.80: {_high*100:.1f}%"
                            f" | 0.60–0.79: {_mid*100:.1f}%"
                            f" | <0.60: {_low*100:.1f}%"
                        )

                    # ── Топ-10 меток ──────────────────────────────────────────
                    _sorted_labels = sorted(
                        summary_dict.items(),
                        key=lambda kv: kv[1]["count"],
                        reverse=True,
                    )
                    _top_n = min(10, len(_sorted_labels))
                    if _top_n > 0:
                        _lbl_lines = "\n".join(
                            f"  {i+1:2d}. «{lbl}»: {d['count']}"
                            f" ({d['count'] / max(total_classified, 1) * 100:.1f}%)"
                            f"  ср.увер={d['prob_sum'] / max(d['prob_count'], 1):.3f}"
                            for i, (lbl, d) in enumerate(_sorted_labels[:_top_n])
                        )
                        self.log_apply(f"[Результат] Топ-{_top_n} меток:\n{_lbl_lines}")

                    self.log_apply(
                        f"Готово ✅ | time={elapsed:.1f}s | rows={total_rows} | "
                        f"labels={len(summary_dict)} | saved={out_path}"
                    )
                    self.last_apply_summary.set(
                        f"saved={out_path.name} | rows={total_rows} | "
                        f"labels={len(summary_dict)} | col={pred_col}"
                    )

                    for it_ in self.apply_tree.get_children():
                        self.apply_tree.delete(it_)
                    _sorted = sorted(summary_dict.items(), key=lambda x: x[1]["count"], reverse=True)
                    for _lbl, _data in _sorted:
                        _cnt = _data["count"]
                        _pct = f"{round(_cnt / max(1, total_classified) * 100, 1)}%"
                        _avg_p = (
                            f"{_data['prob_sum'] / _data['prob_count']:.3f}"
                            if _data["prob_count"] > 0 else "—"
                        )
                        _examples = " | ".join(
                            ex[0] if isinstance(ex, tuple) else ex
                            for ex in _data["examples"]
                        )
                        self.apply_tree.insert("", "end", values=(
                            _lbl, _cnt, _pct, _avg_p, _examples
                        ))

                    # Draw label distribution bar-chart
                    _dist = {lbl: d["count"] for lbl, d in summary_dict.items()}
                    self._draw_result_chart(_dist, total_classified)

                    _lifecycle.complete()

                self.after(0, done_ui)

            except InterruptedError:
                def cancelled_apply(done=done):
                    _lifecycle.cancelled(
                        ui_prog=ui_prog,
                        log_message=f"⏹ Классификация отменена пользователем (обработано {done} строк)",
                    )
                self.after(0, cancelled_apply)

            except (ModelLoadError, FeatureBuildError, PredictPipelineError) as e:
                _tb = _traceback.format_exc()
                _titles = {
                    ModelLoadError:       "Ошибка загрузки модели",
                    FeatureBuildError:    "Ошибка признаков",
                    PredictPipelineError: "Ошибка предсказания",
                }
                _title = _titles.get(type(e), "Ошибка классификации")
                def err(e=e, tb=_tb, title=_title):
                    env = ErrorEnvelope.from_exception(
                        e,
                        error_code="APPLY_PIPELINE_ERROR",
                        stage="run_apply",
                        hint="Проверьте модель, входные колонки и настройки порогов.",
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка",
                        envelope=env,
                        traceback_text=tb,
                    )
                    messagebox.showerror(title, str(e))
                self.after(0, err)

            except Exception as e:
                _tb = _traceback.format_exc()
                def err(e=e, tb=_tb):
                    env = ErrorEnvelope.from_exception(
                        e,
                        error_code="APPLY_UNEXPECTED",
                        stage="run_apply",
                        hint="Проверьте UI-лог и traceback; повторите с тем же файлом для диагностики.",
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка",
                        envelope=env,
                        traceback_text=tb,
                    )
                    _brief = str(e)[:300] + ("…" if len(str(e)) > 300 else "")
                    messagebox.showerror(
                        "Ошибка классификации",
                        f"{_brief}\n\nПодробности смотри в логе классификации.",
                    )
                self.after(0, err)

        threading.Thread(target=worker, daemon=True).start()
