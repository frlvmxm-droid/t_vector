# -*- coding: utf-8 -*-
"""
app_cluster.py — ClusterTabMixin: вкладка «Кластеризация» и логика кластеризации.

Содержит:
  • _build_cluster_tab() — построение UI вкладки
  • run_cluster()        — фоновый поток кластеризации
  • Вспомогательные методы (пикер файла)
"""
from __future__ import annotations

import re
import threading
import time
import traceback as _traceback
from collections import Counter
from dataclasses import dataclass as _dc_dataclass, field as _dc_field
from pathlib import Path
from types import MappingProxyType
from typing import Any, Dict, List, Mapping, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

import numpy as _np
from sklearn.cluster import MiniBatchKMeans
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import normalize as _sk_normalize

from config.ml_constants import KMEANS_BATCH_SIZE, HDBSCAN_NOISE_BATCH_SIZE
from cluster_algo_strategy import (
    cuda_available as _cuda_available,
    cuml_kmeans_available as _cuml_kmeans_available,
    gpu_kmeans as _strategy_gpu_kmeans,
    gpu_umap as _gpu_umap,
)

# Test-patchable overrides for GPU/CPU branch selection. Default None means
# "defer to cluster_algo_strategy's own detection". Contract-tests monkeypatch
# these directly on the `app_cluster` module.
_CUML_KMEANS: Optional[bool] = None
_cuml_cluster = None


def _gpu_kmeans(
    n_clusters: int,
    random_state: int = 42,
    batch_size: int = KMEANS_BATCH_SIZE,
    n_init: int = 10,
    init: str = "k-means++",
    **mb_kwargs,
):
    """Thin wrapper around ``cluster_algo_strategy.gpu_kmeans``.

    Respects module-level overrides (`_CUML_KMEANS`, `_cuml_cluster`) so
    contract-tests can pin the branch without touching the strategy module.
    """
    max_iter = int(mb_kwargs.pop("max_iter", 300))
    if _CUML_KMEANS is True and _cuml_cluster is not None:
        return _cuml_cluster.KMeans(
            n_clusters=n_clusters,
            random_state=random_state,
            init=init if init != "k-means++" else "scalable-k-means++",
            n_init=n_init,
            max_iter=max_iter,
        )
    if _CUML_KMEANS is False:
        return MiniBatchKMeans(
            n_clusters=n_clusters,
            random_state=random_state,
            batch_size=batch_size,
            n_init=n_init,
            init=init,
            max_iter=max_iter,
            **mb_kwargs,
        )
    return _strategy_gpu_kmeans(
        n_clusters=n_clusters,
        random_state=random_state,
        batch_size=batch_size,
        n_init=n_init,
        init=init,
        max_iter=max_iter,
        **mb_kwargs,
    )
from cluster_runtime_service import (
    cleanup_cluster_runtime,
    try_mark_processing,
    clear_processing,
    tune_cluster_runtime_for_input,
)


from constants import CLUST_DIR, RUSSIAN_STOP_WORDS, NOISE_TOKENS, NOISE_PHRASES, now_stamp
from config import DEFAULT_COLS, SBERT_MODELS_LIST, SBERT_DEFAULT, NEURAL_MODELS_LIST
from text_utils import normalize_text
from excel_utils import (
    read_headers, idx_of, estimate_total_rows,
    fmt_eta, fmt_speed, open_tabular, patch_xlsx_col_widths,
)
from feature_builder import build_feature_text, choose_row_profile_weights
from ml_core import (
    PhraseRemover, extract_cluster_keywords,
    extract_cluster_keywords_from_labels,
    extract_cluster_keywords_ctfidf,
    SBERTVectorizer, DeBERTaVectorizer, make_neural_vectorizer, Lemmatizer,
)
from exceptions import FeatureBuildError, ModelLoadError
from cluster_model_loader import (
    ensure_cluster_model_trusted,
)
from model_loader import make_tkinter_confirm_fn
from cluster_incremental_service import load_and_apply_incremental_model
from cluster_export_service import build_cluster_model_bundle
from app_cluster_service import (
    LLMClient,
    ClusterReasonBuilder,
    ClusterModelPersistence,
    ClusterElbowSelector,
    LLMSnapshotDecryptError,
    decrypt_api_key_from_snapshot,
    resolve_api_key,
)
from cluster_ui_builder import build_cluster_primary_sections
from cluster_run_coordinator import prepare_cluster_run_context
from app_cluster_workflow import validate_cluster_preconditions
from cluster_state_adapter import build_cluster_runtime_snapshot
from app_cluster_pipeline import (
    build_t5_source_text,
    prepare_inputs,
    build_vectors,
    run_clustering,
    postprocess_clusters,
    export_cluster_outputs,
)
from snap_utils import freeze_snap as _shared_freeze_snap
from task_runner import ErrorEnvelope, OperationLifecycle, begin_long_task, prepare_long_task_ui
from app_logger import get_logger
from ui_theme import BG, FG, ENTRY_BG, ACCENT, MUTED, BORDER

_log = get_logger(__name__)
from ui_widgets import Tooltip, CollapsibleSection


# endregion


def _analyze_cluster_input(files_snapshot: list, col_names: dict) -> dict:
    """Reads input files and collects text statistics for auto-detect heuristics."""
    SAMPLE_LIMIT = 3000
    total_rows = 0
    n_text_rows = 0
    text_lengths: list = []
    has_dialog_roles = False
    has_chatbot = False
    col_hits: dict = {k: 0 for k in col_names}
    sampled_words: list = []
    _ROLE_PAT = re.compile(r'\b(ОПЕРАТОР|OPERATOR|КЛИЕНТ|CLIENT)\b', re.IGNORECASE)
    _CHATBOT_PAT = re.compile(r'\b(CHATBOT|БОТ|BOT|SYSTEM|СИСТЕМА)\b', re.IGNORECASE)

    for file_str in files_snapshot:
        in_path = Path(file_str)
        with open_tabular(in_path) as it:
            raw_hdr = next(it)
            header = ["" if h is None else str(h) for h in raw_hdr]
            ci: dict = {}
            for key, name in col_names.items():
                ci[key] = idx_of(header, name) if name else None
            for row in it:
                total_rows += 1
                row_vals = list(row)
                if total_rows <= SAMPLE_LIMIT:
                    parts: list = []
                    for key, idx in ci.items():
                        if idx is None or idx >= len(row_vals):
                            continue
                        val = str(row_vals[idx]) if row_vals[idx] is not None else ""
                        val = val.strip()
                        if not val:
                            continue
                        col_hits[key] += 1
                        parts.append(val[:600])
                        if _ROLE_PAT.search(val):
                            has_dialog_roles = True
                        if _CHATBOT_PAT.search(val):
                            has_chatbot = True
                    if parts:
                        combined = " ".join(parts)
                        text_lengths.append(len(combined))
                        n_text_rows += 1
                        if len(sampled_words) < 200_000:
                            sampled_words.extend(
                                w.lower() for w in combined.split() if len(w) > 2
                            )

    return {
        "total_rows": total_rows,
        "n_text_rows": n_text_rows,
        "text_lengths": text_lengths,
        "has_dialog_roles": has_dialog_roles,
        "has_chatbot": has_chatbot,
        "col_hits": col_hits,
        "sampled_words": sampled_words,
        "SAMPLE_LIMIT": SAMPLE_LIMIT,
    }

def _compute_cluster_heuristics(analysis: dict) -> tuple:
    """Computes recommended clustering parameters from file analysis results.
    Returns (params, reasons).
    """
    total_rows       = analysis["total_rows"]
    n_text_rows      = analysis["n_text_rows"]
    text_lengths     = analysis["text_lengths"]
    has_dialog_roles = analysis["has_dialog_roles"]
    has_chatbot      = analysis["has_chatbot"]
    col_hits         = analysis["col_hits"]
    sampled_words    = analysis["sampled_words"]
    SAMPLE_LIMIT     = analysis["SAMPLE_LIMIT"]
    sample_size   = min(total_rows, SAMPLE_LIMIT)
    avg_len       = sum(text_lengths) / max(1, len(text_lengths))
    max_len       = max(text_lengths) if text_lengths else 0
    vocab_sample  = len(set(sampled_words))
    # Экстраполируем словарь на весь датасет (log-масштаб Хипа)
    vocab_est = int(vocab_sample * (total_rows / max(1, sample_size)) ** 0.6)
    # Плотность: доля строк с текстом (в выборке)
    text_density  = n_text_rows / max(1, sample_size)
    # Активные текстовые колонки
    active_cols   = [k for k, v in col_hits.items() if v > sample_size * 0.05]
    has_long_text = avg_len > 500  # длинные транскрипты

    # ── Проверка: есть ли текст по выбранным колонкам ─────────────
    if n_text_rows == 0:
        raise ValueError(
            "В выборке не найдено строк с текстом по указанным колонкам.\n"
            "Проверьте, что названия колонок (Описание, Звонок, Чат …) "
            "совпадают с заголовками файла."
        )

    # ── Наличие GPU (влияет на рекомендацию векторизации) ──────────
    has_gpu = _cuda_available()

    reasons: List[str] = []
    params: Dict[str, Any] = {}

    # ── 1. K кластеров ─────────────────────────────────────────────
    # Эвристика: K ≈ sqrt(n / 2), скорректированная по размеру и разнообразию
    if total_rows < 300:
        k = max(3, min(8, total_rows // 20))
    elif total_rows < 2_000:
        k = max(5, min(20, int(total_rows ** 0.5 / 1.5)))
    elif total_rows < 15_000:
        k = max(10, min(40, int(total_rows ** 0.45)))
    elif total_rows < 100_000:
        k = max(15, min(60, int(total_rows ** 0.38)))
    else:
        k = max(20, min(80, int(total_rows ** 0.32)))
    params["k_clusters"] = k
    reasons.append(
        f"K={k}  (датасет: {total_rows} строк, текстовых: "
        f"{int(text_density * 100)}%)"
    )

    # ── 2. Векторизация (с учётом GPU) ────────────────────────────
    # SBERT/Combo без GPU на больших объёмах слишком медленны —
    # снижаем пороги при отсутствии CUDA.
    if total_rows < 1_000:
        params["cluster_vec_mode"] = "sbert"
        reasons.append(
            "Векторизация: SBERT  "
            "(маленький датасет → семантика важнее скорости)"
        )
    elif total_rows < (20_000 if has_gpu else 5_000):
        params["cluster_vec_mode"] = "combo"
        _gpu_note = "" if has_gpu else "  [GPU не найден — порог снижен до 5k]"
        reasons.append(
            f"Векторизация: Комбо TF-IDF + SBERT{_gpu_note}"
        )
    else:
        params["cluster_vec_mode"] = "tfidf"
        _gpu_note = "  (CUDA не обнаружена)" if not has_gpu and total_rows >= 5_000 else ""
        reasons.append(
            f"Векторизация: TF-IDF{_gpu_note}  "
            "(большой датасет → быстро, надёжно)"
        )

    vec = params["cluster_vec_mode"]

    # ── 3. Алгоритм кластеризации ──────────────────────────────────
    # HDBSCAN лучше KMeans для малых датасетов с неравномерными
    # кластерами: сам находит K и помечает выбросы (-1).
    if total_rows <= 5_000 and has_long_text and vec in ("sbert", "combo"):
        params["cluster_algo"] = "hdbscan"
        reasons.append(
            "Алгоритм: HDBSCAN  "
            "(малый датасет + длинные тексты → лучше KMeans, "
            "автоматически находит K и выбросы)"
        )
    else:
        params["cluster_algo"] = "kmeans"
        reasons.append(
            f"Алгоритм: KMeans  "
            f"({'GPU-ускорение через cuML' if _cuml_kmeans_available() else 'CPU'})"
        )

    algo = params["cluster_algo"]

    # ── 4. Авто-подбор K (только для KMeans) ──────────────────────
    if algo == "kmeans":
        if total_rows < 50_000:
            params["use_elbow"] = True
            if total_rows < 3_000:
                params["k_score_method"] = "silhouette"
                reasons.append(
                    "Авто-подбор K: Silhouette  "
                    "(малый датасет → наиболее точная метрика)"
                )
            else:
                params["k_score_method"] = "calinski"
                reasons.append(
                    "Авто-подбор K: Calinski-Harabász  "
                    "(быстрее Silhouette, хорошо для KMeans)"
                )
        else:
            params["use_elbow"] = False
            reasons.append(
                "Авто-подбор K: выключён  "
                "(датасет большой → перебор K занял бы слишком долго)"
            )
    else:
        # HDBSCAN сам определяет K
        params["use_elbow"] = False

    # ── 5. n_init ──────────────────────────────────────────────────
    if total_rows < 5_000:
        params["n_init_cluster"] = 15
    elif total_rows < 30_000:
        params["n_init_cluster"] = 10
    else:
        params["n_init_cluster"] = 5
    if algo == "kmeans":
        reasons.append(
            f"n_init={params['n_init_cluster']}  "
            "(баланс надёжности и скорости)"
        )

    # ── 6. Косинусная метрика — всегда для текста ─────────────────
    params["use_cosine_cluster"] = True

    # ── 7. UMAP ────────────────────────────────────────────────────
    if vec in ("sbert", "combo") and total_rows > 2_000:
        params["use_umap"] = True
        # dim: минимум 15 для надёжной топологии, максимум 50
        umap_dim = max(15, min(50, k * 2))
        params["umap_n_components"] = umap_dim
        # n_neighbors: локальная структура для малых данных,
        # глобальная — для больших
        umap_nn = max(10, min(50, total_rows // 300))
        params["umap_n_neighbors"] = umap_nn
        reasons.append(
            f"UMAP: включён, dim={umap_dim}, n_neighbors={umap_nn}  "
            f"({'GPU cuML' if _cuml_umap_available() else 'CPU'})"
        )
    else:
        params["use_umap"] = False
        if vec == "tfidf":
            reasons.append("UMAP: выключён  (TF-IDF → SVD достаточно)")
        elif total_rows <= 2_000:
            reasons.append(
                "UMAP: выключён  (малый датасет — UMAP не даёт преимущества)"
            )

    # ── 8. SVD-размерность ─────────────────────────────────────────
    svd_dim = max(100, min(300, int(vocab_est ** 0.28)))
    if vec == "combo":
        params["combo_svd_dim"] = svd_dim
        reasons.append(
            f"Combo SVD-dim={svd_dim}  "
            f"(словарь ~{vocab_est // 1000}k слов)"
        )
    elif vec == "tfidf":
        # SVD перед KMeans на разреженной TF-IDF матрице
        params["svd_components"] = svd_dim
        reasons.append(
            f"TF-IDF SVD-dim={svd_dim}  "
            f"(словарь ~{vocab_est // 1000}k слов)"
        )

    # ── 9. cluster_min_df — отсечение редких слов ─────────────────
    if total_rows >= 50_000:
        params["cluster_min_df"] = 5
    elif total_rows >= 10_000:
        params["cluster_min_df"] = 3
    elif total_rows >= 2_000:
        params["cluster_min_df"] = 2
    else:
        params["cluster_min_df"] = 1

    # ── 10. Потоковый режим для очень больших датасетов ───────────
    if total_rows > 100_000 and vec == "tfidf" and algo == "kmeans":
        params["use_streaming_cluster"] = True
        reasons.append(
            "Потоковый режим: ВКЛ  "
            "(датасет >100k строк → MiniBatchKMeans.partial_fit)"
        )
    else:
        params["use_streaming_cluster"] = False

    # ── 11. Роли диалога ───────────────────────────────────────────
    if has_chatbot:
        params["ignore_chatbot_cluster"] = True
        reasons.append(
            "Игнор. чат-бота: ВКЛ  "
            "(обнаружены метки CHATBOT / BOT / SYSTEM)"
        )
    if has_dialog_roles:
        params["cluster_role_mode"] = "client"
        reasons.append(
            "Источник текста: «Только клиент»  "
            "(обнаружены роли OPERATOR / CLIENT)"
        )
    else:
        params["cluster_role_mode"] = "all"

    # ── 12. Фильтрация шума ────────────────────────────────────────
    params["use_stop_words"]    = True
    params["use_noise_tokens"]  = True
    params["use_noise_phrases"] = True

    # ── Применение параметров в главном потоке ─────────────────────
    return params, reasons


class ClusterTabMixin:
    """Методы вкладки «Кластеризация» и логика кластеризации текстов."""
    _CLUSTER_MODEL_SCHEMA_VERSION = 1

    def _llm_complete_text(
        self,
        provider: str,
        model: str,
        api_key: str,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 128,
        temperature: float | None = None,
    ) -> str:
        """Единый LLM-клиент для cloud и локальных провайдеров."""
        return LLMClient.complete_text(
            provider=provider,
            model=model,
            api_key=api_key,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            max_tokens=max_tokens,
            temperature=temperature,
        )

    @staticmethod
    def _resolve_llm_api_key(snap: Dict[str, Any], provider: str) -> str:
        """Возвращает API-ключ из env или зашифрованного runtime-снимка."""
        encrypted = (snap.get("llm_api_key_encrypted") or "").strip()
        legacy_raw = (snap.get("llm_api_key") or "").strip()
        try:
            snapshot_key = decrypt_api_key_from_snapshot(encrypted or legacy_raw)
        except LLMSnapshotDecryptError:
            snapshot_key = ""
        return resolve_api_key(provider, snapshot_key)

    @staticmethod
    def _llm_provider_ui_defaults(provider: str, current_model: str, current_api_key: str) -> tuple[str, str]:
        """Рекомендуемые значения UI для выбранного LLM-провайдера.

        Для локального Ollama:
          - подставляем дефолт модели, если поле пустое;
          - очищаем API-ключ, так как он не нужен.
        """
        p = (provider or "").strip().lower()
        model = (current_model or "").strip()
        api_key = (current_api_key or "").strip()
        if p == "ollama":
            if not model:
                model = "qwen3:8b"
            api_key = ""
        return model, api_key

    def _on_llm_provider_changed(self, _event=None) -> None:
        """Применяет UX-дефолты при смене LLM-провайдера."""
        model, api_key = self._llm_provider_ui_defaults(
            self.llm_provider.get(),
            self.llm_model.get(),
            self.llm_api_key.get(),
        )
        if self.llm_model.get().strip() != model:
            self.llm_model.set(model)
        if self.llm_api_key.get().strip() != api_key:
            self.llm_api_key.set(api_key)

    def _build_reason_without_llm(self, cluster_name: str, keywords: str, examples: List[str]) -> str:
        """Формирует краткое описание причин кластера без LLM (эвристика)."""
        return ClusterReasonBuilder.build_reason(cluster_name, keywords, examples)

    # region UI — вкладка кластеризации

    def _cw(self, parent, label: str, var, widget_type: str = "spinbox",
            tooltip: str = "", tooltip_title: str = "", style: str = "Card.TLabel",
            managed_list: list = None, **kw) -> object:
        """Widget factory: Label + Control + Tooltip для вкладки кластеризации."""
        row = ttk.Frame(parent, style="Card.TFrame")
        row.pack(fill="x", pady=1)
        lbl = ttk.Label(row, text=label, style=style, anchor="w")
        lbl.pack(side="left")
        if tooltip:
            title = tooltip_title or label.rstrip(":")
            self.attach_help(lbl, title, tooltip)
        if widget_type == "spinbox":
            w = ttk.Spinbox(row, textvariable=var, width=kw.pop("width", 6), **kw)
        elif widget_type == "entry":
            w = ttk.Entry(row, textvariable=var, width=kw.pop("width", 8), **kw)
        elif widget_type == "check":
            w = ttk.Checkbutton(row, variable=var, **kw)
        elif widget_type == "combo":
            w = ttk.Combobox(row, textvariable=var, width=kw.pop("width", 14),
                             state="readonly", **kw)
        else:
            raise ValueError(f"Unknown widget_type: {widget_type!r}")
        w.pack(side="left", padx=(4, 0))
        if managed_list is not None:
            managed_list.append(w)
        return w

    def _build_cluster_columns_card(self, parent: "ttk.Frame") -> None:
        # ── Дополнительные колонки ────────────────────────────────────────────
        _acols = ttk.LabelFrame(parent, text="Дополнительные колонки", padding=12)
        _acols.pack(fill="x", pady=(4, 6))
        self._combobox(_acols, 0, "Суммаризация:", self.summary_col, "Выжимка диалога (если есть)")
        self._combobox(_acols, 1, "Ответ краткий:", self.ans_short_col, "Короткий ответ оператора")
        self._combobox(_acols, 2, "Ответ полный:", self.ans_full_col, "Развёрнутый ответ / HTML")

        _arow_ninit = ttk.Frame(_acols, style="Card.TFrame")
        _arow_ninit.grid(row=3, column=1, sticky="w", padx=10, pady=(8, 0))
        lbl_ninit = ttk.Label(_arow_ninit, text="Инициализаций (n_init):", style="Card.TLabel")
        lbl_ninit.pack(side="left")
        Tooltip(lbl_ninit, "Число запусков KMeans с разными центрами.\n"
                           "Больше — надёжнее результат, медленнее вычисление.\n"
                           "Рекомендуется: 10–20")
        _sp_ninit = ttk.Spinbox(_arow_ninit, from_=1, to=30, textvariable=self.n_init_cluster, width=6)
        _sp_ninit.pack(side="left", padx=8)
        self._cluster_k_widgets.append(_sp_ninit)  # управляется вместе с K и elbow

        lbl_seed = ttk.Label(_arow_ninit, text="  Seed:", style="Card.TLabel")
        lbl_seed.pack(side="left", padx=(16, 0))
        Tooltip(lbl_seed,
                "Фиксированный seed для UMAP, KMeans, SVD и LDA.\n\n"
                "При одинаковом seed и тех же данных результат всегда воспроизводим.\n"
                "Чтобы получить другое разбиение — измените значение.\n\n"
                "Рекомендуется: 42 (стандарт). 0 — случайный seed каждый раз.")
        ttk.Spinbox(_arow_ninit, from_=0, to=9999, textvariable=self.cluster_random_seed, width=6).pack(
            side="left", padx=4
        )

        lbl_sbert_batch = ttk.Label(_arow_ninit, text="  SBERT батч:", style="Card.TLabel")
        lbl_sbert_batch.pack(side="left", padx=(16, 0))
        Tooltip(lbl_sbert_batch,
                "Размер батча при кодировании текстов через SBERT.\n\n"
                "Меньше — меньше памяти GPU/CPU, медленнее.\n"
                "Больше — быстрее, но требует больше RAM/VRAM.\n\n"
                "Рекомендуется: 64–256 для GPU, 16–64 для CPU.")
        ttk.Spinbox(_arow_ninit, from_=1, to=2048, textvariable=self.sbert_batch, width=6).pack(
            side="left", padx=4
        )

        _arow_out = ttk.Frame(_acols, style="Card.TFrame")
        _arow_out.grid(row=4, column=1, sticky="w", padx=10, pady=(8, 0))
        lbl_cid = ttk.Label(_arow_out, text="Столбец ID кластера:", style="Card.TLabel")
        lbl_cid.pack(side="left")
        Tooltip(lbl_cid, "Имя столбца в Excel, куда будет записан номер кластера.\nПо умолчанию: cluster_id")
        ttk.Entry(_arow_out, textvariable=self.cluster_id_col, width=20).pack(side="left", padx=8)
        lbl_kw2 = ttk.Label(_arow_out, text="Столбец ключевых слов:", style="Card.TLabel")
        lbl_kw2.pack(side="left", padx=(12, 0))
        Tooltip(lbl_kw2, "Имя столбца в Excel, куда будут записаны ключевые слова кластера.\nПо умолчанию: cluster_keywords")
        ttk.Entry(_arow_out, textvariable=self.cluster_kw_col, width=20).pack(side="left", padx=8)

        # ── Дополнительная фильтрация шума ────────────────────────────────────
        _anoise = ttk.LabelFrame(parent, text="Дополнительная фильтрация шума", padding=(12, 6))
        _anoise.pack(fill="x", pady=(0, 6))

        cb_normnum = ttk.Checkbutton(
            _anoise, text="Нормализация чисел/дат → __NUM__/__DATE__",
            variable=self.normalize_numbers,
        )
        cb_normnum.pack(side="left")
        self.attach_help(
            cb_normnum,
            "Нормализация чисел и дат",
            "Заменяет числовые артефакты специальными токенами до TF-IDF:\n\n"
            "  ДД.ММ.ГГГГ, ДД/ММ/ГГГГ, ДД-ММ-ГГ → __DATE__\n"
            "  ЧЧ:ММ, ЧЧ:ММ:СС             → __TIME__\n"
            "  Числа из 4+ цифр             → __NUM__\n"
            "  Десятичные (1 234,56)        → __NUM__\n\n"
            "Убирает мусорные токены-даты/суммы из TF-IDF словаря.",
            "Regex-нормализация числовых артефактов",
        )

        cb_lemma_cl = ttk.Checkbutton(
            _anoise, text="Лемматизация (pymorphy2)",
            variable=self.use_lemma_cluster,
        )
        cb_lemma_cl.pack(side="left", padx=(16, 0))
        self._cb_lemma_cl = cb_lemma_cl  # ссылка для управления состоянием из _update_cluster_ui_state
        self.attach_help(
            cb_lemma_cl,
            "Лемматизация текста (кластеризация)",
            "Приводит слова к начальной форме перед TF-IDF:\n\n"
            "  «снятие», «снимаю», «снял» → «снять»\n"
            "  «банке», «банку», «банком» → «банк»\n\n"
            "Объединяет словоформы в один признак, повышая полноту кластеров.\n"
            "Требует: pip install pymorphy2",
            "pymorphy2: морфологическая нормализация",
        )


    def _build_cluster_neural_card(self, parent: "ttk.Frame") -> None:
        # ── Настройки Нейросеть (SBERT / DeBERTa) / Комбо / Ансамбль ──────────
        _avec = ttk.LabelFrame(parent, text="Нейросетевые модели (SBERT / DeBERTa-v2/v3) / Комбо / Ансамбль", padding=(12, 6))
        _avec.pack(fill="x", pady=(0, 6))

        _sbert_row = ttk.Frame(_avec, style="Card.TFrame")
        _sbert_row.pack(fill="x")
        ttk.Label(_sbert_row, text="Нейросеть:", style="Card.TLabel").pack(side="left")
        self.cb_sbert_clust_model = ttk.Combobox(
            _sbert_row, textvariable=self.sbert_model,
            state="readonly", values=NEURAL_MODELS_LIST, width=38,
        )
        self.cb_sbert_clust_model.pack(side="left", padx=(4, 0))
        self.attach_help(
            self.cb_sbert_clust_model,
            "Нейросетевая модель (SBERT / DeBERTa)",
            "── SBERT-модели (sentence-transformers) ──\n"
            "ruBERT-tiny2 (~45 MB)  — быстрый, хорошее качество для RU\n"
            "USER-base/large (DeBERTa) — SOTA для RU, prefix 'clustering:' ⭐\n"
            "USER2-base/large (Matryoshka+DeBERTa) — SOTA RU 2024 ⭐\n"
            "BGE-M3 (~570 MB)  — мультиязычный SOTA 2024, 100+ языков ⭐⭐\n"
            "USER-bge-m3 — BGE-M3 адаптированный для RU ⭐⭐\n"
            "multilingual-e5-large-instruct — инструктивный SOTA ⭐\n\n"
            "── DeBERTa-v2/v3 (transformers + mean pooling) ──\n"
            "deberta-v3-base (~185 MB) — SOTA мультиязычный ⭐\n"
            "deberta-v3-large (~380 MB) — высокое качество ⭐⭐\n"
            "deberta-v1-base (deepvk) — специализированный RU ⭐\n\n"
            "При первом запуске модель скачивается из HuggingFace (~минута).",
            "Выбор нейросетевой модели для кластеризации",
        )
        ttk.Label(_sbert_row, text="  Устройство:", style="Card.TLabel").pack(side="left", padx=(12, 4))
        self.cb_sbert_clust_device = ttk.Combobox(
            _sbert_row, textvariable=self.sbert_device,
            state="readonly", values=self.gpu_device_values, width=7,
        )
        self.cb_sbert_clust_device.pack(side="left")
        self.attach_help(
            self.cb_sbert_clust_device,
            "SBERT устройство",
            "auto — автоматический выбор (GPU если доступен, иначе CPU)\n"
            "cpu  — принудительно CPU (всегда работает)\n"
            "cuda — принудительно GPU (требует torch+CUDA)\n\n"
            "GPU ускоряет кодирование в 5–20× на больших данных.",
            "CPU / GPU для SBERT-кодирования",
        )
        self._sbert_clust_widgets = [self.cb_sbert_clust_model, self.cb_sbert_clust_device]

        _ensemble_row = ttk.Frame(_avec, style="Card.TFrame")
        _ensemble_row.pack(fill="x", pady=(4, 0))
        ttk.Label(_ensemble_row, text="Ансамбль — 2-я модель SBERT:", style="Card.TLabel").pack(side="left")
        self.cb_sbert_ensemble_model2 = ttk.Combobox(
            _ensemble_row, textvariable=self.sbert_model2,
            state="readonly", values=NEURAL_MODELS_LIST, width=38,
        )
        self.cb_sbert_ensemble_model2.pack(side="left", padx=(4, 0))
        self.attach_help(
            self.cb_sbert_ensemble_model2,
            "Ансамбль: вторая SBERT-модель",
            "Вторая нейросетевая модель для режима «Ансамбль».\n\n"
            "Система сравнит TF-IDF, SBERT-1 и SBERT-2 по Silhouette\n"
            "и автоматически выберет лучший результат.",
            "2-я SBERT-модель для ансамбля",
        )
        self._ensemble_widgets = [self.cb_sbert_ensemble_model2]

        _combo_row = ttk.Frame(_avec, style="Card.TFrame")
        _combo_row.pack(fill="x", pady=(4, 0))
        ttk.Label(_combo_row, text="Комбо — SVD-dim:", style="Card.TLabel").pack(side="left")
        sp_svd = ttk.Spinbox(
            _combo_row, from_=50, to=1000,
            textvariable=self.combo_svd_dim, width=6,
        )
        sp_svd.pack(side="left", padx=(4, 0))
        self.attach_help(sp_svd, "SVD-размерность (комбо)",
                         "Число SVD-компонент при сжатии TF-IDF перед конкатенацией с SBERT.\n"
                         "Рекомендуется: 100–300.")
        ttk.Label(_combo_row, text="  Вес SBERT (α):", style="Card.TLabel").pack(side="left", padx=(16, 2))
        sp_alpha = ttk.Spinbox(
            _combo_row, from_=0.0, to=1.0, increment=0.1,
            textvariable=self.combo_alpha, width=5, format="%.1f",
        )
        sp_alpha.pack(side="left")
        self.attach_help(sp_alpha, "Вес SBERT в комбо (α)",
                         "α = 0.5 → равный вес TF-IDF и SBERT (рекомендуется).")
        self._combo_clust_widgets = [sp_svd, sp_alpha]

        _tfidf_svd_row = ttk.Frame(_avec, style="Card.TFrame")
        self._tfidf_svd_row = _tfidf_svd_row
        _tfidf_svd_row.pack(fill="x", pady=(4, 0))
        cb_tfidf_svd = ttk.Checkbutton(
            _tfidf_svd_row, text="SVD для TF-IDF (рекомендуется для KMeans)",
            variable=self.use_tfidf_svd,
            style="Card.TCheckbutton",
        )
        cb_tfidf_svd.pack(side="left")
        self.attach_help(
            cb_tfidf_svd,
            "SVD для TF-IDF кластеризации",
            "TF-IDF → SVD → L2-нормализация → KMeans.\n\n"
            "Снижает размерность разреженной TF-IDF матрицы.\n"
            "Аналог PCA для текстов. Особенно важно при большом словаре.",
            "LSA (Latent Semantic Analysis) перед KMeans",
        )
        ttk.Label(_tfidf_svd_row, text="  SVD-dim:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        sp_tfidf_svd = ttk.Spinbox(
            _tfidf_svd_row, from_=20, to=500,
            textvariable=self.tfidf_svd_dim, width=6,
        )
        sp_tfidf_svd.pack(side="left")
        self.attach_help(sp_tfidf_svd, "SVD-размерность (TF-IDF)",
                         "Рекомендуется: 50–300.\n100 — хороший баланс для большинства датасетов.")


    def _build_cluster_metrics_card(self, parent: "ttk.Frame") -> None:
        # ── Метрика автоподбора K ─────────────────────────────────────────────
        _ak_frm = ttk.LabelFrame(parent, text="Метрика автоподбора K", padding=(12, 6))
        _ak_frm.pack(fill="x", pady=(0, 6))

        self._k_metric_rbs: list = []
        for _text, _val, _tip in [
            (
                "Elbow (инерция)", "elbow",
                "Метод локтя: ищет точку перегиба на кривой инерции.\n\n"
                "Быстрый. Работает всегда. Рекомендуется для >10 000 строк.",
            ),
            (
                "Silhouette", "silhouette",
                "Силуэтный индекс: оценивает компактность и разделённость кластеров.\n\n"
                "Медленнее Elbow. Рекомендуется при K ≤ 30 и датасетах до 50 000 строк.",
            ),
            (
                "Calinski-Harabász", "calinski",
                "Индекс Калинского–Харабаша: отношение дисперсии МЕЖДУ кластерами\n"
                "к дисперсии ВНУТРИ кластеров.\n\n"
                "Быстрее Silhouette. Рекомендуется как быстрая альтернатива Silhouette.",
            ),
        ]:
            rb = ttk.Radiobutton(
                _ak_frm, text=_text,
                variable=self.k_score_method, value=_val,
            )
            rb.pack(side="left", padx=6)
            self.attach_help(rb, f"Метрика K: {_text}", _tip, _tip)
            self._k_metric_rbs.append(rb)

        # ── Метрика расстояния ────────────────────────────────────────────────
        _adist_frm = ttk.LabelFrame(parent, text="Метрика расстояния", padding=(12, 6))
        _adist_frm.pack(fill="x", pady=(0, 6))

        cb_cosine = ttk.Checkbutton(
            _adist_frm, text="Косинусная метрика (L2-нормализация векторов перед кластеризацией)",
            variable=self.use_cosine_cluster,
        )
        cb_cosine.pack(side="left")
        self._cb_cosine = cb_cosine
        self.attach_help(
            cb_cosine,
            "Косинусная метрика",
            "Применяет L2-нормализацию всех векторов перед KMeans / HDBSCAN.\n\n"
            "Рекомендуется для TF-IDF, SBERT и Комбо-режима.",
            "L2-нормализация ≡ косинусное расстояние в KMeans",
        )

        # ── UMAP ─────────────────────────────────────────────────────────────
        _umap_frm = ttk.LabelFrame(parent, text="Снижение размерности (UMAP)", padding=(12, 6))
        _umap_frm.pack(fill="x", pady=(0, 6))

        cb_umap = ttk.Checkbutton(_umap_frm, text="Применить UMAP", variable=self.use_umap)
        cb_umap.pack(side="left")
        self._cb_umap = cb_umap
        self.attach_help(
            cb_umap,
            "UMAP перед кластеризацией",
            "UMAP — нелинейное снижение размерности, сохраняющее локальную структуру.\n\n"
            "Улучшает качество KMeans на высоких размерностях.\n"
            "Требует: pip install umap-learn",
            "Нелинейное снижение размерности (umap-learn)",
        )

        ttk.Label(_umap_frm, text="  Компоненты:").pack(side="left", padx=(16, 2))
        sp_umap = ttk.Spinbox(_umap_frm, from_=5, to=200, textvariable=self.umap_n_components, width=6)
        sp_umap.pack(side="left")
        self.attach_help(sp_umap, "UMAP: число компонент",
                         "Рекомендуется: 5–15 для кластеризации (лучшее разделение кластеров).\n"
                         "50+ — только для промежуточной визуализации, не для кластеризации.\n"
                         "2D (n=2) — только для финального scatter-plot, не для HDBSCAN/KMeans.")

        _umap_params_row = ttk.Frame(_umap_frm, style="Card.TFrame")
        _umap_params_row.pack(fill="x", pady=(4, 0))
        ttk.Label(_umap_params_row, text="n_neighbors:", style="Card.TLabel").pack(side="left")
        sp_umap_nn = ttk.Spinbox(_umap_params_row, from_=2, to=200, textvariable=self.umap_n_neighbors, width=5)
        sp_umap_nn.pack(side="left", padx=(4, 0))
        self.attach_help(sp_umap_nn, "UMAP: n_neighbors",
                         "Число соседей при построении графа близости.\nРекомендуется: 15.")

        ttk.Label(_umap_params_row, text="  min_dist:", style="Card.TLabel").pack(side="left", padx=(12, 0))
        sp_umap_md = ttk.Spinbox(_umap_params_row, from_=0.0, to=0.99, increment=0.05,
                                  textvariable=self.umap_min_dist, width=5, format="%.2f")
        sp_umap_md.pack(side="left", padx=(4, 0))
        self.attach_help(sp_umap_md, "UMAP: min_dist",
                         "0.0–0.1: компактные, чёткие кластеры (рекомендуется для кластеризации).")

        ttk.Label(_umap_params_row, text="  Метрика:", style="Card.TLabel").pack(side="left", padx=(12, 0))
        cb_umap_metric = ttk.Combobox(
            _umap_params_row, textvariable=self.umap_metric,
            state="readonly", width=10,
            values=["cosine", "euclidean", "correlation", "manhattan"],
        )
        cb_umap_metric.pack(side="left", padx=(4, 0))
        self.cb_umap_metric = cb_umap_metric  # защита от перезаписи в _refresh_combobox_values
        self.attach_help(cb_umap_metric, "UMAP: метрика расстояния",
                         "cosine — рекомендуется для TF-IDF и SBERT.")

        self._umap_clust_widgets = [cb_umap, sp_umap, sp_umap_nn, sp_umap_md, cb_umap_metric]

        # ── Параметры алгоритмов ──────────────────────────────────────────────
        _aalgo_frm = ttk.LabelFrame(parent, text="Параметры алгоритмов", padding=(12, 6))
        _aalgo_frm.pack(fill="x", pady=(0, 6))

        _hdb_row = ttk.Frame(_aalgo_frm, style="Card.TFrame")
        _hdb_row.pack(fill="x", pady=(0, 4))
        self._hdb_row = _hdb_row
        ttk.Label(_hdb_row, text="HDBSCAN — min_cluster_size:", style="Card.TLabel").pack(side="left")
        sp_hdb = ttk.Spinbox(_hdb_row, from_=0, to=500,
                              textvariable=self.hdbscan_min_cluster_size, width=6)
        sp_hdb.pack(side="left", padx=(4, 0))
        self.attach_help(sp_hdb, "HDBSCAN: min_cluster_size",
                         "Минимальное число точек в кластере.\n"
                         "0 = авто: max(5, √N) — масштабируется под размер датасета.\n"
                         "Ручное: 5–50 (≈ 0.1–1% от числа строк).")
        ttk.Label(_hdb_row, text="  min_samples:", style="Card.TLabel").pack(side="left", padx=(12, 0))
        sp_hdb_ms = ttk.Spinbox(_hdb_row, from_=0, to=200,
                                 textvariable=self.hdbscan_min_samples, width=5)
        sp_hdb_ms.pack(side="left", padx=(4, 0))
        self.attach_help(sp_hdb_ms, "HDBSCAN: min_samples",
                         "Минимальное число соседей для считающейся плотной точки.\n"
                         "0 = авто (= min_cluster_size). Меньше → крупнее кластеры.")
        ttk.Label(_hdb_row, text="  ε:", style="Card.TLabel").pack(side="left", padx=(12, 0))
        sp_hdb_eps = ttk.Spinbox(_hdb_row, from_=0.0, to=1.0, increment=0.05,
                                  textvariable=self.hdbscan_eps, width=5, format="%.2f")
        sp_hdb_eps.pack(side="left", padx=(4, 0))
        self.attach_help(sp_hdb_eps, "HDBSCAN: cluster_selection_epsilon",
                         "Порог расстояния для слияния «почти одинаковых» кластеров.\n"
                         "0.0 = не применять. 0.1–0.3 объединяют близкие группы.")

        _lda_row = ttk.Frame(_aalgo_frm, style="Card.TFrame")
        _lda_row.pack(fill="x", pady=(0, 4))
        self._lda_row = _lda_row
        ttk.Label(_lda_row, text="LDA — тем:", style="Card.TLabel").pack(side="left")
        ttk.Spinbox(_lda_row, from_=2, to=200, textvariable=self.lda_n_topics, width=6).pack(
            side="left", padx=(4, 0))
        ttk.Label(_lda_row, text="  итераций:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Spinbox(_lda_row, from_=10, to=500, textvariable=self.lda_max_iter, width=6).pack(side="left")
        ttk.Label(_lda_row, text="  столбец тем:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Entry(_lda_row, textvariable=self.lda_topics_col, width=16).pack(side="left")

        _hier_row = ttk.Frame(_aalgo_frm, style="Card.TFrame")
        _hier_row.pack(fill="x", pady=(0, 4))
        self._hier_row = _hier_row
        ttk.Label(_hier_row, text="Иерарх. — K верхн.:", style="Card.TLabel").pack(side="left")
        ttk.Spinbox(_hier_row, from_=2, to=50, textvariable=self.hier_k_top, width=6).pack(
            side="left", padx=(4, 0))
        ttk.Label(_hier_row, text="  K подгруп.:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Spinbox(_hier_row, from_=2, to=30, textvariable=self.hier_k_sub, width=6).pack(side="left")
        ttk.Label(_hier_row, text="  мин. строк:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Spinbox(_hier_row, from_=5, to=500, textvariable=self.hier_min_sub, width=6).pack(side="left")
        ttk.Label(_hier_row, text="  столбец L1:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Entry(_hier_row, textvariable=self.hier_l1_col, width=16).pack(side="left")

        _bert_row = ttk.Frame(_aalgo_frm, style="Card.TFrame")
        _bert_row.pack(fill="x")
        self._bert_row = _bert_row
        ttk.Label(_bert_row, text="BERTopic — мин. размер темы:", style="Card.TLabel").pack(side="left")
        ttk.Spinbox(_bert_row, from_=2, to=500, textvariable=self.bertopic_min_topic_size, width=6).pack(
            side="left", padx=(4, 0))
        ttk.Label(_bert_row, text="  тем (nr_topics):", style="Card.TLabel").pack(side="left", padx=(12, 2))
        _bert_entry = ttk.Entry(_bert_row, textvariable=self.bertopic_nr_topics, width=8)
        _bert_entry.pack(side="left")
        Tooltip(_bert_entry,
                "Число тем для BERTopic.\n'auto' — автоматически определяется HDBSCAN.")


    def _build_cluster_naming_card(self, parent: "ttk.Frame") -> None:
        # ── Семантические якоря ───────────────────────────────────────────────
        _aanch_frm = ttk.LabelFrame(parent, text="Семантические якоря (KMeans / SBERT)", padding=(12, 6))
        _aanch_frm.pack(fill="x", pady=(0, 6))

        _anch_top = ttk.Frame(_aanch_frm, style="Card.TFrame")
        _anch_top.pack(fill="x")
        cb_anchors = ttk.Checkbutton(
            _anch_top, text="Использовать якоря для инициализации KMeans",
            variable=self.use_anchors,
            style="Card.TCheckbutton",
        )
        cb_anchors.pack(side="left")
        self._cb_anchors = cb_anchors
        self.attach_help(
            cb_anchors,
            "Семантические якоря",
            "Инициализирует начальные центры KMeans нейросетевыми\n"
            "эмбеддингами заданных фраз (SBERT).\n\n"
            "Требует режима векторизации SBERT или Комбо.",
            "Задать начальные центры KMeans через SBERT-якоря",
        )
        ttk.Label(_anch_top, text="  Фразы (по одной на строку):",
                  style="Card.Muted.TLabel").pack(side="left", padx=(16, 4))

        self.txt_anchors = tk.Text(
            _aanch_frm, height=4, width=60,
            bg=ENTRY_BG, fg=FG,
            relief="flat", borderwidth=1,
            font=("Segoe UI", 9),
        )
        self.txt_anchors.pack(fill="x", pady=(4, 0))
        Tooltip(self.txt_anchors,
                "Введите по одной фразе-якорю на строку.\n"
                "Каждая фраза будет закодирована SBERT и\n"
                "использована как начальный центр кластера.")

        # ── T5-суммаризация ───────────────────────────────────────────────────
        _at5_frm = ttk.LabelFrame(parent, text="T5-суммаризация текстов (опционально)", padding=(12, 6))
        _at5_frm.pack(fill="x", pady=(0, 6))

        cb_t5 = ttk.Checkbutton(
            _at5_frm, text="Использовать T5-суммаризатор",
            variable=self.use_t5_summary,
            style="Card.TCheckbutton",
        )
        cb_t5.pack(side="left")
        self.attach_help(
            cb_t5,
            "T5-суммаризация",
            "Включает автоматическую суммаризацию текста каждой строки\n"
            "с помощью нейросетевой модели T5 (seq2seq).\n\n"
            "Требует: pip install transformers torch\n"
            "Модель: UrukHan/t5-russian-summarization (~1-2 GB)\n\n"
            "⚠️  Работает медленно на CPU при большом датасете.",
            "Нейросетевая суммаризация каждой строки (T5, seq2seq)",
        )

        ttk.Label(_at5_frm, text="  Модель:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        ent_t5_model = ttk.Entry(_at5_frm, textvariable=self.t5_model_name, width=44)
        ent_t5_model.pack(side="left")
        self.attach_help(ent_t5_model, "T5 модель (HuggingFace ID)",
                         "Имя модели на HuggingFace Hub.\nПо умолчанию: UrukHan/t5-russian-summarization",
                         "HuggingFace model id для T5-суммаризации")

        t5_row2 = ttk.Frame(_at5_frm, style="Card.TFrame")
        t5_row2.pack(side="left", padx=(16, 0))
        ttk.Label(t5_row2, text="Столбец:", style="Card.TLabel").pack(side="left")
        ttk.Entry(t5_row2, textvariable=self.t5_summary_col, width=16).pack(side="left", padx=(4, 0))

        ttk.Label(t5_row2, text="  Вход:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        sp_t5_in = ttk.Spinbox(t5_row2, from_=64, to=2048, textvariable=self.t5_max_input, width=6)
        sp_t5_in.pack(side="left")
        ttk.Label(t5_row2, text="токенов", style="Card.Muted.TLabel").pack(side="left", padx=(2, 0))
        self.attach_help(sp_t5_in, "T5: максимальная длина входа", "Рекомендуется: 256–512 токенов.")

        ttk.Label(t5_row2, text="  Выход:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        sp_t5_out = ttk.Spinbox(t5_row2, from_=32, to=512, textvariable=self.t5_max_output, width=6)
        sp_t5_out.pack(side="left")
        ttk.Label(t5_row2, text="токенов", style="Card.Muted.TLabel").pack(side="left", padx=(2, 0))
        self.attach_help(sp_t5_out, "T5: максимальная длина суммаризации", "Рекомендуется: 64–128 токенов.")

        ttk.Label(t5_row2, text="  Батч:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        sp_t5_batch = ttk.Spinbox(t5_row2, from_=1, to=64, textvariable=self.t5_batch_size, width=5)
        sp_t5_batch.pack(side="left")
        self.attach_help(sp_t5_batch, "T5: размер батча",
                         "GPU: 16–32 для ускорения.\nCPU: оставьте 1–4.")

        # ── LLM-нейминг кластеров ─────────────────────────────────────────────
        _allm_frm = ttk.LabelFrame(parent, text="LLM-нейминг кластеров (Claude / OpenAI / Ollama)", padding=(12, 6))
        _allm_frm.pack(fill="x", pady=(0, 6))

        _llm_row1 = ttk.Frame(_allm_frm, style="Card.TFrame")
        _llm_row1.pack(fill="x")
        cb_llm = ttk.Checkbutton(_llm_row1, text="Использовать LLM для названий кластеров",
                                 variable=self.use_llm_naming, style="Card.TCheckbutton")
        cb_llm.pack(side="left")
        self.attach_help(
            cb_llm,
            "LLM-нейминг кластеров",
            "После кластеризации отправляет топ-ключевые слова + 3 примера\n"
            "каждого кластера в LLM (Claude / OpenAI) и получает\n"
            "читаемое 3-5 словное название.\n\n"
            "Исследование JDS 2024: GPT-3.5-turbo превосходит ручную разметку\n"
            "при правильном промпте.\n\n"
            "Результат записывается в столбец 'cluster_name'.\n"
            "Требует API-ключ и интернет-подключение.",
            "Автоматическое название кластеров через LLM",
        )
        _llm_row2 = ttk.Frame(_allm_frm, style="Card.TFrame")
        _llm_row2.pack(fill="x", pady=(4, 0))
        ttk.Label(_llm_row2, text="Провайдер:", style="Card.TLabel").pack(side="left")
        self.cb_llm_provider = ttk.Combobox(_llm_row2, textvariable=self.llm_provider, state="readonly",
                     values=["anthropic", "openai", "gigachat", "qwen", "ollama"], width=12)
        self.cb_llm_provider.pack(side="left", padx=(4, 0))  # защита от перезаписи в _refresh_combobox_values
        self.cb_llm_provider.bind("<<ComboboxSelected>>", self._on_llm_provider_changed)
        ttk.Label(_llm_row2, text="  Модель:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Entry(_llm_row2, textvariable=self.llm_model, width=26).pack(side="left")
        ttk.Label(_llm_row2, text="  Столбец:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Entry(_llm_row2, textvariable=self.llm_name_col, width=16).pack(side="left")
        _llm_row3 = ttk.Frame(_allm_frm, style="Card.TFrame")
        _llm_row3.pack(fill="x", pady=(4, 0))
        ttk.Label(_llm_row3, text="API-ключ:", style="Card.TLabel").pack(side="left")
        ent_llm_key = ttk.Entry(_llm_row3, textvariable=self.llm_api_key, width=54, show="*")
        ent_llm_key.pack(side="left", padx=(4, 0))
        Tooltip(ent_llm_key,
                "API-ключ провайдера (Anthropic/OpenAI/GigaChat/Qwen).\n"
                "Для Ollama ключ не требуется.\n"
                "Для локальной Ollama: провайдер='ollama', модель как в `ollama list`\n"
                "(например `qwen3:30b`), сервер по умолчанию http://127.0.0.1:11434.\n"
                "Рекомендуется использовать переменную окружения LLM_API_KEY_<PROVIDER>.\n"
                "В runtime-снапшоте ключ хранится в шифрованном виде (при наличии LLM_SNAPSHOT_KEY).")
        self._on_llm_provider_changed()
        ttk.Label(
            _llm_row3,
            text="⚠ Безопаснее хранить ключ в переменной окружения.",
            style="Card.Muted.TLabel",
            foreground="#ffb74d",
        ).pack(side="left", padx=(10, 0))
        _llm_row4 = ttk.Frame(_allm_frm, style="Card.TFrame")
        _llm_row4.pack(fill="x", pady=(4, 0))
        cb_llm_reason = ttk.Checkbutton(
            _llm_row4,
            text="Генерировать обобщённое описание причин кластера",
            variable=self.use_llm_reason_summary,
            style="Card.TCheckbutton",
        )
        cb_llm_reason.pack(side="left")
        ttk.Label(_llm_row4, text="  Столбец:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Entry(_llm_row4, textvariable=self.llm_reason_col, width=20).pack(side="left")
        Tooltip(
            cb_llm_reason,
            "Если включено, для каждого кластера LLM формирует 1–2 предложения\n"
            "с обобщённым описанием причин обращений (не только ключевые слова).",
        )
        _llm_row5 = ttk.Frame(_allm_frm, style="Card.TFrame")
        _llm_row5.pack(fill="x", pady=(2, 0))
        cb_rule_reason = ttk.Checkbutton(
            _llm_row5,
            text="Формировать описание причин без LLM (эвристика)",
            variable=self.use_rule_reason_summary,
            style="Card.TCheckbutton",
        )
        cb_rule_reason.pack(side="left")
        Tooltip(
            cb_rule_reason,
            "Если LLM отключён или недоступен, описание причин будет создано\n"
            "по ключевым словам и примерам текста по правилам.",
        )

        # ── Слияние кластеров + примеры для LLM ──────────────────────────────
        _llm_row6 = ttk.Frame(_allm_frm, style="Card.TFrame")
        _llm_row6.pack(fill="x", pady=(4, 0))
        cb_merge = ttk.Checkbutton(
            _llm_row6,
            text="Объединять похожие кластеры",
            variable=self.merge_similar_clusters,
            style="Card.TCheckbutton",
        )
        cb_merge.pack(side="left")
        self.attach_help(
            cb_merge,
            "Слияние семантически близких кластеров",
            "После кластеризации сравниваются центроиды всех кластеров.\n"
            "Если косинусное сходство ≥ порогу — кластеры сливаются.\n\n"
            "Порог 0.85: объединяются только почти одинаковые кластеры.\n"
            "Порог 0.70: объединяются близкие по смыслу кластеры.\n\n"
            "Лучше всего работает с SBERT/combo-векторизацией.",
            "Слияние по косинусному сходству центроидов",
        )
        ttk.Label(_llm_row6, text="  Порог:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Spinbox(
            _llm_row6, from_=0.50, to=0.99, increment=0.05,
            textvariable=self.merge_threshold, width=5, format="%.2f",
        ).pack(side="left")
        ttk.Label(_llm_row6, text="  Примеров для LLM:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        ttk.Spinbox(
            _llm_row6, from_=3, to=10, increment=1,
            textvariable=self.n_repr_examples, width=4,
        ).pack(side="left")


    def _build_cluster_advanced_card(self, parent: "ttk.Frame") -> None:
        # ── PCA перед UMAP ────────────────────────────────────────────────────
        _apca_frm = ttk.LabelFrame(parent, text="PCA перед UMAP (снижение шума)", padding=(12, 6))
        _apca_frm.pack(fill="x", pady=(0, 6))

        _pca_row = ttk.Frame(_apca_frm, style="Card.TFrame")
        _pca_row.pack(fill="x")
        cb_pca = ttk.Checkbutton(_pca_row, text="PCA-преднормализация перед UMAP",
                                 variable=self.use_pca_before_umap, style="Card.TCheckbutton")
        cb_pca.pack(side="left")
        self.attach_help(
            cb_pca,
            "PCA перед UMAP",
            "Применяет линейное снижение размерности PCA перед нелинейным UMAP.\n\n"
            "Оптимальный пайплайн 2024 (GDELT Project эксперименты):\n"
            "  Эмбеддинги (768 dim) → PCA(50) → UMAP(5-15) → кластеризация\n\n"
            "Преимущества:\n"
            "  • Убирает шум из высокоразмерных эмбеддингов\n"
            "  • Стабилизирует UMAP (меньше артефактов)\n"
            "  • Ускоряет вычисление UMAP\n\n"
            "Рекомендуется при использовании SBERT-эмбеддингов (768+ dim).",
            "Линейное PCA → нелинейное UMAP",
        )
        ttk.Label(_pca_row, text="  Компоненты PCA:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        sp_pca = ttk.Spinbox(_pca_row, from_=10, to=300, textvariable=self.pca_n_components, width=6)
        sp_pca.pack(side="left")
        Tooltip(sp_pca, "Рекомендуется: 50–100. Должно быть > n_components UMAP.")

        # ── Качество кластеров (intra-cohesion) ───────────────────────────────
        _aqlt_frm = ttk.LabelFrame(parent, text="Аннотация качества кластеров", padding=(12, 6))
        _aqlt_frm.pack(fill="x", pady=(0, 6))

        _qlt_row = ttk.Frame(_aqlt_frm, style="Card.TFrame")
        _qlt_row.pack(fill="x")
        cb_qlt = ttk.Checkbutton(
            _qlt_row,
            text="Добавить столбец качества кластера (intra-cohesion)",
            variable=self.show_cluster_quality,
            style="Card.TCheckbutton",
        )
        cb_qlt.pack(side="left")
        self.attach_help(
            cb_qlt,
            "Аннотация качества кластера",
            "Для каждой строки вычисляет среднее косинусное расстояние\n"
            "до других строк в том же кластере.\n\n"
            "Итоговая оценка кластера:\n"
            "  high   — cohesion > 0.5 (плотный, семантически однородный)\n"
            "  medium — cohesion 0.3–0.5 (приемлемо)\n"
            "  low    — cohesion < 0.3 (вероятно шумный кластер)\n\n"
            "Помогает быстро найти проблемные кластеры в Excel.",
            "Оценка внутренней однородности кластеров",
        )
        ttk.Label(_qlt_row, text="  Столбец:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        ttk.Entry(_qlt_row, textvariable=self.cluster_quality_col, width=18).pack(side="left")

        # ── Повторная кластеризация выбросов HDBSCAN ──────────────────────────
        _anoise2_frm = ttk.LabelFrame(parent, text="Повторная кластеризация выбросов HDBSCAN", padding=(12, 6))
        _anoise2_frm.pack(fill="x", pady=(0, 6))

        _n2_row = ttk.Frame(_anoise2_frm, style="Card.TFrame")
        _n2_row.pack(fill="x")
        cb_recluster = ttk.Checkbutton(
            _n2_row,
            text="Кластеризовать выбросы (-1) вторым KMeans",
            variable=self.recluster_noise,
            style="Card.TCheckbutton",
        )
        cb_recluster.pack(side="left")
        self.attach_help(
            cb_recluster,
            "Повторная кластеризация выбросов HDBSCAN",
            "HDBSCAN помечает 10–30% строк как шум (cluster_id = -1).\n\n"
            "Включая этот параметр, вы запустите второй KMeans ТОЛЬКО\n"
            "на выбросах и запишете результат в отдельный столбец.\n\n"
            "Зачем это нужно:\n"
            "  В банковском КЦ 'редкие' обращения часто самые важные:\n"
            "  жалобы на мошенничество, сбои, нестандартные ситуации.\n"
            "  Не стоит их просто отбрасывать.\n\n"
            "Результат — отдельный столбец 'noise_cluster_id'.",
            "Второй KMeans по точкам с label=-1",
        )
        ttk.Label(_n2_row, text="  K выбросов:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        ttk.Spinbox(_n2_row, from_=2, to=50, textvariable=self.noise_k_clusters, width=5).pack(side="left")
        ttk.Label(_n2_row, text="  Столбец:", style="Card.TLabel").pack(side="left", padx=(12, 2))
        ttk.Entry(_n2_row, textvariable=self.noise_cluster_col, width=18).pack(side="left")

        # ── min_df для TF-IDF / семантическая дедупликация / визуализация ─────
        _amisc_frm = ttk.LabelFrame(parent, text="TF-IDF, дедупликация, визуализация", padding=(12, 6))
        _amisc_frm.pack(fill="x", pady=(0, 6))

        _mindf_row = ttk.Frame(_amisc_frm, style="Card.TFrame")
        _mindf_row.pack(fill="x")
        lbl_mindf = ttk.Label(_mindf_row, text="min_df (TF-IDF):", style="Card.TLabel")
        lbl_mindf.pack(side="left")
        self.attach_help(
            lbl_mindf,
            "min_df — минимальная частота слова в TF-IDF",
            "Минимальное число документов, в которых слово должно встречаться.\n\n"
            "0 = авто-адаптивный (1 для <5K строк, 2 для <50K, 3 для 50K+)\n"
            "1 = все слова (включая одиночные опечатки)\n"
            "2–5 = убирает редкие шумные токены\n\n"
            "Рекомендуется: 0 (авто) для большинства случаев.\n"
            "Используйте 1, если ваш домен содержит редкие специальные термины.",
            "TF-IDF: минимальная частота слова",
        )
        ttk.Spinbox(_mindf_row, from_=0, to=20, textvariable=self.cluster_min_df, width=5).pack(
            side="left", padx=(4, 0))
        ttk.Label(_mindf_row, text="  (0 = авто)", style="Card.Muted.TLabel").pack(side="left", padx=(4, 0))

        _dedup_row = ttk.Frame(_amisc_frm, style="Card.TFrame")
        _dedup_row.pack(fill="x", pady=(4, 0))
        cb_dedup = ttk.Checkbutton(
            _dedup_row,
            text="Семантическая дедупликация перед кластеризацией",
            variable=self.use_dedup,
            style="Card.TCheckbutton",
        )
        cb_dedup.pack(side="left")
        self.attach_help(
            cb_dedup,
            "Семантическая дедупликация",
            "Находит пары строк с косинусным сходством > порога и удаляет\n"
            "дубликаты перед кластеризацией.\n\n"
            "Зачем:\n"
            "  Дубликаты перетягивают центроид кластера и занижают Silhouette.\n"
            "  Особенно важно, если в данных много копий одного обращения.\n\n"
            "Требует SBERT-режим (вычисляет эмбеддинги для сравнения).\n"
            "Дублированные строки получают cluster_id оригинала.",
            "Удаление семантических дубликатов перед кластеризацией",
        )
        ttk.Label(_dedup_row, text="  Порог:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        sp_dedup = ttk.Spinbox(_dedup_row, from_=0.5, to=1.0, increment=0.05,
                                textvariable=self.dedup_threshold, width=5, format="%.2f")
        sp_dedup.pack(side="left")
        Tooltip(sp_dedup, "0.95 = очень похожие тексты (почти одинаковые).\n"
                           "0.85 = умеренно похожие тексты.")

        _viz_row = ttk.Frame(_amisc_frm, style="Card.TFrame")
        _viz_row.pack(fill="x", pady=(4, 0))
        cb_viz = ttk.Checkbutton(
            _viz_row,
            text="Интерактивная визуализация кластеров (Plotly HTML)",
            variable=self.use_cluster_viz,
            style="Card.TCheckbutton",
        )
        cb_viz.pack(side="left")
        self.attach_help(
            cb_viz,
            "Интерактивная визуализация",
            "После кластеризации создаёт интерактивный scatter-plot:\n"
            "  UMAP 2D → каждая точка = строка, цвет = кластер,\n"
            "  hover = ключевые слова кластера.\n\n"
            "Сохраняется как .html файл рядом с результатом Excel.\n"
            "Открывается в браузере.\n\n"
            "Требует: pip install plotly umap-learn",
            "Plotly HTML визуализация кластеров",
        )

        # ── #12 Инкрементальная кластеризация (сохранение / применение модели) ──
        _inc_frm = ttk.LabelFrame(parent, text="Инкрементальная кластеризация (#12)", padding=(12, 6))
        _inc_frm.pack(fill="x", pady=(0, 6))

        _inc_save_row = ttk.Frame(_inc_frm, style="Card.TFrame")
        _inc_save_row.pack(fill="x")
        cb_save_model = ttk.Checkbutton(
            _inc_save_row,
            text="Сохранить модель после кластеризации",
            variable=self.save_cluster_model,
            style="Card.TCheckbutton",
        )
        cb_save_model.pack(side="left")
        self.attach_help(
            cb_save_model,
            "Сохранение модели кластеризации",
            "Сохраняет обученные векторайзер + центроиды кластеров в файл .joblib.\n\n"
            "Затем этот файл можно применить к новым данным без переобучения:\n"
            "  новые тексты → тот же векторайзер → ближайший центроид.\n\n"
            "Поддерживаемые алгоритмы: KMeans, GMM, HDBSCAN (прогнозирование).\n"
            "Путь по умолчанию: папка clustered/ с именем модели.",
            "Сохранить векторайзер и центроиды кластеров",
        )

        _inc_path_row = ttk.Frame(_inc_frm, style="Card.TFrame")
        _inc_path_row.pack(fill="x", pady=(4, 0))
        ttk.Label(_inc_path_row, text="Путь к файлу модели (.joblib):", style="Card.TLabel").pack(side="left")
        _inc_path_entry = ttk.Entry(_inc_path_row, textvariable=self.cluster_model_path, width=38)
        _inc_path_entry.pack(side="left", padx=(6, 4))
        Tooltip(_inc_path_entry,
                "Путь для сохранения / загрузки модели кластеризации.\n"
                "Оставьте пустым — файл будет сохранён автоматически в папку clustered/.")
        def _browse_cluster_model():
            from tkinter import filedialog as _fd
            p = _fd.askopenfilename(
                title="Выберите файл модели кластеризации",
                filetypes=[("Joblib файл", "*.joblib"), ("Legacy Pickle", "*.pkl"), ("Все файлы", "*.*")],
            )
            if p:
                self.cluster_model_path.set(p)
        ttk.Button(_inc_path_row, text="📂", command=_browse_cluster_model, width=3).pack(side="left")

        _inc_load_row = ttk.Frame(_inc_frm, style="Card.TFrame")
        _inc_load_row.pack(fill="x", pady=(4, 0))
        cb_load_model = ttk.Checkbutton(
            _inc_load_row,
            text="Применить сохранённую модель (без переобучения)",
            variable=self.use_saved_cluster_model,
            style="Card.TCheckbutton",
        )
        cb_load_model.pack(side="left")
        self.attach_help(
            cb_load_model,
            "Применение сохранённой модели к новым данным",
            "Загружает ранее сохранённую модель и применяет её к новым файлам.\n\n"
            "Новые тексты кодируются тем же векторайзером и относятся к\n"
            "ближайшему существующему кластеру — без пересчёта центроидов.\n\n"
            "Используйте для:\n"
            "  • Ежедневной разметки новых обращений в существующую таксономию\n"
            "  • Проверки новых данных на соответствие старым кластерам\n"
            "  • Инкрементального добавления строк без потери разметки.",
            "Инкрементальная классификация в существующие кластеры",
        )

        _diag_row = ttk.Frame(_inc_frm, style="Card.TFrame")
        _diag_row.pack(fill="x", pady=(4, 0))
        cb_diag = ttk.Checkbutton(
            _diag_row,
            text="Diagnostic mode (расширенные метрики + debug report)",
            variable=self.diagnostic_mode,
            style="Card.TCheckbutton",
        )
        cb_diag.pack(side="left")
        self.attach_help(
            cb_diag,
            "Diagnostic mode",
            "Включает расширенную диагностику инкрементального запуска:\n"
            "  • structured logging с correlation_id,\n"
            "  • сохранение debug-отчёта JSON в папку clustered/.\n\n"
            "Используйте при разборе инцидентов и несовместимых моделей.",
            "Расширенная диагностика и debug-отчёт",
        )

        # ── #20 Потоковый / онлайн-режим (MiniBatchKMeans partial_fit) ───────
        _stream_frm = ttk.LabelFrame(parent, text="Потоковый режим — большие файлы (#20)", padding=(12, 6))
        _stream_frm.pack(fill="x", pady=(0, 6))

        _stream_row = ttk.Frame(_stream_frm, style="Card.TFrame")
        _stream_row.pack(fill="x")
        cb_stream = ttk.Checkbutton(
            _stream_row,
            text="Потоковая кластеризация (MiniBatch partial_fit)",
            variable=self.use_streaming_cluster,
            style="Card.TCheckbutton",
        )
        cb_stream.pack(side="left")
        self.attach_help(
            cb_stream,
            "Потоковый / онлайн-режим кластеризации",
            "Обрабатывает данные чанками — не загружает весь датасет в RAM.\n\n"
            "Как работает:\n"
            "  1. Тексты читаются чанками (chunk_size строк за раз)\n"
            "  2. Для каждого чанка: TF-IDF transform → MiniBatchKMeans.partial_fit()\n"
            "  3. Модель постепенно обновляется без хранения всех данных\n\n"
            "Когда использовать:\n"
            "  • Файлы > 100K строк (не помещаются в RAM)\n"
            "  • Постоянный поток новых данных\n\n"
            "Ограничения: только TF-IDF + KMeans; SBERT/UMAP/HDBSCAN не поддерживаются в потоке.\n"
            "Качество чуть ниже полного KMeans, но работает на любом объёме данных.",
            "Потоковая кластеризация для очень больших файлов",
        )
        ttk.Label(_stream_row, text="  Чанк:", style="Card.TLabel").pack(side="left", padx=(16, 2))
        sp_chunk = ttk.Spinbox(
            _stream_row, from_=500, to=50000, increment=500,
            textvariable=self.streaming_chunk_size, width=7,
        )
        sp_chunk.pack(side="left")
        Tooltip(sp_chunk,
                "Размер чанка (строк за одну итерацию partial_fit).\n"
                "5000 — баланс скорость/качество.\n"
                "Увеличьте до 10 000+ для GPU.")

        # ── #13 ClusterLLM feedback loop ──────────────────────────────────────
        _fb_frm = ttk.LabelFrame(parent, text="ClusterLLM Feedback Loop (#13)", padding=(12, 6))
        _fb_frm.pack(fill="x", pady=(0, 6))

        _fb_row = ttk.Frame(_fb_frm, style="Card.TFrame")
        _fb_row.pack(fill="x")
        cb_fb = ttk.Checkbutton(
            _fb_row,
            text="LLM-анализ кластеров: предложить слияния / разбиения",
            variable=self.use_llm_feedback,
            style="Card.TCheckbutton",
        )
        cb_fb.pack(side="left")
        self.attach_help(
            cb_fb,
            "ClusterLLM Feedback Loop",
            "После именования кластеров LLM анализирует всю таблицу кластеров\n"
            "и предлагает улучшения:\n\n"
            "  MERGE: «Кластеры 3 и 7 похожи — рекомендую объединить»\n"
            "  SPLIT: «Кластер 5 слишком широкий — лучше разбить на 2»\n"
            "  OK: «Разбиение выглядит хорошо»\n\n"
            "Результат записывается в столбец llm_feedback.\n"
            "Слияния применяются автоматически (MERGE → cluster_id обновляется).\n\n"
            "Требует включённый LLM-нейминг (API-ключ выше).",
            "LLM анализирует качество кластеров и предлагает улучшения",
        )

        self._cw(
            _fb_frm, "Столбец обратной связи:", self.llm_feedback_col,
            widget_type="entry", width=20,
            tooltip="Имя столбца в Excel для записи рекомендаций LLM.\nПо умолчанию: llm_feedback",
            tooltip_title="Столбец обратной связи",
        )

        # ── FASTopic параметры ─────────────────────────────────────────────────
        _aft_frm = ttk.LabelFrame(parent, text="Параметры FASTopic", padding=(12, 6))
        _aft_frm.pack(fill="x", pady=(0, 6))
        self._fastopic_frm = _aft_frm
        _ft_row = ttk.Frame(_aft_frm, style="Card.TFrame")
        _ft_row.pack(fill="x")
        ttk.Label(_ft_row, text="Топ-слов на тему:", style="Card.TLabel").pack(side="left")
        ttk.Spinbox(_ft_row, from_=5, to=50, textvariable=self.fastopic_n_top_words, width=5).pack(
            side="left", padx=(4, 0))
        ttk.Label(_ft_row, text="  Theta (разреженность):", style="Card.TLabel").pack(side="left", padx=(16, 2))
        sp_theta = ttk.Spinbox(_ft_row, from_=0.01, to=1.0, increment=0.05,
                                textvariable=self.fastopic_theta, width=6, format="%.2f")
        sp_theta.pack(side="left")
        Tooltip(sp_theta,
                "Параметр разреженности тем.\n"
                "0.1–0.2 = компактные темы (рекомендуется).\n"
                "0.5+ = каждый документ принадлежит нескольким темам.")

    def _build_cluster_tab(self):  # noqa: C901
        # ── Sub-tab frames ─────────────────────────────────────────────────
        _s0 = ttk.Frame(self.tab_cluster)   # Настройки
        _s1 = ttk.Frame(self.tab_cluster)   # Запуск & Лог

        # ══════════════════════════════════════════════════════════════════════
        # ОСНОВНЫЕ элементы управления (всегда видны)
        # ══════════════════════════════════════════════════════════════════════

        # ── Файлы ─────────────────────────────────────────────────────────────
        card = build_cluster_primary_sections(self, _s0)

        # ── Кнопки пресетов ──────────────────────────────────────────────────
        _preset_row_c = ttk.Frame(_s0)
        _preset_row_c.pack(fill="x", pady=(0, 6))
        _btn_save_preset_c = ttk.Button(_preset_row_c, text="💾 Пресет",
                                        command=self._save_preset)
        _btn_save_preset_c.pack(side="left", padx=(0, 4))
        Tooltip(_btn_save_preset_c, "Сохранить текущие настройки кластеризации как пресет.")
        _btn_load_preset_c = ttk.Button(_preset_row_c, text="📂 Пресет ▾",
                                        command=lambda: self._open_preset_menu(_btn_load_preset_c))
        _btn_load_preset_c.pack(side="left")
        Tooltip(_btn_load_preset_c, "Загрузить ранее сохранённый пресет настроек.")

        # ══════════════════════════════════════════════════════════════════════
        # РАСШИРЕННЫЕ НАСТРОЙКИ (сворачиваемый блок)
        # ══════════════════════════════════════════════════════════════════════
        _adv = CollapsibleSection(
            _s0,
            title="Расширенные настройки",
            collapsed=True,
        )
        _adv.pack(fill="x", pady=(0, 6))
        _ab = _adv.body  # удобный псевдоним

        self._build_cluster_columns_card(_ab)
        self._build_cluster_neural_card(_ab)
        self._build_cluster_metrics_card(_ab)
        self._build_cluster_naming_card(_ab)
        self._build_cluster_advanced_card(_ab)

        _run_left = ttk.Frame(_s1, style="Card.TFrame")
        _run_left.pack(fill="both", expand=True)

        self._build_readiness_bar(_run_left, [
            lambda: (
                bool(self.cluster_files),
                f"📄 Файлы: {len(self.cluster_files)} загружено" if self.cluster_files
                else "📄 Файлы: не выбраны",
            ),
        ])
        self.btn_cluster, self.btn_cluster_stop = self._build_action_block(
            _run_left,
            btn_text="▶  Кластеризовать",
            btn_cmd=self.run_cluster,
            progress_var=self.cluster_progress,
            pct_var=self.cluster_pct,
            phase_var=self.cluster_phase,
            speed_var=self.cluster_speed,
            eta_var=self.cluster_eta,
            label="Прогресс кластеризации",
        )
        summ = ttk.Frame(_run_left, style="Card.TFrame", padding=(12, 6))
        summ.pack(fill="x", pady=(0, 6))
        ttk.Label(summ, text="Результат:", style="Card.TLabel").pack(side="left")
        ttk.Label(summ, textvariable=self.last_cluster_summary, style="Card.Muted.TLabel").pack(side="left", padx=10)

        btn_open_clust = ttk.Button(summ, text="📂 Открыть папку результатов",
                                    command=lambda: self._open_directory(CLUST_DIR))
        btn_open_clust.pack(side="right")
        Tooltip(btn_open_clust, f"Открывает папку с файлами кластеризации:\n{CLUST_DIR}")

        _run_hints = CollapsibleSection(_run_left, title="Подсказки по запуску", collapsed=True)
        _run_hints.pack(fill="x", pady=(0, 8))
        ttk.Label(
            _run_hints.body,
            text="1) Сначала выберите файлы.\n2) Подберите алгоритм/режим векторизации.\n"
                 "3) Запустите кластеризацию и проверяйте вкладку «Кластер» справа при необходимости.",
            style="Card.Muted.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=4)

        self.cluster_tree = ttk.Treeview(_run_left, columns=("cluster", "keywords", "count"), show="headings", height=16)
        for c, lbl, w in [("cluster", "Кластер", 90), ("keywords", "Ключевые слова", 900), ("count", "Строк", 120)]:
            self.cluster_tree.heading(c, text=lbl, anchor="w")
            self.cluster_tree.column(c, width=w, anchor="w")
        self.cluster_tree.pack(fill="both", expand=True)

        # ── Register bottom sub-tab strip for Clustering tab ─────────────────
        self._register_sub_tabs(
            2,
            ["Настройки", "Запуск & Лог"],
            [_s0, _s1],
        )

        # Подписка на изменения режима — обновляет доступность несовместимых элементов
        self.cluster_algo.trace_add("write", self._update_cluster_ui_state)
        self.cluster_vec_mode.trace_add("write", self._update_cluster_ui_state)
        self._update_cluster_ui_state()

    # endregion

    # region Вспомогательные методы UI
    def _sync_cluster_file_buttons(self) -> None:
        has_files = bool(getattr(self, "cluster_files", []))
        for _btn_name in ("_cluster_btn_del", "_cluster_btn_clr"):
            _btn = getattr(self, _btn_name, None)
            if _btn is not None:
                _btn.configure(state=("normal" if has_files else "disabled"))

    # ──────────────────────────────────────────────────────────────────────────
    # Динамическое управление состоянием UI кластеризации
    # ──────────────────────────────────────────────────────────────────────────
    def _update_cluster_ui_state(self, *_):
        """Включает / отключает виджеты кластеризации в зависимости от
        выбранного алгоритма и режима векторизации, исключая несовместимые
        комбинации (LDA/BERTopic игнорируют cluster_vec_mode; HDBSCAN/BERTopic
        не используют K и elbow; якоря работают только с SBERT/Combo + KMeans)."""
        algo     = self.cluster_algo.get()
        vec_mode = self.cluster_vec_mode.get()

        # Только KMeans и GMM поддерживают K / n_init / Avto-K / K-metric
        kmeans_only = algo in ("kmeans", "gmm")
        # LDA, BERTopic и FASTopic всегда используют собственную векторизацию
        uses_own_vec = algo in ("lda", "bertopic", "fastopic")
        # BERTopic и FASTopic всегда запускают UMAP внутри — внешний чекбокс неприменим
        bertopic = algo in ("bertopic", "fastopic")
        # HDBSCAN-like алгоритмы (hdbscan / bertopic) требуют HDBSCAN-параметры
        hdbscan_like = algo in ("hdbscan", "bertopic")
        # Ансамбль: TF-IDF + два SBERT, автовыбор лучшего
        use_ensemble = (not uses_own_vec) and vec_mode == "ensemble"
        # Якоря работают только для SBERT/Combo-векторизации И KMeans
        anchors_ok = (vec_mode in ("sbert", "combo")) and (algo == "kmeans")
        # SBERT-comboboxes нужны для SBERT/Combo/Ensemble-режима или BERTopic/FASTopic
        sbert_needed = (not uses_own_vec and vec_mode in ("sbert", "combo", "ensemble")) or bertopic
        # Combo-параметры (SVD-dim, alpha) — только для Combo-режима (не LDA/BERTopic/FASTopic)
        combo_visible = (not uses_own_vec) and vec_mode == "combo"

        def _s(ok: bool) -> str:
            return "normal" if ok else "disabled"

        # K / n_init / Авто-подбор K
        for w in getattr(self, "_cluster_k_widgets", []):
            try:
                w.configure(state=_s(kmeans_only))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # Режим векторизации (TF-IDF / SBERT / Combo)
        for rb in getattr(self, "_cluster_vec_rbs", []):
            try:
                rb.configure(state=_s(not uses_own_vec))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # Ансамбль совместим только с KMeans — отключаем остальные алгоритмы
        for rb in getattr(self, "_cluster_algo_rbs", []):
            try:
                _algo_val = rb.cget("value")
                rb.configure(state=_s(not use_ensemble or _algo_val == "kmeans"))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # Если ансамбль активен, а текущий алгоритм — не KMeans:
        # сбрасываем на KMeans и сообщаем пользователю в лог
        if use_ensemble and self.cluster_algo.get() != "kmeans":
            _prev_algo = self.cluster_algo.get()
            self.cluster_algo.set("kmeans")  # вызовет trace, но рекурсии нет (algo уже kmeans)
            self.log_cluster(
                f"ℹ️  Ансамбль (TF-IDF + 2 SBERT) совместим только с KMeans.\n"
                f"   Алгоритм «{_prev_algo}» сброшен → KMeans автоматически.\n"
                f"   Причина: ансамбль сам выбирает лучшую векторизацию через Silhouette;\n"
                f"   HDBSCAN / LDA / Иерарх. требуют свои специализированные матрицы\n"
                f"   и не совместимы с режимом автовыбора векторизации."
            )

        # SBERT model / device comboboxes
        for w in getattr(self, "_sbert_clust_widgets", []):
            try:
                w.configure(state=_s(sbert_needed))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # Combo SVD-dim / alpha spinboxes
        for w in getattr(self, "_combo_clust_widgets", []):
            try:
                w.configure(state=_s(combo_visible))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # Ensemble second-model combobox
        for w in getattr(self, "_ensemble_widgets", []):
            try:
                w.configure(state=_s(use_ensemble))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # Метрика автоподбора K
        for rb in getattr(self, "_k_metric_rbs", []):
            try:
                rb.configure(state=_s(kmeans_only))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # UMAP — не применим к BERTopic (уже встроен) и к LDA (требует count-матрицу)
        for w in getattr(self, "_umap_clust_widgets", []):
            try:
                w.configure(state=_s(not bertopic and algo != "lda"))
            except Exception as _e:
                _log.debug("UI-state: widget configure failed: %s", _e)

        # SVD для TF-IDF — применим только в чистом TF-IDF режиме (не SBERT/Combo/LDA/BERTopic)
        _tfidf_svd_ok = (not uses_own_vec) and vec_mode == "tfidf"
        try:
            self._set_frame_state(self._tfidf_svd_row, _tfidf_svd_ok)
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # Косинусная метрика — неприменима к LDA (count-матрица) и BERTopic (свой пайплайн)
        try:
            self._cb_cosine.configure(state=_s(algo not in ("lda", "bertopic")))
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # Лемматизация бесполезна при SBERT/Combo/Ensemble/BERTopic — трансформеры сами
        # обрабатывают морфологию. При TF-IDF — полезна и включена.
        _lemma_ok = (uses_own_vec and algo == "lda") or (
            not uses_own_vec and vec_mode == "tfidf"
        )
        try:
            self._cb_lemma_cl.configure(state=_s(_lemma_ok))
            if not _lemma_ok and self.use_lemma_cluster.get():
                # Сбрасываем значение при переключении в режим, где лемматизация
                # не поддерживается — иначе snap["use_lemma_cluster"] останется True
                # и worker запустит Lemmatizer в SBERT/Combo/BERTopic-режиме.
                self.use_lemma_cluster.set(False)
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # HDBSCAN-параметры (min_cluster_size) — только для HDBSCAN / BERTopic
        try:
            self._set_frame_state(self._hdb_row, hdbscan_like)
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # LDA-параметры — только для LDA
        try:
            self._set_frame_state(self._lda_row, algo == "lda")
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # Иерархические параметры — только для hierarchical
        try:
            self._set_frame_state(self._hier_row, algo == "hierarchical")
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # BERTopic-параметры — только для BERTopic
        try:
            self._set_frame_state(self._bert_row, algo == "bertopic")
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # FASTopic-параметры — только для FASTopic
        try:
            self._set_frame_state(getattr(self, "_fastopic_frm", None), algo == "fastopic")
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

        # Якоря
        try:
            self._cb_anchors.configure(state=_s(anchors_ok))
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)
        try:
            self.txt_anchors.configure(state=_s(anchors_ok))
        except Exception as _e:
            _log.debug("UI-state: widget configure failed: %s", _e)

    @staticmethod
    def _set_frame_state(frame, enabled: bool):
        """Рекурсивно включает/отключает все виджеты внутри frame."""
        state = "normal" if enabled else "disabled"
        for child in frame.winfo_children():
            try:
                child.configure(state=state)
            except Exception as _e:
                _log.debug("_set_frame_state: child configure failed: %s", _e)

    def log_cluster(self, msg: str): self._log_to(self.cluster_log, msg)

    # ──────────────────────────────────────────────────────────────────────────
    # Автоопределение оптимальных параметров кластеризации
    # ──────────────────────────────────────────────────────────────────────────
    def _auto_detect_cluster_params(self):
        """Анализирует файлы кластеризации и устанавливает оптимальные параметры.

        Запускает фоновый анализ (без блокировки UI):
          1. Быстро читает все файлы (до SAMPLE_LIMIT строк),
          2. Оценивает размер датасета, длину текстов, разнообразие, диалоговые роли,
          3. Применяет эвристики и устанавливает параметры в главном потоке.
        """
        if not self.cluster_files:
            messagebox.showwarning(
                "Анализ параметров",
                "Сначала добавь хотя бы один файл для кластеризации.",
            )
            return

        self._detect_status_var.set("⏳ Анализ…")
        files_snapshot: List[str] = list(self.cluster_files)

        # Снимаем имена колонок в главном потоке
        col_names = {
            "desc":      self.desc_col.get().strip(),
            "call":      self.call_col.get().strip(),
            "chat":      self.chat_col.get().strip(),
            "summary":   self.summary_col.get().strip(),
            "ans_short": self.ans_short_col.get().strip(),
            "ans_full":  self.ans_full_col.get().strip(),
        }

        def worker():
            try:
                _analysis = _analyze_cluster_input(files_snapshot, col_names)
                total_rows       = _analysis["total_rows"]
                n_text_rows      = _analysis["n_text_rows"]
                text_lengths     = _analysis["text_lengths"]
                has_dialog_roles = _analysis["has_dialog_roles"]
                has_chatbot      = _analysis["has_chatbot"]
                col_hits         = _analysis["col_hits"]
                SAMPLE_LIMIT     = _analysis["SAMPLE_LIMIT"]
                sample_size      = min(total_rows, SAMPLE_LIMIT)

                if n_text_rows == 0:
                    raise ValueError(
                        "В выборке не найдено строк с текстом по указанным колонкам.\n"
                        "Проверьте, что названия колонок (Описание, Звонок, Чат …) "
                        "совпадают с заголовками файла."
                    )

                params, reasons = _compute_cluster_heuristics(_analysis)

                avg_len      = sum(text_lengths) / max(1, len(text_lengths))
                max_len      = max(text_lengths) if text_lengths else 0
                vocab_sample = len(set(_analysis["sampled_words"]))
                vocab_est    = int(vocab_sample * (total_rows / max(1, sample_size)) ** 0.6)
                text_density = n_text_rows / max(1, sample_size)
                active_cols  = [k for k, v in col_hits.items() if v > sample_size * 0.05]

                def apply_params(
                    params=params, reasons=reasons,
                    total_rows=total_rows, sample_size=sample_size,
                    avg_len=avg_len, max_len=max_len,
                    vocab_est=vocab_est, text_density=text_density,
                    active_cols=active_cols,
                    has_dialog_roles=has_dialog_roles, has_chatbot=has_chatbot,
                ):
                    # Проверяем, был ли выбран ансамбль до авто-подбора
                    _was_ensemble = self.cluster_vec_mode.get() == "ensemble"

                    for key, val in params.items():
                        var = getattr(self, key, None)
                        if var is not None:
                            try:
                                var.set(val)
                            except Exception as _e:
                                _log.debug("auto-detect var.set(%s=%r): %s", key, val, _e)

                    # Если пользователь вручную выбрал «Ансамбль» — авто-подбор
                    # заменил его на стандартный режим (tfidf/sbert/combo).
                    # Предупреждаем: авто-подбор не рекомендует ансамбль автоматически.
                    if _was_ensemble:
                        reasons.append(
                            f"⚠️  Режим «Ансамбль» заменён на «{params.get('cluster_vec_mode', '?')}»  "
                            "(авто-подбор выбирает оптимальный одиночный метод; "
                            "ансамбль — ручной выбор для сравнения трёх методов)"
                        )

                    self.log_cluster("\n──────────── Автоопределение параметров ────────────")
                    self.log_cluster(
                        f"Файлов: {len(files_snapshot)} | "
                        f"Всего строк: {total_rows} | "
                        f"Выборка: {sample_size}"
                    )
                    self.log_cluster(
                        f"Ср. длина текста: {avg_len:.0f} симв. | "
                        f"Макс: {max_len} | "
                        f"Словарь ~{vocab_est // 1000}k"
                    )
                    self.log_cluster(
                        f"Текстовая плотность: {text_density:.0%} | "
                        f"Активные поля: {', '.join(active_cols) or '—'}"
                    )
                    self.log_cluster(
                        f"Диалоговые роли: {'да' if has_dialog_roles else 'нет'} | "
                        f"Чат-бот реплики: {'да' if has_chatbot else 'нет'}"
                    )
                    self.log_cluster("Рекомендации:")
                    for r in reasons:
                        self.log_cluster(f"  • {r}")
                    self.log_cluster("────────────────────────────────────────────────────")
                    _algo_short = params.get("cluster_algo", "kmeans")
                    self._detect_status_var.set(
                        f"✅  K={params.get('k_clusters')}, "
                        f"{params.get('cluster_vec_mode')}+{_algo_short}, "
                        f"строк={total_rows}"
                    )

                self.after(0, apply_params)

            except Exception as e:
                self.after(0, lambda e=e: (
                    self._detect_status_var.set(f"⚠️ Ошибка: {e}"),
                    messagebox.showerror("Ошибка анализа параметров", str(e)),
                ))

        threading.Thread(target=worker, daemon=True).start()

    def add_cluster_files(self):
        paths = filedialog.askopenfilenames(
            title="Файлы для кластеризации",
            filetypes=[
                ("Таблицы (xlsx, csv)", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Все файлы", "*.*"),
            ],
        )
        if not paths:
            return
        added = 0
        for p in paths:
            if p not in self.cluster_files:
                self.cluster_files.append(p)
                self.lb_cluster.insert("end", p)
                added += 1
        if self.cluster_files:
            try:
                headers = read_headers(Path(self.cluster_files[0]))
                self._refresh_combobox_values(headers)
                self.log_cluster(f"Файлов: {len(self.cluster_files)} | Заголовков: {len(headers)}")
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))
        self._sync_cluster_file_buttons()

    def add_cluster_folder(self):
        folder = filedialog.askdirectory(title="Выбери папку с файлами для кластеризации")
        if not folder:
            return
        folder_path = Path(folder)
        found = sorted(
            p for p in folder_path.iterdir()
            if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm", ".csv")
        )
        if not found:
            messagebox.showinfo("Папка", f"В папке не найдено xlsx/csv файлов:\n{folder}")
            return
        added = 0
        for p in found:
            s = str(p)
            if s not in self.cluster_files:
                self.cluster_files.append(s)
                self.lb_cluster.insert("end", s)
                added += 1
        if added == 0:
            self.log_cluster(f"Папка: файлы уже добавлены ({len(found)} найдено).")
            return
        try:
            headers = read_headers(Path(self.cluster_files[0]))
            self._refresh_combobox_values(headers)
            self.log_cluster(f"Папка: +{added} файл(ов) | всего: {len(self.cluster_files)} | заголовков: {len(headers)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
        self._sync_cluster_file_buttons()

    def remove_cluster_file(self):
        sel = list(self.lb_cluster.curselection())
        if not sel:
            return
        i = sel[0]
        self.lb_cluster.delete(i)
        del self.cluster_files[i]
        self.log_cluster(f"Файлов кластеризации: {len(self.cluster_files)}")
        self._sync_cluster_file_buttons()

    def clear_cluster_files(self):
        self.cluster_files = []
        self.lb_cluster.delete(0, "end")
        self.log_cluster("Файлы кластеризации очищены.")
        self._sync_cluster_file_buttons()

    def pick_cluster_file(self):
        """Оставлен для обратной совместимости — вызывает add_cluster_files()."""
        self.add_cluster_files()

    # endregion

    # region Запуск кластеризации

    # ── ClusterRunState: общее состояние между стадиями pipeline ──────────────
    from typing import Any as _Any

    @_dc_dataclass(slots=True)
    class ClusterRunState:
        """Передаёт состояние между стадиями pipeline кластеризации.

        ``slots=True`` отказывается от динамических полей — опечатка вида
        ``_crs.labls = ...`` (вместо ``labels``) сразу бросает
        ``AttributeError`` вместо тихого создания левого атрибута.
        Frozen-snapshots стадий доступны через :mod:`cluster_run_stages`.
        """
        # Инициализируется в worker()
        in_paths: list = _dc_field(default_factory=list)
        total_rows: int = 0
        start_ts: float = 0.0
        done: int = 0
        # Стадия 1: чтение файлов
        X_all: list = _dc_field(default_factory=list)
        X_clean: list = _dc_field(default_factory=list)
        file_data: list = _dc_field(default_factory=list)
        raw_texts_all: list = _dc_field(default_factory=list)
        cluster_snap: _Any = None
        n_ok: int = 0
        use_t5: bool = False
        # Стадия 2: векторизация + кластеризация
        Xv: _Any = None
        Xv_tfidf: _Any = None
        labels: _Any = None
        K: int = 0
        kw: list = _dc_field(default_factory=list)
        kw_final: list = _dc_field(default_factory=list)
        vec_kw: _Any = None
        dedup_map: dict = _dc_field(default_factory=dict)
        X_clean_dd: list = _dc_field(default_factory=list)
        dedup_reverse_map: _Any = None
        labels_l1: _Any = None
        noise_labels: _Any = None
        km: _Any = None
        use_hdbscan: bool = False
        use_lda: bool = False
        use_hier: bool = False
        use_bertopic: bool = False
        use_ensemble: bool = False
        use_gmm: bool = False
        use_fastopic: bool = False
        use_fastopic_kw_ready: bool = False
        stop_list: _Any = None
        inc_labels_done: bool = False
        # Стадия 3: постобработка
        cluster_names: dict = _dc_field(default_factory=dict)
        cluster_reasons: dict = _dc_field(default_factory=dict)
        cluster_quality: dict = _dc_field(default_factory=dict)
        t5_summaries: list = _dc_field(default_factory=list)
        # имена полей, используемые в stage 3 → 4
        cluster_name_map: dict = _dc_field(default_factory=dict)
        cluster_reason_map: dict = _dc_field(default_factory=dict)
        cluster_quality_map: dict = _dc_field(default_factory=dict)
        llm_feedback_map: dict = _dc_field(default_factory=dict)
        t5_summaries_all: list = _dc_field(default_factory=list)
        kw_dict: dict = _dc_field(default_factory=dict)
        stamp: str = ""
        use_streaming: bool = False
        use_inc_model: bool = False
        hdbscan_proba: _Any = None   # soft membership [0..1] per sample after HDBSCAN fit

    @staticmethod
    def _freeze_snap(snap: Mapping[str, Any]) -> Mapping[str, Any]:
        """Thin mixin alias for `snap_utils.freeze_snap`.

        Kept as an instance-accessible staticmethod because existing
        stage helpers call `self._freeze_snap(snap)`; the shared
        implementation lives in `snap_utils` so the service layer
        (`ClusteringWorkflow.run`) and any future batch/CLI driver
        share one freeze boundary.
        """
        return _shared_freeze_snap(snap)

    def _cluster_prepare_data(self, files_snapshot: List[str], snap: Dict[str, Any]):
        """Stage: prepare_inputs (pipeline adapter for orchestration layer)."""
        frozen = self._freeze_snap(snap)
        prepared_inputs = prepare_inputs(files_snapshot, frozen)
        role_ctx = prepared_inputs.role_context
        return prepared_inputs, role_ctx.cluster_snap, role_ctx.role_label, role_ctx.ignore_chatbot_label

    def _cluster_fit_predict(self, prepared_inputs, snap: Dict[str, Any]):
        """Stage: fit/predict adapters (no behavior change in this refactor step)."""
        frozen = self._freeze_snap(snap)
        vectors_stage = build_vectors(prepared_inputs, frozen)
        return run_clustering(vectors_stage, frozen)

    def _cluster_postprocess(self, cluster_stage, prepared_inputs, snap: Dict[str, Any]):
        """Stage: postprocess adapter (no behavior change in this refactor step)."""
        return postprocess_clusters(cluster_stage, prepared_inputs, self._freeze_snap(snap))

    def _cluster_export(self, post_stage, snap: Dict[str, Any]):
        """Stage: export adapter (no behavior change in this refactor step)."""
        return export_cluster_outputs(post_stage, self._freeze_snap(snap))

    def export_cluster_results(self):
        """Копирует последние файлы кластеризации в выбранную папку."""
        paths = getattr(self, "_last_cluster_out_paths", None)
        if not paths:
            messagebox.showinfo(
                "Экспорт кластеров",
                "Сначала выполните кластеризацию — затем можно экспортировать результат."
            )
            return
        dst_dir = filedialog.askdirectory(title="Куда сохранить файлы кластеризации?")
        if not dst_dir:
            return
        import shutil
        copied = 0
        for src in paths:
            try:
                shutil.copy(src, dst_dir)
                copied += 1
            except Exception as e:
                self.log_cluster(f"Ошибка копирования {Path(src).name}: {e}")
        self.log_cluster(f"Экспорт: {copied} файл(ов) → {dst_dir}")

    def _prepare_cluster_run_context(self):
        """Готовит snapshot и preconditions для run_cluster.

        Preflight резолвится через модульный атрибут `validate_cluster_preconditions`,
        что позволяет тестам монки-патчить его на уровне `app_cluster`.
        """
        if not validate_cluster_preconditions(self):
            with self._proc_lock:
                self._processing = False
            return None, None
        return prepare_cluster_run_context(self)

    def _cluster_step_llm_naming(self, snap: dict, labels: list, X_clean: list, Xv, cluster_name_map: dict) -> None:
        # ── LLM-нейминг кластеров (#4) ────────────────────────────
        # ── Центроид-ближайшие тексты для LLM ───────────────────
        # Вычисляем один раз и используем везде вместо случайного отбора.
        _n_repr = int(snap.get("n_repr_examples", 5))
        _repr_texts: dict = {}
        try:
            from ml_diagnostics import find_cluster_representative_texts as _find_repr
            import scipy.sparse as _sp_repr
            _Xv_repr = (Xv.toarray() if _sp_repr.issparse(Xv) else _np.asarray(Xv))
            _repr_texts = _find_repr(X_clean, _np.asarray(labels), _Xv_repr, n_top=_n_repr)
        except Exception as _repr_e:
            _log.debug("Representative texts fallback: %s", _repr_e)
            # Fallback: первые N по порядку
            _la_fb = _np.asarray(labels)
            for _fi in range(K):
                _ii = [j for j, c in enumerate(_la_fb) if int(c) == _fi][:_n_repr]
                _repr_texts[_fi] = [str(X_clean[j])[:300] for j in _ii if j < len(X_clean)]

        cluster_name_map: dict = {}
        if snap.get("use_llm_naming", False) and kw:
            try:
                _llm_provider = snap.get("llm_provider", "anthropic").lower()
                _llm_model    = snap.get("llm_model", "claude-sonnet-4-6")
                _llm_api_key  = self._resolve_llm_api_key(snap, _llm_provider)
                if (not _llm_api_key) and _llm_provider != "ollama":
                    self.after(0, lambda: self.log_cluster(
                        "⚠️ LLM-нейминг: API-ключ не задан — пропускаю"
                    ))
                else:
                    self.after(0, lambda: self.log_cluster(
                        f"LLM-нейминг кластеров: {_llm_provider}/{_llm_model}…"
                    ))
                    # Few-shot примеры в системном промпте дают модели эталон
                    # формата «глагол/существительное-проблема + уточнение»
                    # и единый тон по всем кластерам — без них модель часто
                    # выбирает слишком общие («Вопросы по картам») или слишком
                    # длинные названия. Temperature=0.2 фиксирует результат:
                    # у одних и тех же входов название воспроизводимо между
                    # запусками, что важно для сравнения экспериментов.
                    _sys_prompt = (
                        "Ты — аналитик клиентских обращений банка. "
                        "Сформулируй краткое название (3–6 слов) на русском языке, "
                        "отражающее КОНКРЕТНУЮ причину или проблему обращения клиента. "
                        "Используй глагол или существительное-проблему, "
                        "избегай общих слов («вопрос», «проблема» без уточнения). "
                        "Отвечай только названием, без пояснений и без кавычек.\n\n"
                        "Примеры корректных названий:\n"
                        "• «Задержка зачисления платежа»\n"
                        "• «Блокировка карты при оплате»\n"
                        "• «Списание комиссии за обслуживание»\n"
                        "• «Оспаривание операции по карте»\n"
                        "• «Невозможно войти в мобильное приложение»\n\n"
                        "Плохие (слишком общие) названия: "
                        "«Вопрос по карте», «Обращение клиента», «Проблема с банком»."
                    )
                    for _ci, _kw_str in enumerate(kw):
                        if not _kw_str:
                            continue
                        _ex_for_naming = _repr_texts.get(_ci, [])
                        _user_msg = (
                            f"Ключевые признаки: {_kw_str}\n\n"
                            "Типичные обращения:\n" +
                            "\n".join(f"• {ex}" for ex in _ex_for_naming)
                        )
                        try:
                            _name = self._llm_complete_text(
                                provider=_llm_provider,
                                model=_llm_model,
                                api_key=_llm_api_key,
                                system_prompt=_sys_prompt,
                                user_prompt=_user_msg,
                                max_tokens=64,
                                temperature=0.2,
                            )
                            if _name:
                                cluster_name_map[_ci] = _name
                        except Exception as _llm_row_e:
                            self.after(0, lambda e=_llm_row_e, i=_ci: self.log_cluster(
                                f"⚠️ LLM кластер {i}: {e}"
                            ))
                    self.after(0, lambda n=len(cluster_name_map): self.log_cluster(
                        f"LLM-нейминг: названы {n} кластеров ✅"
                    ))
            except Exception as _llm_e:
                self.after(0, lambda e=_llm_e: self.log_cluster(
                    f"⚠️ LLM-нейминг: {e}"
                ))

        # ── #13 ClusterLLM Feedback Loop ──────────────────────────

    def _cluster_step_llm_feedback(self, snap: dict, labels: list, cluster_name_map: dict, cluster_reason_map: dict, cluster_quality_map: dict) -> None:
        # Второй LLM-вызов: анализ кластеров → MERGE/SPLIT/OK советы
        cluster_reason_map: dict = {}
        if snap.get("use_llm_reason_summary", False) and kw and snap.get("use_llm_naming", False):
            try:
                _llm_provider_r = snap.get("llm_provider", "anthropic").lower()
                _llm_model_r    = snap.get("llm_model", "claude-sonnet-4-6")
                _llm_api_key_r  = self._resolve_llm_api_key(snap, _llm_provider_r)
                if (not _llm_api_key_r) and _llm_provider_r != "ollama":
                    self.after(0, lambda: self.log_cluster(
                        "⚠️ LLM-описание причин: API-ключ не задан — пропускаю"
                    ))
                else:
                    self.after(0, lambda: self.log_cluster(
                        "LLM: формирую обобщённые описания причин по кластерам…"
                    ))
                    for _ci, _kw_str in enumerate(kw):
                        _examples = _repr_texts.get(_ci, [])
                        _title = cluster_name_map.get(_ci, "")
                        _sys_r = (
                            "Ты — аналитик качества клиентского сервиса банка. "
                            "Ответь на вопрос: по какой конкретной причине клиенты "
                            "обращаются в этом кластере? "
                            "Напиши 2–3 предложения: что случилось, чего хотел клиент, "
                            "какова типичная ситуация. "
                            "Не пересказывай ключевые слова — делай вывод из реальных примеров. "
                            "Пиши конкретно, без общих фраз."
                        )
                        _user_r = (
                            f"Тема кластера: «{_title}»\n"
                            f"Ключевые признаки: {_kw_str}\n\n"
                            "Наиболее характерные обращения:\n" +
                            "\n".join(f"• {ex}" for ex in _examples)
                        )
                        try:
                            _reason = self._llm_complete_text(
                                provider=_llm_provider_r,
                                model=_llm_model_r,
                                api_key=_llm_api_key_r,
                                system_prompt=_sys_r,
                                user_prompt=_user_r,
                                max_tokens=220,
                                temperature=0.3,
                            )
                            if _reason:
                                cluster_reason_map[_ci] = _reason
                        except Exception as _llm_reason_e:
                            self.after(0, lambda e=_llm_reason_e, i=_ci: self.log_cluster(
                                f"⚠️ LLM-описание причин кластер {i}: {e}"
                            ))
                    self.after(0, lambda n=len(cluster_reason_map): self.log_cluster(
                        f"LLM-описание причин: заполнено {n} кластеров ✅"
                    ))
            except Exception as _llm_reason_outer_e:
                self.after(0, lambda e=_llm_reason_outer_e: self.log_cluster(
                    f"⚠️ LLM-описание причин: {e}"
                ))

        if snap.get("use_rule_reason_summary", True) and kw:
            try:
                _filled = 0
                for _ci, _kw_str in enumerate(kw):
                    if cluster_reason_map.get(_ci):
                        continue
                    _examples2 = _repr_texts.get(_ci, [])
                    _title2 = cluster_name_map.get(_ci, "")
                    cluster_reason_map[_ci] = self._build_reason_without_llm(
                        cluster_name=_title2,
                        keywords=_kw_str,
                        examples=_examples2,
                    )
                    _filled += 1
                self.after(0, lambda n=_filled: self.log_cluster(
                    f"Эвристическое описание причин: заполнено {n} кластеров ✅"
                ))
            except Exception as _rule_reason_e:
                self.after(0, lambda e=_rule_reason_e: self.log_cluster(
                    f"⚠️ Эвристическое описание причин: {e}"
                ))

        llm_feedback_map: dict = {}
        if snap.get("use_llm_feedback", False) and kw and snap.get("use_llm_naming", False):
            try:
                _llm_provider2 = snap.get("llm_provider", "anthropic").lower()
                _llm_model2    = snap.get("llm_model", "claude-sonnet-4-6")
                _llm_api_key2  = self._resolve_llm_api_key(snap, _llm_provider2)
                if (not _llm_api_key2) and _llm_provider2 != "ollama":
                    self.after(0, lambda: self.log_cluster(
                        "⚠️ ClusterLLM Feedback: API-ключ не задан — пропускаю"
                    ))
                else:
                    self.after(0, lambda: self.log_cluster(
                        "ClusterLLM Feedback: анализ кластеров…"
                    ))
                    # Собираем сводную таблицу кластеров для LLM
                    _clust_summary_lines = []
                    _labels_cnt = {}
                    for _lbl in labels:
                        _labels_cnt[int(_lbl)] = _labels_cnt.get(int(_lbl), 0) + 1
                    for _ci, _kw_str in enumerate(kw):
                        _cname = cluster_name_map.get(_ci, "")
                        _cnt = _labels_cnt.get(_ci, 0)
                        _name_part = f" ({_cname})" if _cname else ""
                        _clust_summary_lines.append(
                            f"  Кластер {_ci}{_name_part}: {_cnt} строк | {_kw_str[:80]}"
                        )
                    _clust_table = "\n".join(_clust_summary_lines[:60])  # не перегружаем LLM
                    _fb_sys = (
                        "Ты — эксперт по кластеризации текстов банковских обращений. "
                        "Проанализируй таблицу кластеров и для каждого выдай рекомендацию:\n"
                        "  MERGE X Y — слить кластеры X и Y (слишком похожи)\n"
                        "  SPLIT X — разбить кластер X на 2 (слишком широкий)\n"
                        "  OK X — кластер X хорошего качества\n"
                        "Одна рекомендация на строку. Только рекомендации, без пояснений."
                    )
                    _fb_user = f"Кластеры:\n{_clust_table}"
                    try:
                        _fb_text = self._llm_complete_text(
                            provider=_llm_provider2,
                            model=_llm_model2,
                            api_key=_llm_api_key2,
                            system_prompt=_fb_sys,
                            user_prompt=_fb_user,
                            max_tokens=512,
                        )
                        # Разбираем рекомендации
                        _merge_map: dict = {}  # {old_id: new_id}
                        for _line in _fb_text.splitlines():
                            _line = _line.strip()
                            if _line.upper().startswith("MERGE"):
                                _parts = _line.split()
                                if len(_parts) >= 3:
                                    try:
                                        _a, _b = int(_parts[1]), int(_parts[2])
                                        if 0 <= _a < K and 0 <= _b < K and _a != _b:
                                            _merge_map[max(_a, _b)] = min(_a, _b)
                                            llm_feedback_map[_a] = f"MERGE→{min(_a,_b)}"
                                            llm_feedback_map[_b] = f"MERGE→{min(_a,_b)}"
                                    except ValueError:
                                        _log.debug("ClusterLLM Feedback: MERGE — нечисловой токен, строка пропущена: %r", _line)
                            elif _line.upper().startswith("SPLIT"):
                                _parts = _line.split()
                                if len(_parts) >= 2:
                                    try:
                                        _cidx = int(_parts[1])
                                        if 0 <= _cidx < K:
                                            llm_feedback_map[_cidx] = "SPLIT"
                                    except ValueError:
                                        _log.debug("ClusterLLM Feedback: SPLIT — нечисловой токен, строка пропущена: %r", _line)
                            elif _line.upper().startswith("OK"):
                                _parts = _line.split()
                                if len(_parts) >= 2:
                                    try:
                                        _cidx = int(_parts[1])
                                        if 0 <= _cidx < K:
                                            llm_feedback_map[_cidx] = "OK"
                                    except ValueError:
                                        _log.debug("ClusterLLM Feedback: OK — нечисловой токен, строка пропущена: %r", _line)
                        # Применяем слияния к labels
                        if _merge_map:
                            _labels_arr_fb = _np.asarray(labels, dtype=int)
                            for _old, _new in _merge_map.items():
                                _labels_arr_fb[_labels_arr_fb == _old] = _new
                            labels = _labels_arr_fb
                            self.after(0, lambda n=len(_merge_map): self.log_cluster(
                                f"ClusterLLM Feedback: применено {n} слияний ✅"
                            ))
                        self.after(0, lambda t=_fb_text[:300]: self.log_cluster(
                            f"ClusterLLM ответ:\n{t}"
                        ))
                    except Exception as _fb_call_e:
                        self.after(0, lambda e=_fb_call_e: self.log_cluster(
                            f"⚠️ ClusterLLM вызов: {e}"
                        ))
            except Exception as _fb_e:
                self.after(0, lambda e=_fb_e: self.log_cluster(
                    f"⚠️ ClusterLLM Feedback: {e}"
                ))

        # ── Plotly-визуализация (2D UMAP → HTML) (#11) ────────────

    def _cluster_step_plotly(self, snap: dict, Xv, labels: list, X_clean: list) -> None:
        if snap.get("use_cluster_viz", False):
            try:
                import plotly.express as _px
                import pandas as _pd_viz
                self.after(0, lambda: self.log_cluster("Plotly: строю 2D-визуализацию…"))
                import scipy.sparse as _sp_viz
                _Xviz = Xv.toarray() if _sp_viz.issparse(Xv) else _np.asarray(Xv)
                # Ограничиваем выборку для скорости
                _viz_max = 10_000
                _viz_n = min(_viz_max, len(_Xviz))
                if len(_Xviz) > _viz_max:
                    _viz_idx = _np.random.choice(len(_Xviz), _viz_max, replace=False)
                    _Xviz = _Xviz[_viz_idx]
                    _viz_lbl = _np.asarray(labels)[_viz_idx]
                    _viz_texts = [X_clean[i] for i in _viz_idx]
                else:
                    _viz_idx = None
                    _viz_lbl = _np.asarray(labels)
                    _viz_texts = list(X_clean)
                # UMAP 2D
                try:
                    import umap as _umap_viz
                    _umap2d = _umap_viz.UMAP(
                        n_components=2,
                        n_neighbors=15,
                        min_dist=0.1,
                        random_state=snap.get("cluster_random_seed", 42),
                        low_memory=True,
                    )
                    _emb2d = _umap2d.fit_transform(_Xviz)
                except ImportError:
                    # Fallback: PCA 2D
                    from sklearn.decomposition import PCA as _PCA2
                    _emb2d = _PCA2(n_components=2, random_state=snap.get("cluster_random_seed", 42)).fit_transform(_Xviz)
                _df_viz = _pd_viz.DataFrame({
                    "x": _emb2d[:, 0],
                    "y": _emb2d[:, 1],
                    "cluster": [str(int(l)) for l in _viz_lbl],
                    "text": [str(t)[:120] for t in _viz_texts],
                })
                _fig = _px.scatter(
                    _df_viz, x="x", y="y",
                    color="cluster",
                    hover_data={"text": True, "x": False, "y": False},
                    title=f"Кластеры: K={K}, n={_viz_n}",
                    width=1100, height=750,
                )
                _fig.update_traces(marker=dict(size=5, opacity=0.7))
                _viz_path = CLUST_DIR / f"clusters_viz_{now_stamp()}.html"
                _fig.write_html(str(_viz_path))
                self.after(0, lambda p=_viz_path: self.log_cluster(
                    f"Визуализация сохранена: {p.name} ✅"
                ))
            except ImportError as _viz_ie:
                self.after(0, lambda e=_viz_ie: self.log_cluster(
                    f"⚠️ Визуализация пропущена — {e}"
                ))
            except Exception as _viz_e:
                self.after(0, lambda e=_viz_e: self.log_cluster(
                    f"⚠️ Визуализация: {e}"
                ))

        # ── T5-суммаризация (опционально) ─────────────────────────

    def _cluster_step_t5(self, snap: dict, raw_texts_all: list, t5_summaries_all: list) -> None:
        t5_summaries_all: List[str] = []
        if use_t5 and raw_texts_all:
            try:
                from t5_summarizer import T5RussianSummarizer
                _t5_model_name = snap.get("t5_model_name", "UrukHan/t5-russian-summarization")
                _t5_batch      = snap.get("t5_batch_size", 4)
                _t5_max_in     = snap.get("t5_max_input", 512)
                _t5_max_out    = snap.get("t5_max_output", 128)

                def _t5_prog_cb(frac: float, status: str):
                    pct = 90.0 + frac * 7.0   # 90..97 %
                    self.after(0, lambda p=pct, s=status: ui_prog(p, s))

                t5 = T5RussianSummarizer(
                    model_name=_t5_model_name,
                    max_input_length=_t5_max_in,
                    max_target_length=_t5_max_out,
                    batch_size=_t5_batch,
                    device=snap.get("sbert_device", "auto"),
                    log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                    progress_cb=_t5_prog_cb,
                )
                self.after(0, lambda: ui_prog(90.0, "T5: загрузка модели суммаризации…"))
                t5_summaries_all = t5.summarize(raw_texts_all)
                self.after(0, lambda n=len(t5_summaries_all): self.log_cluster(
                    f"T5: суммаризировано {n} строк ✅"
                ))
            except ImportError as _t5_ie:
                self.after(0, lambda e=_t5_ie: self.log_cluster(
                    f"⚠️ T5 пропущен — {e}"
                ))
            except Exception as _t5_e:
                self.after(0, lambda e=_t5_e: self.log_cluster(
                    f"⚠️ T5 ошибка: {e}"
                ))

        # ═══════════════════════════════════════════════════════════════
        # СТАДИЯ 4: Экспорт результатов в XLSX, сводная таблица

    def _cluster_step_quality_metrics(self, snap: dict, Xv, labels: list, X_clean: list) -> None:
        # ── Метрики качества кластеризации (Silhouette / CH / DB) ──
        _labels_arr_q = _np.asarray(labels)
        _valid_mask   = (_labels_arr_q >= 0)  # HDBSCAN может дать -1 (шум)
        if _valid_mask.sum() > 1 and len(set(_labels_arr_q[_valid_mask])) > 1:
            try:
                import scipy.sparse as _sp_q
                _Xq = Xv[_valid_mask]
                _Xq = _Xq.toarray() if _sp_q.issparse(_Xq) else _np.asarray(_Xq)
                _lq = _labels_arr_q[_valid_mask]
                _sample_q = min(5000, len(_lq))
                from sklearn.metrics import (
                    silhouette_score as _sil_q,
                    calinski_harabasz_score as _ch_q,
                    davies_bouldin_score as _db_q,
                )
                _sil = _sil_q(_Xq, _lq, sample_size=_sample_q, random_state=snap.get("cluster_random_seed", 42))
                _ch  = _ch_q(_Xq, _lq)
                _db  = _db_q(_Xq, _lq)
                def _sil_hint(s):
                    if s > 0.5:  return "отлично"
                    if s > 0.25: return "хорошо"
                    if s > 0.0:  return "слабо"
                    return "плохо (кластеры перекрываются)"

                def _db_hint(d):
                    if d < 0.5:  return "отлично"
                    if d < 1.0:  return "хорошо"
                    if d < 2.0:  return "приемлемо"
                    return "плохо (кластеры размыты)"

                self.after(0, lambda s=_sil, c=_ch, d=_db: self.log_cluster(
                    f"Метрики качества:\n"
                    f"  Silhouette:       {s:+.3f}  [{_sil_hint(s)}]  (-1..+1, выше=лучше)\n"
                    f"  Calinski-Harabász:{c:>10.1f}  (выше=лучше, нет верхней границы)\n"
                    f"  Davies-Bouldin:   {d:.3f}  [{_db_hint(d)}]  (ниже=лучше, 0=идеал)"
                ))
            except Exception as _qe:
                self.after(0, lambda e=_qe: self.log_cluster(
                    f"⚠️ Метрики качества: {e}"
                ))

        # ── C_NPMI: тематическая когерентность кластеров ──────────
        # Вычисляем только если gensim доступен и > 1 кластер
        try:
            import gensim.corpora as _gensim_corpora
            from gensim.models.coherencemodel import CoherenceModel as _CoherenceModel
            _labels_c = _np.asarray(labels)
            _n_valid_c = int((_labels_c >= 0).sum())
            if _n_valid_c > 10 and len(set(_labels_c[_labels_c >= 0])) > 1 and kw:
                # Токенизируем X_clean для gensim
                _corpus_texts = [str(t).split() for t in X_clean]
                _dictionary = _gensim_corpora.Dictionary(_corpus_texts)
                _dictionary.filter_extremes(no_below=2, no_above=0.95, keep_n=30000)
                # Топ-слова по кластерам → список списков
                _topics_tokens = [str(kw[i]).split()[:10] for i in range(len(kw)) if kw[i]]
                _topics_tokens = [t for t in _topics_tokens if len(t) >= 2]
                if _topics_tokens:
                    _cm = _CoherenceModel(
                        topics=_topics_tokens,
                        texts=_corpus_texts,
                        dictionary=_dictionary,
                        coherence="c_npmi",
                    )
                    _npmi_score = float(_cm.get_coherence())
                    def _npmi_hint(s):
                        if s > 0.1:  return "хорошо"
                        if s > 0.0:  return "приемлемо"
                        return "слабо (темы размыты)"
                    self.after(0, lambda s=_npmi_score: self.log_cluster(
                        f"  C_NPMI когерентность: {s:+.4f}  [{_npmi_hint(s)}]  (-1..+1, выше=лучше)"
                    ))
        except ImportError as _e:
            _log.debug("NPMI coherence skipped (gensim not installed): %s", _e)
        except Exception as _npmi_e:
            self.after(0, lambda e=_npmi_e: self.log_cluster(
                f"⚠️ C_NPMI: {e}"
            ))

        # ── Слияние семантически похожих кластеров ───────────────

    def _cluster_step_keywords(self, snap: dict, labels: list, X_clean: list, Xv_tfidf, kw: list, vec_kw, use_hdbscan: bool, use_lda: bool, use_hier: bool, use_bertopic: bool, use_fastopic: bool, use_fastopic_kw_ready: bool) -> None:
        # ── Извлечение ключевых слов ──────────────────────────────
        # c-TF-IDF: специфичные для кластера слова (по умолчанию)
        # обычный TF-IDF mean: usable as fallback
        if not use_fastopic_kw_ready:
            _use_label_kw = (
                use_sbert_clust
                or use_combo_clust
                or use_ensemble
                or use_hdbscan
                or use_lda
                or use_hier
                or use_bertopic
                or use_gmm
                or use_fastopic
                or snap.get("use_cosine_cluster", False)
                or snap.get("use_umap", False)
                or snap.get("use_tfidf_svd", False)
            )
            _use_ctfidf = snap.get("use_ctfidf_keywords", True)
            if _use_ctfidf:
                kw = extract_cluster_keywords_ctfidf(
                    X_clean, labels, K,
                    stop_words=_stop_list,
                    top_n=12,
                    use_lemma=bool(snap.get("use_lemma_cluster", True)),
                )
            elif _use_label_kw:
                kw = extract_cluster_keywords_from_labels(
                    vec_kw, Xv_tfidf, labels, K, top_n=12,
                )
            else:
                kw = extract_cluster_keywords(vec_kw, km.cluster_centers_, top_n=12)

        if use_fastopic_kw_ready:
            _kw_method = "FASTopic"
        else:
            _kw_method = ("c-TF-IDF" if _use_ctfidf
                          else "по меткам" if _use_label_kw
                          else "центроидный TF-IDF")
        _kw_n = len(kw)  # kw is always List[str]
        _kw_avg = sum(len(v) for v in kw) // max(1, _kw_n)
        self.after(0, lambda m=_kw_method, n=_kw_n, a=_kw_avg: self.log_cluster(
            f"Ключевые слова извлечены: метод={m} | кластеров={n} | ~{a} слов/кластер ✅"
        ))

        # ── Повторная кластеризация шума HDBSCAN (#19) ────────────

    def _cluster_step_hdbscan_reclust(self, snap: dict, Xv, labels: list, noise_labels: list) -> None:
        noise_labels: _np.ndarray = _np.full(len(labels), -1, dtype=int)
        if use_hdbscan and snap.get("recluster_noise", True):
            _noise_mask = (_np.asarray(labels) == -1)
            _n_noise = int(_noise_mask.sum())
            if _n_noise >= 10:
                try:
                    self.after(0, lambda n=_n_noise: self.log_cluster(
                        f"Повторная кластеризация {n} шумовых точек HDBSCAN…"
                    ))
                    import scipy.sparse as _sp_noise
                    _Xv_noise = Xv[_noise_mask]
                    _Xv_noise_d = _Xv_noise.toarray() if _sp_noise.issparse(_Xv_noise) else _np.asarray(_Xv_noise)
                    _K_noise = min(snap.get("noise_k_clusters", 5), _n_noise // 2)
                    _K_noise = max(2, _K_noise)
                    _km_noise = _gpu_kmeans(
                        n_clusters=_K_noise,
                        random_state=snap.get("cluster_random_seed", 42),
                        batch_size=min(HDBSCAN_NOISE_BATCH_SIZE, _n_noise),
                        n_init="auto",
                    )
                    _noise_sub_labels = _km_noise.fit_predict(_Xv_noise_d)
                    noise_labels[_noise_mask] = _noise_sub_labels
                    self.after(0, lambda k=_K_noise: self.log_cluster(
                        f"Шум разбит на {k} под-кластеров ✅"
                    ))
                except Exception as _ne:
                    self.after(0, lambda e=_ne: self.log_cluster(
                        f"⚠️ Повторная кластеризация шума: {e}"
                    ))

        # ── Аннотация качества кластеров (intra-cohesion) (#7) ────

    def _cluster_step_cohesion(self, snap: dict, Xv, labels: list, X_clean: list) -> None:
        cluster_quality_map: dict = {}
        if snap.get("show_cluster_quality", True):
            try:
                import scipy.sparse as _sp_coh
                from sklearn.preprocessing import normalize as _coh_norm
                _labels_coh = _np.asarray(labels)
                _Xcoh = Xv
                _Xcoh_d = _Xcoh.toarray() if _sp_coh.issparse(_Xcoh) else _np.asarray(_Xcoh)
                _Xcoh_n = _coh_norm(_Xcoh_d, norm="l2")
                _cluster_ids = sorted(int(x) for x in set(_labels_coh) if int(x) >= 0)
                # Показываем качество для всех кластеров (по умолчанию low),
                # чтобы в выгрузке не было "пустых" категорий качества.
                for _cid in _cluster_ids:
                    cluster_quality_map[_cid] = "low"

                # Per-cluster silhouette: в отличие от mean intra-cluster cosine
                # учитывает и близость к соседним кластерам (inter-cluster).
                # «Плотный, но пересекающийся с соседом» кластер правильно
                # классифицируется как низко-качественный.
                _use_silhouette = len(_cluster_ids) >= 2
                _sil_per_cluster: dict[int, float] = {}
                if _use_silhouette:
                    try:
                        from sklearn.metrics import silhouette_samples as _sil_samples
                        _valid_mask = (_labels_coh >= 0)
                        _Xsil = _Xcoh_n[_valid_mask]
                        _lsil = _labels_coh[_valid_mask]
                        # Для датасетов > 10k сэмплируем (silhouette — O(n²)).
                        _n_sil = len(_lsil)
                        if _n_sil > 10_000:
                            _rng = _np.random.default_rng(
                                snap.get("cluster_random_seed", 42)
                            )
                            _idx = _rng.choice(_n_sil, size=10_000, replace=False)
                            _sil_vals = _sil_samples(_Xsil[_idx], _lsil[_idx],
                                                     metric="cosine")
                            _lsil_eff = _lsil[_idx]
                        else:
                            _sil_vals = _sil_samples(_Xsil, _lsil, metric="cosine")
                            _lsil_eff = _lsil
                        for _cid in _cluster_ids:
                            _mask_cid = (_lsil_eff == _cid)
                            if _mask_cid.sum() > 0:
                                _sil_per_cluster[_cid] = float(_sil_vals[_mask_cid].mean())
                    except Exception as _sil_exc:
                        _log.debug(
                            "silhouette_samples failed (%s), fallback to cosine: %s",
                            type(_sil_exc).__name__, _sil_exc,
                        )
                        _use_silhouette = False
                        _sil_per_cluster = {}

                for _cid in _cluster_ids:
                    if _cid < 0:
                        continue
                    _mask_coh = (_labels_coh == _cid)
                    _pts = _Xcoh_n[_mask_coh]
                    if len(_pts) < 2:
                        cluster_quality_map[_cid] = "single"
                        continue
                    if _cid in _sil_per_cluster:
                        # silhouette ∈ [-1, +1]: +1 идеально разделён,
                        # 0 на границе, <0 скорее принадлежит соседу.
                        _s = _sil_per_cluster[_cid]
                        if _s >= 0.35:
                            cluster_quality_map[_cid] = "high"
                        elif _s >= 0.10:
                            cluster_quality_map[_cid] = "medium"
                        else:
                            cluster_quality_map[_cid] = "low"
                    else:
                        # Fallback на mean intra-cluster cosine, если silhouette
                        # недоступен (<2 кластеров или исключение).
                        _sim_sum = float(_np.sum(_pts @ _pts.T))
                        _n_pts = len(_pts)
                        _mean_sim = (_sim_sum - _n_pts) / (_n_pts * (_n_pts - 1))
                        if _mean_sim >= 0.50:
                            cluster_quality_map[_cid] = "high"
                        elif _mean_sim >= 0.30:
                            cluster_quality_map[_cid] = "medium"
                        else:
                            cluster_quality_map[_cid] = "low"
                _q_counts = {}
                for v in cluster_quality_map.values():
                    _q_counts[v] = _q_counts.get(v, 0) + 1
                _metric_name = "silhouette" if _sil_per_cluster else "cosine"
                self.after(0, lambda c=dict(_q_counts), m=_metric_name: self.log_cluster(
                    f"Качество кластеров ({m}): "
                    + ", ".join(f"{k}={v}" for k, v in sorted(c.items()))
                ))
            except Exception as _coh_e:
                self.after(0, lambda e=_coh_e: self.log_cluster(
                    f"⚠️ Аннотация качества: {e}"
                ))

        self._cluster_step_llm_naming(snap, labels, X_clean, Xv, cluster_name_map)
        self._cluster_step_llm_feedback(snap, labels, cluster_name_map, cluster_reason_map, cluster_quality_map)

    def _cluster_step_umap(self, snap: dict, Xv, X_clean: list) -> "Any":
        # ── UMAP снижение размерности ─────────────────────────────
        # Пропускаем для BERTopic (UMAP уже применён в пути векторизации)
        # Пропускаем для LDA: LDA ожидает неотрицательные целые числа
        # (count-матрицу), UMAP даёт float и может давать отрицательные значения —
        # это сломало бы LDA и привело бы к некорректным результатам.
        if snap.get("use_umap", False) and not use_bertopic and not use_lda and not use_ensemble:
            try:
                n_umap   = snap.get("umap_n_components", 50)
                _umap_nn = snap.get("umap_n_neighbors", 15)
                _umap_md = snap.get("umap_min_dist", 0.1)
                _umap_mt = snap.get("umap_metric", "cosine")
                _umap_src = "GPU (cuML)" if _cuml_umap_available() else "umap-learn"
                self.after(0, lambda n=n_umap, nn=_umap_nn, md=_umap_md, mt=_umap_mt, s=_umap_src: self.log_cluster(
                    f"UMAP [{s}]: {len(X_clean)} текстов → {n}д "
                    f"(n_neighbors={nn}, min_dist={md:.2f}, metric={mt})…"
                ))
                import scipy.sparse as _sp_umap
                _Xv_dense = Xv.toarray() if _sp_umap.issparse(Xv) else _np.asarray(Xv)

                # PCA-преднормализация перед UMAP (опционально)
                if snap.get("use_pca_before_umap", False):
                    _pca_dim = min(
                        snap.get("pca_n_components", 50),
                        _Xv_dense.shape[1] - 1,
                        _Xv_dense.shape[0] - 1,
                    )
                    _pca_dim = max(n_umap + 1, _pca_dim)
                    self.after(0, lambda d=_pca_dim: self.log_cluster(
                        f"PCA: снижаю размерность до {d} перед UMAP…"
                    ))
                    from sklearn.decomposition import PCA as _PCA
                    _pca = _PCA(n_components=_pca_dim,
                                random_state=snap.get("cluster_random_seed", 42))
                    _Xv_dense = _pca.fit_transform(_Xv_dense)
                    self.after(0, lambda d=_pca_dim: self.log_cluster(
                        f"PCA готово ✅  форма={_Xv_dense.shape}"
                    ))

                reducer = _gpu_umap(
                    n_components=n_umap,
                    n_neighbors=_umap_nn,
                    min_dist=_umap_md,
                    metric=_umap_mt,
                    random_state=snap.get("cluster_random_seed", 42),
                )
                Xv = reducer.fit_transform(_Xv_dense)
                _umap_shape = Xv.shape
                self.after(0, lambda sh=_umap_shape: self.log_cluster(
                    f"UMAP готово ✅  итоговая форма={sh}"
                ))
            except ImportError:
                self.after(0, lambda: self.log_cluster(
                    "⚠️ UMAP не установлен — пропускаю. Установите: pip install umap-learn"
                ))

        self.after(0, lambda n=len(X_all): self.cluster_status.set(
            f"Стадия 2/4: Кластеризация {n:,} векторов…"
        ))
        K_user          = snap["k_clusters"]
        k_score_method  = snap.get("k_score_method", "elbow")
        use_hdbscan     = use_hdbscan_algo or use_bertopic  # HDBSCAN also drives BERTopic
        km              = None   # заполняется только для KMeans
        labels_l1: Optional[List[int]] = None  # для иерархической кластеризации

        # ── Защита: K не может превышать число строк с текстом ────

        return Xv

    def _cluster_step_autok(self, snap: dict, Xv, X_clean: list, n_ok: int) -> int:
        # ── Авто-подбор K (только для KMeans, не для ансамбля) ────
        if snap["use_elbow"] and not use_hdbscan and not use_lda and not use_hier and not use_ensemble:
            lo = max(2, K - 8)
            hi = min(60, K + 8, n_ok - 1)  # не больше n_ok-1
            if lo > hi:
                lo = 2
                hi = min(n_ok - 1, lo + 8)
            ks = list(range(lo, hi + 1))
            if len(ks) < 2:
                return K

            import scipy.sparse as _sp_ks
            _Xv_ks = Xv.toarray() if _sp_ks.issparse(Xv) else _np.asarray(Xv)

            if k_score_method == "silhouette":
                from sklearn.metrics import silhouette_score as _sil_score
                _sample_sz = min(5000, len(X_clean))
                self.after(0, lambda: ui_prog(_elbow_start, f"Silhouette: K={lo}..{hi}"))
                _scores: List[float] = []
                for i, k in enumerate(ks):
                    km_tmp = _gpu_kmeans(n_clusters=k, random_state=snap.get("cluster_random_seed", 42),
                                         batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                                         n_init="auto", reassignment_ratio=0.01)
                    _lbl_tmp = km_tmp.fit_predict(_Xv_ks)
                    _sc = _sil_score(_Xv_ks, _lbl_tmp,
                                     sample_size=_sample_sz, random_state=snap.get("cluster_random_seed", 42))
                    _scores.append(_sc)
                    if self._cancel_event.is_set():
                        raise InterruptedError()
                    if (i + 1) % 3 == 0:
                        pct = _elbow_start + (_km_start - _elbow_start) * (i + 1) / len(ks)
                        self.after(0, lambda pct=pct, k=k, sc=_sc: ui_prog(
                            pct, f"Silhouette K={k}: {sc:.3f}"))
                K = ks[int(_np.argmax(_scores))]
                self.after(0, lambda: self.log_cluster(
                    f"Silhouette выбрал K={K} "
                    f"(score={max(_scores):.3f}, user K={K_user})"
                ))

            elif k_score_method == "calinski":
                from sklearn.metrics import calinski_harabasz_score as _ch_score
                self.after(0, lambda: ui_prog(_elbow_start, f"Calinski-Harabász: K={lo}..{hi}"))
                _scores = []
                for i, k in enumerate(ks):
                    km_tmp = _gpu_kmeans(n_clusters=k, random_state=snap.get("cluster_random_seed", 42),
                                         batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                                         n_init="auto", reassignment_ratio=0.01)
                    _lbl_tmp = km_tmp.fit_predict(_Xv_ks)
                    _sc = _ch_score(_Xv_ks, _lbl_tmp)
                    _scores.append(_sc)
                    if self._cancel_event.is_set():
                        raise InterruptedError()
                    if (i + 1) % 3 == 0:
                        pct = _elbow_start + (_km_start - _elbow_start) * (i + 1) / len(ks)
                        self.after(0, lambda pct=pct, k=k: ui_prog(pct, f"Calinski K={k}"))
                K = ks[int(_np.argmax(_scores))]
                self.after(0, lambda: self.log_cluster(
                    f"Calinski-Harabász выбрал K={K} "
                    f"(score={max(_scores):.1f}, user K={K_user})"
                ))

            else:
                # Elbow (inertia) — исходный алгоритм
                inertias: List[float] = []
                self.after(0, lambda: ui_prog(_elbow_start, f"Elbow: проба K={lo}..{hi}"))
                for i, k in enumerate(ks):
                    km_tmp = _gpu_kmeans(n_clusters=k, random_state=snap.get("cluster_random_seed", 42),
                                         batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                                         n_init="auto", reassignment_ratio=0.01)
                    km_tmp.fit(Xv)
                    inertias.append(float(km_tmp.inertia_))
                    if self._cancel_event.is_set():
                        raise InterruptedError()
                    if (i + 1) % 3 == 0:
                        pct = _elbow_start + (_km_start - _elbow_start) * (i + 1) / len(ks)
                        self.after(0, lambda pct=pct, k=k: ui_prog(pct, f"Elbow: K={k}"))
                K = ClusterElbowSelector.pick_elbow_k(inertias, ks)
                self.after(0, lambda: self.log_cluster(
                    f"Elbow выбрал K={K} (user K={K_user})"
                ))

        # ── Кластеризация ──────────────────────────────────────────

        return K

    def _cluster_worker_stage1(self, _crs: "ClusterRunState", snap: dict) -> None:
        """СТАДИЯ 1: Чтение файлов, очистка текста, дедупликация.

        Reads all input files, builds _crs.X_all / _crs.file_data / _crs.raw_texts_all.
        Inputs from _crs: in_paths, total_rows, start_ts, use_t5, cluster_snap.
        Outputs to _crs: X_all, file_data, raw_texts_all, done, n_ok.
        """
        # СТАДИЯ 1: Чтение файлов, очистка текста, дедупликация
        # ═══════════════════════════════════════════════════════════════
        # ── Чтение всех файлов ────────────────────────────────────────
        for fi, in_path in enumerate(_crs.in_paths):
            row_has_text_file: List[bool] = []
            header: List[str] = []

            with open_tabular(in_path) as it:
                header = ["" if h is None else str(h) for h in next(it)]
                header_index = {col: i for i, col in enumerate(header)}

                has_any = any(
                    col and col in header
                    for col in [
                        snap["desc_col"], snap["call_col"], snap["chat_col"],
                        (snap["summary_col"] if snap["use_summary"] else ""),
                        snap["ans_short_col"], snap["ans_full_col"],
                    ]
                )
                if not has_any:
                    raise ValueError(
                        f"Не найдены колонки текста в файле {in_path.name}\n"
                        "(проверь заголовки и настройки колонок)."
                    )

                _fi_label = f"[{fi+1}/{_n_files}] {in_path.name}"
                for row in it:
                    row_vals = list(row)
                    xfeat = self._row_to_feature_text(row_vals, header, _crs.cluster_snap, header_index=header_index)
                    ok = bool(xfeat.strip())
                    row_has_text_file.append(ok)
                    if ok:
                        _crs.X_all.append(xfeat)
                        if _crs.use_t5:
                            _crs.raw_texts_all.append(
                                build_t5_source_text(
                                    row_vals,
                                    header,
                                    snap,
                                    _crs.cluster_snap,
                                    header_index=header_index,
                                )
                            )
                    _crs.done += 1
                    if _crs.done % 2000 == 0:
                        pct = 10.0 * _crs.done / float(_crs.total_rows)
                        msg = f"Чтение {_fi_label}: {_crs.done}/{_crs.total_rows} | {fmt_speed(_crs.start_ts, _crs.done)} | {fmt_eta(_crs.start_ts, _crs.done, _crs.total_rows)}"
                        self.after(0, lambda pct=pct, msg=msg: ui_prog(pct, msg))
                        if self._cancel_event.is_set():
                            raise InterruptedError()

            _crs.file_data.append((in_path, header, row_has_text_file))

        _crs.n_ok = len(_crs.X_all)
        n_skip = _crs.done - _crs.n_ok
        self.after(0, lambda n_ok=_crs.n_ok, n_skip=n_skip: self.log_cluster(
            f"Строк прочитано: {_crs.done} | с текстом: {_crs.n_ok} | пропущено (пустые): {n_skip}"
        ))

        if _crs.n_ok < 2:
            raise FeatureBuildError(
                f"Слишком мало строк с текстом ({_crs.n_ok}). "
                "Для кластеризации нужно минимум 2 строки."
            )

        if self._cancel_event.is_set():
            raise InterruptedError()

        self.after(0, lambda: ui_prog(12.0, f"Очистка текстов (rows={len(_crs.X_all)})…"))

        # ── Удаляем шаблонные фразы до токенизации ───────────────
        if snap.get("use_noise_phrases", True):
            _user_ph = [
                p.lower() for p in (snap.get("extra_noise_phrases") or [])
                if p.strip()
            ]
            # Пользовательские фразы идут первыми — матчатся раньше коротких встроенных
            pr = PhraseRemover(_user_ph + list(NOISE_PHRASES))
            X_clean = pr.fit_transform(_crs.X_all)
        else:
            X_clean = list(_crs.X_all)

        # ── Нормализация чисел и дат → __NUM__ / __DATE__ / __TIME__
        if snap.get("normalize_numbers", True):
            import re as _re_num
            _DATE_RE  = _re_num.compile(
                r'\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b')
            _TIME_RE  = _re_num.compile(r'\b\d{1,2}:\d{2}(?::\d{2})?\b')
            _NUM4_RE  = _re_num.compile(r'\b\d{4,}\b')
            _NUMDEC_RE = _re_num.compile(r'\b\d+[.,]\d+\b')
            _NUM1_RE  = _re_num.compile(r'\b\d{1,3}\b')
            def _norm_nums(t: str) -> str:
                t = _DATE_RE.sub('__DATE__', t)
                t = _TIME_RE.sub('__TIME__', t)
                # ВАЖНО: _NUMDEC_RE должен идти ДО _NUM4_RE,
                # иначе "1234.56" → _NUM4_RE заменяет "1234" → "__NUM__.56",
                # и _NUMDEC_RE уже не находит десятичный разделитель.
                t = _NUMDEC_RE.sub('__NUM__', t)
                t = _NUM4_RE.sub('__NUM__', t)
                t = _NUM1_RE.sub('__NUM__', t)
                return t
            X_clean = [_norm_nums(t) for t in X_clean]
            self.after(0, lambda: self.log_cluster(
                "Нормализация чисел/дат → __NUM__/__DATE__/__TIME__ ✅"
            ))

        # ── Лемматизация (pymorphy2) ──────────────────────────────
        if snap.get("use_lemma_cluster", False):
            try:
                lemm = Lemmatizer()
                lemm.fit(X_clean)
                if getattr(lemm, "is_active_", False):
                    self.after(0, lambda: self.log_cluster(
                        "Лемматизация: pymorphy2 запущен…"
                    ))
                    X_clean = lemm.transform(X_clean)
                    self.after(0, lambda: self.log_cluster(
                        "Лемматизация: готово ✅"
                    ))
                else:
                    self.after(0, lambda: self.log_cluster(
                        "⚠️ Лемматизация: pymorphy2 не установлен — пропуск. "
                        "Установите: pip install pymorphy2"
                    ))
            except Exception as _lemm_ex:
                _log.warning("lemmatizer disabled due to runtime error: %s", _lemm_ex, exc_info=True)
                self.after(0, lambda e=_lemm_ex: self.log_cluster(
                    f"⚠️ Лемматизация отключена из-за ошибки pymorphy2: {type(e).__name__}: {e}"
                ))

        # ── Стоп-слова / токены (общий набор для обоих путей) ─────
        _stop_set = (
            (RUSSIAN_STOP_WORDS if snap.get("use_stop_words", True) else set())
            | (NOISE_TOKENS if snap.get("use_noise_tokens", True) else set())
            | set(snap.get("extra_stop_words") or [])
            | {t.lower() for t in (snap.get("extra_noise_tokens") or [])}
        )
        _stop_list = sorted(_stop_set) or None

        _vec_mode       = snap.get("cluster_vec_mode", "tfidf")
        _algo           = snap.get("cluster_algo", "kmeans")
        use_combo_clust  = (_vec_mode == "combo")
        use_sbert_clust  = (_vec_mode == "sbert")
        use_ensemble     = (_vec_mode == "ensemble")
        use_hdbscan_algo = (_algo == "hdbscan")
        use_lda          = (_algo == "lda")
        use_hier         = (_algo == "hierarchical")
        use_bertopic     = (_algo == "bertopic")
        use_gmm          = (_algo == "gmm")
        use_fastopic     = (_algo == "fastopic")

        # ── Семантическая дедупликация (#10) ─────────────────────
        # Удаляем почти-дублирующиеся тексты перед кластеризацией.
        # Хранит обратный маппинг для восстановления меток всех строк.
        _dedup_reverse_map: "_np.ndarray | None" = None
        if snap.get("use_dedup", False) and len(X_clean) > 10:
            try:
                _dedup_thr = snap.get("dedup_threshold", 0.95)
                self.after(0, lambda t=_dedup_thr, n=len(X_clean): self.log_cluster(
                    f"Дедупликация: порог={t}, строк={n}…"
                ))
                from sklearn.feature_extraction.text import TfidfVectorizer as _TfDd
                from sklearn.preprocessing import normalize as _norm_dd
                _dd_vec = _TfDd(max_features=20000, min_df=1, analyzer="word",
                                ngram_range=(1, 1), sublinear_tf=True)
                _Xdd = _dd_vec.fit_transform(X_clean)
                _Xdd_n = _norm_dd(_Xdd, norm="l2")
                # Greedy dedup: iterate, mark if too similar to any already-kept text
                _kept_idx: list = []
                _reverse_map = list(range(len(X_clean)))  # maps each orig → kept slot
                _kept_vecs_list: list = []
                import scipy.sparse as _sp_dd
                for _i in range(len(X_clean)):
                    _v = _Xdd_n[_i]
                    if _kept_vecs_list:
                        import scipy.sparse as _sp_dd2
                        _kept_mat = _sp_dd2.vstack(_kept_vecs_list)
                        _sims = (_kept_mat @ _v.T).toarray().ravel()
                        _best = float(_sims.max())
                        if _best >= _dedup_thr:
                            _reverse_map[_i] = _kept_idx[int(_sims.argmax())]
                            continue
                    _reverse_map[_i] = len(_kept_idx)
                    _kept_idx.append(_i)
                    _kept_vecs_list.append(_v)
                if len(_kept_idx) < len(X_clean):
                    _n_removed = len(X_clean) - len(_kept_idx)
                    X_clean = [X_clean[i] for i in _kept_idx]
                    _dedup_reverse_map = _np.array(_reverse_map, dtype=int)
                    self.after(0, lambda r=_n_removed, k=len(X_clean): self.log_cluster(
                        f"Дедупликация: удалено {r} дублей, осталось {k} ✅"
                    ))
                else:
                    self.after(0, lambda: self.log_cluster("Дедупликация: дублей не найдено"))
            except Exception as _dd_e:
                self.after(0, lambda e=_dd_e: self.log_cluster(
                    f"⚠️ Дедупликация: {e}"
                ))

        # min_df: пользовательское значение (0 = авто-адаптивный)
        _n_clean = len(X_clean)
        _user_min_df = snap.get("cluster_min_df", 0)
        if _user_min_df > 0:
            _adaptive_min_df = _user_min_df
            self.after(0, lambda m=_adaptive_min_df: self.log_cluster(
                f"TF-IDF: min_df={m} (задан вручную)"
            ))
        else:
            _adaptive_min_df = 1 if _n_clean < 5_000 else (2 if _n_clean < 50_000 else 3)
            self.after(0, lambda m=_adaptive_min_df, n=_n_clean: self.log_cluster(
                f"TF-IDF: адаптивный min_df={m} (строк={n})"
            ))

        _tfidf_params = dict(
            analyzer="word",
            ngram_range=snap["word_ng"],
            min_df=_adaptive_min_df,
            max_features=max(50000, snap["max_features"] // 3),
            sublinear_tf=snap["sublinear_tf"],
            stop_words=_stop_list,
        )

        def _safe_tfidf(params: dict, texts: list):
            """fit_transform TfidfVectorizer; при пустом словаре повторяет без стоп-слов."""
            try:
                _vec = TfidfVectorizer(**params)
                _mat = _vec.fit_transform(texts)
                return _mat, _vec
            except ValueError as _ve:
                if "empty vocabulary" not in str(_ve):
                    raise
            self.after(0, lambda: self.log_cluster(
                "⚠️ Пустой словарь со стоп-словами — повтор без стоп-слов"
            ))
            _p2 = dict(params)
            _p2["stop_words"] = None
            _vec2 = TfidfVectorizer(**_p2)
            return _vec2.fit_transform(texts), _vec2

        # ═══════════════════════════════════════════════════════════════


    def _cluster_worker_stage4(
        self,
        _crs: "ClusterRunState",
        snap: dict,
        t0: float,
        ui_prog,
        _lifecycle,
    ) -> None:
        """СТАДИЯ 4: Экспорт результатов в XLSX, сводная таблица.

        Writes clustered XLSX output files and fires done_ui callback.
        Reads from _crs: labels, kw_dict, kw, file_data, cluster_name_map,
          cluster_reason_map, cluster_quality_map, llm_feedback_map,
          t5_summaries_all, labels_l1, noise_labels, K, total_rows,
          use_hdbscan, use_hier, use_lda.
        Params: t0 (wall-clock start for elapsed), ui_prog, _lifecycle.
        """
        # ═══════════════════════════════════════════════════════════════
        # ── Раскладываем метки обратно по файлам ──────────────────
        cid_col   = snap["cluster_id_col"]
        ckw_col   = snap["cluster_kw_col"]
        t5_col    = snap.get("t5_summary_col", "t5_summary")
        hier_l1_col = snap.get("hier_l1_col", "cluster_l1") if _crs.use_hier else None
        lda_topics_col = snap.get("lda_topics_col", "lda_topics") if _crs.use_lda else None
        cname_col   = snap.get("llm_name_col", "cluster_name") if _crs.cluster_name_map else None
        reason_col  = snap.get("llm_reason_col", "cluster_reason") if _crs.cluster_reason_map else None
        cqual_col   = snap.get("cluster_quality_col", "cluster_quality") if _crs.cluster_quality_map else None
        noise_col   = snap.get("noise_cluster_col", "noise_cluster_id") if (_crs.use_hdbscan and snap.get("recluster_noise", True) and _np.any(_crs.noise_labels >= 0)) else None
        fb_col      = snap.get("llm_feedback_col", "llm_feedback") if _crs.llm_feedback_map else None
        conf_col    = snap.get("cluster_confidence_col", "cluster_confidence") if (
            _crs.use_hdbscan
            and _crs.hdbscan_proba is not None
            and len(_crs.hdbscan_proba) > 0
        ) else None

        def _clu_col_w(col: str) -> float:
            if col == cid_col:        return 15.0
            if col == ckw_col:        return 55.0
            if col == t5_col:         return 60.0
            if col == hier_l1_col:    return 15.0
            if col == lda_topics_col: return 55.0
            if col == cname_col:      return 35.0
            if col == reason_col:     return 80.0
            if col == cqual_col:      return 15.0
            if col == noise_col:      return 18.0
            if col == fb_col:         return 30.0
            if col == conf_col:       return 18.0
            return 22.0

        _crs.kw_dict: Dict[int, str] = {i: _crs.kw[i] for i in range(len(_crs.kw))}

        stamp = now_stamp()
        out_paths: List[Path] = []
        label_offset = 0
        for in_path, header, row_has_text_file in _crs.file_data:
            # Собираем full_cluster, full_l1, full_t5, full_noise для этого файла
            full_cluster: List[int] = []
            full_l1: List[int] = []
            full_t5: List[str] = []
            full_noise: List[int] = []
            full_conf: List[float] = []
            label_cursor = label_offset
            for ok in row_has_text_file:
                if ok:
                    full_cluster.append(int(_crs.labels[label_cursor]))
                    if _crs.labels_l1 is not None:
                        full_l1.append(
                            _crs.labels_l1[label_cursor]
                            if label_cursor < len(_crs.labels_l1) else -1
                        )
                    if _crs.t5_summaries_all:
                        full_t5.append(
                            _crs.t5_summaries_all[label_cursor]
                            if label_cursor < len(_crs.t5_summaries_all) else ""
                        )
                    full_noise.append(
                        int(_crs.noise_labels[label_cursor])
                        if label_cursor < len(_crs.noise_labels) else -1
                    )
                    if conf_col is not None:
                        full_conf.append(
                            float(_crs.hdbscan_proba[label_cursor])
                            if label_cursor < len(_crs.hdbscan_proba) else 0.0
                        )
                    label_cursor += 1
                else:
                    full_cluster.append(-1)
                    if _crs.labels_l1 is not None:
                        full_l1.append(-1)
                    if _crs.t5_summaries_all:
                        full_t5.append("")
                    full_noise.append(-1)
                    if conf_col is not None:
                        full_conf.append(0.0)
            label_offset = label_cursor

            # Ограничиваем долю неопределённых кластеров (cid < 0) до 10%:
            # избыточные строки переводим в доминирующий кластер файла.
            _assigned = [int(c) for c in full_cluster if int(c) >= 0]
            _dominant_cid = Counter(_assigned).most_common(1)[0][0] if _assigned else 0
            _unassigned_idx = [ii for ii, c in enumerate(full_cluster) if int(c) < 0]
            _max_unassigned = int(len(full_cluster) * 0.10)
            if len(_unassigned_idx) > _max_unassigned:
                _to_reassign = len(_unassigned_idx) - _max_unassigned
                for _ii in _unassigned_idx[:_to_reassign]:
                    full_cluster[_ii] = _dominant_cid
                self.after(0, lambda n=_to_reassign, c=_dominant_cid, p=in_path.name: self.log_cluster(
                    f"⚠️ {p}: {n} строк без кластера переназначены в кластер {c} (лимит 10%)"
                ))
            _fallback_kw = _crs.kw_dict.get(_dominant_cid, "общая тема")

            wb_out = Workbook(write_only=True)
            ws_out = wb_out.create_sheet(title=in_path.stem[:31])

            with open_tabular(in_path) as it2:
                header2 = ["" if h is None else str(h) for h in next(it2)]
                header_out = header2[:]
                for c in (cid_col, ckw_col):
                    if c not in header_out:
                        header_out.append(c)
                # LLM-имя кластера
                if cname_col and cname_col not in header_out:
                    header_out.append(cname_col)
                if reason_col and reason_col not in header_out:
                    header_out.append(reason_col)
                # Качество кластера
                if cqual_col and cqual_col not in header_out:
                    header_out.append(cqual_col)
                # Иерархический L1-столбец
                if hier_l1_col and hier_l1_col not in header_out:
                    header_out.append(hier_l1_col)
                # LDA темы
                if lda_topics_col and lda_topics_col not in header_out:
                    header_out.append(lda_topics_col)
                # Добавляем столбец T5-суммаризации только если она выполнялась
                if _crs.t5_summaries_all and t5_col not in header_out:
                    header_out.append(t5_col)
                # Под-кластер шума HDBSCAN
                if noise_col and noise_col not in header_out:
                    header_out.append(noise_col)
                # LLM feedback (MERGE/SPLIT/OK)
                if fb_col and fb_col not in header_out:
                    header_out.append(fb_col)
                # HDBSCAN soft-cluster confidence (0.0–1.0)
                if conf_col and conf_col not in header_out:
                    header_out.append(conf_col)

                ws_out.freeze_panes = "A2"
                ws_out.auto_filter.ref = f"A1:{get_column_letter(len(header_out))}1"
                ws_out.append(header_out)

                idx_map = {col: idx_of(header2, col) for col in header2}
                for i, row in enumerate(it2):
                    row_vals = list(row)
                    cid = full_cluster[i]
                    _undef_kw = _fallback_kw
                    _undef_name = "Неопределённый кластер"
                    _undef_reason = "Недостаточно данных для устойчивой кластеризации (пустой/шумовой текст)."
                    _undef_quality = "unassigned"
                    kws = _undef_kw if cid < 0 else _crs.kw_dict.get(cid, _undef_kw)
                    out_row = []
                    for col in header_out:
                        if col == cid_col:
                            out_row.append(cid)
                        elif col == ckw_col:
                            out_row.append(kws)
                        elif col == cname_col:
                            out_row.append(_undef_name if cid < 0 else _crs.cluster_name_map.get(cid, _undef_name))
                        elif col == reason_col:
                            out_row.append(_undef_reason if cid < 0 else _crs.cluster_reason_map.get(cid, _undef_reason))
                        elif col == cqual_col:
                            out_row.append(_undef_quality if cid < 0 else _crs.cluster_quality_map.get(cid, _undef_quality))
                        elif col == hier_l1_col and full_l1:
                            out_row.append(full_l1[i] if i < len(full_l1) else -1)
                        elif col == lda_topics_col:
                            out_row.append(kws)  # те же ключевые слова темы
                        elif col == t5_col and full_t5:
                            out_row.append(full_t5[i] if i < len(full_t5) else "")
                        elif col == noise_col:
                            out_row.append(full_noise[i] if i < len(full_noise) else -1)
                        elif col == fb_col:
                            out_row.append("REVIEW" if cid < 0 else _crs.llm_feedback_map.get(cid, "OK"))
                        elif col == conf_col:
                            out_row.append(round(full_conf[i], 4) if i < len(full_conf) else None)
                        else:
                            j = idx_map.get(col)
                            out_row.append(row_vals[j] if (j is not None and j < len(row_vals)) else None)
                    ws_out.append(out_row)

            out_path = CLUST_DIR / f"{in_path.stem}_clustered_{stamp}.xlsx"
            wb_out.save(out_path)
            cluster_widths = [_clu_col_w(c) for c in header_out]
            patch_xlsx_col_widths(out_path, [(0, cluster_widths)])
            out_paths.append(out_path)
            self.after(0, lambda p=out_path: self.log_cluster(f"  -> saved: {p.name}"))

        counts  = Counter([int(lbl) for lbl in _crs.labels])
        rows = [
            (
                cid,
                _crs.kw_dict.get(cid, "(шум / нет меток)" if cid < 0 else ""),
                counts[cid],
            )
            for cid in sorted(counts.keys())
        ]

        elapsed = time.time() - t0
        _out_summary = out_paths[-1].name if len(out_paths) == 1 else f"{len(out_paths)} файлов"
        self._last_cluster_out_paths = list(out_paths)

        def done_ui():
            ui_prog(100.0, "Готово ✅")
            self.cluster_speed.set("")
            self.cluster_eta.set(f"Итого: {elapsed:.0f}с")

            # ── Таблица кластеров ─────────────────────────────────────
            _total_lbl = sum(counts[cid] for cid in counts if cid >= 0)
            _noise_cnt = counts.get(-1, 0)
            _cl_lines = []
            for cid, kw_text, cnt in rows[:20]:
                _pct = cnt / max(_total_lbl + _noise_cnt, 1) * 100
                _name = (kw_text[:38] + "…") if len(kw_text) > 38 else kw_text
                _prefix = "Шум  " if cid < 0 else f"К{cid:<3d}"
                _cl_lines.append(f"  {_prefix} «{_name}»: {cnt} ({_pct:.1f}%)")
            if len(rows) > 20:
                _cl_lines.append(f"  … и ещё {len(rows)-20} кластеров")
            _header = (
                f"[Кластеры] {_crs.K} кластеров | {_total_lbl} строк"
                + (f" + {_noise_cnt} шум" if _noise_cnt else "")
            )
            self.log_cluster(_header + ":\n" + "\n".join(_cl_lines))

            self.log_cluster(f"Готово ✅ | time={elapsed:.1f}s | файлов сохранено: {len(out_paths)}")
            self.last_cluster_summary.set(f"saved={_out_summary} | rows={_crs.total_rows} | _crs.K={_crs.K}")

            for it_ in self.cluster_tree.get_children():
                self.cluster_tree.delete(it_)
            for r in rows[:200]:
                self.cluster_tree.insert("", "end", values=r)

            _lifecycle.complete()

        self.after(0, done_ui)


    def run_cluster(self):
        with self._proc_lock:
            if self._processing:
                return

        snap, files_snapshot = self._prepare_cluster_run_context()
        if snap is None:
            return

        # Preflight passed — safe to touch UI.
        self._right_tabs.select(self._log_tab_indices["cluster"])
        tune_cluster_runtime_for_input(
            files_snapshot=files_snapshot,
            snap=snap,
            hw=self._hw,
            log_fn=self.log_cluster,
        )
        if not try_mark_processing(self):
            return

        begin_long_task(
            cancel_event=self._cancel_event,
            run_button=self.btn_cluster,
            run_button_busy_text="⏳ Кластеризация…",
            stop_button=self.btn_cluster_stop,
            progress_var=self.cluster_progress,
            status_var=self.cluster_status,
            pct_var=self.cluster_pct,
            phase_var=self.cluster_phase,
            speed_var=self.cluster_speed,
            eta_var=self.cluster_eta,
            clear_summary_var=self.last_cluster_summary,
            start_phase="Чтение файла…",
            start_log=self.log_cluster,
        )
        self.log_cluster("==== CLUSTER START ====")
        t0 = time.time()

        _task_ui = prepare_long_task_ui(
            owner=self,
            progress_var=self.cluster_progress,
            status_var=self.cluster_status,
            pct_var=self.cluster_pct,
            phase_var=self.cluster_phase,
            speed_var=self.cluster_speed,
            eta_var=self.cluster_eta,
            run_button=self.btn_cluster,
            run_button_idle_text="▶  Кластеризовать",
            stop_button=self.btn_cluster_stop,
            log_fn=self.log_cluster,
        )
        _ctrl = _task_ui.controller
        ui_prog = _task_ui.ui_prog
        _lifecycle = _task_ui.lifecycle

        def worker():
            try:
                in_paths = [Path(p) for p in files_snapshot]
                total_rows = estimate_total_rows(in_paths)
                start_ts = time.time()
                done = 0

                _crs = ClusterRunState()

                X_all: List[str] = []
                # per-file tracking: (path, header, row_has_text_list)
                file_data: List[tuple] = []

                use_t5 = snap.get("use_t5_summary", False)
                # Исходные тексты для T5 — параллельны X_all (только непустые строки)
                raw_texts_all: List[str] = []

                # ── Роль-специфичные настройки (pipeline stage: prepare_inputs) ─
                prepared_inputs, cluster_snap, _role_str, _bot_str = self._cluster_prepare_data(
                    files_snapshot,
                    snap,
                )
                _n_files  = len(in_paths)
                self.after(0, lambda: self.log_cluster(
                    f"Файлов: {_n_files} | Источник текста: {_role_str} | Игнор. чат-бота: {_bot_str}"
                ))
                # ═══════════════════════════════════════════════════════════════
                # ═══════════════════════════════════════════════════════════════
                # СТАДИЯ 1: Чтение файлов, очистка текста, дедупликация
                # ═══════════════════════════════════════════════════════════════
                # Checkpoint inputs into _crs
                _crs.in_paths = list(in_paths)
                _crs.total_rows = total_rows
                _crs.start_ts = start_ts
                _crs.use_t5 = use_t5
                _crs.cluster_snap = cluster_snap
                # Run stage 1
                self._cluster_worker_stage1(_crs, snap)
                # Unpack outputs
                X_all = _crs.X_all
                file_data = _crs.file_data
                raw_texts_all = _crs.raw_texts_all
                done = _crs.done
                n_ok = _crs.n_ok

                # СТАДИЯ 2: Векторизация (TF-IDF / SBERT / combo / ensemble) и кластеризация
                # ═══════════════════════════════════════════════════════════════
                self.after(0, lambda: self.cluster_status.set("Стадия 2/4: Векторизация текстов…"))
                # ── #12 Применение сохранённой модели (инкрементальный режим) ─
                _use_inc_model = snap.get("use_saved_cluster_model", False)
                _inc_model_path = snap.get("cluster_model_path", "").strip()
                _inc_labels_done = False  # флаг: метки уже выставлены, пропустить обучение
                if _use_inc_model and _inc_model_path:
                    try:
                        if not ensure_cluster_model_trusted(
                            self._trust_store, _inc_model_path,
                            confirm_fn=make_tkinter_confirm_fn(self),
                            logger=_log,
                        ):
                            raise FeatureBuildError("Загрузка модели отменена пользователем.")
                        self.after(0, lambda p=_inc_model_path: self.log_cluster(
                            f"Инкременталь: загрузка модели из {Path(p).name}…"
                        ))
                        _inc = load_and_apply_incremental_model(
                            _inc_model_path,
                            schema_version=self._CLUSTER_MODEL_SCHEMA_VERSION,
                            texts=X_clean,
                            trusted_paths=tuple(self._trust_store.trusted_canonical_paths()),
                            logger=_log,
                            correlation_id=getattr(self, "_cluster_correlation_id", None),
                            diagnostic_mode=bool(getattr(self, "diagnostic_mode", False)),
                            diagnostic_report_path=str(
                                Path(CLUST_DIR) / "cluster_incremental_diagnostic_report.json"
                            ) if bool(getattr(self, "diagnostic_mode", False)) else None,
                        )
                        _inc_vec, _inc_algo, K, kw = _inc.vectorizer, _inc.algo, _inc.k_clusters, _inc.kw
                        self.after(0, lambda k=K, a=_inc_algo: self.log_cluster(
                            f"Инкременталь: K={k}, algo={a}"
                        ))
                        Xv = _inc.xv
                        vec_kw = _inc_vec
                        Xv_tfidf = Xv
                        labels = _inc.labels
                        _inc_labels_done = True
                        use_fastopic_kw_ready = _inc.use_fastopic_kw_ready
                        self.after(0, lambda n=len(labels): self.log_cluster(
                            f"Инкременталь: {n} строк размечены по существующим кластерам ✅"
                        ))
                    except (ModelLoadError, FeatureBuildError) as _inc_e:
                        self.after(0, lambda e=_inc_e: self.log_cluster(
                            f"⚠️ Загрузка модели не удалась: {e} — продолжаю обычную кластеризацию"
                        ))
                        _inc_labels_done = False
                    except Exception as _inc_e:  # pragma: no cover - defensive fallback
                        _log.exception("incremental cluster model load unexpected failure: %s", _inc_e)
                        self.after(0, lambda e=_inc_e: self.log_cluster(
                            f"⚠️ Внутренняя ошибка инкрементальной модели: {e} — "
                            "перехожу к обычной кластеризации"
                        ))
                        _inc_labels_done = False

                # ── #20 Потоковый / онлайн-режим (HashingVectorizer + partial_fit) ─
                _use_streaming = snap.get("use_streaming_cluster", False) and not _inc_labels_done
                if _use_streaming and not use_sbert_clust and not use_combo_clust and not use_ensemble:
                    try:
                        _chunk_sz  = snap.get("streaming_chunk_size", 5000)
                        _K_stream  = min(K_user, _n_clean)
                        self.after(0, lambda k=_K_stream, c=_chunk_sz, n=_n_clean: self.log_cluster(
                            f"Потоковый режим: K={k} | chunk={c} | строк={n}"
                        ))
                        from sklearn.feature_extraction.text import HashingVectorizer as _HV
                        _hv = _HV(
                            n_features=2 ** 18,
                            ngram_range=snap["word_ng"],
                            alternate_sign=False,
                            norm="l2",
                        )
                        _km_stream = MiniBatchKMeans(
                            n_clusters=_K_stream,
                            random_state=snap.get("cluster_random_seed", 42),
                            batch_size=min(_chunk_sz, _n_clean),
                            n_init="auto",
                            max_iter=10,
                        )
                        _n_chunks = max(1, (_n_clean + _chunk_sz - 1) // _chunk_sz)
                        for _ci in range(_n_chunks):
                            _s, _e = _ci * _chunk_sz, min((_ci + 1) * _chunk_sz, _n_clean)
                            _chunk = X_clean[_s:_e]
                            _Xchunk = _hv.transform(_chunk)
                            _km_stream.partial_fit(_Xchunk)
                            _pct = 15.0 + 55.0 * (_ci + 1) / _n_chunks
                            self.after(0, lambda p=_pct, i=_ci+1, t=_n_chunks: ui_prog(
                                p, f"Поток: чанк {i}/{t}…"
                            ))
                            if self._cancel_event.is_set():
                                raise InterruptedError()
                        # Финальный predict на всех данных
                        Xv_stream_parts = [_hv.transform(X_clean[_s:_s+_chunk_sz])
                                           for _s in range(0, _n_clean, _chunk_sz)]
                        import scipy.sparse as _sp_st
                        Xv = _sp_st.vstack(Xv_stream_parts)
                        labels = _km_stream.predict(Xv)
                        K  = _K_stream
                        km = _km_stream
                        vec_kw = None  # для keyword extraction нужен отдельный tfidf
                        Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)
                        _inc_labels_done = True  # пропустить обычную кластеризацию
                        self.after(0, lambda k=K: self.log_cluster(
                            f"Потоковая кластеризация завершена: K={k} ✅"
                        ))
                    except Exception as _st_e:
                        self.after(0, lambda e=_st_e: self.log_cluster(
                            f"⚠️ Потоковый режим: {e} — переключаюсь на обычный"
                        ))
                        _inc_labels_done = False
                        _use_streaming = False

                if use_combo_clust and not _inc_labels_done:
                    # ── Комбо: TF-IDF → SVD → L2-норм  +  SBERT → L2-норм → конкатенация ──
                    self.after(0, lambda: ui_prog(13.0, f"Комбо: TF-IDF (rows={len(X_all)})…"))
                    Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)

                    svd_dim = min(
                        snap.get("combo_svd_dim", 200),
                        Xv_tfidf.shape[1] - 1,
                        Xv_tfidf.shape[0] - 1,
                    )
                    svd_dim = max(10, svd_dim)
                    self.after(0, lambda d=svd_dim: ui_prog(18.0, f"Комбо: SVD dim={d}…"))
                    svd          = TruncatedSVD(n_components=svd_dim, random_state=snap.get("cluster_random_seed", 42))
                    Xv_tfidf_svd = svd.fit_transform(Xv_tfidf)
                    Xv_tfidf_n   = _sk_normalize(Xv_tfidf_svd, norm="l2")

                    def _sbert_prog_combo(raw_pct, status):
                        # SBERT 51..90 → 22..62
                        p = 22.0 + max(0.0, min(1.0, (raw_pct - 51.0) / 39.0)) * 40.0
                        self.after(0, lambda p=p, s=status: ui_prog(p, s))

                    sbert_combo = make_neural_vectorizer(
                        model_name=snap["sbert_model"],
                        batch_size=snap["sbert_batch"],
                        log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                        progress_cb=_sbert_prog_combo,
                        device=snap["sbert_device"],
                    )
                    self.after(0, lambda: ui_prog(22.0, "Комбо: нейросеть загрузка модели…"))
                    Xv_sbert   = sbert_combo.fit_transform(X_clean)
                    Xv_sbert_n = _sk_normalize(Xv_sbert, norm="l2")

                    alpha = snap.get("combo_alpha", 0.5)
                    self.after(0, lambda: ui_prog(63.0, "Комбо: конкатенация TF-IDF + SBERT…"))
                    Xv = _np.hstack([
                        Xv_tfidf_n * (1.0 - alpha),
                        Xv_sbert_n * alpha,
                    ])
                    self.after(0, lambda d=svd_dim, a=alpha, s=Xv.shape: self.log_cluster(
                        f"Комбо: TF-IDF-SVD({d}д) + SBERT, α={a:.1f}, итоговая форма={s}"
                    ))
                    _elbow_start = 70.0
                    _km_start    = 82.0

                elif use_sbert_clust and not _inc_labels_done:
                    # ── Путь SBERT: нейросетевые эмбеддинги ──────────────
                    def _sbert_prog(raw_pct, status):
                        # SBERT 51..90 → 20..58
                        p = 20.0 + max(0.0, min(1.0, (raw_pct - 51.0) / 39.0)) * 38.0
                        self.after(0, lambda p=p, s=status: ui_prog(p, s))

                    sbert = make_neural_vectorizer(
                        model_name=snap["sbert_model"],
                        batch_size=snap["sbert_batch"],
                        log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                        progress_cb=_sbert_prog,
                        device=snap["sbert_device"],
                    )
                    self.after(0, lambda: ui_prog(20.0, "Нейросеть: загрузка модели…"))
                    Xv = sbert.fit_transform(X_clean)

                    # TF-IDF только для извлечения ключевых слов
                    self.after(0, lambda: ui_prog(60.0, "TF-IDF для ключевых слов…"))
                    Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)
                    _elbow_start = 62.0
                    _km_start    = 76.0

                elif use_lda and not _inc_labels_done:
                    # ── Путь LDA: CountVectorizer + LatentDirichletAllocation ─
                    from sklearn.feature_extraction.text import CountVectorizer
                    _cnt_params = dict(
                        analyzer="word",
                        ngram_range=snap["word_ng"],
                        min_df=1,
                        max_features=max(50000, snap["max_features"] // 3),
                        stop_words=_stop_list,
                    )
                    self.after(0, lambda: ui_prog(20.0, f"LDA: CountVectorizer (rows={len(X_all)})…"))
                    try:
                        cnt_vec = CountVectorizer(**_cnt_params)
                        Xv_cnt  = cnt_vec.fit_transform(X_clean)
                    except ValueError as _cv_err:
                        if "empty vocabulary" not in str(_cv_err):
                            raise
                        self.after(0, lambda: self.log_cluster(
                            "⚠️ CountVectorizer: пустой словарь — повтор без стоп-слов"
                        ))
                        _cnt2 = dict(_cnt_params)
                        _cnt2["stop_words"] = None
                        cnt_vec = CountVectorizer(**_cnt2)
                        Xv_cnt  = cnt_vec.fit_transform(X_clean)
                    # TF-IDF для ключевых слов
                    Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)
                    Xv       = Xv_cnt  # LDA работает с count-матрицей
                    _elbow_start = 35.0
                    _km_start    = 50.0

                elif use_bertopic and not _inc_labels_done:
                    # ── BERTopic-like: SBERT → UMAP → HDBSCAN ────────────────
                    def _sbert_prog_bt(raw_pct, status):
                        p = 20.0 + max(0.0, min(1.0, (raw_pct - 51.0) / 39.0)) * 30.0
                        self.after(0, lambda p=p, s=status: ui_prog(p, s))

                    sbert_bt = make_neural_vectorizer(
                        model_name=snap["sbert_model"],
                        batch_size=snap["sbert_batch"],
                        log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                        progress_cb=_sbert_prog_bt,
                        device=snap["sbert_device"],
                    )
                    self.after(0, lambda: ui_prog(20.0, "BERTopic: нейросеть загрузка модели…"))
                    Xv = sbert_bt.fit_transform(X_clean)

                    # UMAP снижение размерности
                    _bt_umap_dim = min(snap.get("umap_n_components", 10), len(X_clean) - 2, 50)
                    _bt_umap_dim = max(2, _bt_umap_dim)
                    try:
                        _umap_src = "GPU (cuML)" if _cuml_umap_available() else "umap-learn"
                        self.after(0, lambda d=_bt_umap_dim, s=_umap_src: ui_prog(
                            55.0, f"BERTopic: UMAP {s} dim={d}…"
                        ))
                        _Xv_dense_bt = _np.asarray(Xv)
                        reducer_bt = _gpu_umap(n_components=_bt_umap_dim, random_state=snap.get("cluster_random_seed", 42))
                        Xv = reducer_bt.fit_transform(_Xv_dense_bt)
                        self.after(0, lambda sh=Xv.shape: self.log_cluster(
                            f"BERTopic UMAP готово ✅  форма={sh}"
                        ))
                    except ImportError:
                        self.after(0, lambda: self.log_cluster(
                            "⚠️ UMAP не установлен — пропускаю. "
                            "Установите: pip install umap-learn"
                        ))

                    # TF-IDF для ключевых слов
                    self.after(0, lambda: ui_prog(62.0, "BERTopic: TF-IDF для ключевых слов…"))
                    Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)
                    _elbow_start = 0.0  # HDBSCAN, no elbow needed
                    _km_start    = 65.0

                elif use_ensemble and not _inc_labels_done:
                    # ── Ансамбль: TF-IDF + SBERT1 + SBERT2 → выбор лучшего ──
                    from sklearn.metrics import silhouette_score as _sil_ens

                    K_ens = min(snap["k_clusters"], n_ok)

                    def _vec_to_dense(Xv_):
                        import scipy.sparse as _sp_e
                        Xd = Xv_.toarray() if _sp_e.issparse(Xv_) else _np.asarray(Xv_)
                        if snap.get("use_cosine_cluster", False):
                            Xd = _sk_normalize(Xd, norm="l2")
                        return Xd

                    def _kmeans_and_score(Xd_, K_, label_str_):
                        """KMeans + Silhouette для одного метода."""
                        km_ = _gpu_kmeans(
                            n_clusters=K_, random_state=snap.get("cluster_random_seed", 42),
                            batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                            n_init=snap["n_init_cluster"] if snap["n_init_cluster"] > 0 else "auto",
                            reassignment_ratio=0.01,
                        )
                        lbl_ = km_.fit_predict(Xd_)
                        _sample = min(5000, len(lbl_))
                        try:
                            sc_ = float(_sil_ens(Xd_, lbl_, sample_size=_sample, random_state=snap.get("cluster_random_seed", 42)))
                        except Exception as _e:
                            _log.debug("silhouette score failed: %s", _e)
                            sc_ = -1.0
                        self.after(0, lambda s=label_str_, v=sc_: self.log_cluster(
                            f"  {s}: Silhouette = {v:+.3f}"
                        ))
                        return lbl_, sc_

                    # 1. TF-IDF
                    self.after(0, lambda: ui_prog(18.0, f"Ансамбль: TF-IDF (rows={len(X_clean)})…"))
                    Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)
                    _Xd_tfidf = _vec_to_dense(Xv_tfidf)
                    self.after(0, lambda: self.log_cluster("Ансамбль: оцениваю варианты…"))
                    _lbl_tfidf, _sc_tfidf = _kmeans_and_score(_Xd_tfidf, K_ens, "TF-IDF")

                    if self._cancel_event.is_set():
                        raise InterruptedError()

                    # 2. SBERT модель 1
                    _m1 = snap["sbert_model"]

                    def _sbert_prog_ens1(raw_pct, status):
                        p = 35.0 + max(0.0, min(1.0, (raw_pct - 51.0) / 39.0)) * 18.0
                        self.after(0, lambda p=p, s=status: ui_prog(p, s))

                    self.after(0, lambda m=_m1: ui_prog(
                        35.0, f"Ансамбль: SBERT-1 ({m.split('/')[-1]})…"
                    ))
                    _sbert1 = make_neural_vectorizer(
                        model_name=_m1,
                        batch_size=snap["sbert_batch"],
                        log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                        progress_cb=_sbert_prog_ens1,
                        device=snap["sbert_device"],
                    )
                    Xv_sbert1 = _sbert1.fit_transform(X_clean)
                    _Xd_s1 = _vec_to_dense(Xv_sbert1)
                    _lbl_s1, _sc_s1 = _kmeans_and_score(
                        _Xd_s1, K_ens, f"SBERT-1 ({_m1.split('/')[-1]})"
                    )

                    if self._cancel_event.is_set():
                        raise InterruptedError()

                    # 3. SBERT модель 2
                    _m2 = snap.get("sbert_model2", _m1)

                    def _sbert_prog_ens2(raw_pct, status):
                        p = 55.0 + max(0.0, min(1.0, (raw_pct - 51.0) / 39.0)) * 18.0
                        self.after(0, lambda p=p, s=status: ui_prog(p, s))

                    self.after(0, lambda m=_m2: ui_prog(
                        55.0, f"Ансамбль: SBERT-2 ({m.split('/')[-1]})…"
                    ))
                    _sbert2 = make_neural_vectorizer(
                        model_name=_m2,
                        batch_size=snap["sbert_batch"],
                        log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                        progress_cb=_sbert_prog_ens2,
                        device=snap["sbert_device"],
                    )
                    Xv_sbert2 = _sbert2.fit_transform(X_clean)
                    _Xd_s2 = _vec_to_dense(Xv_sbert2)
                    _lbl_s2, _sc_s2 = _kmeans_and_score(
                        _Xd_s2, K_ens, f"SBERT-2 ({_m2.split('/')[-1]})"
                    )

                    if self._cancel_event.is_set():
                        raise InterruptedError()

                    # 4. Сравниваем и выбираем победителя
                    _candidates = [
                        ("TF-IDF",                       _Xd_tfidf, _lbl_tfidf, _sc_tfidf),
                        (f"SBERT-1 ({_m1.split('/')[-1]})", _Xd_s1, _lbl_s1, _sc_s1),
                        (f"SBERT-2 ({_m2.split('/')[-1]})", _Xd_s2, _lbl_s2, _sc_s2),
                    ]
                    _best = max(_candidates, key=lambda x: x[3])
                    _best_name, Xv, labels, _best_sc = _best

                    self.after(0, lambda n=_best_name, s=_best_sc: self.log_cluster(
                        f"Ансамбль: победитель → {n}  "
                        f"(Silhouette={s:+.3f}) ✅"
                    ))
                    K          = K_ens
                    _elbow_start = 75.0
                    _km_start    = 85.0

                elif not _inc_labels_done:
                    # ── Путь TF-IDF ───────────────────────────────────────
                    self.after(0, lambda: ui_prog(20.0, f"TF-IDF fit (rows={len(X_all)})…"))
                    Xv_tfidf, vec_kw = _safe_tfidf(_tfidf_params, X_clean)
                    Xv       = Xv_tfidf
                    _elbow_start = 35.0
                    _km_start    = 60.0

                    # ── SVD для TF-IDF → LSA (опционально) ───────────────
                    if snap.get("use_tfidf_svd", False):
                        _tsvd_dim = min(
                            snap.get("tfidf_svd_dim", 100),
                            Xv_tfidf.shape[1] - 1,
                            Xv_tfidf.shape[0] - 1,
                        )
                        _tsvd_dim = max(20, _tsvd_dim)
                        self.after(0, lambda d=_tsvd_dim: ui_prog(
                            28.0, f"TF-IDF SVD dim={d}…"
                        ))
                        _tsvd = TruncatedSVD(n_components=_tsvd_dim, random_state=snap.get("cluster_random_seed", 42))
                        Xv = _sk_normalize(_tsvd.fit_transform(Xv_tfidf), norm="l2")
                        self.after(0, lambda d=_tsvd_dim, sh=Xv.shape: self.log_cluster(
                            f"TF-IDF SVD({d}д) → L2-норм готово ✅  форма={sh}"
                        ))
                        _elbow_start = 38.0
                        _km_start    = 63.0

                # ── Косинусная нормализация (L2) ─────────────────────────
                # Для ансамбля нормализация уже применена внутри пайплайна
                if snap.get("use_cosine_cluster", False) and not use_lda and not use_bertopic and not use_ensemble:
                    Xv = _sk_normalize(Xv, norm="l2")
                    self.after(0, lambda: self.log_cluster(
                        "Косинусная метрика: L2-нормализация применена"
                    ))

                Xv = self._cluster_step_umap(snap, Xv, X_clean)
                if not use_hdbscan and not use_lda and not use_hier and not use_ensemble:
                    K = min(K_user, n_ok)
                    if K < K_user:
                        _k_warn = K
                        self.after(0, lambda k=_k_warn: self.log_cluster(
                            f"⚠️ K уменьшен с {K_user} до {k} "
                            f"(строк с текстом: {n_ok})"
                        ))
                else:
                    K = K_user

                K = self._cluster_step_autok(snap, Xv, X_clean, n_ok)
                # region LDA
                if use_lda and not use_ensemble:
                    # ── LDA тематическое моделирование ───────────────────────
                    from sklearn.decomposition import LatentDirichletAllocation
                    _lda_n      = snap.get("lda_n_topics", 15)
                    _lda_iter   = snap.get("lda_max_iter", 50)
                    _lda_n      = min(_lda_n, n_ok)
                    self.after(0, lambda n=_lda_n, it=_lda_iter: ui_prog(
                        _km_start, f"LDA fit (topics={n}, iter={it})…"
                    ))
                    import scipy.sparse as _sp_lda
                    _Xv_lda = Xv.toarray() if _sp_lda.issparse(Xv) else _np.asarray(Xv, dtype="float32")
                    lda_model = LatentDirichletAllocation(
                        n_components=_lda_n,
                        max_iter=_lda_iter,
                        random_state=snap.get("cluster_random_seed", 42),
                        learning_method="online",
                        n_jobs=self._hw.n_jobs_cv,
                    )
                    doc_topics = lda_model.fit_transform(_Xv_lda)
                    labels     = _np.argmax(doc_topics, axis=1)
                    K          = _lda_n
                    self.after(0, lambda k=K: self.log_cluster(
                        f"LDA: {k} тем ✅  (доминирующая тема → cluster_id)"
                    ))

                # endregion
                # region Иерархическая кластеризация
                elif use_hier and not use_ensemble:
                    # ── Иерархическая (2 уровня) ─────────────────────────────
                    _k_top   = min(snap.get("hier_k_top", 8), n_ok)
                    _k_sub   = snap.get("hier_k_sub", 5)
                    _min_sub = snap.get("hier_min_sub", 50)
                    n_init_val = snap["n_init_cluster"]

                    self.after(0, lambda k=_k_top: ui_prog(
                        _km_start, f"Иерарх. L1: KMeans (K={k})…"
                    ))
                    import scipy.sparse as _sp_hier
                    _Xv_dense_h = Xv.toarray() if _sp_hier.issparse(Xv) else _np.asarray(Xv)
                    km_l1 = _gpu_kmeans(
                        n_clusters=_k_top, random_state=snap.get("cluster_random_seed", 42),
                        batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                        n_init="auto", reassignment_ratio=0.01,
                    )
                    labels_l1_arr = km_l1.fit_predict(_Xv_dense_h)
                    labels_l1 = labels_l1_arr.tolist()

                    # Уровень 2 — разбиваем крупные кластеры
                    labels_hier = list(labels_l1_arr)
                    sub_offset = int(max(labels_l1_arr)) + 1
                    for l1_id in range(_k_top):
                        indices = [i for i, l in enumerate(labels_l1_arr) if l == l1_id]
                        if len(indices) < _min_sub:
                            continue
                        k_sub_cur = min(_k_sub, len(indices))
                        if k_sub_cur < 2:
                            continue
                        self.after(0, lambda l=l1_id, n=len(indices), k=k_sub_cur: self.log_cluster(
                            f"  L1={l}: {n} строк → {k} подкластеров"
                        ))
                        Xv_sub = _Xv_dense_h[indices]
                        km_sub = _gpu_kmeans(
                            n_clusters=k_sub_cur, random_state=snap.get("cluster_random_seed", 42),
                            n_init="auto", reassignment_ratio=0.01,
                        )
                        sub_lbl = km_sub.fit_predict(Xv_sub)
                        for ii, idx in enumerate(indices):
                            labels_hier[idx] = sub_offset + int(sub_lbl[ii])
                        sub_offset += k_sub_cur
                        if self._cancel_event.is_set():
                            raise InterruptedError()

                    labels = _np.array(labels_hier)
                    K = sub_offset
                    self.after(0, lambda k=K, kt=_k_top: self.log_cluster(
                        f"Иерарх. ✅  L1={kt} групп, итого {k} подкластеров"
                    ))

                # endregion
                # region HDBSCAN / BERTopic
                elif use_hdbscan and not use_ensemble:
                    # ── HDBSCAN / BERTopic ────────────────────────────────────
                    _hdb_label = "BERTopic" if use_bertopic else "HDBSCAN"
                    self.after(0, lambda lbl=_hdb_label: ui_prog(_km_start, f"{lbl} кластеризация…"))
                    import scipy.sparse as _sp_hdb
                    _Xv_hdb = Xv.toarray() if _sp_hdb.issparse(Xv) else _np.asarray(Xv)
                    _hdb_min = (
                        snap.get("bertopic_min_topic_size", 10)
                        if use_bertopic
                        else snap.get("hdbscan_min_cluster_size", 10)
                    )
                    # Sentinel 0 (или отрицательное) = «auto»: масштабируется
                    # sqrt(N). Для 100 строк даёт 10, для 10 000 — 100;
                    # избавляет пользователя от ручного подбора для разных
                    # размеров датасетов.
                    try:
                        _hdb_min_int = int(_hdb_min)
                    except (TypeError, ValueError):
                        _hdb_min_int = 0
                    if _hdb_min_int <= 0:
                        _hdb_auto = max(5, int(len(_Xv_hdb) ** 0.5))
                        self.after(0, lambda v=_hdb_auto, n=len(_Xv_hdb): self.log_cluster(
                            f"HDBSCAN min_cluster_size=auto → {v} (√{n})"
                        ))
                        _hdb_min = _hdb_auto
                    else:
                        _hdb_min = _hdb_min_int
                    _hdbscan_ok = False
                    try:
                        try:
                            from sklearn.cluster import HDBSCAN as _HDBSCAN_cls
                        except ImportError:
                            import hdbscan as _hdbscan_mod
                            _HDBSCAN_cls = _hdbscan_mod.HDBSCAN
                        _hdb_min_samples = snap.get("hdbscan_min_samples", 0) or None
                        _hdb_eps = float(snap.get("hdbscan_eps", 0.0))
                        _hdb_kwargs: dict = {"min_cluster_size": _hdb_min}
                        if _hdb_min_samples:
                            _hdb_kwargs["min_samples"] = _hdb_min_samples
                        if _hdb_eps > 0.0:
                            _hdb_kwargs["cluster_selection_epsilon"] = _hdb_eps
                        _hdb_model = _HDBSCAN_cls(**_hdb_kwargs)
                        labels     = _hdb_model.fit_predict(_Xv_hdb)
                        _crs.hdbscan_proba = _np.asarray(
                            getattr(_hdb_model, "probabilities_", [])
                        )
                        _valid     = [l for l in labels if l >= 0]
                        K          = int(max(labels)) + 1 if _valid else 1
                        _n_noise   = int(sum(1 for l in labels if l < 0))
                        self.after(0, lambda lbl=_hdb_label: self.log_cluster(
                            f"{lbl}: K={K} кластеров, шум={_n_noise} точек "
                            f"({100 * _n_noise / max(1, len(labels)):.1f}%)"
                        ))
                        _hdbscan_ok = True

                        # BERTopic: опциональное объединение тем до nr_topics
                        if use_bertopic:
                            _nr_raw = snap.get("bertopic_nr_topics", "auto")
                            _nr_str = str(_nr_raw).strip() if _nr_raw is not None else "auto"
                            if _nr_str and _nr_str.lower() != "auto":
                                try:
                                    _nr_int = int(_nr_str)
                                    if 2 <= _nr_int < K:
                                        self.after(0, lambda n=_nr_int, k=K: self.log_cluster(
                                            f"BERTopic: объединяю {k} тем → {n} (nr_topics)…"
                                        ))
                                        _km_merge = MiniBatchKMeans(
                                            n_clusters=_nr_int,
                                            random_state=snap.get("cluster_random_seed", 42),
                                            batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                                            init="k-means++", n_init="auto",
                                            reassignment_ratio=0.01,
                                        )
                                        labels = _km_merge.fit_predict(_Xv_hdb)
                                        K = _nr_int
                                        _n_noise = 0  # KMeans не создаёт шум
                                        self.after(0, lambda n=_nr_int: self.log_cluster(
                                            f"BERTopic: объединение завершено, итого тем={n} ✅"
                                        ))
                                except Exception as _merge_e:
                                    self.after(0, lambda e=_merge_e: self.log_cluster(
                                        f"⚠️ BERTopic nr_topics merge ошибка: {e} — пропуск"
                                    ))
                    except Exception as _hdb_e:
                        self.after(0, lambda e=_hdb_e: self.log_cluster(
                            f"⚠️ HDBSCAN ошибка: {e}\n"
                            "Установите sklearn ≥ 1.3 или pip install hdbscan\n"
                            "Переключаюсь на MiniBatchKMeans…"
                        ))
                    if not _hdbscan_ok:
                        use_hdbscan = False   # fallback → KMeans

                use_fastopic_kw_ready = False  # FASTopic может выдать kw сам

                if not _inc_labels_done and use_gmm and not use_hdbscan and not use_lda and not use_hier and not use_ensemble:
                    # ── Gaussian Mixture Models ────────────────────────────────
                    from sklearn.mixture import GaussianMixture as _GMM
                    import scipy.sparse as _sp_gmm
                    _K_gmm = min(K_user, n_ok)
                    self.after(0, lambda k=_K_gmm: ui_prog(_km_start, f"GMM (K={k})…"))
                    _Xv_gmm = Xv.toarray() if _sp_gmm.issparse(Xv) else _np.asarray(Xv)
                    _gmm = _GMM(
                        n_components=_K_gmm,
                        covariance_type="diag",
                        random_state=snap.get("cluster_random_seed", 42),
                        max_iter=200, n_init=1,
                    )
                    labels = _gmm.fit_predict(_Xv_gmm)
                    K = _K_gmm
                    self.after(0, lambda k=K: self.log_cluster(
                        f"GMM ✅  K={k} кластеров (covariance=diag)"
                    ))

                elif not _inc_labels_done and use_fastopic and not use_hdbscan and not use_lda and not use_hier and not use_ensemble:
                    # ── FASTopic ──────────────────────────────────────────────
                    _K_ft = min(K_user, n_ok)
                    self.after(0, lambda k=_K_ft: ui_prog(_km_start, f"FASTopic (K={k})…"))
                    try:
                        from fastopic import FASTopic as _FASTopic
                        _ft_model = _FASTopic(
                            num_topics=_K_ft,
                            num_top_words=snap.get("fastopic_n_top_words", 15),
                        )
                        _ft_model.fit(X_clean)
                        _ft_doc_topic = _ft_model.get_doc_topic_distribution()
                        labels = _np.argmax(_np.asarray(_ft_doc_topic), axis=1)
                        _ft_kw_raw = _ft_model.get_topics(snap.get("fastopic_n_top_words", 15))
                        kw = [" ".join(words) for words in _ft_kw_raw]
                        K = _K_ft
                        use_fastopic_kw_ready = True
                        self.after(0, lambda k=K: self.log_cluster(f"FASTopic ✅  K={k} тем"))
                    except ImportError:
                        self.after(0, lambda: self.log_cluster(
                            "⚠️ fastopic не установлен — переключаюсь на KMeans.\n"
                            "Установите: pip install fastopic"
                        ))
                        use_fastopic = False

                # endregion
                # region KMeans (MiniBatchKMeans, основной алгоритм)
                if (not _inc_labels_done and not use_gmm and not use_fastopic and
                        not use_hdbscan and not use_lda and not use_hier and not use_ensemble):
                    # ── MiniBatchKMeans (основной алгоритм) ───────────────────
                    n_init_val = snap["n_init_cluster"]
                    _use_anch  = snap.get("use_anchors", False)
                    _anch_phrases = snap.get("anchor_phrases", [])

                    if _use_anch and _anch_phrases and (use_sbert_clust or use_combo_clust):
                        # Семантические якоря: SBERT-кодирование фраз → init_centers
                        self.after(0, lambda n=len(_anch_phrases): ui_prog(
                            _km_start, f"Якоря: SBERT кодирование {n} фраз…"
                        ))
                        try:
                            sbert_anch = make_neural_vectorizer(
                                model_name=snap["sbert_model"],
                                batch_size=snap["sbert_batch"],
                                log_cb=lambda m: self.after(0, lambda m=m: self.log_cluster(m)),
                                device=snap["sbert_device"],
                            )
                            anch_vecs = sbert_anch.fit_transform(_anch_phrases)
                            if use_combo_clust:
                                # нужно добавить TF-IDF компонент — упрощение: дополняем нулями
                                # ВАЖНО: используем реальный svd_dim (мог быть обрезан по shape),
                                # а не snap["combo_svd_dim"] — иначе shape-mismatch при hstack.
                                svd_part = _np.zeros((len(_anch_phrases), svd_dim))
                                alpha = snap.get("combo_alpha", 0.5)
                                anch_vecs_n = _sk_normalize(_np.asarray(anch_vecs), norm="l2")
                                anch_centers = _np.hstack([
                                    svd_part * (1.0 - alpha),
                                    anch_vecs_n * alpha,
                                ])
                            else:
                                anch_centers = _sk_normalize(_np.asarray(anch_vecs), norm="l2")
                            # Подгоняем число центров к K
                            K_anch = min(len(anch_centers), K)
                            if K_anch < K:
                                # Дополняем случайными строками из Xv
                                import scipy.sparse as _sp_anch
                                _Xvd = Xv.toarray() if _sp_anch.issparse(Xv) else _np.asarray(Xv)
                                rand_idx = _np.random.choice(_Xvd.shape[0], K - K_anch, replace=False)
                                anch_centers = _np.vstack([anch_centers[:K_anch], _Xvd[rand_idx]])
                            else:
                                anch_centers = anch_centers[:K]
                            self.after(0, lambda: self.log_cluster(
                                f"Якоря: {len(anch_centers)} центров → KMeans init"
                            ))
                            km = MiniBatchKMeans(
                                n_clusters=K, random_state=snap.get("cluster_random_seed", 42),
                                batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                                n_init=1, init=anch_centers,
                            )
                        except Exception as _anch_e:
                            self.after(0, lambda e=_anch_e: self.log_cluster(
                                f"⚠️ Якоря: ошибка {e} — использую стандартный KMeans"
                            ))
                            km = None

                    if km is None:
                        _km_label = "GPU KMeans" if _cuml_kmeans_available() else "MiniBatchKMeans"
                        self.after(0, lambda lbl=_km_label: ui_prog(_km_start, f"{lbl} (K={K})…"))
                        km = _gpu_kmeans(
                            n_clusters=K,
                            random_state=snap.get("cluster_random_seed", 42),
                            batch_size=snap.get("kmeans_batch", KMEANS_BATCH_SIZE),
                            n_init=n_init_val if n_init_val > 0 else "auto",
                            reassignment_ratio=0.01,
                        )
                    labels = km.fit_predict(Xv)

                self.after(0, lambda: ui_prog(88.0, "Keywords…"))

                # ═══════════════════════════════════════════════════════════════
                # СТАДИЯ 3: Постобработка — ключевые слова, LLM-нейминг, качество
                # ═══════════════════════════════════════════════════════════════
                # ── Восстановление меток дедупликации ─────────────────────
                # Если были удалены дубли, разворачиваем метки на все исходные строки
                if _dedup_reverse_map is not None:
                    _labels_dedup = _np.asarray(labels)
                    labels = _labels_dedup[_dedup_reverse_map]

                # ── #12 Сохранение модели кластеризации ───────────────────
                if snap.get("save_cluster_model", False) and not _inc_labels_done:
                    try:
                        _mpath = ClusterModelPersistence.normalize_model_path(
                            snap.get("cluster_model_path", "").strip(),
                            str(CLUST_DIR / f"cluster_model_{now_stamp()}.joblib"),
                        )
                        _save_algo = _algo
                        _save_centers = km.cluster_centers_ if hasattr(km, "cluster_centers_") else None
                        _saved_bundle = build_cluster_model_bundle(
                            schema_version=self._CLUSTER_MODEL_SCHEMA_VERSION,
                            vectorizer=vec_kw,
                            algo=_save_algo,
                            k_clusters=K,
                            kw=list(kw) if kw else [],
                            centers=_save_centers,
                            model=hdb if use_hdbscan and "hdb" in dir() else None,
                        )
                        _saved_path = ClusterModelPersistence.save_bundle(_saved_bundle, _mpath)
                        self.after(0, lambda p=_saved_path: self.log_cluster(
                            f"Модель сохранена: {Path(p).name} ✅"
                        ))
                    except Exception as _se:
                        self.after(0, lambda e=_se: self.log_cluster(
                            f"⚠️ Сохранение модели: {e}"
                        ))

                self._cluster_step_quality_metrics(snap, Xv, labels, X_clean)
                if snap.get("merge_similar_clusters", False) and K > 1:
                    try:
                        from ml_diagnostics import merge_similar_clusters as _merge_clusters
                        _merge_thr = float(snap.get("merge_threshold", 0.85))
                        import scipy.sparse as _sp_merge
                        _Xv_merge = (Xv.toarray() if _sp_merge.issparse(Xv) else _np.asarray(Xv))
                        _merged_labels, _minfo = _merge_clusters(
                            labels, _Xv_merge, _merge_thr,
                            log_fn=lambda m: self.after(0, lambda msg=m: self.log_cluster(msg)),
                        )
                        if _minfo["n_after"] < _minfo["n_before"]:
                            labels = _merged_labels
                            K = _minfo["n_after"]
                            self.after(0, lambda i=_minfo: self.log_cluster(
                                f"Слияние кластеров (порог={_merge_thr:.2f}): "
                                f"{i['n_before']}→{i['n_after']} ✅"
                            ))
                    except Exception as _merge_e:
                        self.after(0, lambda e=_merge_e: self.log_cluster(
                            f"⚠️ Слияние кластеров: {e}"
                        ))

                self._cluster_step_keywords(snap, labels, X_clean, Xv_tfidf, kw, vec_kw, use_hdbscan, use_lda, use_hier, use_bertopic, use_fastopic, use_fastopic_kw_ready)
                self._cluster_step_hdbscan_reclust(snap, Xv, labels, noise_labels)
                self._cluster_step_cohesion(snap, Xv, labels, X_clean)
                self._cluster_step_plotly(snap, Xv, labels, X_clean)
                self._cluster_step_t5(snap, raw_texts_all, t5_summaries_all)

                # ═══════════════════════════════════════════════════════════════
                # СТАДИЯ 4: Экспорт результатов в XLSX, сводная таблица
                # ═══════════════════════════════════════════════════════════════
                # Checkpoint state for export
                _crs.labels = labels
                _crs.kw_dict = kw_dict
                _crs.kw = kw
                _crs.file_data = file_data
                _crs.cluster_name_map = cluster_name_map
                _crs.cluster_reason_map = cluster_reason_map
                _crs.cluster_quality_map = cluster_quality_map
                _crs.llm_feedback_map = llm_feedback_map
                _crs.t5_summaries_all = t5_summaries_all
                _crs.labels_l1 = labels_l1
                _crs.noise_labels = noise_labels
                _crs.K = K
                self._cluster_worker_stage4(_crs, snap, t0, ui_prog, _lifecycle)

            except InterruptedError:
                def cancelled_cluster():
                    _lifecycle.cancelled(
                        ui_prog=ui_prog,
                        log_message="⏹ Кластеризация отменена пользователем",
                    )
                self.after(0, cancelled_cluster)

            except FeatureBuildError as e:
                _tb = _traceback.format_exc()
                def _feat_err(e=e, tb=_tb):
                    env = ErrorEnvelope.from_exception(
                        e,
                        error_code="CLUSTER_FEATURE_BUILD",
                        stage="run_cluster",
                        hint="Проверьте текстовые колонки и параметры подготовки/фильтрации.",
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка данных",
                        envelope=env,
                        traceback_text=tb,
                    )
                    messagebox.showerror("Ошибка данных", str(e))
                self.after(0, _feat_err)

            except ValueError as e:
                _emsg = str(e)
                _tb = _traceback.format_exc()
                if "empty vocabulary" in _emsg or "vocabulary" in _emsg.lower():
                    _err_prefix = "ОШИБКА (пустой словарь TF-IDF)"
                    _hint = (
                        "Словарь TF-IDF пуст — все слова отфильтрованы стоп-словами.\n\n"
                        "Попробуйте:\n"
                        "  • Отключить «Стоп-слова» в разделе «Фильтрация шума»\n"
                        "  • Отключить «Шумовые токены»\n"
                        "  • Уменьшить пользовательские стоп-слова\n"
                        "  • Проверить, что текстовые столбцы содержат данные"
                    )
                else:
                    _err_prefix = "ОШИБКА (ValueError)"
                    _hint = _emsg
                def _vocab_err(hint=_hint, prefix=_err_prefix, tb=_tb):
                    env = ErrorEnvelope.from_exception(
                        ValueError(prefix),
                        error_code="CLUSTER_VALUE_ERROR",
                        stage="run_cluster",
                        hint=hint.split("\n")[0],
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка",
                        envelope=env,
                        traceback_text=tb,
                    )
                    messagebox.showerror("Ошибка кластеризации", hint)
                self.after(0, _vocab_err)

            except Exception as e:
                _tb_exc = _traceback.format_exc()
                def err(e=e, tb=_tb_exc):
                    env = ErrorEnvelope.from_exception(
                        e,
                        error_code="CLUSTER_UNEXPECTED",
                        stage="run_cluster",
                        hint="Проверьте traceback и повторите запуск с более узкими параметрами.",
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка",
                        envelope=env,
                        traceback_text=tb,
                    )
                    _brief = str(e)[:300] + ("…" if len(str(e)) > 300 else "")
                    messagebox.showerror(
                        "Ошибка кластеризации",
                        f"{_brief}\n\nПодробности смотри в логе кластеризации.",
                    )
                self.after(0, err)
            finally:
                cleanup_cluster_runtime(lambda msg: _log.debug(msg))
                clear_processing(self)

        threading.Thread(target=worker, daemon=True).start()

    # endregion  # Запуск кластеризации
