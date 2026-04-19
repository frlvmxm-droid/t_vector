# -*- coding: utf-8 -*-
"""
app_train.py — TrainTabMixin: вкладка «Обучение» и логика обучения модели.

Содержит:
  • _build_train_tab()         — построение UI вкладки
  • run_training()             — фоновый поток обучения
  • run_grid_search_c()        — подбор параметра C через CV
  • _build_training_dataset()  — чтение и подготовка обучающей выборки
  • Вспомогательные методы (файл-пикеры, пресеты весов, стоп-слова)
"""
from __future__ import annotations

import json
import subprocess
import sys
import threading
import time
import traceback as _traceback
from collections import Counter, defaultdict
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from sklearn.pipeline import FeatureUnion, Pipeline as _SklPipeline
from sklearn.preprocessing import Normalizer as _SklNormalizer

from constants import (
    MODEL_DIR, CLASS_DIR, CLUST_DIR,
    RUSSIAN_STOP_WORDS,
    NOISE_TOKENS, NOISE_PHRASES,
    ModelConfig, now_stamp, upgrade_config_dict,
)
from config import (
    DEFAULT_COLS,
    PRESET_WEIGHTS, PRESET_WEIGHTS_DESC, PRESET_ALGO_PARAMS,
    SBERT_MODELS, SBERT_MODELS_LIST, SBERT_DEFAULT,
    SETFIT_MODELS, SETFIT_MODELS_LIST, SETFIT_DEFAULT,
    load_exclusions, save_exclusions,
)
from text_utils import normalize_text, parse_dialog_roles, clean_answer_text
from feature_builder import build_feature_text, choose_row_profile_weights
from excel_utils import (
    read_headers, idx_of, estimate_total_rows,
    fmt_eta, fmt_speed, open_tabular,
)
from ml_core import (
    make_hybrid_vectorizer, SBERTVectorizer,
    find_best_c, dataset_health_checks, clean_training_data,
    find_setfit_classifier, detect_near_duplicate_conflicts,
    detect_mislabeled_examples, optuna_tune,
)
from exceptions import FeatureBuildError
from model_loader import ensure_trusted_model_path, load_model_artifact, get_trusted_model_hash, make_tkinter_confirm_fn
from app_train_service import TrainingWorkflow
from ml_training import TrainingOptions
from app_train_view import build_train_files_card as _build_train_files_card_view
from workflow_controller import WorkflowProgressController
from app_train_workflow import validate_train_preconditions, build_validated_train_snapshot
from task_runner import ErrorEnvelope, OperationLifecycle, begin_long_task, prepare_long_task_ui
from app_logger import get_logger
from ui_theme import BG, FG, ENTRY_BG, ACCENT, ACCENT2, MUTED, MUTED2, BORDER, PANEL, SUCCESS, WARNING, ERROR, _best_font
from ui_widgets import Tooltip, ToggleSwitch, RoundedCard, CollapsibleSection
from artifact_contracts import TRAIN_MODEL_ARTIFACT_TYPE
from dataset_analyzer import analyze_dataset, build_param_changes, _fmt

_log = get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Диалог «Анализ датасета»
# ─────────────────────────────────────────────────────────────────────────────

class DatasetAnalysisDialog(tk.Toplevel):
    """Модальный диалог с результатами анализа датасета и кнопкой авто-настройки."""

    def __init__(self, owner: Any, result: Dict[str, Any], apply_cb: Callable):
        super().__init__(owner)
        self.title("📊 Анализ датасета — рекомендации")
        self.resizable(True, True)
        self.configure(bg=BG)
        self.geometry("780x700")
        self.minsize(640, 480)
        self._apply_cb = apply_cb
        self._result   = result
        self._build(result)
        self.grab_set()
        self.focus_force()
        self.transient(owner)

    def _build(self, result: Dict[str, Any]) -> None:
        stats  = result["stats"]
        issues = result["issues"]
        rec    = result["recommendations"]
        cur    = result.get("current_values", {})
        changes = build_param_changes(rec, cur)

        outer = ttk.Frame(self, style="TFrame")
        outer.pack(fill="both", expand=True, padx=12, pady=10)

        # ── Scrollable body ───────────────────────────────────────────────────
        canvas = tk.Canvas(outer, bg=BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        body = ttk.Frame(canvas, style="TFrame")
        body_id = canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_resize(evt):
            canvas.itemconfig(body_id, width=evt.width)
        canvas.bind("<Configure>", _on_resize)
        body.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))

        # ── 1. Статистика ─────────────────────────────────────────────────────
        lf_stat = ttk.LabelFrame(body, text="📋 Статистика датасета", padding=10)
        lf_stat.pack(fill="x", pady=(0, 8))

        _pairs = [
            ("Примеров в датасете",    f"{stats['n_samples']:,}"),
            ("Классов",                f"{stats['n_classes']}"),
            ("Дисбаланс макс/мин",     f"{stats['imbalance_ratio']}x"),
            ("Редких классов (<10)",   f"{len(stats['rare_classes'])}"),
            ("Ср. длина текста",       f"{stats['avg_text_len']:.0f} символов"),
            ("Медиана длины текста",   f"{stats['median_text_len']:.0f} символов"),
            ("Рекомендуемая модель",   rec.get("model_label", "LinearSVC")),
        ]
        for r_i, (lbl, val) in enumerate(_pairs):
            c = r_i % 2
            row_fr = lf_stat if c == 0 else lf_stat
            ttk.Label(lf_stat, text=lbl + ":", style="Card.Muted.TLabel",
                      width=28, anchor="w").grid(row=r_i, column=0, sticky="w", pady=1)
            ttk.Label(lf_stat, text=val, style="Card.TLabel",
                      anchor="w").grid(row=r_i, column=1, sticky="w", padx=(8, 0), pady=1)
        lf_stat.columnconfigure(1, weight=1)

        # Топ-5 классов
        if stats.get("top5_classes"):
            ttk.Separator(lf_stat, orient="horizontal").grid(
                row=len(_pairs), column=0, columnspan=2, sticky="ew", pady=(6, 4))
            ttk.Label(lf_stat, text="Топ-5 крупнейших классов:",
                      style="Card.Muted.TLabel").grid(
                row=len(_pairs)+1, column=0, columnspan=2, sticky="w")
            for j, (lbl, cnt) in enumerate(stats["top5_classes"][:5]):
                total = stats["n_samples"]
                pct   = cnt / max(total, 1) * 100
                bar_w = max(4, int(pct / 1.5))
                bar   = "█" * min(bar_w, 40)
                ttk.Label(lf_stat,
                          text=f"  {lbl[:35]:<36} {cnt:>5}  ({pct:.1f}%)  {bar}",
                          style="Card.TLabel", font=("Courier New", 9)).grid(
                    row=len(_pairs)+2+j, column=0, columnspan=2, sticky="w")

        # Поля с низким покрытием
        cov = stats.get("field_coverage", {})
        if cov:
            ttk.Separator(lf_stat, orient="horizontal").grid(
                row=len(_pairs)+8, column=0, columnspan=2, sticky="ew", pady=(6, 4))
            ttk.Label(lf_stat, text="Заполненность полей:",
                      style="Card.Muted.TLabel").grid(
                row=len(_pairs)+9, column=0, columnspan=2, sticky="w")
            for cj, (field, pct) in enumerate(sorted(cov.items(), key=lambda x: x[1])):
                color = SUCCESS if pct > 0.7 else (WARNING if pct > 0.4 else ERROR)
                bar_w = max(1, int(pct * 30))
                bar   = "█" * bar_w + "░" * (30 - bar_w)
                ttk.Label(lf_stat,
                          text=f"  {field[:20]:<21} {pct*100:>5.1f}%  {bar}",
                          style="Card.TLabel", font=("Courier New", 9),
                          foreground=color).grid(
                    row=len(_pairs)+10+cj, column=0, columnspan=2, sticky="w")

        # ── 2. Проблемы ───────────────────────────────────────────────────────
        if issues:
            lf_iss = ttk.LabelFrame(body, text="⚠  Выявленные проблемы", padding=10)
            lf_iss.pack(fill="x", pady=(0, 8))
            _level_colors = {"critical": ERROR, "warning": WARNING, "info": ACCENT2}
            _level_icons  = {"critical": "🔴", "warning": "🟡", "info": "🔵"}
            for iss in issues:
                color = _level_colors.get(iss["level"], MUTED)
                icon  = _level_icons.get(iss["level"], "•")
                fr = ttk.Frame(lf_iss, style="TFrame")
                fr.pack(fill="x", pady=2)
                ttk.Label(fr, text=icon + "  " + iss["msg"],
                          foreground=color, background=BG,
                          wraplength=680, justify="left", anchor="w").pack(
                    side="left", fill="x", expand=True)

        # ── 3. Таблица рекомендуемых параметров ───────────────────────────────
        lf_rec = ttk.LabelFrame(body, text="⚙  Рекомендуемые параметры", padding=10)
        lf_rec.pack(fill="x", pady=(0, 8))

        # Заголовок таблицы
        hdr_font = ("Segoe UI", 9, "bold")
        for col_i, hdr in enumerate(("Параметр", "Сейчас", "Рекомендуется")):
            ttk.Label(lf_rec, text=hdr, foreground=ACCENT2, background=BG,
                      font=hdr_font, anchor="w").grid(
                row=0, column=col_i, sticky="w", padx=(0 if col_i else 0, 12), pady=(0, 4))
        ttk.Separator(lf_rec, orient="horizontal").grid(
            row=1, column=0, columnspan=3, sticky="ew", pady=(0, 4))

        changed_count = 0
        for row_i, ch in enumerate(changes, start=2):
            cur_str = _fmt(ch["current"])
            rec_str = _fmt(ch["recommended"])
            changed = ch["changed"]
            if changed:
                changed_count += 1
            rec_color = ACCENT2 if changed else FG

            ttk.Label(lf_rec, text=ch["label"], style="Card.Muted.TLabel",
                      anchor="w", width=32).grid(row=row_i, column=0, sticky="w", pady=1)
            ttk.Label(lf_rec, text=cur_str, style="Card.TLabel",
                      anchor="w", width=14).grid(row=row_i, column=1, sticky="w", padx=(0, 12), pady=1)
            rec_lbl = ttk.Label(lf_rec, text=rec_str, anchor="w", width=14,
                                foreground=rec_color, background=PANEL)
            rec_lbl.grid(row=row_i, column=2, sticky="w", pady=1)
            if changed:
                rec_lbl.configure(font=("Segoe UI", 9, "bold"))

        lf_rec.columnconfigure(0, weight=1)

        if rec.get("suggest_anchor_texts"):
            ttk.Separator(lf_rec, orient="horizontal").grid(
                row=len(changes)+2, column=0, columnspan=3, sticky="ew", pady=(8, 4))
            ttk.Label(lf_rec,
                      text="💡 Совет: есть редкие классы — заполните раздел «Якорные тексты» "
                           "для улучшения качества на них.",
                      foreground=ACCENT2, background=BG,
                      wraplength=660, justify="left").grid(
                row=len(changes)+3, column=0, columnspan=3, sticky="w")

        # ── Кнопки ────────────────────────────────────────────────────────────
        btn_row = ttk.Frame(outer, style="TFrame")
        btn_row.pack(fill="x", pady=(8, 0))

        changed_txt = f" ({changed_count} изменений)" if changed_count else " (параметры уже оптимальны)"
        apply_btn = ttk.Button(
            btn_row,
            text=f"✅  Применить рекомендации{changed_txt}",
            command=self._on_apply,
            style="Accent.TButton" if changed_count else "TButton",
        )
        apply_btn.pack(side="left", padx=(0, 8))
        ttk.Button(btn_row, text="Закрыть", command=self.destroy).pack(side="left")

    def _on_apply(self) -> None:
        self._apply_cb(self._result["recommendations"])
        self.destroy()


class TrainTabMixin:
    """Методы вкладки «Обучение» и логика обучения модели."""

    def _lbl_ctrl(self, parent, label: str, var, widget_type: str = "spinbox",
                  tooltip: str = "", padx_lbl=(0, 0), padx_ctrl=(6, 0), **kw):
        """Label + control widget + optional tooltip on the label. Returns the control."""
        lbl = ttk.Label(parent, text=label, anchor="w")
        lbl.pack(side="left", padx=padx_lbl)
        if tooltip:
            Tooltip(lbl, tooltip)
        if widget_type == "spinbox":
            w = ttk.Spinbox(parent, textvariable=var, width=kw.pop("width", 6), **kw)
        elif widget_type == "entry":
            w = ttk.Entry(parent, textvariable=var, width=kw.pop("width", 8), **kw)
        elif widget_type == "check":
            w = ttk.Checkbutton(parent, variable=var, **kw)
        elif widget_type == "combo":
            w = ttk.Combobox(parent, textvariable=var, width=kw.pop("width", 12), **kw)
        else:
            raise ValueError(f"Unknown widget_type: {widget_type!r}")
        w.pack(side="left", padx=padx_ctrl)
        return w

    def _build_train_files_card(self, parent: tk.Widget) -> None:
        """Секция выбора и управления обучающими файлами."""
        _build_train_files_card_view(self, parent)

    # --------------------------------------------------- exclusion words I/O
    def _save_exclusions_from_ui(self) -> None:
        """Читает все три редактора исключений, сохраняет в user_exclusions.json."""
        def _read(widget) -> List[str]:
            raw = widget.get("1.0", "end").strip()
            return [w.strip() for w in raw.splitlines() if w.strip()]

        data = {
            "stop_words":    _read(self._excl_sw_text),
            "noise_tokens":  _read(self._excl_tok_text),
            "noise_phrases": _read(self._excl_ph_text),
        }
        save_exclusions(data)
        # Синхронизируем _user_exclusions на App-уровне
        self._user_exclusions = data
        self._custom_stop_words = data["stop_words"]
        # Обновляем статус-строку
        total = sum(len(v) for v in data.values())
        self._excl_status.set(
            f"✓ Сохранено: стоп-слов {len(data['stop_words'])}, "
            f"токенов {len(data['noise_tokens'])}, "
            f"фраз {len(data['noise_phrases'])}"
        )

    # --------------------------------------------------------- UI toggle helpers
    def _toggle_guide(self):
        """Разворачивает/сворачивает памятку по обучению модели."""
        if self._guide_collapsed:
            self._guide_frame.pack(fill="x", pady=(4, 0))
            self._guide_toggle_btn.configure(text="▼  Скрыть памятку")
            self._guide_collapsed = False
        else:
            self._guide_frame.pack_forget()
            self._guide_toggle_btn.configure(text="▶  Показать памятку")
            self._guide_collapsed = True

    def _toggle_params(self):
        """Разворачивает/сворачивает блок параметров модели."""
        if self._params_collapsed:
            self._params_frame.pack(fill="x", pady=(4, 4))
            self._params_toggle_btn.configure(text="▼  Свернуть параметры")
            self._params_collapsed = False
        else:
            self._params_frame.pack_forget()
            self._params_toggle_btn.configure(text="▶  Развернуть параметры")
            self._params_collapsed = True

    def _browse_pseudo_label_file(self) -> None:
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Файл для псевдо-разметки",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("All files", "*.*")],
        )
        if path:
            self.pseudo_label_file.set(path)

    def _toggle_sw_editor(self):
        if self._sw_collapsed:
            self._sw_frame.pack(fill="x", pady=(4, 0))
            self._sw_toggle_btn.configure(text="▼ Свернуть редактор исключений")
            self._sw_collapsed = False
        else:
            self._sw_frame.pack_forget()
            self._sw_toggle_btn.configure(text="▶ Редактировать слова-исключения")
            self._sw_collapsed = True

    def _on_excl_focusout(self, _event=None) -> None:
        """Авто-сохранение при потере фокуса любым редактором исключений."""
        self._save_exclusions_from_ui()

    def _clear_excl_section(self, widget) -> None:
        """Очищает указанный редактор и сохраняет."""
        widget.delete("1.0", "end")
        self._save_exclusions_from_ui()

    # ─────────────────────────── train tab section builders ──────────────────

    def _build_train_analysis_card(self, parent: "ttk.Frame") -> None:
        """Секция «Анализ датасета + Памятка» для вкладки Данные."""
        _analyze_lf = ttk.LabelFrame(
            parent,
            text="🔍 Предварительный анализ датасета",
            padding=(10, 8),
        )
        _analyze_lf.pack(fill="x", pady=(0, 8))

        _a_row1 = ttk.Frame(_analyze_lf)
        _a_row1.pack(fill="x")

        self._btn_analyze = ttk.Button(
            _a_row1,
            text="🔍 Анализировать датасет и авто-настроить",
            command=self.run_dataset_analysis,
        )
        self._btn_analyze.pack(side="left")
        Tooltip(self._btn_analyze,
                "Читает обучающие файлы, вычисляет ключевые метрики датасета\n"
                "и предлагает оптимальные параметры модели:\n\n"
                "• max_features, C, SVD — по размеру датасета\n"
                "• n-gram диапазон — по средней длине текста\n"
                "• SMOTE, balanced — по дисбалансу классов\n"
                "• Иерархия, K-fold — по числу классов\n"
                "• Field-dropout — по заполненности полей\n\n"
                "После анализа откроется диалог с таблицей\n"
                "«текущее → рекомендуется» и кнопкой «Применить».")

        ttk.Label(_a_row1,
                  text="  ←  Запустите перед обучением для выбора оптимальных параметров",
                  style="Muted.TLabel").pack(side="left", padx=(8, 0))

        _preset_frame = ttk.Frame(_a_row1)
        _preset_frame.pack(side="right")
        _btn_save_preset = ttk.Button(_preset_frame, text="💾 Пресет",
                                      command=self._save_preset)
        _btn_save_preset.pack(side="left", padx=(0, 2))
        Tooltip(_btn_save_preset, "Сохранить текущие настройки как пресет.\nПресеты хранятся в ~/.classification_tool/presets/")
        _btn_load_preset = ttk.Button(_preset_frame, text="📂 Пресет ▾",
                                      command=lambda: self._open_preset_menu(_btn_load_preset))
        _btn_load_preset.pack(side="left")
        Tooltip(_btn_load_preset, "Загрузить ранее сохранённый пресет настроек.")

        # ---- Памятка: как правильно обучать модель ----
        guide_lf = ttk.LabelFrame(
            parent,
            text="📚 Памятка: как правильно обучать модель",
            padding=10,
        )
        guide_lf.pack(fill="x", pady=(0, 10))

        self._guide_collapsed = True
        self._guide_toggle_btn = ttk.Button(
            guide_lf, text="▶  Показать памятку",
            command=self._toggle_guide,
        )
        self._guide_toggle_btn.pack(anchor="w")
        Tooltip(self._guide_toggle_btn,
                "Рекомендации по размеру датасета, балансу классов,\n"
                "признакам переобучения и качеству разметки.")

        self._guide_frame = ttk.Frame(guide_lf)

        _gt = tk.Text(
            self._guide_frame,
            wrap="word", state="normal",
            bg=PANEL, fg=FG, relief="flat",
            font=("Segoe UI", 9),
            height=17,
            cursor="arrow",
            highlightthickness=0,
            padx=6, pady=4,
        )
        _gt.pack(fill="x")

        _gt.tag_configure("h",    foreground=ACCENT2, font=("Segoe UI", 9, "bold"))
        _gt.tag_configure("good", foreground=SUCCESS)
        _gt.tag_configure("warn", foreground=WARNING)
        _gt.tag_configure("bad",  foreground=ERROR)
        _gt.tag_configure("dim",  foreground=MUTED)
        _gt.tag_configure("bold", font=("Segoe UI", 9, "bold"))

        def _g(text, *tags):
            _gt.insert("end", text, tags)

        # ─── Объём данных на класс ───
        _g("📊 Объём данных на класс\n", "h")
        _g("  ✓ Минимум: ", "dim");       _g("30–50 строк", "warn");  _g(" — базовая работоспособность, метрики нестабильны\n", "dim")
        _g("  ✓ Рекомендуется: ", "dim"); _g("100–200 строк", "good"); _g(" — стабильные метрики, разумная точность\n", "dim")
        _g("  ✓ Отлично: ", "dim");       _g("300+ строк", "good");   _g(" — высокая точность и обобщаемость\n", "dim")
        _g("  ✗ Не создавайте класс если примеров < 30 — ", "dim");   _g("добавьте в «Другое»\n", "warn")
        _g("  ✗ Минимум датасета: ", "dim"); _g("N_классов × 50 строк", "bold")
        _g(";  оптимально: ", "dim");        _g("N_классов × 150 строк\n\n", "bold")

        # ─── Баланс классов ───
        _g("⚖ Баланс классов\n", "h")
        _g("  ✓ Допустимо: ", "dim");         _g("до 5:1", "good"); _g(" (крупный к малому) — модель справится\n", "dim")
        _g("  ⚠ Пограничная зона: ", "dim");  _g("5:1 – 10:1", "warn")
        _g(" → включите «Балансировка классов» + SMOTE\n", "dim")
        _g("  ✗ Опасная зона: ", "dim");      _g(">10:1", "bad")
        _g(" → приоритет — дособрать данные; редкий класс будет плохо находиться\n", "dim")
        _g("  ✗ Перекос ", "dim");            _g(">20:1", "bad"); _g(" — модель почти игнорирует редкий класс\n\n", "dim")

        # ─── Переобучение ───
        _g("🔁 Признаки и лечение переобучения\n", "h")
        _g("  ✗ Признак: ", "dim"); _g("Train F1 >> Test F1", "bad")
        _g(" (разница >10–15%) — модель запомнила примеры, не обобщает\n", "dim")
        _g("  ✓ Решения: ", "dim")
        _g("уменьшить C", "good"); _g(" (0.01–0.3)  ·  ", "dim")
        _g("увеличить долю теста", "good"); _g(" (0.25–0.30)  ·  ", "dim")
        _g("добавить данных\n", "good")
        _g("  ✓ Включите лемматизацию и стоп-слова — убирают шум из словаря\n", "dim")
        _g("  ✓ «Подобрать C…» автоматически найдёт оптимальную регуляризацию\n\n", "dim")

        # ─── Качество разметки ───
        _g("🗂 Качество разметки\n", "h")
        _g("  ✓ Каждый класс — единственная чёткая причина без пересечений с другими\n", "dim")
        _g("  ✗ Удалите «мусорные» метки: ", "dim"); _g("прочерки, пустые, НД, неизвестно\n", "bad")
        _g("  ✗ Один и тот же текст с разными метками — шум, требует перепроверки\n", "dim")
        _g("  ✓ Кнопка «📊 Статистика» покажет топ дублей, мельчайшие классы и перекос\n", "dim")

        _gt.configure(state="disabled")

    def _build_train_mode_card(self, parent: "ttk.Frame") -> None:
        """Секция «Режим обучения» (New / Finetune / Ensemble) для вкладки Данные."""
        mode = ttk.Frame(parent, style="Card.TFrame", padding=12)
        mode.pack(fill="x", pady=(0, 10))
        ttk.Label(mode, text="Режим обучения:").grid(row=0, column=0, sticky="w")

        # ── ToggleSwitch radio-group for train mode ───────────────────────────
        _toggle_row = ttk.Frame(mode)
        _toggle_row.grid(row=0, column=1, columnspan=4, sticky="w", padx=10)

        _mv_new = tk.BooleanVar(value=self.train_mode.get() == "new")
        _mv_ft  = tk.BooleanVar(value=self.train_mode.get() == "finetune")
        _mv_ens = tk.BooleanVar(value=self.train_mode.get() == "ensemble")

        def _sync_mode_vars(*_):
            m = self.train_mode.get()
            _mv_new.set(m == "new")
            _mv_ft.set(m == "finetune")
            _mv_ens.set(m == "ensemble")

        self.train_mode.trace_add("write", _sync_mode_vars)

        def _set_mode_exclusive(mode_val, main_var, *others):
            if not main_var.get():
                main_var.set(True)   # prevent turning off the active toggle
                return
            self.train_mode.set(mode_val)
            for v in others:
                v.set(False)

        # New model
        _cell_new = ttk.Frame(_toggle_row)
        _cell_new.pack(side="left", padx=(0, 6))
        ToggleSwitch(_cell_new, variable=_mv_new, bg=PANEL,
                     command=lambda: _set_mode_exclusive("new", _mv_new, _mv_ft, _mv_ens)
                     ).pack(side="left")
        ttk.Label(_cell_new, text="Новая модель").pack(side="left", padx=(5, 0))

        # Finetune
        _cell_ft = ttk.Frame(_toggle_row)
        _cell_ft.pack(side="left", padx=(10, 6))
        ToggleSwitch(_cell_ft, variable=_mv_ft, bg=PANEL,
                     command=lambda: _set_mode_exclusive("finetune", _mv_ft, _mv_new, _mv_ens)
                     ).pack(side="left")
        _lbl_ft = ttk.Label(_cell_ft, text="Дообучение")
        _lbl_ft.pack(side="left", padx=(5, 0))
        self.attach_help(_lbl_ft, "Дообучение",
                         "TF‑IDF обычно требует переобучения для расширения словаря.\n"
                         "Здесь режим «дообучение» = перенять конфиг базовой модели + обучение на новом датасете.",
                         "Дообучение = переобучение по конфигу")

        # Ensemble
        _cell_ens = ttk.Frame(_toggle_row)
        _cell_ens.pack(side="left", padx=(10, 6))
        ToggleSwitch(_cell_ens, variable=_mv_ens, bg=PANEL,
                     command=lambda: _set_mode_exclusive("ensemble", _mv_ens, _mv_new, _mv_ft)
                     ).pack(side="left")
        _lbl_ens = ttk.Label(_cell_ens, text="Ансамбль (2 модели)")
        _lbl_ens.pack(side="left", padx=(5, 0))
        self.attach_help(
            _lbl_ens, "Ансамбль (2 модели)",
            "Обучает две модели последовательно на одном датасете.\n\n"
            "Модель 1 — использует текущие настройки (TF-IDF / SBERT / Гибрид).\n"
            "Модель 2 — использует отдельно выбранный векторизатор (см. блок ниже).\n\n"
            "Каждая модель сохраняется с отдельным именем:\n"
            "  marker1_model_{stamp}_M1_{vec}.joblib\n"
            "  marker1_model_{stamp}_M2_{vec}.joblib\n\n"
            "В логе отображается прогресс каждой модели с префиксом [М1] / [М2].\n"
            "После обучения выводятся метрики обеих моделей.",
            "Обучение двух моделей на одних данных",
        )

        _lbl_base = ttk.Label(mode, text="Базовая модель (.joblib):")
        _lbl_base.grid(row=1, column=0, sticky="w", pady=(8, 0))
        _ent_base = ttk.Entry(mode, textvariable=self.base_model_file, width=80)
        _ent_base.grid(row=1, column=1, columnspan=2, sticky="we", padx=10, pady=(8, 0))
        _btn_base_pick = ttk.Button(mode, text="Выбрать…", command=self.pick_base_model)
        _btn_base_pick.grid(row=1, column=3, pady=(8, 0))
        _btn_base_clr = ttk.Button(mode, text="Очистить",
                                   command=lambda: self.base_model_file.set(""))
        _btn_base_clr.grid(row=1, column=4, pady=(8, 0), padx=(8, 0))
        mode.columnconfigure(1, weight=1)

        _base_widgets = [_lbl_base, _ent_base, _btn_base_pick, _btn_base_clr]

        # ── Блок настроек модели 2 (виден только в режиме ансамбля) ─────────
        _lbl_ens2_vec = ttk.Label(mode, text="Модель 2 — векторизатор:")
        _lbl_ens2_vec.grid(row=2, column=0, sticky="w", pady=(8, 0))

        _ens2_vec_row = ttk.Frame(mode)
        _ens2_vec_row.grid(row=2, column=1, columnspan=3, sticky="w", padx=10, pady=(8, 0))

        for _txt, _val, _tip in [
            ("TF-IDF", "tfidf",
             "Стандартный TF-IDF (char + word N-граммы).\n"
             "Быстро, хорошо, не требует GPU."),
            ("SBERT", "sbert",
             "Нейросетевые эмбеддинги (Sentence-BERT).\n"
             "Медленнее, требует sentence-transformers."),
            ("Гибрид SBERT+TF-IDF", "hybrid",
             "SBERT + TF-IDF через FeatureUnion.\n"
             "Лучшее качество, но самый медленный вариант.\n"
             "SVD для TF-IDF включается автоматически."),
        ]:
            _rb2 = ttk.Radiobutton(_ens2_vec_row, text=_txt,
                                   variable=self.ensemble_vec2, value=_val)
            _rb2.pack(side="left", padx=6)
            self.attach_help(_rb2, f"М2: {_txt}", _tip, _tip)

        _lbl_ens2_sbert = ttk.Label(mode, text="Модель 2 — SBERT модель:")
        _lbl_ens2_sbert.grid(row=3, column=0, sticky="w", pady=(4, 0))

        _ens2_sbert_row = ttk.Frame(mode)
        _ens2_sbert_row.grid(row=3, column=1, columnspan=3, sticky="w", padx=10, pady=(4, 0))
        _cb_ens2_sbert = ttk.Combobox(
            _ens2_sbert_row, textvariable=self.sbert_model2,
            state="readonly", values=SBERT_MODELS_LIST, width=48,
        )
        _cb_ens2_sbert.pack(side="left")
        self.cb_ens2_sbert = _cb_ens2_sbert  # защита от перезаписи в _refresh_combobox_values
        self.attach_help(
            _cb_ens2_sbert,
            "Модель 2: SBERT-модель",
            "SBERT-модель для второй модели ансамбля.\n"
            "Используется только если M2 = SBERT или Гибрид.\n\n"
            "Совет: выбирайте модель, отличную от той, что выбрана для М1,\n"
            "чтобы две модели дополняли друг друга.",
            "Выбор SBERT-модели для М2",
        )

        _ens2_widgets = [_lbl_ens2_vec, _ens2_vec_row, _lbl_ens2_sbert, _ens2_sbert_row]

        def _toggle_base_model_row(*_):
            _mode = self.train_mode.get()
            if _mode == "finetune":
                for w in _base_widgets:
                    w.grid()
                for w in _ens2_widgets:
                    w.grid_remove()
            elif _mode == "ensemble":
                for w in _base_widgets:
                    w.grid_remove()
                for w in _ens2_widgets:
                    w.grid()
            else:  # "new"
                for w in _base_widgets:
                    w.grid_remove()
                for w in _ens2_widgets:
                    w.grid_remove()

        self.train_mode.trace_add("write", _toggle_base_model_row)
        _toggle_base_model_row()  # применить сразу при старте

    def _build_train_params_card(self, parent: "ttk.Frame") -> None:
        """Секция «Параметры модели» + «Стоп-слова» для вкладки Параметры."""

        # ── И1: Авто / Ручная настройка ──────────────────────────────────────
        self._expert_mode_var = tk.BooleanVar(value=False)

        _mode_row = ttk.Frame(parent)
        _mode_row.pack(fill="x", pady=(0, 6))
        ttk.Label(_mode_row, text="Авто-режим", style="Card.TLabel").pack(side="left", padx=(2, 0))
        _tgl = ToggleSwitch(_mode_row, variable=self._expert_mode_var, bg=PANEL)
        _tgl.pack(side="left", padx=8)
        ttk.Label(_mode_row, text="Ручная настройка", style="Card.TLabel").pack(side="left")
        Tooltip(_tgl, "Авто: анализирует датасет и авто-настраивает параметры.\n"
                      "Ручная: полный доступ к настройкам TF-IDF, SVM и аугментации.")

        # Авто-фрейм: показывается в Авто-режиме
        _auto_frame = ttk.Frame(parent)
        ttk.Label(
            _auto_frame,
            text="✨  Нажмите кнопку ниже — система проанализирует датасет и подберёт\n"
                 "     оптимальные параметры автоматически. Затем нажмите «Обучить».",
            style="Card.Muted.TLabel",
            justify="left",
        ).pack(anchor="w", pady=(0, 8))
        ttk.Button(
            _auto_frame,
            text="🔍  Анализировать датасет и авто-настроить",
            command=self.run_dataset_analysis,
        ).pack(anchor="w")

        # ---- параметры модели (Ручной режим) ----
        _params_box = ttk.LabelFrame(parent,
                                     text="Шаг 3 — Параметры модели (дополнительно)",
                                     padding=12)
        self._params_collapsed = True
        self._params_toggle_btn = ttk.Button(
            _params_box, text="▶  Развернуть параметры",
            command=self._toggle_params,
        )
        self._params_toggle_btn.pack(anchor="w", pady=(0, 2))
        Tooltip(self._params_toggle_btn,
                "Дополнительные параметры TF-IDF и LinearSVC.\n"
                "По умолчанию скрыты — значения подобраны разумно.\n"
                "Открывайте только если хотите тонкую настройку модели.")
        self._params_frame = ttk.Frame(_params_box)
        # не пакуем до первого разворота — блок скрыт изначально

        def _update_mode(*_):
            if self._expert_mode_var.get():
                _auto_frame.pack_forget()
                _params_box.pack(fill="x", pady=(0, 10))
            else:
                _params_box.pack_forget()
                _auto_frame.pack(fill="x", pady=(0, 10))

        self._expert_mode_var.trace_add("write", _update_mode)
        _update_mode()  # начальное состояние

        pbox = self._params_frame  # все row* создаём внутри _params_frame

        # --- Строка: характеристики ПК + кнопка сброса рекомендаций ---
        hw_row = ttk.Frame(pbox); hw_row.pack(fill="x", pady=(0, 6))
        lbl_hw = ttk.Label(hw_row, text="ПК:", style="Card.Muted.TLabel")
        lbl_hw.pack(side="left")
        lbl_hw_info = ttk.Label(hw_row, text=self._hw.summary(),
                                style="Card.Muted.TLabel")
        lbl_hw_info.pack(side="left", padx=(6, 16))
        Tooltip(lbl_hw_info, self._hw.rec_summary())

        def _apply_hw_all_rec():
            import hw_profile as _hw_mod
            from tkinter import messagebox as _mb
            self._hw = _hw_mod.detect()
            self._apply_all_hw_params()
            lbl_hw_info.configure(text=self._hw.summary())

            hw = self._hw
            lines = [
                "Конфигурация ПК определена:",
                "",
                f"  CPU : {hw.cpu_cores} ядер",
                f"  ОЗУ : {hw.ram_gb:.0f} ГБ",
            ]
            if hw.gpu_count >= 2:
                for _i, _n in enumerate(hw.gpu_names):
                    lines.append(f"  GPU {_i}: {_n}")
                _vr_str = f"{hw.gpu_vram_gb:.1f}" if hw.gpu_vram_gb is not None else "?"
                lines.append(f"  VRAM: {_vr_str} ГБ × {hw.gpu_count}")
            elif hw.gpu_name:
                lines.append(f"  GPU : {hw.gpu_name}")
                lines.append(f"  VRAM: {hw.gpu_vram_gb:.1f} ГБ")
            else:
                lines.append("  GPU : не обнаружен")
            if hw.gpu_compute_major > 0:
                lines.append(
                    f"  CC  : {hw.gpu_compute_major}.{hw.gpu_compute_minor}  "
                    f"TF32={'да' if hw.gpu_supports_tf32 else 'нет'}  "
                    f"BF16={'да' if hw.gpu_supports_bf16 else 'нет'}"
                )
            lines += [
                "",
                "Применённые параметры:",
                f"  TF-IDF max_features : {hw.max_features:,}",
                f"  SBERT батч          : {hw.sbert_batch}",
                f"  SetFit батч         : {hw.setfit_batch}",
                f"  T5 батч             : {hw.t5_batch}",
            ]
            _mb.showinfo("Конфигурация ПК", "\n".join(lines), parent=self)

        btn_hw_rec = ttk.Button(hw_row, text="Определить конфигурацию ПК",
                                command=_apply_hw_all_rec)
        btn_hw_rec.pack(side="left")
        Tooltip(btn_hw_rec, "Re-определяет железо и устанавливает оптимальные параметры")
        ttk.Separator(pbox, orient="horizontal").pack(fill="x", pady=(0, 6))

        def _sec_header(text: str, top: int = 8):
            _fnt = (_best_font(), 9, "bold")
            _r = ttk.Frame(pbox)
            _r.pack(fill="x", pady=(top, 2))
            tk.Label(_r, text=text, font=_fnt, bg=PANEL, fg=ACCENT).pack(side="left")
            tk.Frame(_r, bg=MUTED2, height=1).pack(side="left", fill="x", expand=True, padx=(6, 0))

        _sec_header("⚙  Векторизация", top=0)
        row0 = ttk.Frame(pbox); row0.pack(fill="x")
        lbl_char = ttk.Label(row0, text="Симв. N-граммы:", anchor="w")
        lbl_char.pack(side="left")
        Tooltip(lbl_char, "Диапазон символьных N-грамм для TF-IDF.\n"
                          "Символьные N-граммы ловят морфологию, опечатки и суффиксы.\n"
                          "Например, (2,4) — биграммы, триграммы и 4-граммы из символов.\n"
                          "Рекомендуется: от 2 до 4")
        ttk.Spinbox(row0, from_=1, to=10, textvariable=self.char_ng_min, width=4).pack(side="left", padx=(8, 4))
        ttk.Label(row0, text="—", anchor="w").pack(side="left")
        ttk.Spinbox(row0, from_=1, to=10, textvariable=self.char_ng_max, width=4).pack(side="left", padx=4)
        lbl_word = ttk.Label(row0, text="Слов. N-граммы:", anchor="w")
        lbl_word.pack(side="left", padx=(14, 0))
        Tooltip(lbl_word, "Диапазон словарных N-грамм для TF-IDF.\n"
                          "Словарные N-граммы ловят устойчивые словосочетания.\n"
                          "Например, (1,2) — юниграммы и биграммы из слов.\n"
                          "Рекомендуется: от 1 до 2")
        ttk.Spinbox(row0, from_=1, to=5, textvariable=self.word_ng_min, width=4).pack(side="left", padx=(8, 4))
        ttk.Label(row0, text="—", anchor="w").pack(side="left")
        ttk.Spinbox(row0, from_=1, to=5, textvariable=self.word_ng_max, width=4).pack(side="left", padx=4)

        row1 = ttk.Frame(pbox); row1.pack(fill="x", pady=(8, 0))
        lbl_mindf = ttk.Label(row1, text="Мин. частота слова:", anchor="w")
        lbl_mindf.pack(side="left")
        Tooltip(lbl_mindf, "Минимальное число документов, в которых слово должно встречаться.\n"
                           "Слова реже этого порога игнорируются — снижает шум.\n"
                           "Рекомендуется: 2–5")
        ttk.Spinbox(row1, from_=1, to=50, textvariable=self.min_df, width=6).pack(side="left", padx=8)
        lbl_mxf = ttk.Label(row1, text="Макс. признаков:", anchor="w")
        lbl_mxf.pack(side="left", padx=(14, 0))
        Tooltip(lbl_mxf, "Максимальное число признаков в словаре TF-IDF.\n"
                         "Оставьте пустым — без ограничений.\n"
                         "Ограничение ускоряет обучение и снижает потребление памяти.")
        ttk.Entry(row1, textvariable=self.max_features, width=10).pack(side="left", padx=8)
        cb_stf = ttk.Checkbutton(row1, text="Логарифм TF", variable=self.sublinear_tf)
        cb_stf.pack(side="left", padx=(14, 0))
        self.attach_help(cb_stf, "Логарифм TF (sublinear_tf)",
                         "sublinear_tf=True заменяет TF на 1+log(TF).\n"
                         "Уменьшает влияние сверхчастых слов — улучшает качество.\n"
                         "Обычно даёт +точность и +устойчивость модели.",
                         "Рекомендуется: включено")

        _sec_header("🤖  Классификатор")
        row2 = ttk.Frame(pbox); row2.pack(fill="x", pady=(0, 0))
        cb_cw = ttk.Checkbutton(row2, text="Балансировка классов", variable=self.class_weight_balanced)
        cb_cw.pack(side="left")
        self.attach_help(cb_cw, "Балансировка классов (class_weight=balanced)",
                         "Автоматически взвешивает каждый класс обратно пропорционально\n"
                         "его частоте в обучающей выборке.\n\n"
                         "КАК РАБОТАЕТ:\n"
                         "  weight[класс] = N_строк / (N_классов × count[класс])\n"
                         "  Редкие классы получают вес >> 1, частые — << 1.\n"
                         "  LinearSVC учитывает эти веса при вычислении ошибки:\n"
                         "  ошибка на редком классе «стоит» модели дороже.\n\n"
                         "НА ЧТО ВЛИЯЕТ:\n"
                         "  + Macro F1 растёт — редкие причины классифицируются точнее\n"
                         "  + Recall редких классов улучшается значительно\n"
                         "  − Precision частых классов может немного снизиться\n"
                         "  − Accuracy (общая) может снизиться на 1–3%\n\n"
                         "КОГДА ВКЛЮЧАТЬ:\n"
                         "  Всегда, если некоторые классы встречаются в 10+ раз реже\n"
                         "  других (типичная ситуация в банковских тематиках).\n\n"
                         "КОГДА ОТКЛЮЧАТЬ:\n"
                         "  Если классы равномерно распределены (разница ≤ 2x),\n"
                         "  или если важна Accuracy, а не Macro F1.",
                         "Рекомендуется: включено при неравномерных данных")
        lbl_c = ttk.Label(row2, text="C (регуляризация):", anchor="w")
        lbl_c.pack(side="left", padx=(14, 0))
        Tooltip(lbl_c, "Параметр регуляризации LinearSVC.\n"
                       "Малые значения (0.01–0.1) — сильная регуляризация, устойчивость.\n"
                       "Большие значения (1–10) — слабая регуляризация, подгонка под данные.\n"
                       "Рекомендуется: 0.1–1.0")
        ttk.Entry(row2, textvariable=self.C, width=8).pack(side="left", padx=8)
        self.btn_grid_search_c = ttk.Button(row2, text="Подобрать C…", command=self.run_grid_search_c)
        self.btn_grid_search_c.pack(side="left", padx=(0, 4))
        btn_gs = self.btn_grid_search_c
        Tooltip(btn_gs, "Перебирает C ∈ [0.05, 0.1, 0.3, 1, 3, 10, 30] через\n"
                        "5-fold стратифицированную CV (метрика: macro F1).\n"
                        "Читает те же файлы обучения и те же параметры что и обучение.\n"
                        "По завершению автоматически выставляет лучший C в поле выше.\n"
                        "⚠ Занимает ~5–15 минут на больших датасетах.")
        self.btn_grid_search_stop = ttk.Button(row2, text="⏹", width=3,
                                               command=self._stop_grid_search, state="disabled")
        self.btn_grid_search_stop.pack(side="left", padx=(0, 8))
        Tooltip(self.btn_grid_search_stop, "Остановить GridSearch после текущей итерации.\n"
                                           "Результаты уже завершённых итераций будут сохранены.")
        lbl_mi = ttk.Label(row2, text="Макс. итераций:", anchor="w")
        lbl_mi.pack(side="left", padx=(14, 0))
        Tooltip(lbl_mi, "Максимальное число итераций оптимизатора LinearSVC.\n"
                        "Увеличьте при предупреждении «ConvergenceWarning».\n"
                        "Рекомендуется: 1000–5000")
        ttk.Entry(row2, textvariable=self.max_iter, width=8).pack(side="left", padx=8)
        lbl_ts = ttk.Label(row2, text="Доля теста:", anchor="w")
        lbl_ts.pack(side="left", padx=(14, 0))
        Tooltip(lbl_ts, "Доля данных, выделяемых в тестовую выборку для валидации.\n"
                        "Например, 0.2 = 20% данных — тест, 80% — обучение.\n"
                        "Установите 0 — обучение без валидации (все данные идут в модель).\n"
                        "Рекомендуется: 0.15–0.25")
        ttk.Entry(row2, textvariable=self.test_size, width=6).pack(side="left", padx=8)

        row2b = ttk.Frame(pbox); row2b.pack(fill="x", pady=(6, 0))
        lbl_cm = ttk.Label(row2b, text="Калибровка вероятностей:", anchor="w")
        lbl_cm.pack(side="left")
        cb_cm = ttk.Combobox(row2b, textvariable=self.calib_method, state="readonly",
                             width=12, values=["auto", "sigmoid", "isotonic"])
        cb_cm.pack(side="left", padx=8)
        self.cb_calib_method = cb_cm  # защита от перезаписи в _refresh_combobox_values
        self.attach_help(cb_cm, "Метод калибровки вероятностей",
                         "Управляет тем, как LinearSVC → вероятности класса.\n\n"
                         "auto (рекомендуется):\n"
                         "  • Автоматически выбирает метод по среднему числу сэмплов/класс\n"
                         "  • >= 1000 сэмплов/класс → isotonic, иначе → sigmoid\n\n"
                         "sigmoid (Platt scaling):\n"
                         "  • Подходит для любого размера датасета\n"
                         "  • Быстрее и надёжнее при < 1000 сэмплов/класс\n"
                         "  • Вероятности могут быть слегка переоптимистичны\n\n"
                         "isotonic (монотонная регрессия):\n"
                         "  • Даёт лучше откалиброванные вероятности\n"
                         "  • Рекомендуется при > 1000 сэмплов/класс\n"
                         "  • Риск переобучения на малых датасетах\n\n"
                         "НА ЧТО ВЛИЯЕТ:\n"
                         "  pred_proba в Excel — насколько точно отражает уверенность\n"
                         "  review_threshold — пороговая фильтрация по уверенности\n"
                         "  Macro F1 почти не меняется — влияет только качество вероятностей",
                         "auto / sigmoid / isotonic")

        row3 = ttk.Frame(pbox); row3.pack(fill="x", pady=(8, 0))
        cb_sw = ttk.Checkbutton(row3, text="Стоп-слова (рус.)", variable=self.use_stop_words)
        cb_sw.pack(side="left")
        self.attach_help(cb_sw, "Стоп-слова (рус.)",
                         "Удаляет частотные русские стоп-слова из словарного TF-IDF.\n"
                         "Полезно, если в данных много шумных служебных слов (и, в, на, для…).\n"
                         "Может ухудшить качество, если причина выражается короткими словами.\n"
                         "Не влияет на символьный TF-IDF — только на словарный (word).",
                         "По умолчанию: включено")

        row4 = ttk.Frame(pbox); row4.pack(fill="x", pady=(4, 0))
        cb_nt = ttk.Checkbutton(row4, text="Шумовые токены", variable=self.use_noise_tokens)
        cb_nt.pack(side="left")
        self.attach_help(cb_nt, "Шумовые токены",
                         "Исключает встроенный список шумовых токенов из словарного TF-IDF.\n"
                         "Список: маски ПДн (fio, num…), SDK-артефакты (sdk, request),\n"
                         "URL-фрагменты (sberbank, appeals), HTML-остатки (class),\n"
                         "фатические филлеры (угу, ага), коды шаблонов (7000, 7001).\n"
                         "Отключите, если хотите, чтобы эти токены влияли на веса модели.",
                         "По умолчанию: включено")
        cb_np = ttk.Checkbutton(row4, text="Шумовые фразы", variable=self.use_noise_phrases)
        cb_np.pack(side="left", padx=(16, 0))
        self.attach_help(cb_np, "Шумовые фразы",
                         "Удаляет встроенный список шаблонных фраз до токенизации.\n"
                         "Работает для обоих векторизаторов (char и word).\n"
                         "Список: скрипты операторов, шаблоны ответов банка, фразы ожидания.\n"
                         "Отключите, если фразы несут смысл в вашем датасете.",
                         "По умолчанию: включено")
        cb_pf = ttk.Checkbutton(row4, text="TF-IDF по полям", variable=self.use_per_field)
        cb_pf.pack(side="left", padx=(16, 0))
        self.attach_help(cb_pf, "Раздельный TF-IDF по полям",
                         "Включено: каждое поле (Описание, Клиент, Оператор, …) получает\n"
                         "свой независимый словарь TF-IDF. Веса применяются как\n"
                         "скалярный множитель нормированной матрицы — без артефактов\n"
                         "повторения текста.\n\n"
                         "Отключено: единый TF-IDF на весь объединённый текст\n"
                         "(совместимость с моделями, обученными до этой версии).\n\n"
                         "Рекомендуется: включено (точность выше на 3–7%)",
                         "Рекомендуется: включено")

        row4b = ttk.Frame(pbox); row4b.pack(fill="x", pady=(4, 0))
        cb_svd = ttk.Checkbutton(row4b, text="SVD (LSA)", variable=self.use_svd)
        cb_svd.pack(side="left")
        self.attach_help(cb_svd, "TruncatedSVD — Латентный семантический анализ",
                         "Сжимает пространство TF-IDF до N латентных тем.\n"
                         "Эффект: схожие слова ('перевод', 'перевести', 'переводить')\n"
                         "группируются в одну латентную тему → меньше шума.\n\n"
                         "После SVD применяется L2-нормализация (нужна для LinearSVC).\n\n"
                         "⚠ Обучение и классификация замедляются при N > 500.\n"
                         "Рекомендуется: 200–500 компонент для начала.")
        ttk.Label(row4b, text="Компоненты:").pack(side="left", padx=(8, 0))
        ttk.Spinbox(row4b, from_=50, to=3000, increment=50,
                    textvariable=self.svd_components, width=6).pack(side="left", padx=4)
        ttk.Label(row4b, text="(рек. 200)", style="Card.Muted.TLabel").pack(side="left")
        cb_smote = ttk.Checkbutton(row4b, text="SMOTE", variable=self.use_smote)
        cb_smote.pack(side="left", padx=(16, 0))
        self.attach_help(cb_smote, "SMOTE — балансировка редких классов",
                         "Текстовый оверсэмплинг: если один класс встречается\n"
                         "в 3+ раза реже мажоритарного — дублирует его примеры\n"
                         "до 1/3 мажоритарного (не более чем втрое).\n\n"
                         "Пример: 1000 примеров vs 100 → 100 дублируется до 300.\n\n"
                         "Работает ДО обучения (на текстах), не требует\n"
                         "дополнительных библиотек.\n\n"
                         "Ожидаемый прирост: +8–15% macro F1 на редких классах.\n"
                         "Рекомендуется: включено (особенно при дисбалансе классов).",
                         "Балансировка при дисбалансе ≥ 3:1")
        cb_drop_conf = ttk.Checkbutton(row4b, text="Убирать конфликты", variable=self.drop_conflicts)
        cb_drop_conf.pack(side="left", padx=(16, 0))
        self.attach_help(cb_drop_conf, "Удалять конфликтные примеры",
                         "Конфликт: один и тот же текст имеет разные метки в датасете.\n"
                         "Такие строки — противоречивый обучающий сигнал,\n"
                         "они ухудшают точность модели.\n\n"
                         "При включении: ВСЕ строки конфликтных текстов\n"
                         "удаляются до обучения (до дедупликации).\n\n"
                         "Ожидаемый прирост: +3–10% macro F1 при наличии конфликтов.\n"
                         "Рекомендуется: включено.",
                         "Удаляет тексты с противоречивой разметкой")
        cb_lem = ttk.Checkbutton(row4b, text="Лемматизация", variable=self.use_lemma)
        cb_lem.pack(side="left", padx=(16, 0))
        self.attach_help(cb_lem, "Лемматизация (pymorphy2)",
                         "Приводит все слова к базовой форме перед TF-IDF.\n\n"
                         "Примеры:\n"
                         "  снял / снимаю / снятие / снять → «снять»\n"
                         "  банке / банку / банком / банков → «банк»\n"
                         "  перевод / перевести / переводит → «переводить»\n\n"
                         "Особенно полезно для русского языка с богатой морфологией.\n"
                         "Ожидаемый прирост: +4–10% macro F1.\n\n"
                         "⚠ Требует: pip install pymorphy2\n"
                         "⚠ Заметно замедляет обучение и классификацию\n"
                         "   (≈ x3–5 раз при больших датасетах).",
                         "Требует pymorphy2")

        row4c = ttk.Frame(pbox); row4c.pack(fill="x", pady=(4, 0))
        cb_meta = ttk.Checkbutton(row4c, text="Мета-признаки диалога", variable=self.use_meta)
        cb_meta.pack(side="left")
        self.attach_help(cb_meta, "Мета-признаки диалога (MetaFeatureExtractor)",
                         "Добавляет 15 числовых признаков структуры диалога\n"
                         "поверх TF-IDF через sklearn FeatureUnion.\n\n"
                         "ЧТО ИЗВЛЕКАЕТСЯ:\n"
                         "  • Наличие секций: [DESC], [CLIENT], [OPERATOR],\n"
                         "    [SUMMARY], [ANSWER_SHORT], [ANSWER_FULL]  (0/1)\n"
                         "  • Тип канала: call, chat, call+chat            (0/1)\n"
                         "  • Длины секций в словах (log-масштаб)\n"
                         "  • Число реплик клиента и оператора\n"
                         "  • Доля слов клиента в диалоге (0..1)\n\n"
                         "НА ЧТО ВЛИЯЕТ:\n"
                         "  + TF-IDF не видит «кто говорит больше» — мета-признаки\n"
                         "    добавляют именно этот сигнал\n"
                         "  + Помогает различать похожие тематики с разной структурой\n"
                         "    (напр. простой вопрос vs длинная жалоба)\n"
                         "  + Ортогонален TF-IDF — не конкурирует, а дополняет\n"
                         "  − Минимальный оверхед (~0.01 сек на 10к строк)\n\n"
                         "СОВМЕСТИМОСТЬ:\n"
                         "  Работает с PerFieldVectorizer и с legacy TF-IDF.\n"
                         "  При use_svd=True SVD применяется к TF-IDF части;\n"
                         "  мета-признаки добавляются после SVD без сжатия.\n\n"
                         "Рекомендуется: включить для улучшения Macro F1 на +2–5%.",
                         "Числовые признаки структуры диалога")

        _sec_header("⚖  Аугментация данных")
        row4d = ttk.Frame(pbox); row4d.pack(fill="x", pady=(0, 0))
        cb_llm_aug = ttk.Checkbutton(row4d, text="LLM-аугментация редких классов",
                                      variable=self.use_llm_augment)
        cb_llm_aug.pack(side="left")
        self.attach_help(cb_llm_aug, "LLM-аугментация: перефразировки для редких классов",
                         "Для классов с числом примеров ниже порога — генерирует\n"
                         "перефразировки через LLM (Anthropic/OpenAI/Ollama).\n\n"
                         "Требует настроенного LLM-провайдера (вкладка Кластеризация).\n\n"
                         "Ожидаемый прирост: +5–15% macro F1 на редких классах.\n"
                         "⚠ Замедляет обучение на время LLM-запросов.",
                         "Генерирует новые примеры через LLM")
        ttk.Label(row4d, text="Мин. примеров:").pack(side="left", padx=(16, 0))
        ttk.Spinbox(row4d, from_=5, to=200, increment=5,
                    textvariable=self.augment_min_samples, width=5).pack(side="left", padx=(2, 0))
        ttk.Label(row4d, text="Перефразировок:").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row4d, from_=1, to=10, increment=1,
                    textvariable=self.augment_n_paraphrases, width=4).pack(side="left", padx=(2, 0))

        row4e = ttk.Frame(pbox); row4e.pack(fill="x", pady=(4, 0))
        cb_nd = ttk.Checkbutton(row4e, text="Искать почти-дубли с конфликтами",
                                variable=self.detect_near_dups)
        cb_nd.pack(side="left")
        self.attach_help(cb_nd, "Детекция почти-дублей с разными метками",
                         "Находит пары текстов с косинусным сходством ≥ порога,\n"
                         "которым присвоены разные метки.\n\n"
                         "Такие пары сигнализируют о:\n"
                         "  • Противоречивой разметке (одна ситуация → разные метки)\n"
                         "  • Слишком похожих категориях (стоит объединить)\n\n"
                         "Результат выводится в лог — пары не удаляются автоматически.\n"
                         "Порог 0.92: почти идентичные. 0.80: очень похожие.\n\n"
                         "⚠ Для датасетов > 5 000 строк анализируется выборка 5 000.",
                         "Диагностика противоречивой разметки")
        ttk.Label(row4e, text="Порог сходства:").pack(side="left", padx=(16, 0))
        ttk.Spinbox(row4e, from_=0.70, to=0.99, increment=0.02,
                    textvariable=self.near_dup_threshold, width=5,
                    format="%.2f").pack(side="left", padx=(2, 0))

        row4f = ttk.Frame(pbox); row4f.pack(fill="x", pady=(4, 0))
        cb_hn = ttk.Checkbutton(row4f, text="Hard negative oversampling",
                                variable=self.use_hard_negatives)
        cb_hn.pack(side="left")
        self.attach_help(cb_hn, "Оверсэмплинг граничных примеров",
                         "Для пар классов с близкими центроидами (TF-IDF косинус ≥ 0.50)\n"
                         "находит примеры каждого класса, наиболее похожие на другой класс,\n"
                         "и дублирует их в обучающей выборке.\n\n"
                         "Это помогает классификатору лучше различать категории,\n"
                         "которые он чаще всего путает.\n\n"
                         "Ожидаемый прирост: +3–8% macro F1 на близких классах.\n"
                         "⚠ Работает только с TF-IDF (без SBERT).\n"
                         "⚠ Немного замедляет обучение (~5–15 сек на большом датасете).",
                         "Усиливает обучение на граничных примерах")

        row4g = ttk.Frame(pbox); row4g.pack(fill="x", pady=(4, 0))
        cb_fd = ttk.Checkbutton(row4g, text="Field dropout (пропуски полей)",
                                variable=self.use_field_dropout)
        cb_fd.pack(side="left")
        self.attach_help(cb_fd, "Аугментация: устойчивость к пропущенным полям",
                         "Создаёт обучающие копии с случайно удалёнными секциями\n"
                         "([DESC], [CLIENT], [OPERATOR] и др.).\n\n"
                         "Если у вас часть строк не заполнена (например, 3 из 6\n"
                         "колонок пусты) — модель обучается на похожих примерах\n"
                         "и не теряется при неполных данных.\n\n"
                         "Вероятность удаления поля: задаётся ниже (15% = каждое\n"
                         "поле удаляется в 15% копий).\n"
                         "Копий на пример: сколько аугментированных версий создавать.\n\n"
                         "Ожидаемый прирост: +5–15% macro F1 при >30% пропусков.\n"
                         "⚠ Увеличивает обучающую выборку в (1 + Копий) раз.",
                         "Обучает модель на частичных данных")
        ttk.Label(row4g, text="Вер-ть дропаута:").pack(side="left", padx=(16, 0))
        ttk.Spinbox(row4g, from_=0.05, to=0.50, increment=0.05,
                    textvariable=self.field_dropout_prob, width=5,
                    format="%.2f").pack(side="left", padx=(2, 0))
        ttk.Label(row4g, text="Копий:").pack(side="left", padx=(10, 0))
        ttk.Spinbox(row4g, from_=1, to=5, increment=1,
                    textvariable=self.field_dropout_copies, width=3).pack(side="left", padx=(2, 0))

        row4h = ttk.Frame(pbox); row4h.pack(fill="x", pady=(4, 0))
        cb_en = ttk.Checkbutton(row4h, text="Нормализация сущностей",
                                variable=self.use_entity_norm)
        cb_en.pack(side="left")
        self.attach_help(cb_en, "Нормализация банковских сущностей",
                         "Заменяет конкретные числа/даты/номера на токены:\n"
                         "  5 000 руб → [СУММА]\n"
                         "  +7 999 123-45-67 → [ТЕЛЕФОН]\n"
                         "  4276 **** 1234 → [КАРТА]\n"
                         "  23.04.2024 → [ДАТА]\n"
                         "  договор №1234567 → [ДОГОВОР]\n\n"
                         "TF-IDF перестаёт тратить словарный бюджет на вариации\n"
                         "одной сущности → больше места для смысловых слов.\n\n"
                         "Ожидаемый прирост: +3–8% macro F1.\n"
                         "⚠ Модель должна быть переобучена после включения.",
                         "Заменяет суммы/даты/номера на токены")

        cb_ml = ttk.Checkbutton(row4h, text="Искать ошибки разметки",
                                variable=self.detect_mislabeled)
        cb_ml.pack(side="left", padx=(16, 0))
        self.attach_help(cb_ml, "Детекция неправильно размеченных примеров",
                         "После обучения прогоняет модель на обучающих данных.\n"
                         "Примеры, где вер-ть истинного класса < порога — подозрительные:\n"
                         "возможная ошибка разметки или пограничный случай.\n\n"
                         "Результат: лист «Подозрительные метки» в Excel-отчёте,\n"
                         "отсортированный по возрастанию уверенности.\n\n"
                         "Ожидаемый эффект: +5–12% после ручной коррекции меток.",
                         "Флагирует возможно ошибочные метки")
        ttk.Label(row4h, text="Порог:").pack(side="left", padx=(8, 0))
        ttk.Spinbox(row4h, from_=0.05, to=0.60, increment=0.05,
                    textvariable=self.mislabeled_threshold, width=5,
                    format="%.2f").pack(side="left", padx=(2, 0))

        row4i = ttk.Frame(pbox); row4i.pack(fill="x", pady=(4, 0))
        cb_hier = ttk.Checkbutton(row4i, text="Иерархическая классификация",
                                  variable=self.use_hierarchical)
        cb_hier.pack(side="left")
        self.attach_help(cb_hier, "Автоматическая иерархия классов",
                         "Кластеризует классы в √N групп по схожести текстов.\n"
                         "Обучает «групповую» модель поверх основной.\n\n"
                         "При классификации: вероятности класса умножаются на\n"
                         "вероятность его группы → путаница между группами\n"
                         "подавляется, различение внутри группы усиливается.\n\n"
                         "Особенно полезно при ≥ 10 классах с иерархической\n"
                         "структурой (напр., «карты» / «переводы» / «кредиты»).\n\n"
                         "Ожидаемый прирост: +5–15% macro F1 на ≥ 15 классах.\n"
                         "⚠ Требует минимум 6 классов в обучающем датасете.",
                         "Автогруппировка: √N групп из N классов")

        _sec_header("🔬  Расширенные методы")
        row4j = ttk.Frame(pbox); row4j.pack(fill="x", pady=(0, 0))
        cb_pl = ttk.Checkbutton(row4j, text="Псевдо-разметка",
                                variable=self.use_pseudo_label)
        cb_pl.pack(side="left")
        self.attach_help(cb_pl, "Псевдо-разметка неразмеченного файла",
                         "После первичного обучения прогоняет модель по указанному\n"
                         "файлу и добавляет строки с уверенностью ≥ порога в\n"
                         "обучающий датасет, затем обучает итоговую модель.\n\n"
                         "Позволяет использовать тысячи неразмеченных обращений\n"
                         "для улучшения качества модели.\n\n"
                         "Ожидаемый прирост: +4–10% macro F1 на редких классах.\n"
                         "⚠ Высокий порог (≥ 0.92) снижает риск шумовых меток.\n"
                         "⚠ Файл должен иметь те же колонки текста, что и обучающий.",
                         "Добавляет высокоуверенные предсказания в обучение")
        ttk.Label(row4j, text="Порог:").pack(side="left", padx=(8, 0))
        ttk.Spinbox(row4j, from_=0.70, to=0.99, increment=0.02,
                    textvariable=self.pseudo_label_threshold, width=5,
                    format="%.2f").pack(side="left", padx=(2, 0))
        ttk.Label(row4j, text="Файл:").pack(side="left", padx=(12, 0))
        ttk.Entry(row4j, textvariable=self.pseudo_label_file,
                  width=28).pack(side="left", padx=(2, 0))
        ttk.Button(row4j, text="…",
                   command=lambda: self._browse_pseudo_label_file()).pack(side="left", padx=(2, 0))

        # --- якорные тексты классов ---
        row4k = ttk.Frame(pbox); row4k.pack(fill="x", pady=(4, 0))
        cb_anc = ttk.Checkbutton(row4k, text="Якорные тексты классов",
                                 variable=self.use_anchor_texts)
        cb_anc.pack(side="left")
        self.attach_help(cb_anc, "Якорные тексты классов",
                         "Для каждого класса напишите 1–3 предложения, описывающие типичное обращение.\n"
                         "Эти тексты добавятся в обучающую выборку как синтетические примеры.\n\n"
                         "Особенно помогает для редких классов (< 20 примеров).\n\n"
                         "Формат: каждая строка = «Название класса: текст описания»\n"
                         "Пример:\n"
                         "  Задержка зачисления: Клиент ждёт перевод уже несколько дней, деньги не пришли.\n"
                         "  Блокировка карты: Карта заблокирована при попытке оплаты за границей.")
        ttk.Label(row4k, text="Копий:").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row4k, from_=1, to=10, increment=1,
                    textvariable=self.anchor_copies, width=4).pack(side="left")

        row4k2 = ttk.Frame(pbox); row4k2.pack(fill="x", pady=(2, 4))
        ttk.Label(row4k2, text="  Якоря (Класс: текст):", style="Card.TLabel").pack(side="left", anchor="nw")
        _anc_frame = ttk.Frame(row4k2)
        _anc_frame.pack(side="left", fill="x", expand=True, padx=(8, 0))
        _anc_vscroll = ttk.Scrollbar(_anc_frame, orient="vertical")
        self.anchor_texts_widget = tk.Text(_anc_frame, height=4, wrap="word",
                                           yscrollcommand=_anc_vscroll.set)
        _anc_vscroll.config(command=self.anchor_texts_widget.yview)
        self.anchor_texts_widget.pack(side="left", fill="x", expand=True)
        _anc_vscroll.pack(side="right", fill="y")

        # --- Confident Learning ---
        row4l = ttk.Frame(pbox); row4l.pack(fill="x", pady=(4, 0))
        cb_cl = ttk.Checkbutton(row4l, text="Confident Learning (ошибки разметки)",
                                variable=self.use_confident_learning)
        cb_cl.pack(side="left")
        self.attach_help(cb_cl, "Confident Learning",
                         "Запускает 5-fold кросс-валидацию и для каждого примера\n"
                         "вычисляет вероятность истинного класса.\n\n"
                         "Если p(данная метка) < среднего по классу × порог —\n"
                         "пример считается потенциально ошибочно размеченным.\n\n"
                         "Результат: лист «Confident Learning» в Excel-отчёте.\n\n"
                         "Порог < 1.0 = строже (меньше подозрительных).\n"
                         "Порог > 1.0 = мягче (больше подозрительных).")
        ttk.Label(row4l, text="Порог:").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row4l, from_=0.3, to=2.0, increment=0.1,
                    textvariable=self.confident_learning_threshold,
                    width=5, format="%.1f").pack(side="left")

        # --- K-fold ансамбль ---
        row4n = ttk.Frame(pbox); row4n.pack(fill="x", pady=(4, 0))
        cb_kf = ttk.Checkbutton(row4n, text="K-fold ансамбль",
                                variable=self.use_kfold_ensemble)
        cb_kf.pack(side="left")
        self.attach_help(cb_kf, "K-fold ансамбль",
                         "Обучает K независимых моделей на разных фолдах.\n"
                         "При классификации усредняются вероятности всех K моделей\n"
                         "с весом 50% (основная модель) + 50% (K-fold среднее).\n\n"
                         "Уменьшает дисперсию предсказаний, повышает стабильность.\n\n"
                         "K=5 — хороший баланс качества и скорости.\n"
                         "Требует ≥ K×10 обучающих примеров.")
        ttk.Label(row4n, text="K =").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row4n, from_=2, to=10, increment=1,
                    textvariable=self.kfold_k, width=4).pack(side="left")

        # --- Optuna автоподбор гиперпараметров ---
        row4m = ttk.Frame(pbox); row4m.pack(fill="x", pady=(4, 0))
        cb_opt = ttk.Checkbutton(row4m, text="Optuna: авто-подбор гиперпараметров",
                                 variable=self.use_optuna)
        cb_opt.pack(side="left")
        self.attach_help(cb_opt, "Optuna — автоматический подбор гиперпараметров",
                         "Перед основным обучением запускается Optuna (TPE-сэмплер),\n"
                         "которая подбирает оптимальное значение C для LinearSVC.\n\n"
                         "Требует: pip install optuna\n\n"
                         "Рекомендуемое число испытаний: 20–50.\n"
                         "Больше испытаний = лучше качество, но дольше обучение.")
        ttk.Label(row4m, text="Испытаний:").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row4m, from_=10, to=100, increment=5,
                    textvariable=self.n_optuna_trials, width=5).pack(side="left")

        self._sw_collapsed = True
        self._sw_toggle_btn = ttk.Button(row3, text="▶ Редактировать слова-исключения",
                                         command=self._toggle_sw_editor)
        self._sw_toggle_btn.pack(side="left", padx=(16, 0))

        # ── Редактор слов-исключений (изначально скрыт) ──────────────────────
        self._sw_frame = ttk.Frame(pbox)

        # Загружаем текущие пользовательские исключения
        _excl = self._user_exclusions

        def _make_builtin_viewer(parent, label: str, content: str, height: int = 3):
            """Создаёт read-only текстовый блок для встроенных исключений."""
            lf = ttk.LabelFrame(parent, text=label, padding=4)
            lf.pack(fill="x", pady=(0, 4))
            sc = ttk.Scrollbar(lf, orient="vertical")
            txt = tk.Text(lf, height=height, wrap="word", state="normal",
                          bg=PANEL, fg=MUTED, relief="flat",
                          yscrollcommand=sc.set)
            sc.configure(command=txt.yview)
            sc.pack(side="right", fill="y")
            txt.pack(fill="x")
            txt.insert("1.0", content)
            txt.configure(state="disabled")
            return txt

        def _make_user_editor(parent, label: str, initial: List[str], height: int = 4):
            """Создаёт редактируемый блок для пользовательских исключений."""
            lf = ttk.LabelFrame(parent, text=label, padding=4)
            lf.pack(fill="x", pady=(0, 4))
            sc = ttk.Scrollbar(lf, orient="vertical")
            txt = tk.Text(lf, height=height, wrap="word", undo=True,
                          bg=ENTRY_BG, fg=FG, insertbackground=FG, relief="flat",
                          yscrollcommand=sc.set)
            sc.configure(command=txt.yview)
            sc.pack(side="right", fill="y")
            txt.pack(fill="x")
            if initial:
                txt.insert("1.0", "\n".join(initial))
            # Авто-сохранение при потере фокуса
            txt.bind("<FocusOut>", self._on_excl_focusout)
            return txt

        # ── Секция 1: Стоп-слова ─────────────────────────────────────────────
        ttk.Label(
            self._sw_frame,
            text="Стоп-слова исключаются из словаря TF-IDF (применяются при чекбоксе «Стоп-слова»).",
            style="Card.Muted.TLabel",
        ).pack(anchor="w", pady=(6, 2))

        _builtin_tokens = sorted(RUSSIAN_STOP_WORDS | NOISE_TOKENS)
        _make_builtin_viewer(
            self._sw_frame,
            f"Встроенные стоп-слова + токены-шум ({len(_builtin_tokens)} шт.) — только просмотр",
            ",  ".join(_builtin_tokens),
            height=3,
        )
        _sw_user_hdr = ttk.Frame(self._sw_frame)
        _sw_user_hdr.pack(fill="x")
        ttk.Label(_sw_user_hdr, text="+ Ваши стоп-слова (одно на строку):").pack(side="left")
        ttk.Button(
            _sw_user_hdr, text="Очистить",
            command=lambda: self._clear_excl_section(self._excl_sw_text),
        ).pack(side="right")
        self._excl_sw_text = _make_user_editor(
            self._sw_frame,
            f"Пользовательские стоп-слова ({len(_excl.get('stop_words', []))} шт.)",
            _excl.get("stop_words", []),
        )

        # ── Секция 2: Токены-шум ─────────────────────────────────────────────
        ttk.Separator(self._sw_frame, orient="horizontal").pack(fill="x", pady=(4, 4))
        ttk.Label(
            self._sw_frame,
            text="Токены-шум добавляются в стоп-слова ВСЕГДА (независимо от чекбокса). "
                 "Применяется при чекбоксе «Токены-шум».",
            style="Card.Muted.TLabel",
        ).pack(anchor="w", pady=(0, 2))

        _make_builtin_viewer(
            self._sw_frame,
            f"Встроенные токены-шум ({len(NOISE_TOKENS)} шт.) — только просмотр",
            ",  ".join(sorted(NOISE_TOKENS)),
            height=3,
        )
        _tok_user_hdr = ttk.Frame(self._sw_frame)
        _tok_user_hdr.pack(fill="x")
        ttk.Label(_tok_user_hdr, text="+ Ваши токены-шум (одно на строку):").pack(side="left")
        ttk.Button(
            _tok_user_hdr, text="Очистить",
            command=lambda: self._clear_excl_section(self._excl_tok_text),
        ).pack(side="right")
        self._excl_tok_text = _make_user_editor(
            self._sw_frame,
            f"Пользовательские токены-шум ({len(_excl.get('noise_tokens', []))} шт.)",
            _excl.get("noise_tokens", []),
        )

        # ── Секция 3: Фразы-шум ──────────────────────────────────────────────
        ttk.Separator(self._sw_frame, orient="horizontal").pack(fill="x", pady=(4, 4))
        ttk.Label(
            self._sw_frame,
            text="Фразы-шум удаляются из текста ПЕРЕД токенизацией (regex-замена). "
                 "Применяется при чекбоксе «Фразы-шум».",
            style="Card.Muted.TLabel",
        ).pack(anchor="w", pady=(0, 2))

        _make_builtin_viewer(
            self._sw_frame,
            f"Встроенные фразы-шум ({len(NOISE_PHRASES)} шт.) — только просмотр",
            "\n".join(NOISE_PHRASES),
            height=4,
        )
        _ph_user_hdr = ttk.Frame(self._sw_frame)
        _ph_user_hdr.pack(fill="x")
        ttk.Label(_ph_user_hdr, text="+ Ваши фразы-шум (одна на строку):").pack(side="left")
        ttk.Button(
            _ph_user_hdr, text="Очистить",
            command=lambda: self._clear_excl_section(self._excl_ph_text),
        ).pack(side="right")
        self._excl_ph_text = _make_user_editor(
            self._sw_frame,
            f"Пользовательские фразы-шум ({len(_excl.get('noise_phrases', []))} шт.)",
            _excl.get("noise_phrases", []),
            height=4,
        )

        # ── Кнопки управления + статус ───────────────────────────────────────
        sw_btns = ttk.Frame(self._sw_frame)
        sw_btns.pack(fill="x", pady=(6, 4))
        ttk.Button(
            sw_btns, text="Сохранить всё",
            command=self._save_exclusions_from_ui,
        ).pack(side="left")
        _total_excl = sum(len(v) for v in _excl.values())
        self._excl_status = tk.StringVar(
            value=(
                f"✓ Сохранено: стоп-слов {len(_excl.get('stop_words', []))}, "
                f"токенов {len(_excl.get('noise_tokens', []))}, "
                f"фраз {len(_excl.get('noise_phrases', []))}"
            )
        )
        ttk.Label(sw_btns, textvariable=self._excl_status,
                  style="Card.Muted.TLabel").pack(side="left", padx=(12, 0))

        # ---- SBERT section ----

    def _build_sbert_install_card(self, parent: "ttk.Frame") -> None:
        """Секция «SBERT + SetFit» для вкладки Параметры."""
        sbox = ttk.LabelFrame(parent, text="Шаг 4 — SBERT векторизация (опционально)", padding=12)
        sbox.pack(fill="x", pady=(0, 10))

        srow0 = ttk.Frame(sbox); srow0.pack(fill="x")
        cb_sbert = ttk.Checkbutton(srow0, text="SBERT вместо TF-IDF", variable=self.use_sbert)
        cb_sbert.pack(side="left")
        self.attach_help(cb_sbert, "SBERT векторизация (вместо TF-IDF)",
                         "SBERT (Sentence-BERT) создаёт dense-эмбеддинги через нейросеть.\n"
                         "Полностью заменяет TF-IDF: все параметры ngram/min_df игнорируются.\n\n"
                         "КАК РАБОТАЕТ:\n"
                         "  Каждая строка → вектор 768 чисел (семантическое представление).\n"
                         "  Синонимы и перефразировки оказываются близко в пространстве.\n\n"
                         "НА ЧТО ВЛИЯЕТ:\n"
                         "  + Понимает семантику: «снять наличные» = «получить деньги»\n"
                         "  + Независим от словаря (нет OOV-проблемы)\n"
                         "  + Лучше на коротких текстах и описаниях (DESC)\n"
                         "  − Медленнее TF-IDF (особенно без GPU)\n"
                         "  − Первый запуск = скачивание модели из интернета\n"
                         "  − TF-IDF признаки (точные n-граммы) теряются\n\n"
                         "Рекомендуется: rubert-tiny2 для старта, sbert_large — макс. качество.\n"
                         "Для гибрида (SBERT + TF-IDF) используй чекбокс «SBERT + TF-IDF» ниже.\n\n"
                         "⚠ Несовместимо с «SBERT + TF-IDF (гибрид)» — при включении гибрида\n"
                         "этот режим отключается автоматически.",
                         "SBERT вместо TF-IDF (полная замена)")

        ttk.Separator(srow0, orient="vertical").pack(side="left", fill="y", padx=(14, 0))
        ttk.Label(srow0, text="или", style="Card.Muted.TLabel").pack(side="left", padx=4)
        cb_sbert_hybrid = ttk.Checkbutton(srow0, text="SBERT + TF-IDF (гибрид)",
                                           variable=self.use_sbert_hybrid)
        cb_sbert_hybrid.pack(side="left", padx=(4, 0))
        self.attach_help(cb_sbert_hybrid, "SBERT + TF-IDF гибрид (E2)",
                         "Объединяет TF-IDF (точные n-граммы) и SBERT (семантику)\n"
                         "через sklearn FeatureUnion — лучший результат обоих подходов.\n\n"
                         "КАК РАБОТАЕТ:\n"
                         "  1. TF-IDF пайплайн → SVD (обязателен, даёт dense-вектор)\n"
                         "     → вектор ~200–300 чисел (точные слова и n-граммы)\n"
                         "  2. SBERT → вектор 768 чисел (семантика)\n"
                         "  3. FeatureUnion конкатенирует: итого ~970 признаков\n"
                         "  4. LinearSVC обучается на объединённом пространстве\n\n"
                         "НА ЧТО ВЛИЯЕТ:\n"
                         "  + Точные совпадения терминов (TF-IDF) + семантика (SBERT)\n"
                         "  + Стабильно превосходит каждый подход по отдельности\n"
                         "  + Macro F1 обычно выше на 3–8% чем лучший из двух\n"
                         "  − Обучение в 2–4x медленнее (кодирование SBERT + SVD)\n"
                         "  − Требует sentence-transformers + GPU желательна\n\n"
                         "ВАЖНО:\n"
                         "  SVD автоматически включается для TF-IDF части (обязательно\n"
                         "  для получения dense-вектора пригодного для конкатенации).\n"
                         "  Если «SBERT вместо TF-IDF» тоже включён — гибрид приоритетнее.\n\n"
                         "Рекомендуется: включить если GPU доступна и качество критично.\n\n"
                         "⚠ Несовместимо с «SBERT вместо TF-IDF» — при включении гибрида\n"
                         "режим «только SBERT» отключается автоматически.",
                         "SBERT + TF-IDF — гибрид двух подходов")

        # Взаимоисключение: включение одного режима снимает другой
        def _on_sbert_enabled(*_):
            if self.use_sbert.get() and self.use_sbert_hybrid.get():
                self.use_sbert_hybrid.set(False)

        def _on_hybrid_enabled(*_):
            if self.use_sbert_hybrid.get() and self.use_sbert.get():
                self.use_sbert.set(False)

        self.use_sbert.trace_add("write", _on_sbert_enabled)
        self.use_sbert_hybrid.trace_add("write", _on_hybrid_enabled)

        srow1 = ttk.Frame(sbox); srow1.pack(fill="x", pady=(8, 0))
        ttk.Label(srow1, text="Модель:").pack(side="left")
        cb_sm = ttk.Combobox(srow1, textvariable=self.sbert_model, state="readonly",
                             width=48, values=SBERT_MODELS_LIST)
        cb_sm.pack(side="left", padx=8)
        self.cb_sbert_model_combo = cb_sm  # защита от перезаписи в _refresh_combobox_values
        self.attach_help(cb_sm, "SBERT модель",
                         "\n".join(f"• {k}:\n  {v}" for k, v in SBERT_MODELS.items()),
                         "Выбор SBERT-модели")

        # статус установки + кнопка install
        self._sbert_status = tk.StringVar(value="")
        self._sbert_status_lbl = ttk.Label(srow1, textvariable=self._sbert_status, style="Card.Muted.TLabel")
        self._sbert_status_lbl.pack(side="left", padx=12)

        def _install_sbert():

            self._sbert_status.set("Устанавливаю sentence-transformers…")
            btn_inst.configure(state="disabled")
            def _run():
                try:
                    result = subprocess.run(
                        [sys.executable, "-m", "pip", "install",
                         "--upgrade", "sentence-transformers"],
                        timeout=300,
                    )
                    if result.returncode == 0:
                        self.after(0, lambda: self._sbert_status.set(
                            "✅ Установлено! Перезапустите приложение."))
                    else:
                        self.after(0, lambda: self._sbert_status.set(
                            f"❌ pip ошибка (code={result.returncode}). "
                            f"Попробуйте: pip install --upgrade sentence-transformers"))
                except subprocess.TimeoutExpired:
                    self.after(0, lambda: self._sbert_status.set(
                        "❌ Timeout. Выполните вручную:\n"
                        "pip install --upgrade sentence-transformers"))
                except Exception as ex:
                    self.after(0, lambda ex=ex: self._sbert_status.set(f"❌ Ошибка: {ex}"))
                finally:
                    self.after(0, lambda: btn_inst.configure(state="normal"))
            threading.Thread(target=_run, daemon=True).start()

        btn_inst = ttk.Button(srow1, text="Установить sentence-transformers", command=_install_sbert)
        btn_inst.pack(side="left", padx=8)

        # --- строка 2: выбор устройства (CPU / GPU) ---
        srow_dev = ttk.Frame(sbox); srow_dev.pack(fill="x", pady=(8, 0))
        lbl_dev = ttk.Label(srow_dev, text="Устройство:")
        lbl_dev.pack(side="left")
        Tooltip(lbl_dev,
                "Устройство для выполнения SBERT-инференса.\n\n"
                "• auto — SentenceTransformer сам выбирает (GPU если доступен)\n"
                "• cpu  — всегда CPU (медленнее, но универсально)\n"
                "• cuda — GPU через CUDA (быстрее, требует NVIDIA + torch-cuda)\n\n"
                "При обучении и классификации используется один параметр.")
        cb_dev = ttk.Combobox(
            srow_dev, textvariable=self.sbert_device,
            state="readonly", width=10,
            values=self.gpu_device_values,
        )
        cb_dev.pack(side="left", padx=8)
        self.cb_sbert_device_combo = cb_dev   # защита от _refresh_combobox_values
        self.attach_help(cb_dev, "SBERT Устройство",
                         "Управляет тем, на каком железе выполняется кодирование SBERT.\n\n"
                         "auto   — авто-выбор (рекомендуется: GPU если есть).\n"
                         "cpu    — явный CPU. Подходит для машин без GPU.\n"
                         "cuda   — GPU (NVIDIA). Значительно ускоряет обработку "
                         "больших файлов (10k+ строк).\n"
                         "cuda:0 — явно первый GPU (при нескольких видеокартах).\n"
                         "cuda:1 — явно второй GPU.\n\n"
                         "Требования: NVIDIA GPU + PyTorch с CUDA-поддержкой.\n"
                         "Проверка: предупреждение появится если CUDA недоступна.",
                         "auto / cpu / cuda / cuda:0 / cuda:1")
        # динамически показываем GPU-статус
        self._sbert_gpu_status = tk.StringVar(value="")
        ttk.Label(srow_dev, textvariable=self._sbert_gpu_status,
                  style="Card.Muted.TLabel").pack(side="left", padx=12)

        # --- строка 3: размер батча SBERT ---
        srow_batch = ttk.Frame(sbox); srow_batch.pack(fill="x", pady=(8, 0))
        lbl_sb = ttk.Label(srow_batch, text="Батч SBERT:")
        lbl_sb.pack(side="left")
        Tooltip(lbl_sb,
                "Количество текстов, кодируемых за один вызов SentenceTransformer.encode().\n\n"
                "• GPU: больше батч → выше загрузка VRAM → быстрее.\n"
                "  Если CUDA out of memory — уменьшите значение.\n"
                "• CPU: 16–64 (выше почти не ускоряет).\n\n"
                f"Рекомендовано для вашего ПК: {self._hw.sbert_batch}")
        ttk.Spinbox(srow_batch, from_=8, to=1024, increment=8,
                    textvariable=self.sbert_batch, width=6).pack(side="left", padx=8)
        ttk.Label(srow_batch,
                  text=f"(рек. {self._hw.sbert_batch})",
                  style="Card.Muted.TLabel").pack(side="left")

        # --- строка torch+CUDA-исправления ---
        srow_cuda = ttk.Frame(sbox)
        srow_cuda.pack(fill="x", pady=(4, 0))
        self._sbert_torch_hint = tk.StringVar(value="Проверяю GPU…")
        ttk.Label(srow_cuda, textvariable=self._sbert_torch_hint,
                  style="Card.Muted.TLabel").pack(side="left")

        def _install_torch_cuda():

            btn_tc.configure(state="disabled")

            def _run():
                self.after(0, lambda: self._sbert_torch_hint.set(
                    "Устанавливаю torch + torchvision + torchaudio (cu124)…"))
                try:
                    result = subprocess.run(
                        [sys.executable, "-m", "pip", "install",
                         "--force-reinstall",
                         "torch", "torchvision", "torchaudio",
                         "--index-url", "https://download.pytorch.org/whl/cu124"],
                        stderr=subprocess.PIPE, text=True, timeout=900,
                    )
                    if result.returncode == 0:
                        self.after(0, lambda: self._sbert_torch_hint.set(
                            "✅ Установлено. Перезапустите приложение для активации GPU."))
                        self.after(0, lambda: btn_tc.configure(state="disabled"))
                    else:
                        err = (result.stderr or "неизвестная ошибка")[-400:]
                        self.after(0, lambda err=err: self._sbert_torch_hint.set(f"❌ {err}"))
                        self.after(0, lambda: btn_tc.configure(state="normal"))
                except subprocess.TimeoutExpired:
                    self.after(0, lambda: self._sbert_torch_hint.set(
                        "❌ Timeout. Установите вручную:\n"
                        "pip install --force-reinstall torch torchvision torchaudio "
                        "--index-url https://download.pytorch.org/whl/cu124"))
                    self.after(0, lambda: btn_tc.configure(state="normal"))
                except Exception as ex:
                    self.after(0, lambda ex=ex: self._sbert_torch_hint.set(f"❌ {ex}"))
                    self.after(0, lambda: btn_tc.configure(state="normal"))
            threading.Thread(target=_run, daemon=True).start()

        btn_tc = ttk.Button(srow_cuda, text="⚡ Установить torch+CUDA (cu124)",
                            command=_install_torch_cuda)
        btn_tc.pack(side="left", padx=(8, 0))
        Tooltip(btn_tc,
                "Устанавливает PyTorch с поддержкой CUDA 12.4 (RTX 30xx/40xx).\n"
                "Заменяет CPU-сборки torch/torchvision/torchaudio на CUDA-сборки.\n"
                "Это устраняет DLL-конфликты при смешанных версиях пакетов.\n\n"
                "Эквивалент команды:\n"
                "  pip install --force-reinstall\n"
                "      torch torchvision torchaudio\n"
                "      --index-url https://download.pytorch.org/whl/cu124\n\n"
                "После установки необходим перезапуск приложения.")

        def _check_cuda_status():
            import importlib.util as _ilu

            def _diag(msg: str):
                self.after(0, lambda m=msg: self.log_train(m))

            _diag("[SBERT ДИАГНОСТИКА] ==========================================")
            _diag(f"  Python : {sys.executable}")
            _st_spec = _ilu.find_spec("sentence_transformers")
            _tv_spec = _ilu.find_spec("torchvision")
            _diag(f"  sentence-transformers: {'установлен' if _st_spec else 'НЕ установлен'}")
            _diag(f"  torchvision          : {'установлен' if _tv_spec else 'НЕ установлен'}")
            if _st_spec:
                try:
                    import sentence_transformers as _st
                    _diag(f"  sentence-transformers import: OK  v{_st.__version__}")
                except Exception as _e:
                    _diag(f"  sentence-transformers import: ОШИБКА — {_e}")
            if _tv_spec:
                try:
                    import torchvision as _tv
                    _diag(f"  torchvision import          : OK  v{_tv.__version__}")
                except Exception as _e:
                    _diag(f"  torchvision import          : ОШИБКА — {_e}")
            try:
                import torch as _t
                _diag(f"  torch     : v{_t.__version__}  CUDA-сборка={_t.version.cuda}  "
                      f"is_available={_t.cuda.is_available()}")
            except Exception as _e:
                _diag(f"  torch     : ОШИБКА — {_e}")
            _diag("  Команды для ручного исправления:")
            _diag("    pip install --upgrade sentence-transformers")
            _diag("    pip install --force-reinstall torch torchvision torchaudio "
                  "--index-url https://download.pytorch.org/whl/cu124")
            _diag("[SBERT ДИАГНОСТИКА] ==========================================\n")

            cuda_ok = SBERTVectorizer.is_cuda_available()
            sbert_ok = SBERTVectorizer.is_available()
            sbert_installed = SBERTVectorizer.is_installed()

            if cuda_ok:
                try:
                    import torch
                    name = torch.cuda.get_device_name(0)
                    self.after(0, lambda: self._sbert_gpu_status.set(f"✅ CUDA: {name}"))
                except (RuntimeError, AttributeError) as _e:
                    _log.debug("CUDA device name probe: %s", _e)
                    self.after(0, lambda: self._sbert_gpu_status.set("✅ CUDA доступна"))

                if sbert_ok:
                    self.after(0, srow_cuda.pack_forget)
                elif sbert_installed:
                    self.after(0, lambda: self._sbert_torch_hint.set(
                        "⚠ sentence-transformers установлен, но не импортируется"
                        " (DLL-конфликт torchvision). Переустановите torch-стек:"))
                else:
                    self.after(0, lambda: self._sbert_torch_hint.set(
                        "sentence-transformers не установлен. Нажмите «Установить»"
                        " выше, затем переустановите torch-стек:"))
                return

            try:
                import torch
                cuda_ver = torch.version.cuda
                torch_ver = torch.__version__
                if cuda_ver is None:
                    msg = f"— torch {torch_ver} (CPU-сборка, нет поддержки CUDA)"
                    self.after(0, lambda msg=msg: self._sbert_gpu_status.set(msg))
                    self.after(0, lambda: self._sbert_torch_hint.set(
                        "Для GPU нужна CUDA-сборка PyTorch:"))
                else:
                    msg = (f"— CUDA недоступна  "
                           f"(torch {torch_ver}, скомпилирован под CUDA {cuda_ver},"
                           f" проверьте драйвер NVIDIA)")
                    self.after(0, lambda msg=msg: self._sbert_gpu_status.set(msg))
                    self.after(0, lambda: self._sbert_torch_hint.set(
                        "Переустановите torch-стек (или проверьте драйвер NVIDIA):"))
            except ImportError:
                self.after(0, lambda: self._sbert_gpu_status.set("— torch не установлен"))
                self.after(0, lambda: self._sbert_torch_hint.set(
                    "Установите torch с поддержкой CUDA:"))
            except Exception as _e:
                _es = str(_e).lower()
                if "triton" in _es:
                    msg = "— torch+CUDA: triton не установлен (нужен только для torch.compile)"
                elif "dll" in _es or "entry point" in _es or "0x" in _es:
                    msg = f"— DLL-конфликт torchvision (несовместимая версия): {_e}"
                else:
                    msg = f"— Ошибка torch: {_e}"
                self.after(0, lambda msg=msg: self._sbert_gpu_status.set(msg))
                self.after(0, lambda: self._sbert_torch_hint.set(
                    "Переустановите torch-стек для устранения конфликта:"))
        threading.Thread(target=_check_cuda_status, daemon=True).start()

        # --- строка 3: кэш-статус + кнопка скачать заранее ---
        srow2 = ttk.Frame(sbox); srow2.pack(fill="x", pady=(8, 0))
        self._sbert_cache_status = tk.StringVar(value="")
        ttk.Label(srow2, textvariable=self._sbert_cache_status, style="Card.Muted.TLabel").pack(side="left")

        def _download_sbert_now():

            model = self.sbert_model.get().strip() or SBERT_DEFAULT
            btn_dl.configure(state="disabled")

            def _run():
                if not SBERTVectorizer.is_available():
                    self.after(0, lambda: self._sbert_cache_status.set(
                        "Устанавливаю sentence-transformers…"
                    ))
                    self.after(0, lambda: self._sbert_status.set(
                        "Устанавливаю sentence-transformers…"
                    ))
                    try:
                        result = subprocess.run(
                            [sys.executable, "-m", "pip", "install", "sentence-transformers"],
                            timeout=300,
                        )
                        if result.returncode != 0:
                            self.after(0, lambda: self._sbert_cache_status.set(
                                f"❌ pip завершился с ошибкой (code={result.returncode})"
                            ))
                            self.after(0, lambda: btn_dl.configure(state="normal"))
                            return
                        self.after(0, lambda: self._sbert_status.set("✅ Установлено!"))
                    except subprocess.TimeoutExpired:
                        self.after(0, lambda: self._sbert_cache_status.set(
                            "❌ Установка зависла (timeout 5 мин). Попробуйте вручную."
                        ))
                        self.after(0, lambda: btn_dl.configure(state="normal"))
                        return
                    except Exception as ex:
                        self.after(0, lambda ex=ex: self._sbert_cache_status.set(
                            f"❌ Ошибка установки: {ex}"
                        ))
                        self.after(0, lambda: btn_dl.configure(state="normal"))
                        return

                self.after(0, lambda: self._sbert_cache_status.set(f"Скачиваю {model}…"))
                try:
                    vec = SBERTVectorizer(
                        model_name=model,
                        log_cb=lambda msg: self.after(0, lambda m=msg: self.log_train(m)),
                        progress_cb=None,
                    )
                    cached = vec._check_cache()
                    if cached:
                        self.after(0, lambda: self._sbert_cache_status.set(
                            f"✅ {model} — уже в кэше"
                        ))
                    else:
                        vec._download_model()
                        self.after(0, lambda: self._sbert_cache_status.set(
                            f"✅ {model} — скачано"
                        ))
                except Exception as ex:
                    self.after(0, lambda ex=ex: self._sbert_cache_status.set(
                        f"❌ Ошибка скачивания: {ex}"
                    ))
                finally:
                    self.after(0, lambda: btn_dl.configure(state="normal"))

            threading.Thread(target=_run, daemon=True).start()

        btn_dl = ttk.Button(srow2, text="Скачать модель заранее ↓", command=_download_sbert_now)
        btn_dl.pack(side="left", padx=(12, 0))
        self.attach_help(btn_dl, "Скачать SBERT модель",
                         "Скачивает выбранную модель в папку ./sbert_models/ рядом с приложением.\n\n"
                         "При обучении прогресс отображается в лог-окне внизу.\n"
                         "Можно скачать заранее, чтобы не ждать во время обучения.",
                         "Предварительное скачивание модели")

        # авто-обновить статус при первом рендере
        def _check_sbert_installed():
            if SBERTVectorizer.is_available():
                self._sbert_status.set("✅ sentence-transformers доступен")
                self.after(0, btn_inst.pack_forget)
            else:
                self._sbert_status.set("⚠️ sentence-transformers не установлен")
            def _check_cache():
                if SBERTVectorizer.is_available():
                    model = self.sbert_model.get().strip() or SBERT_DEFAULT
                    vec = SBERTVectorizer(model_name=model)
                    cached = vec._check_cache()
                    msg = f"✅ {model}: в кэше" if cached else f"⬇ {model}: требует скачивания"
                    self.after(0, lambda: self._sbert_cache_status.set(msg))
            threading.Thread(target=_check_cache, daemon=True).start()
        self.after(200, _check_sbert_installed)

        # обновлять кэш-статус при смене модели
        def _on_model_change(*_):
            if not SBERTVectorizer.is_available():
                return
            model = self.sbert_model.get().strip() or SBERT_DEFAULT
            self._sbert_cache_status.set(f"Проверяю кэш для {model}…")
            def _chk():
                vec = SBERTVectorizer(model_name=model)
                cached = vec._check_cache()
                msg = f"✅ {model}: в кэше" if cached else f"⬇ {model}: требует скачивания"
                self.after(0, lambda: self._sbert_cache_status.set(msg))
            threading.Thread(target=_chk, daemon=True).start()
        self.sbert_model.trace_add("write", _on_model_change)

        # ---- SetFit нейросетевой классификатор ----
        sfbox = ttk.LabelFrame(
            parent,
            text="Шаг 5 — SetFit нейросетевой классификатор (опционально, заменяет LinearSVC)",
            padding=12,
        )
        sfbox.pack(fill="x", pady=(0, 10))

        sfrow0 = ttk.Frame(sfbox); sfrow0.pack(fill="x")
        cb_sf = ttk.Checkbutton(
            sfrow0, text="Использовать SetFit вместо LinearSVC",
            variable=self.use_setfit,
        )
        cb_sf.pack(side="left")
        self.attach_help(
            cb_sf,
            "SetFit нейросетевой классификатор",
            "SetFit (HuggingFace) — нейросетевой классификатор few-shot:\n"
            "  Обучает контрастную голову поверх SBERT-энкодера.\n\n"
            "ПРЕИМУЩЕСТВА:\n"
            "  + Ожидаемый прирост macro F1: +4–10% vs LinearSVC+TF-IDF\n"
            "  + Особенно эффективен для few-shot классов (50–300 примеров)\n"
            "  + Понимает семантику: синонимы, перефразировки\n"
            "  + Работает с теми же данными без изменений\n\n"
            "ТРЕБОВАНИЯ:\n"
            "  • pip install setfit>=0.9\n"
            "  • GPU (NVIDIA RTX) значительно ускоряет обучение\n"
            "  • torch CUDA: pip install torch --index-url .../whl/cu124\n\n"
            "РЕКОМЕНДАЦИИ ПО ЖЕЛЕЗУ:\n"
            "  • RTX 4070 Ti 12 ГБ: USER2-large, batch=32\n"
            "  • RTX 4060 Ti 8 ГБ:  USER2-base, batch=16\n\n"
            "⚠ При включении SetFit опции LinearSVC (C, max_iter) игнорируются.\n"
            "  TF-IDF параметры тоже не используются — SetFit работает напрямую\n"
            "  с текстом через SBERT-энкодер.",
            "SetFit — few-shot нейросетевой классификатор",
        )

        # Статус установки setfit
        self._setfit_status = tk.StringVar(value="")
        ttk.Label(sfrow0, textvariable=self._setfit_status,
                  style="Card.Muted.TLabel").pack(side="left", padx=12)

        def _check_setfit_installed():
            import importlib.util as _ilu
            ok = _ilu.find_spec("setfit") is not None
            msg = "✅ setfit установлен" if ok else "⚠ setfit не установлен: pip install setfit>=0.9"
            self.after(0, lambda: self._setfit_status.set(msg))
        self.after(300, _check_setfit_installed)

        # --- Строка 1: выбор модели ---
        sfrow1 = ttk.Frame(sfbox); sfrow1.pack(fill="x", pady=(8, 0))
        ttk.Label(sfrow1, text="Модель:").pack(side="left")
        cb_sfm = ttk.Combobox(
            sfrow1, textvariable=self.setfit_model,
            state="readonly", width=48,
            values=SETFIT_MODELS_LIST,
        )
        cb_sfm.pack(side="left", padx=8)
        self.cb_setfit_model_combo = cb_sfm  # защита от перезаписи в _refresh_combobox_values
        self.attach_help(
            cb_sfm, "SetFit модель (backbone)",
            "\n".join(f"• {k}:\n  {v}" for k, v in SETFIT_MODELS.items()),
            "Выбор SBERT-энкодера для SetFit",
        )

        # Рекомендация по VRAM
        _vram = self._hw.gpu_vram_gb
        if _vram is not None and _vram >= 12:
            _sf_rec = f"Рек. для вашего GPU ({_vram:.0f} ГБ): USER2-large, batch={self._hw.setfit_batch}"
        elif _vram is not None and _vram >= 8:
            _sf_rec = f"Рек. для вашего GPU ({_vram:.0f} ГБ): USER2-base, batch={self._hw.setfit_batch}"
        elif _vram is not None:
            _sf_rec = f"GPU {_vram:.0f} ГБ VRAM — рек. USER-base или rubert-tiny2"
        else:
            _sf_rec = "GPU не обнаружен — рек. rubert-tiny2 (CPU)"
        ttk.Label(sfrow1, text=_sf_rec, style="Card.Muted.TLabel").pack(side="left", padx=12)

        # --- Строка 2: параметры обучения ---
        sfrow2 = ttk.Frame(sfbox); sfrow2.pack(fill="x", pady=(8, 0))
        ttk.Label(sfrow2, text="Эпох:").pack(side="left")
        spn_sf_ep = ttk.Spinbox(
            sfrow2, from_=1, to=20, increment=1,
            textvariable=self.setfit_epochs, width=5,
        )
        spn_sf_ep.pack(side="left", padx=(4, 12))
        Tooltip(spn_sf_ep,
                "Количество эпох обучения классификационной головы SetFit.\n"
                "Рекомендуется: 1–5 эпох (больше = риск переобучения на малых данных).")

        ttk.Label(sfrow2, text="Итераций пар:").pack(side="left")
        spn_sf_iter = ttk.Spinbox(
            sfrow2, from_=5, to=100, increment=5,
            textvariable=self.setfit_num_iterations, width=5,
        )
        spn_sf_iter.pack(side="left", padx=(4, 12))
        Tooltip(spn_sf_iter,
                "Количество контрастных пар на класс при обучении энкодера.\n"
                "Больше = лучше качество, но дольше обучение.\n"
                "Рекомендуется: 20 (дефолт). Для больших датасетов: 10–15.")

        ttk.Label(sfrow2, text="Батч:").pack(side="left")
        spn_sf_b = ttk.Spinbox(
            sfrow2, from_=1, to=256, increment=4,
            textvariable=self.setfit_batch, width=5,
        )
        spn_sf_b.pack(side="left", padx=(4, 12))
        ttk.Label(
            sfrow2, text=f"(рек. {self._hw.setfit_batch})",
            style="Card.Muted.TLabel",
        ).pack(side="left")
        Tooltip(spn_sf_b,
                f"Размер батча контрастного обучения SetFit.\n"
                f"Рекомендуется для вашего GPU: {self._hw.setfit_batch}\n\n"
                "При CUDA OOM — уменьшите значение вдвое.")

        # --- Строка 3: fp16 ---
        sfrow3 = ttk.Frame(sfbox); sfrow3.pack(fill="x", pady=(6, 0))
        cb_fp16 = ttk.Checkbutton(
            sfrow3, text="fp16 (mixed precision, только GPU)",
            variable=self.setfit_fp16,
        )
        cb_fp16.pack(side="left")
        Tooltip(cb_fp16,
                "Включает смешанную точность (fp16) при обучении SetFit.\n"
                "Уменьшает потребление VRAM и ускоряет обучение на GPU.\n"
                "Автоматически отключается на CPU.")


    # -------------------------------------------------------- tab: Обучение
    def _build_train_tab(self):
        # ── Sub-tab frames ────────────────────────────────────────────────────
        _s0 = ttk.Frame(self.tab_train)   # Данные
        _s1 = ttk.Frame(self.tab_train)   # Параметры
        _s2 = ttk.Frame(self.tab_train)   # Запуск

        self._build_train_files_card(_s0)
        self._build_train_analysis_card(_s0)
        self._build_train_mode_card(_s0)

        cols = ttk.LabelFrame(_s0, text="Шаг 1 — выбрать колонки", padding=12)
        cols.pack(fill="x", pady=(0, 10))
        self._combobox(cols, 0, "Описание:", self.desc_col, "Описание обращения")
        self._combobox(cols, 1, "Текст звонка:", self.call_col, "Транскрипт звонка")
        self._combobox(cols, 2, "Текст чата:", self.chat_col, "Транскрипт чата")
        self._combobox(cols, 3, "Суммаризация:", self.summary_col, "Выжимка (если есть)")
        self._combobox(cols, 4, "Ответ краткий:", self.ans_short_col, "Короткий ответ оператора (если есть)")
        self._combobox(cols, 5, "Ответ полный:", self.ans_full_col, "Развёрнутый ответ / HTML (если есть)")
        self._combobox(cols, 6, "Label / причина:", self.label_col, "Ручная разметка причины")

        flags = ttk.Frame(cols)
        flags.grid(row=7, column=1, sticky="w", padx=10, pady=(8, 0))
        ttk.Checkbutton(flags, text="Использовать суммаризацию", variable=self.use_summary).pack(side="left")
        ttk.Checkbutton(flags, text="Игнорировать CHATBOT", variable=self.ignore_chatbot).pack(side="left", padx=14)

        ap = ttk.Frame(cols)
        ap.grid(row=8, column=1, sticky="w", padx=10, pady=(8, 0))
        ttk.Label(ap, text="Авто-профиль строки:").pack(side="left")
        cb_ap = ttk.Combobox(ap, textvariable=self.auto_profile, state="readonly", width=12, values=["off", "smart", "strict"])
        cb_ap.pack(side="left", padx=8)
        self.cb_auto_profile = cb_ap  # защита от перезаписи в _refresh_combobox_values
        self.attach_help(cb_ap, "Auto profile per row",
                         "Автологика весов для каждой строки по заполненности и наличию разметки ролей.\n"
                         "- off: только твои веса\n- smart (реком.): усиливает лучший профиль\n- strict: жёстко переключает профили",
                         "Реком.: smart")

        wbox = ttk.LabelFrame(_s0, text="Шаг 2 — веса", padding=12)
        wbox.pack(fill="x", pady=(0, 10))

        # ── Группа 1: универсальные пресеты ─────────────────────────────────
        pres = ttk.Frame(wbox); pres.pack(fill="x", pady=(0, 4))
        ttk.Label(pres, text="Универсальные:").pack(side="left")
        btn_balanced = ttk.Button(pres, text="Баланс", command=lambda: self.apply_weight_preset("balanced"))
        btn_balanced.pack(side="left", padx=6)
        Tooltip(btn_balanced, "Пресет «Баланс»:\nВсе источники равномерно — desc и summary как главные сигналы, клиент усилен.\ndesc=3, client=2, operator=1, summary=3, ans_short=2, ans_full=1")
        btn_client = ttk.Button(pres, text="Клиент", command=lambda: self.apply_weight_preset("client"))
        btn_client.pack(side="left", padx=6)
        Tooltip(btn_client, "Пресет «Клиент»:\nДиалог без summary — клиент и desc основа, оператор фон.\ndesc=3, client=3, operator=1, summary=2, ans_short=1, ans_full=0")
        btn_summary = ttk.Button(pres, text="Суммаризация", command=lambda: self.apply_weight_preset("summary"))
        btn_summary.pack(side="left", padx=6)
        Tooltip(btn_summary, "Пресет «Суммаризация»:\nНет живого диалога — summary + desc + ответ как главные источники.\ndesc=3, client=1, operator=0, summary=3, ans_short=2, ans_full=1")
        btn_answers = ttk.Button(pres, text="Ответы", command=lambda: self.apply_weight_preset("answers"))
        btn_answers.pack(side="left", padx=6)
        Tooltip(btn_answers, "Пресет «Ответы»:\nНет диалога и summary — опора на ответы банка и desc.\ndesc=2, client=1, operator=0, summary=2, ans_short=3, ans_full=2")

        # ── Группа 2: пресеты для жалоб ─────────────────────────────────────
        pres_complaints = ttk.Frame(wbox); pres_complaints.pack(fill="x", pady=(0, 4))
        ttk.Label(pres_complaints, text="Жалобы:").pack(side="left")
        btn_complaints = ttk.Button(
            pres_complaints, text="★ Жалобы (полный)",
            command=lambda: self.apply_weight_preset("complaints"),
            style="Accent.TButton",
        )
        btn_complaints.pack(side="left", padx=6)
        Tooltip(
            btn_complaints,
            "Пресет «Жалобы (полный)» — оптимален для файлов жалоб с\n"
            "полным набором колонок: Описание + Диалог + Суммаризация + Ответы банка.\n\n"
            "Веса:\n"
            "  desc=4    (формулировка жалобы — ключевой сигнал)\n"
            "  summary=4 (сжатый итог обращения — ключевой сигнал)\n"
            "  ans_short=3 (ответ банка отражает категорию)\n"
            "  ans_full=2 (развёрнутый ответ — дополнительный контекст)\n"
            "  client=2  (контекст жалобы)\n"
            "  operator=1 (фоновый)\n\n"
            "Параметры TF-IDF / LinearSVC:\n"
            "  char n-gram 2–9, word n-gram 1–3\n"
            "  min_df=2, C=3.0, SVD=250\n"
            "  SMOTE вкл., лемматизация вкл.",
        )
        btn_complaints_no_ans = ttk.Button(
            pres_complaints, text="★ Жалобы (без ответов)",
            command=lambda: self.apply_weight_preset("complaints_no_answers"),
            style="Accent.TButton",
        )
        btn_complaints_no_ans.pack(side="left", padx=6)
        Tooltip(
            btn_complaints_no_ans,
            "Пресет «Жалобы (без ответов)» — для файлов жалоб без ответов банка:\n"
            "Описание + Диалог + Суммаризация (поля ответов не заполнены).\n\n"
            "Веса:\n"
            "  desc=4    (формулировка жалобы — ключевой сигнал)\n"
            "  summary=4 (сжатый итог — компенсирует отсутствие ответов)\n"
            "  client=3  (слова клиента усилены — дополнительный источник)\n"
            "  operator=1 (фоновый)\n"
            "  ans_short=0, ans_full=0 (отключены)\n\n"
            "Параметры TF-IDF / LinearSVC:\n"
            "  char n-gram 2–9, word n-gram 1–3\n"
            "  min_df=2, C=2.0, SVD=220\n"
            "  SMOTE вкл., лемматизация вкл.",
        )

        # ── Группа 3: пресет для консультаций ───────────────────────────────
        pres2 = ttk.Frame(wbox); pres2.pack(fill="x", pady=(0, 4))
        ttk.Label(pres2, text="Консультации:").pack(side="left")
        btn_consult = ttk.Button(
            pres2, text="★ Консультации",
            command=lambda: self.apply_weight_preset("consultation"),
            style="Accent.TButton",
        )
        btn_consult.pack(side="left", padx=6)
        Tooltip(
            btn_consult,
            "Пресет «Консультации» — максимальная точность для файлов с\n"
            "звонками/чатами и суммаризацией (без desc и ответов банка).\n\n"
            "Веса:\n"
            "  summary=5 (главный сигнал, 98% заполнен)\n"
            "  client=3  (слова клиента — суть обращения)\n"
            "  operator=1 (фоновый)\n"
            "  desc=0, ans_short=0, ans_full=0\n\n"
            "Параметры TF-IDF / LinearSVC:\n"
            "  char n-gram 2–8, word n-gram 1–2\n"
            "  min_df=2, C=5.0, SVD=200\n"
            "  SMOTE вкл., лемматизация вкл.",
        )

        # ── Группа 4: оптимизированные пресеты под конкретные файлы ─────────
        pres3 = ttk.Frame(wbox); pres3.pack(fill="x", pady=(0, 8))
        ttk.Label(pres3, text="Оптимальные:").pack(side="left")

        btn_fullset = ttk.Button(
            pres3, text="⭐ Жалобы (полный набор, 30+ классов)",
            command=lambda: self.apply_weight_preset("complaints_fullset"),
            style="Accent.TButton",
        )
        btn_fullset.pack(side="left", padx=6)
        Tooltip(
            btn_fullset,
            "Пресет «Жалобы — полный набор, 30+ классов»\n\n"
            "Когда использовать:\n"
            "  • Файл содержит все 6 колонок:\n"
            "    Описание + Диалог + Суммаризация + Ответы банка\n"
            "  • Классов 30 и больше\n"
            "  • Возможен сильный дисбаланс между классами\n\n"
            "Веса:\n"
            "  desc=4    (краткая точная формулировка — сильный сигнал)\n"
            "  summary=4 (концентрат обращения — ключевой источник)\n"
            "  ans_short=3 (ответ банка маркирует тип проблемы)\n"
            "  ans_full=2  (развёрнутый ответ — дополнительный контекст)\n"
            "  client=2  (контекст)\n"
            "  operator=1 (фоновый)\n\n"
            "Параметры TF-IDF / LinearSVC:\n"
            "  char n-gram 2–9, word n-gram 1–3\n"
            "  min_df=2  (сохраняет редкие значимые термины)\n"
            "  C=2.5     (мягче — малочисленные классы, риск переобучения)\n"
            "  SVD=300   (30+ классов требуют широкого латентного пространства)\n"
            "  SMOTE вкл., лемматизация вкл.",
        )

        btn_dialog = ttk.Button(
            pres3, text="⭐ Консультации (диалоги+summary, 5–20 классов)",
            command=lambda: self.apply_weight_preset("consultation_dialog"),
            style="Accent.TButton",
        )
        btn_dialog.pack(side="left", padx=6)
        Tooltip(
            btn_dialog,
            "Пресет «Консультации — диалоги + суммаризация, 5–20 классов»\n\n"
            "Когда использовать:\n"
            "  • Файл содержит только Звонки/Чаты + Суммаризация\n"
            "  • Описание и Ответы банка отсутствуют / не заполнены\n"
            "  • Классов 5–20, диалоги длинные (>2 000 символов)\n\n"
            "Веса:\n"
            "  summary=5 (краткая выжимка диалога — главный сигнал)\n"
            "  client=3  (суть запроса от клиента)\n"
            "  operator=1 (фоновый)\n"
            "  desc=0, ans_short=0, ans_full=0 (отключены)\n\n"
            "Параметры TF-IDF / LinearSVC:\n"
            "  char n-gram 2–8, word n-gram 1–2\n"
            "  min_df=3  (убирает шум длинных диалогов при 1000+ строках)\n"
            "  C=6.0     (строже — длинные диалоги порождают шумный словарь)\n"
            "  SVD=150   (5–20 классов — меньшего пространства достаточно)\n"
            "  SMOTE вкл., лемматизация вкл.",
        )

        sliders = ttk.Frame(wbox); sliders.pack(fill="x")
        self._weight_slider(sliders, 0, "Описание", self.w_desc)
        self._weight_slider(sliders, 1, "Клиент", self.w_client)
        self._weight_slider(sliders, 2, "Оператор", self.w_operator)
        self._weight_slider(sliders, 3, "Суммаризация", self.w_summary)
        self._weight_slider(sliders, 4, "Ответ краткий", self.w_ans_short)
        self._weight_slider(sliders, 5, "Ответ полный", self.w_ans_full)

        self._build_train_params_card(_s1)
        self._build_sbert_install_card(_s1)
        self._build_readiness_bar(_s2, [
            lambda: (
                bool(self.train_files),
                f"📄 Файлы: {len(self.train_files)} загружено" if self.train_files
                else "📄 Файлы: не выбраны",
            ),
            lambda: (
                bool(self.label_col.get()),
                f"🏷 Метки: «{self.label_col.get()}»" if self.label_col.get()
                else "🏷 Метки: колонка не указана",
            ),
        ])
        self.btn_train, self.btn_train_stop = self._build_action_block(
            _s2,
            btn_text="▶  Обучить / сохранить модель",
            btn_cmd=self.run_training,
            progress_var=self.train_progress,
            pct_var=self.train_pct,
            phase_var=self.train_phase,
            speed_var=self.train_speed,
            eta_var=self.train_eta,
            label="Прогресс обучения",
        )
        self.log_train(f"Автосохранение:\n- {MODEL_DIR}\n- {CLASS_DIR}\n- {CLUST_DIR}\n")

        # ── Advanced ML Tools ─────────────────────────────────────────────────
        _adv_lf = ttk.LabelFrame(_s2, text="Продвинутые ML-инструменты", padding=(10, 6))
        _adv_lf.pack(fill="x", pady=(12, 0))
        ttk.Label(_adv_lf,
                  text="Дополнительные инструменты для улучшения качества модели. "
                       "Требуют наличия обученной модели или размеченного датасета.",
                  wraplength=700, justify="left").pack(anchor="w", pady=(0, 8))

        _adv_row1 = ttk.Frame(_adv_lf); _adv_row1.pack(fill="x", pady=(0, 4))

        _btn_mlm = ttk.Button(_adv_row1, text="🔤 MLM Pretrain",
                              command=self._run_mlm_pretrain_dialog)
        _btn_mlm.pack(side="left", padx=(0, 8))
        Tooltip(_btn_mlm,
                "Дообучить SBERT-модель на доменных текстах (Masked Language Modeling).\n"
                "Улучшает качество векторизации для банковской тематики.\n"
                "Требует: transformers, torch, datasets.\n"
                "Время: 10–60 мин в зависимости от объёма данных.")

        _btn_distill = ttk.Button(_adv_row1, text="🎓 Дистилляция",
                                  command=self._run_distillation_dialog)
        _btn_distill.pack(side="left", padx=(0, 8))
        Tooltip(_btn_distill,
                "Передать знания большой модели (учителя) в компактную (ученика).\n"
                "Например: SetFit → TF-IDF LinearSVC.\n"
                "Результат: скорость компактной модели при точности близкой к большой.")

        _btn_al = ttk.Button(_adv_row1, text="🎯 Active Learning",
                             command=self._run_active_learning_dialog)
        _btn_al.pack(side="left", padx=(0, 8))
        Tooltip(_btn_al,
                "Найти строки, которые модель классифицирует с наименьшей уверенностью.\n"
                "Разметка именно этих строк даст максимальный прирост качества модели.\n"
                "Результат: Excel-файл с приоритетами для разметки.")

        # ── Register bottom sub-tab strip for Training tab ────────────────────
        self._register_sub_tabs(
            0,
            ["Данные", "Параметры", "Запуск"],
            [_s0, _s1, _s2],
        )

    # ----------------------------------------------------------- log
    def log_train(self, msg: str):
        self._log_to(self.train_log, msg)

    # --------------------------------------------------------- weight presets
    def apply_weight_preset(self, name: str):
        pw = PRESET_WEIGHTS[name]
        self.w_desc.set(pw["w_desc"])
        self.w_client.set(pw["w_client"])
        self.w_operator.set(pw["w_operator"])
        self.w_summary.set(pw["w_summary"])
        self.w_ans_short.set(pw["w_answer_short"])
        self.w_ans_full.set(pw["w_answer_full"])

        # Применяем параметры алгоритма, если пресет их задаёт
        ap = PRESET_ALGO_PARAMS.get(name)
        if ap:
            if "char_ng_min"    in ap: self.char_ng_min.set(ap["char_ng_min"])
            if "char_ng_max"    in ap: self.char_ng_max.set(ap["char_ng_max"])
            if "word_ng_min"    in ap: self.word_ng_min.set(ap["word_ng_min"])
            if "word_ng_max"    in ap: self.word_ng_max.set(ap["word_ng_max"])
            if "min_df"         in ap: self.min_df.set(ap["min_df"])
            if "C"              in ap: self.C.set(ap["C"])
            if "use_svd"        in ap: self.use_svd.set(ap["use_svd"])
            if "svd_components" in ap: self.svd_components.set(ap["svd_components"])
            if "use_smote"      in ap: self.use_smote.set(ap["use_smote"])
            if "sublinear_tf"   in ap: self.sublinear_tf.set(ap["sublinear_tf"])
            if "use_lemma"      in ap: self.use_lemma.set(ap["use_lemma"])

        self.set_help(f"Пресет: {name}", PRESET_WEIGHTS_DESC.get(name, ""))

    # ----------------------------------------------------- file pickers: train
    def add_train_files(self):
        paths = filedialog.askopenfilenames(
            title="Файлы для обучения",
            filetypes=[
                ("Таблицы (xlsx, csv)", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("Все файлы", "*.*"),
            ],
        )
        if not paths:
            return
        for p in paths:
            if p not in self.train_files:
                self.train_files.append(p)
                self.lb_train.insert("end", p)
                self._add_recent("train_files", p)
        try:
            headers = read_headers(Path(self.train_files[0]))
            self._refresh_combobox_values(headers)
            self.log_train(f"Файлов обучения: {len(self.train_files)} | Заголовков: {len(headers)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def add_train_folder(self):
        """Добавляет все .xlsx/.xlsm/.csv-файлы из выбранной папки (рекурсивно)."""
        d = filedialog.askdirectory(title="Папка с файлами обучения (xlsx/csv)")
        if not d:
            return
        exts = {".xlsx", ".xlsm", ".csv"}
        added = 0
        for p in sorted(Path(d).rglob("*")):
            if p.suffix.lower() in exts:
                sp = str(p)
                if sp not in self.train_files:
                    self.train_files.append(sp)
                    self.lb_train.insert("end", sp)
                    self._add_recent("train_files", sp)
                    added += 1
        if not added:
            messagebox.showinfo("Папка пуста",
                                f"В папке не найдено xlsx/xlsm/csv файлов:\n{d}")
            return
        try:
            headers = read_headers(Path(self.train_files[0]))
            self._refresh_combobox_values(headers)
            self.log_train(
                f"Папка: +{added} файл(ов) | всего: {len(self.train_files)} | "
                f"заголовков: {len(headers)}"
            )
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _add_train_file_from_path(self, p: str) -> None:
        """Добавляет файл обучения из недавних без диалога."""
        if p in self.train_files:
            return
        if not Path(p).exists():
            messagebox.showwarning("Файл не найден", f"Файл не найден:\n{p}")
            return
        self.train_files.append(p)
        self.lb_train.insert("end", p)
        try:
            from excel_utils import read_headers
            headers = read_headers(Path(p))
            self._refresh_combobox_values(headers)
            self.log_train(f"Добавлен: {Path(p).name} | Заголовков: {len(headers)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def remove_train_file(self):
        sel = list(self.lb_train.curselection())
        if not sel:
            return
        i = sel[0]
        self.lb_train.delete(i)
        del self.train_files[i]
        self.log_train(f"Файлов обучения: {len(self.train_files)}")

    def clear_train_files(self):
        self.train_files = []
        self.lb_train.delete(0, "end")
        self.log_train("Файлы обучения очищены.")

    def show_dataset_stats(self):
        """Читает файлы обучения и выводит статистику датасета в лог."""
        if not self.train_files:
            messagebox.showwarning("Статистика датасета", "Добавь файлы обучения сначала.")
            return

        label_col = self.label_col.get().strip()
        if not label_col:
            messagebox.showwarning("Статистика датасета", "Укажи колонку метки класса.")
            return

        self.log_train("\n==== СТАТИСТИКА ДАТАСЕТА ====")

        def _worker():
            from collections import Counter as _Counter
            total_rows = 0
            label_counts: _Counter = _Counter()
            empty_text = 0
            file_rows = {}

            for fpath in self.train_files:
                p = Path(fpath)
                file_total = 0
                try:
                    with open_tabular(p) as it:
                        headers = [str(h) if h is not None else "" for h in next(it)]
                        lbl_idx = idx_of(headers, label_col)
                        if lbl_idx is None:
                            self.after(0, lambda f=p.name: self.log_train(
                                f"  ⚠ {f}: колонка «{label_col}» не найдена, пропускаем"
                            ))
                            continue
                        for row in it:
                            file_total += 1
                            lbl = str(row[lbl_idx] or "").strip() if lbl_idx < len(row) else ""
                            if lbl:
                                label_counts[lbl] += 1
                            else:
                                empty_text += 1
                except Exception as e:
                    self.after(0, lambda f=p.name, e=e: self.log_train(f"  ⚠ {f}: {e}"))
                    continue
                file_rows[p.name] = file_total
                total_rows += file_total

            def _show():
                lines = []
                lines.append(f"Файлов: {len(file_rows)}")
                for fname, cnt in file_rows.items():
                    lines.append(f"  {fname}: {cnt} строк")
                lines.append(f"Строк всего: {total_rows}")
                lines.append(f"Без метки: {empty_text}")
                lines.append(f"Классов: {len(label_counts)}")
                if label_counts:
                    most = label_counts.most_common(5)
                    least = label_counts.most_common()[:-6:-1]
                    total_lbl = sum(label_counts.values())
                    max_cnt = most[0][1]
                    min_cnt = min(label_counts.values())
                    imbalance = round(max_cnt / max(1, min_cnt), 1)
                    lines.append(f"Соотношение макс/мин классов: {imbalance}x")
                    lines.append("Топ-5 крупных:")
                    for lbl, cnt in most:
                        lines.append(f"  {lbl}: {cnt} ({cnt/max(1,total_lbl)*100:.1f}%)")
                    lines.append("Топ-5 мелких:")
                    for lbl, cnt in sorted(least, key=lambda x: x[1]):
                        lines.append(f"  {lbl}: {cnt} ({cnt/max(1,total_lbl)*100:.1f}%)")
                self.log_train("\n".join(lines))
                self.log_train("==== КОНЕЦ СТАТИСТИКИ ====")

            self.after(0, _show)

        threading.Thread(target=_worker, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────────────
    # Анализ датасета и авто-настройка параметров
    # ─────────────────────────────────────────────────────────────────────────

    def run_dataset_analysis(self) -> None:
        """
        Запускает предварительный анализ обучающего датасета в фоновом потоке.
        По завершении открывает диалог с результатами и кнопкой «Применить».
        """
        if not self.train_files:
            messagebox.showwarning("Анализ датасета",
                                   "Добавьте файлы обучения перед анализом.")
            return
        label_col = self.label_col.get().strip()
        if not label_col:
            messagebox.showwarning("Анализ датасета",
                                   "Укажите колонку с метками классов.")
            return

        # Блокируем кнопку
        if hasattr(self, "_btn_analyze"):
            self._btn_analyze.configure(state="disabled",
                                        text="⏳ Анализ…")
        self.log_train("==== АНАЛИЗ ДАТАСЕТА ====")

        def _worker():
            try:
                from collections import Counter as _C

                # ── Конфигурация колонок из текущих UI-значений ───────────────
                desc_col    = self.desc_col.get().strip()
                call_col    = self.call_col.get().strip()
                chat_col    = self.chat_col.get().strip()
                summary_col = self.summary_col.get().strip() if hasattr(self, "summary_col") else ""
                as_col      = self.ans_short_col.get().strip() if hasattr(self, "ans_short_col") else ""
                af_col      = self.ans_full_col.get().strip() if hasattr(self, "ans_full_col") else ""

                text_fields = {
                    k: v for k, v in {
                        "Описание":  desc_col,
                        "Звонок":    call_col,
                        "Чат":       chat_col,
                        "Суммаризация": summary_col,
                        "Ответ кр.": as_col,
                        "Ответ полн.": af_col,
                    }.items() if v
                }

                X: List[str] = []
                y: List[str] = []
                field_hits   = {f: 0 for f in text_fields}
                total_rows   = 0

                for fpath in self.train_files:
                    p = Path(fpath)
                    try:
                        with open_tabular(p) as it:
                            headers = [str(h) if h is not None else "" for h in next(it)]
                            lbl_idx = idx_of(headers, label_col)
                            col_idx = {fname: idx_of(headers, col)
                                       for fname, col in text_fields.items()}
                            for row in it:
                                rv = list(row)
                                total_rows += 1
                                # метка
                                lbl = ""
                                if lbl_idx is not None and lbl_idx < len(rv):
                                    lbl = str(rv[lbl_idx] or "").strip()
                                if not lbl:
                                    continue
                                y.append(lbl)
                                # текст для длин
                                parts = []
                                for fname, ci in col_idx.items():
                                    if ci is not None and ci < len(rv):
                                        val = str(rv[ci] or "").strip()
                                        if val:
                                            parts.append(val)
                                            field_hits[fname] = field_hits.get(fname, 0) + 1
                                X.append(" ".join(parts)[:2000])
                    except Exception as _e:
                        self.after(0, lambda e=_e, f=p.name: self.log_train(
                            f"  ⚠ {f}: {e}"))

                n_labeled = len(y)
                field_coverage = {
                    f: (field_hits[f] / max(n_labeled, 1))
                    for f in text_fields
                } if n_labeled else {}

                if n_labeled == 0:
                    self.after(0, lambda: messagebox.showwarning(
                        "Анализ датасета",
                        "Не удалось прочитать данные. Проверьте колонку меток и файлы."))
                    return

                # ── Анализ ────────────────────────────────────────────────────
                result = analyze_dataset(X, y, field_coverage)

                # ── Текущие значения параметров ───────────────────────────────
                current_values = {
                    "max_features":          int(self.max_features.get())
                                             if hasattr(self, "max_features") else None,
                    "C":                     float(self.C.get())
                                             if hasattr(self, "C") else None,
                    "use_svd":               bool(self.use_svd.get())
                                             if hasattr(self, "use_svd") else None,
                    "svd_components":        int(self.svd_components.get())
                                             if hasattr(self, "svd_components") else None,
                    "char_ng_min":           int(self.char_ng_min.get())
                                             if hasattr(self, "char_ng_min") else None,
                    "char_ng_max":           int(self.char_ng_max.get())
                                             if hasattr(self, "char_ng_max") else None,
                    "word_ng_min":           int(self.word_ng_min.get())
                                             if hasattr(self, "word_ng_min") else None,
                    "word_ng_max":           int(self.word_ng_max.get())
                                             if hasattr(self, "word_ng_max") else None,
                    "balanced":              bool(self.class_weight_balanced.get())
                                             if hasattr(self, "class_weight_balanced") else None,
                    "use_smote":             bool(self.use_smote.get())
                                             if hasattr(self, "use_smote") else None,
                    "use_hierarchical":      bool(self.use_hierarchical.get())
                                             if hasattr(self, "use_hierarchical") else None,
                    "use_hard_negatives":    bool(self.use_hard_negatives.get())
                                             if hasattr(self, "use_hard_negatives") else None,
                    "use_field_dropout":     bool(self.use_field_dropout.get())
                                             if hasattr(self, "use_field_dropout") else None,
                    "use_kfold_ensemble":    bool(self.use_kfold_ensemble.get())
                                             if hasattr(self, "use_kfold_ensemble") else None,
                    "use_confident_learning": bool(self.use_confident_learning.get())
                                             if hasattr(self, "use_confident_learning") else None,
                }
                result["current_values"] = current_values

                # ── Лог-строка ────────────────────────────────────────────────
                stats = result["stats"]
                n_issues = len(result["issues"])
                self.after(0, lambda s=stats, ni=n_issues: self.log_train(
                    f"  Примеров: {s['n_samples']:,} | "
                    f"Классов: {s['n_classes']} | "
                    f"Дисбаланс: {s['imbalance_ratio']}x | "
                    f"Проблем: {ni}"
                ))
                self.after(0, lambda: self.log_train("==== АНАЛИЗ ЗАВЕРШЁН ===="))

                # ── Открыть диалог ────────────────────────────────────────────
                self.after(0, lambda r=result: self._open_analysis_dialog(r))

            except Exception as _e:
                self.after(0, lambda e=_e: (
                    self.log_train(f"  ⚠ Ошибка анализа: {e}"),
                    messagebox.showerror("Анализ датасета", f"Ошибка: {e}")))
            finally:
                self.after(0, self._reset_analyze_btn)

        threading.Thread(target=_worker, daemon=True).start()

    def _open_analysis_dialog(self, result: Dict[str, Any]) -> None:
        DatasetAnalysisDialog(self, result,
                              apply_cb=self._apply_analysis_recommendations)

    def _reset_analyze_btn(self) -> None:
        if hasattr(self, "_btn_analyze"):
            self._btn_analyze.configure(state="normal",
                                        text="🔍 Анализировать датасет и авто-настроить")

    def _apply_analysis_recommendations(self, rec: Dict[str, Any]) -> None:
        """Применяет рекомендации анализатора — устанавливает tk.Var."""
        applied = []

        def _set(attr: str, val: Any) -> None:
            if hasattr(self, attr):
                try:
                    getattr(self, attr).set(val)
                    applied.append(attr)
                except Exception as _set_exc:
                    _log.debug("_apply_analysis_recommendations: skip attr=%s: %s", attr, _set_exc)

        _set("max_features",          rec.get("max_features"))
        _set("C",                     rec.get("C"))
        _set("use_svd",               rec.get("use_svd"))
        _set("svd_components",        rec.get("svd_components"))
        _set("char_ng_min",           rec.get("char_ng_min"))
        _set("char_ng_max",           rec.get("char_ng_max"))
        _set("word_ng_min",           rec.get("word_ng_min"))
        _set("word_ng_max",           rec.get("word_ng_max"))
        _set("class_weight_balanced", rec.get("balanced"))
        _set("use_smote",             rec.get("use_smote"))
        _set("use_hierarchical",      rec.get("use_hierarchical"))
        _set("use_hard_negatives",    rec.get("use_hard_negatives"))
        _set("use_field_dropout",     rec.get("use_field_dropout"))
        _set("use_kfold_ensemble",    rec.get("use_kfold_ensemble"))
        _set("kfold_k",               rec.get("kfold_k"))
        _set("use_confident_learning", rec.get("use_confident_learning"))

        self.log_train(
            f"  ✅ Рекомендации применены: {len(applied)} параметров обновлено. "
            "Модель: " + rec.get("model_label", "LinearSVC"))

    def pick_base_model(self):
        p = filedialog.askopenfilename(title="Базовая модель .joblib", filetypes=[("Joblib", "*.joblib"), ("All files", "*.*")])
        if not p:
            return
        is_trusted = ensure_trusted_model_path(
            self._trust_store, p, "Базовая модель",
            confirm_fn=make_tkinter_confirm_fn(self),
            logger=_log,
        )
        if not is_trusted:
            return
        try:
            load_pkg = getattr(self, "_load_model_pkg", None)
            if callable(load_pkg):
                pkg = load_pkg(p, log_fn=self.log_train)
            else:
                pkg = load_model_artifact(
                    p,
                    precomputed_sha256=self._trust_store.get_hash(p),
                    supported_schema_version=1,
                    expected_artifact_types=(TRAIN_MODEL_ARTIFACT_TYPE,),
                    required_keys=("pipeline",),
                    allowed_extensions=(".joblib",),
                    require_trusted=True,
                    trusted_paths=self._trust_store.trusted_canonical_paths(),
                    allow_missing_schema=True,
                    log_fn=self.log_train,
                    logger=_log,
                )
            cfg = upgrade_config_dict(pkg.get("config", {}))
            if not pkg.get("config"):
                raise FeatureBuildError("В модели нет config (нужен .joblib из этого приложения).")
            self.base_model_file.set(p)
            self._apply_config_to_ui(cfg)
            self.log_train(f"Базовая модель загружена: {p}")
        except Exception as e:
            self.base_model_file.set("")
            messagebox.showerror("Ошибка", f"Не удалось загрузить базовую модель:\n{e}")

    # --------------------------------------------------- build training dataset
    def _build_training_dataset(
        self,
        xlsx_paths: List[Path],
        snap: Dict[str, Any],
        progress_cb: Optional[Callable[[float, str], None]] = None,
    ) -> Tuple[List[str], List[str], Dict[str, Any]]:
        desc_col = snap["desc_col"]
        call_col = snap["call_col"]
        chat_col = snap["chat_col"]
        sum_col  = snap["summary_col"]
        as_col   = snap["ans_short_col"]
        af_col   = snap["ans_full_col"]
        lab_col  = snap["label_col"]

        use_sum = snap["use_summary"]
        ign_bot = snap["ignore_chatbot"]
        auto_p  = snap["auto_profile"]
        base_w  = snap["base_w"]

        stats = defaultdict(int)
        total_rows = estimate_total_rows(xlsx_paths)
        start_ts = time.time()
        done = 0

        X_all: List[str] = []
        y_all: List[str] = []

        for path in xlsx_paths:
            self.after(0, lambda p=path: self.log_train(f"  Файл: {p.name}"))
            with open_tabular(path) as it:
                header = next(it)
                headers = ["" if h is None else str(h) for h in header]

                i_desc = idx_of(headers, desc_col) if desc_col else None
                i_call = idx_of(headers, call_col) if call_col else None
                i_chat = idx_of(headers, chat_col) if chat_col else None
                i_sum  = idx_of(headers, sum_col) if (use_sum and sum_col) else None
                i_as   = idx_of(headers, as_col) if as_col else None
                i_af   = idx_of(headers, af_col) if af_col else None
                i_lab  = idx_of(headers, lab_col) if lab_col else None

                if i_lab is None:
                    raise FeatureBuildError(f"Не найден label-столбец '{lab_col}' в файле: {path.name}")

                def get(i, row):
                    return row[i] if (i is not None and i < len(row)) else None

                for row in it:
                    stats["rows_raw"] += 1
                    row = list(row)

                    lab = get(i_lab, row)
                    lab_s = "" if lab is None else str(lab).strip()
                    if not lab_s or lab_s.lower() == "nan":
                        stats["rows_skipped_no_label"] += 1
                        done += 1
                        continue

                    desc = normalize_text(get(i_desc, row)) if i_desc is not None else ""
                    call_raw = get(i_call, row) if i_call is not None else None
                    chat_raw = get(i_chat, row) if i_chat is not None else None
                    summ = normalize_text(get(i_sum, row)) if i_sum is not None else ""
                    ans_s = clean_answer_text(get(i_as, row)) if i_as is not None else ""
                    ans_f = clean_answer_text(get(i_af, row)) if i_af is not None else ""

                    call_clean, call_client, call_oper, r1 = parse_dialog_roles(call_raw, ignore_chatbot=ign_bot)
                    chat_clean, chat_client, chat_oper, r2 = parse_dialog_roles(chat_raw, ignore_chatbot=ign_bot)
                    roles_found = bool(r1 or r2)
                    if roles_found:
                        stats["roles_found_rows"] += 1

                    has_dialog = bool(call_clean or chat_clean)
                    if has_dialog:
                        stats["has_dialog_rows"] += 1

                    client_text = "\n".join([t for t in [call_client, chat_client] if t]).strip()
                    operator_text = "\n".join([t for t in [call_oper, chat_oper] if t]).strip()

                    channel = (
                        "call" if (call_clean and not chat_clean) else
                        "chat" if (chat_clean and not call_clean) else
                        "call+chat" if (call_clean and chat_clean) else
                        "none"
                    )

                    weights = choose_row_profile_weights(
                        base=base_w,
                        auto_profile=auto_p,
                        has_desc=bool(desc),
                        has_dialog=has_dialog,
                        roles_found=roles_found,
                        has_summary=bool(summ),
                        has_ans_s=bool(ans_s),
                        has_ans_f=bool(ans_f),
                    )

                    has_any_content = bool(
                        desc or client_text or operator_text or summ or ans_s or ans_f
                    )
                    if not has_any_content:
                        stats["rows_skipped_empty_text"] += 1
                        done += 1
                        continue

                    x = build_feature_text(
                        channel=channel,
                        desc=desc,
                        client_text=client_text,
                        operator_text=operator_text,
                        summary=summ,
                        ans_short=ans_s,
                        ans_full=ans_f,
                        weights=weights,
                        normalize_entities=bool(snap.get("use_entity_norm", False)),
                    )

                    if not x.strip():
                        stats["rows_skipped_empty_text"] += 1
                        done += 1
                        continue

                    X_all.append(x)
                    y_all.append(lab_s)
                    stats["rows_used"] += 1

                    done += 1
                    if progress_cb and (done % 200 == 0 or done == 1 or done == total_rows):
                        pct = 45.0 * done / float(total_rows)
                        progress_cb(pct, f"Чтение: {done}/{total_rows} | {fmt_speed(start_ts, done)} | {fmt_eta(start_ts, done, total_rows)}")

        return X_all, y_all, dict(stats)

    # ----------------------------------------------------------- run_grid_search_c
    def run_grid_search_c(self):
        """Запускает GridSearch по C в фоновом потоке, затем ставит лучший C в UI."""
        with self._proc_lock:
            if self._processing:
                return
            self._processing = True

        if not self.train_files:
            with self._proc_lock:
                self._processing = False
            messagebox.showwarning("GridSearch C", "Добавь файлы обучения сначала.")
            return

        snap = self._snap_params()
        try:
            TrainWorkflowConfig.from_snapshot(snap)
        except Exception as ex:
            with self._proc_lock:
                self._processing = False
            messagebox.showerror("Обучение", f"Некорректные параметры: {ex}")
            return
        self._cancel_event.clear()
        self.btn_train.configure(state="disabled", text="⏳ Обучение…")
        self.btn_grid_search_c.configure(state="disabled")
        self.btn_grid_search_stop.configure(state="normal")
        self.train_progress.set(0.0)
        self.train_status.set("GridSearch C — старт…")
        self.train_pct.set("0%")
        self.train_phase.set("GridSearch C…")
        self.train_speed.set("")
        self.train_eta.set("")
        self.log_train("\n==== GridSearch C — START ====")

        xlsx_paths = [Path(p) for p in self.train_files]
        _ctrl = WorkflowProgressController(
            self.train_progress, self.train_status, self.train_pct,
            self.train_phase, self.train_speed, self.train_eta,
        )

        def ui_prog(pct: float, status: str):
            _ctrl.update(pct, status)

        _lifecycle = OperationLifecycle(
            owner=self,
            run_button=self.btn_train,
            run_button_idle_text="▶  Обучить / сохранить модель",
            stop_button=self.btn_train_stop,
            log_fn=self.log_train,
        )

        def worker():
            try:
                self.after(0, lambda: self.train_progress.set(0.0))
                self.after(0, lambda: self.train_status.set("Чтение данных…"))

                X, y, stats = self._build_training_dataset(
                    xlsx_paths, snap,
                    progress_cb=lambda pct, s: self.after(0, lambda p=pct, s=s: ui_prog(p * 0.3, s)),
                )
                if not X:
                    self.after(0, lambda: messagebox.showerror("GridSearch C", "Не удалось прочитать данные."))
                    return

                self.after(0, lambda: self.log_train(f"Данные: {len(X)} строк, {len(set(y))} классов"))

                if snap.get("use_sbert_hybrid") or snap.get("use_sbert"):
                    self.after(0, lambda: self.log_train(
                        "ℹ GridSearch: SBERT не используется (только TF-IDF) для скорости.\n"
                        "  Найденный C — приближение. Проверьте качество вручную после обучения с SBERT."
                    ))

                features = make_hybrid_vectorizer(
                    char_ng=snap["char_ng"],
                    word_ng=snap["word_ng"],
                    min_df=snap["min_df"],
                    max_features=snap["max_features"],
                    sublinear_tf=snap["sublinear_tf"],
                    use_stop_words=snap["use_stop_words"],
                    extra_stop_words=snap["extra_stop_words"] or None,
                    extra_noise_tokens=snap.get("extra_noise_tokens") or None,
                    extra_noise_phrases=snap.get("extra_noise_phrases") or None,
                    use_noise_tokens=snap["use_noise_tokens"],
                    use_noise_phrases=snap["use_noise_phrases"],
                    use_per_field=snap.get("use_per_field", True),
                    base_weights=snap.get("base_w"),
                    use_svd=snap.get("use_svd", False),
                    svd_components=snap.get("svd_components", 300),
                    use_lemma=snap.get("use_lemma", False),
                    use_pos_tags=snap.get("use_pos_tags", False),
                    use_meta=snap.get("use_meta", False),
                )

                def _prog(pct, s):
                    self.after(0, lambda p=pct, s=s: ui_prog(30.0 + p * 0.7, s))

                best_c, scores = find_best_c(
                    X, y,
                    features=features,
                    balanced=snap["balanced"],
                    max_iter=snap["max_iter"],
                    progress_cb=_prog,
                    cancel_event=self._cancel_event,
                    n_jobs=self._hw.n_jobs_cv,
                )

                cancelled = self._cancel_event.is_set()
                no_results = not scores
                lines = ["GridSearch C — результаты (прервано):" if cancelled else "GridSearch C — результаты:"]
                for c in sorted(scores):
                    mark = " ← лучший" if c == best_c else ""
                    lines.append(f"  C={c:<6}  macro F1={scores[c]:.4f}{mark}")
                result_text = "\n".join(lines)

                def _done(cancelled=cancelled, no_results=no_results):
                    self.log_train(result_text)
                    self.btn_train.configure(state="normal", text="▶  Обучить / сохранить модель")
                    self.btn_grid_search_c.configure(state="normal")
                    self.btn_grid_search_stop.configure(state="disabled")
                    if no_results:
                        self.log_train("==== GridSearch C — ПРЕРВАНО (нет результатов) ====")
                        self.train_status.set("Прервано до первой итерации.")
                        messagebox.showwarning("GridSearch C",
                                               "GridSearch прерван до завершения первой итерации.\n"
                                               "Результатов нет. Значение C не изменено.")
                        return
                    self.C.set(round(best_c, 6))
                    if cancelled:
                        self.log_train("==== GridSearch C — ПРЕРВАНО ====")
                        self.train_status.set(f"Прервано. Лучший C из завершённых: {best_c}")
                    else:
                        self.log_train("==== GridSearch C — DONE ====")
                        self.train_progress.set(100.0)
                        self.train_status.set(f"Готово: C={best_c}")
                    msg = (
                        f"GridSearch прерван. Лучший C из {len(scores)} итераций = {best_c}\n"
                        f"macro F1 = {scores[best_c]:.4f}\n\n"
                        f"Значение выставлено в поле C автоматически."
                        if cancelled else
                        f"Лучший C = {best_c}\n"
                        f"macro F1 = {scores[best_c]:.4f}\n\n"
                        f"Значение выставлено в поле C автоматически."
                    )
                    messagebox.showinfo("GridSearch C", msg)

                self.after(0, _done)

            except Exception as exc:
                def _on_error(e=exc):
                    self.log_train(f"[GridSearch C ERROR] {e}")
                    self.train_progress.set(0.0)
                    self.btn_train.configure(state="normal", text="▶  Обучить / сохранить модель")
                    self.btn_grid_search_c.configure(state="normal")
                    self.btn_grid_search_stop.configure(state="disabled")
                self.after(0, _on_error)
            finally:
                with self._proc_lock:
                    self._processing = False

        threading.Thread(target=worker, daemon=True).start()

    def _stop_grid_search(self):
        """Устанавливает cancel_event для остановки GridSearch после текущей итерации."""
        self._cancel_event.set()
        self.btn_grid_search_stop.configure(state="disabled")
        self.log_train("⏹ Остановка GridSearch после текущей итерации…")

    def _log_train_config_snapshot(self, snap: Dict[str, Any]) -> None:
        """Логирует ключевую конфигурацию обучения перед стартом worker."""
        vec_parts = []
        if snap.get("use_setfit"):
            vec_parts.append("SetFit")
        elif snap.get("use_sbert"):
            vec_parts.append("SBERT+TF-IDF" if snap.get("use_sbert_hybrid") else "SBERT")
        else:
            vec_parts.append("TF-IDF")
        if snap.get("use_svd"):
            vec_parts.append(f"SVD({snap.get('svd_components', 300)})")
        if snap.get("use_lemma"):
            vec_parts.append("Лемм")
        self.log_train(
            f"[Конфиг] Векторизатор: {' + '.join(vec_parts)}"
            f" | C={snap.get('C', '?')} | max_iter={snap.get('max_iter', '?')}"
            f" | max_features={snap.get('max_features', 0):,}"
            f" | test_size={snap.get('test_size', '?')}"
            f" | seed={snap.get('random_state', '?')}"
        )
        self.log_train(
            f"[Конфиг] SMOTE={'да' if snap.get('use_smote') else 'нет'}"
            f" | balanced={'да' if snap.get('balanced') else 'нет'}"
            f" | calib={snap.get('calib_method', 'sigmoid')}"
            f" | файлов={len(self.train_files)}"
        )

    def _log_dataset_snapshot(self, stats: Dict[str, Any], y: List[str]) -> None:
        """Логирует базовую статистику собранного датасета."""
        label_counter = Counter(y)
        n_classes = len(label_counter)
        sorted_labels = sorted(label_counter.items(), key=lambda x: -x[1])
        shown = sorted_labels[:25]
        tail = f"\n    … и ещё {n_classes - 25} классов" if n_classes > 25 else ""
        label_lines = "\n".join(f"    {lbl}: {cnt}" for lbl, cnt in shown) + tail
        self.log_train(
            f"\nДатасет прочитан:\n"
            f"  Строк всего:              {stats.get('rows_raw', 0)}\n"
            f"  Использовано для обучения: {stats.get('rows_used', 0)}\n"
            f"  Пропущено (нет метки):     {stats.get('rows_skipped_no_label', 0)}\n"
            f"  Пропущено (пустой текст):  {stats.get('rows_skipped_empty_text', 0)}\n"
            f"  Строк с диалогом:          {stats.get('has_dialog_rows', 0)}\n"
            f"  Строк с ролями CLIENT/OP:  {stats.get('roles_found_rows', 0)}\n"
            f"Классов: {n_classes}\n"
            f"Распределение меток:\n{label_lines}"
        )

    def _log_cleaning_report(self, cleaning_report: Dict[str, Any]) -> None:
        """Логирует результат очистки датасета."""
        lines = ["\n--- Проверка и очистка датасета ---"]
        if cleaning_report["n_duplicates"] > 0:
            lines.append(f"  Дублирующих строк убрано  : {cleaning_report['n_duplicates']}")
        else:
            lines.append("  Дублирующих строк         : не обнаружено")
        if cleaning_report["n_conflicts"] > 0:
            dropped = cleaning_report.get("n_conflict_rows_dropped", 0)
            if dropped > 0:
                lines.append(
                    f"  Конфликты меток убраны    : {cleaning_report['n_conflicts']} уникальных "
                    f"текстов → {dropped} строк удалено"
                )
            else:
                lines.append(
                    f"  ⚠ Конфликты меток         : {cleaning_report['n_conflicts']} уникальных "
                    f"текстов имеют разные метки (строки оставлены)"
                )
        if cleaning_report["excluded_classes"]:
            lines.append(
                f"  Классов исключено (< {cleaning_report['min_samples']} прим.): "
                f"{len(cleaning_report['excluded_classes'])}"
            )
            for lab, cnt_v in sorted(cleaning_report["excluded_classes"].items(), key=lambda kv: kv[1]):
                lines.append(f"    - «{lab}»: {cnt_v} прим.")
            lines.append(f"  Строк исключено            : {cleaning_report['n_excluded_rows']}")
        else:
            lines.append(f"  Классов с мало примеров   : нет (порог: {cleaning_report['min_samples']})")
        lines.append(f"  Итого строк для обучения  : {cleaning_report['n_final']}")
        self.log_train("\n".join(lines))

    # ──────────────────────────────── worker step helpers ─────────────────────

    def _train_step_near_dups(self, snap: dict, X: list, y: list) -> None:
        """Detect near-duplicate conflicts in training data and log results."""
        if not snap.get("detect_near_dups", False):
            return
        _nd_thr = float(snap.get("near_dup_threshold", 0.92))
        self.after(0, lambda thr=_nd_thr: self.log_train(
            f"  Поиск почти-дублей с конфликтами (порог косинуса ≥{thr:.2f})…"
        ))
        _nd_pairs = detect_near_duplicate_conflicts(
            X, y, threshold=_nd_thr, max_pairs=200,
            log_fn=lambda m: self.after(0, lambda msg=m: self.log_train(msg)),
        )
        if _nd_pairs:
            _nd_lines = "\n".join(
                f"    [{s:.3f}] «{t1[:80]}» ({l1}) ≈ «{t2[:80]}» ({l2})"
                for t1, t2, l1, l2, s in _nd_pairs[:10]
            )
            self.after(0, lambda n=len(_nd_pairs), lines=_nd_lines: self.log_train(
                f"  ⚠ Почти-дубли с разными метками: {n} пар\n{lines}"
                + (f"\n    … и ещё {n - 10} пар" if n > 10 else "")
            ))
        else:
            self.after(0, lambda thr=_nd_thr: self.log_train(
                f"  ✓ Почти-дубли с конфликтами не найдены (порог {thr:.2f})"
            ))

    def _train_step_llm_augment(
        self, snap: dict, X: list, y: list
    ) -> tuple[list, list]:
        """Augment rare classes via LLM paraphrasing; returns (possibly extended) X, y."""
        if not snap.get("use_llm_augment", False):
            return X, y
        _aug_min  = int(snap.get("augment_min_samples", 30))
        _aug_n    = int(snap.get("augment_n_paraphrases", 3))
        _llm_prov = snap.get("llm_provider", "anthropic").lower()
        _llm_mod  = snap.get("llm_model", "claude-sonnet-4-6")
        _llm_key  = self._resolve_llm_api_key(snap, _llm_prov)
        if _llm_key or _llm_prov == "ollama":
            from ml_augment import augment_rare_classes as _aug_fn
            self.after(0, lambda: self.log_train(
                f"  LLM-аугментация: порог ≤{_aug_min} примеров, "
                f"+{_aug_n} перефразировок на класс…"
            ))
            X, y, _aug_rep = _aug_fn(
                X=X, y=y,
                min_samples_threshold=_aug_min,
                n_paraphrases=_aug_n,
                llm_complete_fn=self._llm_complete_text,
                provider=_llm_prov,
                model=_llm_mod,
                api_key=_llm_key,
                log_fn=lambda m: self.after(0, lambda msg=m: self.log_train(msg)),
                cancel_event=self._cancel_event,
            )
            if _aug_rep["rows_added"] > 0:
                self.after(0, lambda r=_aug_rep: self.log_train(
                    f"  ✅ LLM-аугментация: +{r['rows_added']} строк "
                    f"({r['classes_augmented']} классов)"
                ))
        else:
            self.after(0, lambda: self.log_train(
                "  ⚠ LLM-аугментация пропущена: не задан API-ключ"
            ))
        return X, y

    def _train_step_svd_autocap(self, snap: dict, n_final: int) -> dict:
        """Cap SVD components to dataset size; returns (possibly new) snap dict."""
        if not snap.get("use_svd", False):
            return snap
        _svd_user = snap.get("svd_components", 200)
        _svd_safe = min(_svd_user, max(50, n_final // 5))
        if _svd_safe < _svd_user:
            snap = dict(snap)
            snap["svd_components"] = _svd_safe
            self.after(0, lambda u=_svd_user, s=_svd_safe: self.log_train(
                f"  SVD auto-cap: {u} → {s} компонент "
                f"(датасет {n_final} строк, макс. {n_final // 5})"
            ))
        return snap

    def _train_step_pseudo_label(self, snap: dict, X: list, y: list) -> None:
        """Extend X, y with high-confidence pseudo-labelled rows (in-place)."""
        if not snap.get("use_pseudo_label", False):
            return
        _pl_file = str(snap.get("pseudo_label_file", "")).strip()
        _pl_thr  = float(snap.get("pseudo_label_threshold", 0.92))
        if not (_pl_file and Path(_pl_file).exists()):
            self.after(0, lambda: self.log_train(
                "  ⚠ Псевдо-разметка: файл не задан или не найден"
            ))
            return
        self.after(0, lambda f=Path(_pl_file).name, t=_pl_thr: self.log_train(
            f"  [Псевдо-разметка] Файл: «{f}» | порог: {t:.2f}"
        ))
        try:
            from ml_core import make_classifier
            from sklearn.pipeline import Pipeline as _PLPipeline
            from sklearn.feature_extraction.text import TfidfVectorizer as _TVPL
            _pl_tfidf = _TVPL(
                analyzer="char_wb", ngram_range=(2, 5),
                max_features=80_000, sublinear_tf=True,
            )
            _pl_clf, _ = make_classifier(y, C=float(snap.get("C", 1.0)),
                                         max_iter=500, balanced=True)
            _pl_pipe = _PLPipeline([("tfidf", _pl_tfidf), ("clf", _pl_clf)])
            _pl_pipe.fit(X, y)
            _pl_X_new: List[str] = []
            with open_tabular(Path(_pl_file)) as _pl_it:
                _pl_header = ["" if h is None else str(h) for h in next(_pl_it)]
                _pl_idx = {c: i for i, c in enumerate(_pl_header)}
                for _pl_row in _pl_it:
                    _pl_rv = list(_pl_row)
                    _pl_feat = self._row_to_feature_text(
                        _pl_rv, _pl_header, snap, header_index=_pl_idx
                    )
                    if _pl_feat.strip():
                        _pl_X_new.append(_pl_feat)
            if _pl_X_new:
                _pl_proba = _pl_pipe.predict_proba(_pl_X_new)
                _pl_classes = list(_pl_pipe.classes_)
                _pl_added = 0
                for _pi, _prow in enumerate(_pl_proba):
                    _pb = int(_prow.argmax())
                    if float(_prow[_pb]) >= _pl_thr:
                        X.append(_pl_X_new[_pi])
                        y.append(_pl_classes[_pb])
                        _pl_added += 1
                self.after(0, lambda n=_pl_added, tot=len(_pl_X_new): self.log_train(
                    f"  [Псевдо-разметка] +{n} строк из {tot} "
                    f"→ датасет: {len(X)} примеров"
                ))
        except Exception as _pl_e:
            self.after(0, lambda e=_pl_e: self.log_train(
                f"  ⚠ Псевдо-разметка: ошибка — {e}"
            ))

    def _train_step_anchor_texts(self, snap: dict, X: list, y: list) -> None:
        """Inject anchor texts for classes with few examples (in-place)."""
        if not snap.get("use_anchor_texts", False):
            return
        _anc_raw = str(snap.get("anchor_text_lines", "")).strip()
        _anc_copies = int(snap.get("anchor_copies", 3))
        _anc_added = 0
        _known_labels = set(y)
        for _anc_line in _anc_raw.splitlines():
            _anc_line = _anc_line.strip()
            if ":" not in _anc_line:
                continue
            _anc_sep = _anc_line.index(":")
            _anc_lbl = _anc_line[:_anc_sep].strip()
            _anc_txt = _anc_line[_anc_sep + 1:].strip()
            if not _anc_lbl or not _anc_txt:
                continue
            if _anc_lbl not in _known_labels:
                self.after(0, lambda l=_anc_lbl: self.log_train(
                    f"  ⚠ Якорь: класс «{l}» не найден в обучающих данных — пропущен"
                ))
                continue
            for _ in range(_anc_copies):
                X.append(_anc_txt)
                y.append(_anc_lbl)
            _anc_added += 1
        if _anc_added:
            self.after(0, lambda n=_anc_added, c=_anc_copies, tot=len(X): self.log_train(
                f"  [Якорные тексты] +{n} классов × {c} копий → датасет: {tot} примеров"
            ))

    # --------------------------------------------------------------- run_training
    def _run_training_setfit(
        self, snap: dict, X: list, y: list, workflow,
        ui_prog, log_fn, t0: float, _t_phase: float, _t_read: float,
    ) -> bool:
        """Run SetFit training path. Returns True if SetFit ran (caller should return)."""
        if not snap.get("use_setfit", False):
            return False
        import importlib.util as _sf_ilu
        if not _sf_ilu.find_spec("setfit"):
            raise FeatureBuildError(
                "Пакет setfit не установлен.\n"
                "Выполните: pip install setfit>=0.9\n"
                "Или установите через вкладку «Зависимости»."
            )
        from ml_setfit import train_model_setfit as _tmsf
        from ml_core import SBERT_LOCAL_DIR as _sf_cache

        _sf_model   = snap.get("setfit_model") or SETFIT_DEFAULT
        _sf_epochs  = snap.get("setfit_epochs", 3)
        _sf_iters   = snap.get("setfit_num_iterations", 20)
        _sf_batch   = snap.get("setfit_batch", 16)
        _sf_fp16    = snap.get("setfit_fp16", True)
        _sf_dev     = snap.get("sbert_device", "auto")
        _sf_cache   = str(_sf_cache)

        self.after(0, lambda: ui_prog(45.0, f"[SetFit] Запуск обучения {_sf_model}…"))

        pipe, clf_type, report, labels, cm, train_extras = _tmsf(
            X, y,
            model_name=_sf_model,
            num_iterations=_sf_iters,
            num_epochs=_sf_epochs,
            batch_size=_sf_batch,
            fp16=_sf_fp16,
            test_size=snap["test_size"],
            random_state=42,
            use_smote=snap.get("use_smote", True),
            device=_sf_dev,
            cache_dir=_sf_cache,
            progress_cb=lambda p, s: self.after(0, lambda p=p, s=s: ui_prog(p, s)),
            log_cb=_sbert_log,
        )

        if self._cancel_event.is_set():
            raise InterruptedError()

        self.after(0, lambda: ui_prog(95.0, "Сохранение SetFit модели…"))
        cfg = ModelConfig(
            version=11,
            created_at=datetime.now().isoformat(timespec="seconds"),
            language="ru",
            classifier_type=clf_type,
            desc_col=snap["desc_col"],
            call_col=snap["call_col"],
            chat_col=snap["chat_col"],
            summary_col=snap["summary_col"],
            answer_short_col=snap["ans_short_col"],
            answer_full_col=snap["ans_full_col"],
            label_col=snap["label_col"],
            use_summary=snap["use_summary"],
            ignore_chatbot=snap["ignore_chatbot"],
            auto_profile=snap["auto_profile"],
            use_stop_words=snap["use_stop_words"],
            w_desc=snap["w_desc"],
            w_client=snap["w_client"],
            w_operator=snap["w_operator"],
            w_summary=snap["w_summary"],
            w_answer_short=snap["w_ans_short"],
            w_answer_full=snap["w_ans_full"],
            char_ng_min=snap["char_ng"][0],
            char_ng_max=snap["char_ng"][1],
            word_ng_min=snap["word_ng"][0],
            word_ng_max=snap["word_ng"][1],
            min_df=snap["min_df"],
            max_features=snap["max_features"],
            sublinear_tf=snap["sublinear_tf"],
            C=snap["C"],
            max_iter=snap["max_iter"],
            class_weight_balanced=snap["balanced"],
            test_size=snap["test_size"],
            parent_model=snap["base_model_file"] if snap["train_mode"] == "finetune" else "",
            use_noise_tokens=snap["use_noise_tokens"],
            use_noise_phrases=snap["use_noise_phrases"],
            use_per_field=snap.get("use_per_field", True),
            use_svd=snap.get("use_svd", False),
            svd_components=snap.get("svd_components", 300),
            use_lemma=snap.get("use_lemma", False),
            use_sbert=False,
            sbert_model=_sf_model,
            use_sbert_hybrid=False,
            use_meta=False,
            calib_method=snap.get("calib_method", "sigmoid"),
            use_smote=snap.get("use_smote", True),
            classifier_backend="setfit",
            nn_model=_sf_model,
            nn_epochs=_sf_epochs,
            nn_batch_size=_sf_batch,
            nn_num_iterations=_sf_iters,
            nn_fp16=_sf_fp16,
            nn_gradient_checkpointing=False,
        )

        model_path = MODEL_DIR / f"marker1_model_{now_stamp()}.joblib"
        workflow.persist_artifact(
            {
                "artifact_type": TRAIN_MODEL_ARTIFACT_TYPE,
                "pipeline": pipe,
                "config": asdict(cfg),
                "schema_version": 1,
                "per_class_thresholds": train_extras.get("per_class_thresholds", {}),
            },
            str(model_path),
        )
        elapsed = time.time() - t0

        _t_fit = time.time() - _t_phase - _t_read

        def _done_setfit():
            self.after(0, lambda: ui_prog(100.0, "✅ SetFit обучение завершено"))
            self.log_train(f"\n{'═'*44}")
            self.log_train(f"  МОДЕЛЬ СОХРАНЕНА: {model_path.name}")
            self.log_train(f"  Классификатор:    {clf_type}")
            self.log_train(f"  Классов:          {len(set(y))}")
            self.log_train(f"  Времени:          {elapsed:.1f} с")
            try:
                import os as _os
                _sz = _os.path.getsize(str(model_path)) / (1024 * 1024)
                self.log_train(f"  Размер файла:     {_sz:.1f} МБ")
            except OSError as _e:
                _log.debug("model file size stat: %s", _e)
            self.log_train(
                f"[Время] Чтение: {_t_read:.1f}с"
                f" | Обучение+Вектор.: {_t_fit:.1f}с"
                f" | Итого: {elapsed:.1f}с"
            )
            if train_extras.get("thresh_75") is not None:
                _t75 = float(train_extras["thresh_75"])
                self._rec_threshold_75.set(_t75)
                self._rec_thr_label.set(f"← рек. {_t75:.3f}")
            if report and "ПРОПУЩЕНА" not in report:
                self.log_train(f"\n{report}")
            if train_extras.get("roc_auc_macro") is not None:
                self.log_train(f"  ROC-AUC macro:    {train_extras['roc_auc_macro']:.4f}")
            self.log_train(f"{'═'*44}")
            self._processing = False
            self.btn_train.configure(state="normal", text="▶  Обучить / сохранить модель")
            self.btn_train_stop.configure(state="disabled")

        self.after(0, _done_setfit)
        return True

    def _run_training_ensemble(
        self, snap: dict, X: list, y: list, workflow,
        ui_prog, sbert_log, t0: float,
        use_sbert_val: bool, use_sbert_hybrid: bool,
        sbert_model_val: str, make_tfidf,
    ) -> bool:
        if snap["train_mode"] != "ensemble":
            return False
        _ens_stamp = now_stamp()
        _ens_results: List[tuple] = []

        # Конфигурации двух моделей
        _vec2_val = snap.get("ensemble_vec2", "tfidf")
        _ens_models = [
            {
                "label": "М1",
                "use_sbert":  use_sbert_val,
                "use_hybrid": use_sbert_hybrid,
                "sbert_model": sbert_model_val,
                "pct_base":   0.0,
            },
            {
                "label": "М2",
                "use_sbert":  _vec2_val == "sbert",
                "use_hybrid": _vec2_val == "hybrid",
                "sbert_model": snap.get("sbert_model2", sbert_model_val),
                "pct_base":   50.0,
            },
        ]

        def _vec_tag(use_hyb, use_sb):
            if use_hyb: return "hybrid"
            if use_sb:  return "sbert"
            return "tfidf"

        def _vec_display(use_hyb, use_sb, model_name):
            if use_hyb: return f"Гибрид SBERT+TF-IDF [{model_name}]"
            if use_sb:  return f"SBERT [{model_name}]"
            _mn = " + мета-признаки" if snap.get("use_meta") else ""
            return f"TF-IDF{_mn}"

        for _em in _ens_models:
            _lbl       = _em["label"]
            _pct_base  = _em["pct_base"]
            _em_hyb    = _em["use_hybrid"]
            _em_sb     = _em["use_sbert"]
            _em_smod   = _em["sbert_model"]

            if self._cancel_event.is_set():
                raise InterruptedError()

            self.after(0, lambda l=_lbl: (
                self.log_train(f"\n{'═'*44}"),
                self.log_train(f"    ОБУЧЕНИЕ МОДЕЛИ {l} / 2"),
                self.log_train(f"{'═'*44}"),
            ))

            def _prog_em(pct, status, _b=_pct_base, _l=_lbl):
                self.after(0, lambda p=pct, s=status, b=_b, l=_l:
                    ui_prog(b + p * 0.5, f"[{l}] {s}"))

            _em_slog = (lambda msg, _l=_lbl:
                self.after(0, lambda m=msg, l=_l: self.log_train(f"[{l}] {m}")))

            _vi = _vec_display(_em_hyb, _em_sb, _em_smod)
            self.after(0, lambda l=_lbl, vi=_vi:
                self.log_train(f"[{l}] Векторизатор: {vi}"))

            self.after(0, lambda _b=_pct_base, _l=_lbl:
                ui_prog(_b, f"[{_l}] Подготовка признаков…"))

            if _em_hyb:
                _tfidf_p = make_tfidf(force_svd=True)
                _sbert_p = SBERTVectorizer(
                    model_name=_em_smod,
                    device=snap["sbert_device"],
                    batch_size=snap.get("sbert_batch", 32),
                    log_cb=_em_slog,
                    progress_cb=_prog_em,
                )
                # TF-IDF здесь проходит SVD + L2 (force_svd=True), SBERT отдаёт
                # ненормированные эмбеддинги — выравниваем масштаб вручную.
                # Вес sbert > tfidf, чтобы 384 dense-признака не тонули в
                # TF-IDF(SVD-200) при последующем LinearSVC.
                _sbert_pipe = _SklPipeline([
                    ("sbert", _sbert_p),
                    ("l2", _SklNormalizer(norm="l2", copy=False)),
                ])
                _em_feats = FeatureUnion(
                    [("tfidf", _tfidf_p), ("sbert", _sbert_pipe)],
                    transformer_weights={"tfidf": 1.0, "sbert": 3.0},
                )
            elif _em_sb:
                _em_feats = SBERTVectorizer(
                    model_name=_em_smod,
                    device=snap["sbert_device"],
                    batch_size=snap.get("sbert_batch", 32),
                    log_cb=_em_slog,
                    progress_cb=_prog_em,
                )
            else:
                _em_feats = make_tfidf()

            # Ensemble-члены наследуют ту же конфигурацию аугментации, что и
            # главная модель, кроме run_cv: каждый K-fold член — уже сам по
            # себе CV, второй слой кросс-валидации был бы избыточен.
            _em_options = TrainingOptions(
                calib_method=snap.get("calib_method", "sigmoid"),
                use_smote=snap.get("use_smote", True),
                oversample_strategy=snap.get("oversample_strategy", "augment_light"),
                max_dup_per_sample=int(snap.get("max_dup_per_sample", 5)),
                use_hard_negatives=snap.get("use_hard_negatives", False),
                use_field_dropout=snap.get("use_field_dropout", False),
                field_dropout_prob=float(snap.get("field_dropout_prob", 0.15)),
                field_dropout_copies=int(snap.get("field_dropout_copies", 2)),
                use_label_smoothing=bool(snap.get("use_label_smoothing", False)),
                label_smoothing_eps=float(snap.get("label_smoothing_eps", 0.05)),
                use_fuzzy_dedup=bool(snap.get("use_fuzzy_dedup", False)),
                fuzzy_dedup_threshold=int(snap.get("fuzzy_dedup_threshold", 92)),
            )
            _em_pipe, _em_ctype, _em_rep, _em_lbls, _em_cm, _em_ext = workflow.fit_and_evaluate(
                X, y, _em_feats,
                C=snap["C"],
                max_iter=snap["max_iter"],
                balanced=snap["balanced"],
                test_size=snap["test_size"],
                random_state=42,
                options=_em_options,
                progress_cb=lambda p, s, _b=_pct_base, _l=_lbl:
                    self.after(0, lambda p=p, s=s, b=_b, l=_l:
                        ui_prog(b + p * 0.5, f"[{l}] {s}")),
                log_cb=_em_slog,
            )

            if self._cancel_event.is_set():
                raise InterruptedError()

            _em_cfg = ModelConfig(
                version=11,
                created_at=datetime.now().isoformat(timespec="seconds"),
                language="ru",
                classifier_type=_em_ctype,
                desc_col=snap["desc_col"],
                call_col=snap["call_col"],
                chat_col=snap["chat_col"],
                summary_col=snap["summary_col"],
                answer_short_col=snap["ans_short_col"],
                answer_full_col=snap["ans_full_col"],
                label_col=snap["label_col"],
                use_summary=snap["use_summary"],
                ignore_chatbot=snap["ignore_chatbot"],
                auto_profile=snap["auto_profile"],
                use_stop_words=snap["use_stop_words"],
                w_desc=snap["w_desc"],
                w_client=snap["w_client"],
                w_operator=snap["w_operator"],
                w_summary=snap["w_summary"],
                w_answer_short=snap["w_ans_short"],
                w_answer_full=snap["w_ans_full"],
                char_ng_min=snap["char_ng"][0],
                char_ng_max=snap["char_ng"][1],
                word_ng_min=snap["word_ng"][0],
                word_ng_max=snap["word_ng"][1],
                min_df=snap["min_df"],
                max_features=snap["max_features"],
                sublinear_tf=snap["sublinear_tf"],
                C=snap["C"],
                max_iter=snap["max_iter"],
                class_weight_balanced=snap["balanced"],
                test_size=snap["test_size"],
                parent_model="",
                use_noise_tokens=snap["use_noise_tokens"],
                use_noise_phrases=snap["use_noise_phrases"],
                use_per_field=snap.get("use_per_field", True),
                use_svd=snap.get("use_svd", False) or _em_hyb,
                svd_components=snap.get("svd_components", 300),
                use_lemma=snap.get("use_lemma", False),
                use_sbert=_em_sb,
                sbert_model=_em_smod if (_em_sb or _em_hyb) else SBERT_DEFAULT,
                use_sbert_hybrid=_em_hyb,
                use_meta=snap.get("use_meta", False),
                calib_method=snap.get("calib_method", "sigmoid"),
                use_smote=snap.get("use_smote", True),
                classifier_backend="linearsvc",
            )

            _tag = _vec_tag(_em_hyb, _em_sb)
            _em_model_path = MODEL_DIR / (
                f"marker1_model_{_ens_stamp}_{_lbl}_{_tag}.joblib"
            )
            workflow.persist_artifact(
                {
                    "artifact_type": TRAIN_MODEL_ARTIFACT_TYPE,
                    "pipeline": _em_pipe,
                    "config": asdict(_em_cfg),
                    "schema_version": 1,
                    "per_class_thresholds": _em_ext.get("per_class_thresholds", {}),
                },
                str(_em_model_path),
            )

            # Метрики Excel
            _em_mpath: Optional[Path] = None
            _em_repdict = _em_ext.get("report_dict")
            if _em_repdict and _em_lbls is not None and _em_cm is not None:
                try:
                    from openpyxl import Workbook as _WB2
                    _mwb = _WB2(); _mwb.remove(_mwb.active)
                    _ms1 = _mwb.create_sheet("Метрики по классам")
                    _ms1.append(["Класс", "Precision", "Recall", "F1", "Support"])
                    for _c in _em_lbls:
                        _r = _em_repdict.get(str(_c), {})
                        _ms1.append([_c,
                            round(_r.get("precision", 0), 4),
                            round(_r.get("recall", 0), 4),
                            round(_r.get("f1-score", 0), 4),
                            int(_r.get("support", 0))])
                    for _k in ("macro avg", "weighted avg"):
                        _r = _em_repdict.get(_k, {})
                        if _r:
                            _ms1.append([_k,
                                round(_r.get("precision", 0), 4),
                                round(_r.get("recall", 0), 4),
                                round(_r.get("f1-score", 0), 4),
                                int(_r.get("support", 0))])
                    _ms2 = _mwb.create_sheet("Матрица ошибок")
                    _ms2.append(["TRUE / PRED"] + [str(l) for l in _em_lbls])
                    for _i2, _lb2 in enumerate(_em_lbls):
                        _ms2.append([_lb2] + [int(_em_cm[_i2][_j2]) for _j2 in range(len(_em_lbls))])
                    _em_mpath = MODEL_DIR / (
                        _em_model_path.stem.replace("marker1_model_", "marker1_metrics_") + ".xlsx"
                    )
                    _mwb.save(_em_mpath)
                except OSError as _e:
                    _log.debug("ensemble metrics save: %s", _e)
                    _em_mpath = None

            _ens_results.append((_em_model_path, _em_mpath, _em_rep, _vi, _em_ctype, _em_ext, _em_pipe))

            self.after(0, lambda l=_lbl, mp=_em_model_path: self.log_train(
                f"[{l}] Модель сохранена: {mp}"))
            self.after(0, lambda l=_lbl, r=_em_rep: self.log_train(
                f"\n[{l}] ── МЕТРИКИ ──\n{r}"))

        # ── Мета-классификатор: LogisticRegression на стеке вероятностей ──
        if len(_ens_results) == 2:
            try:
                import numpy as _np_meta
                from sklearn.model_selection import train_test_split as _tts_meta
                from sklearn.linear_model import LogisticRegression as _LR_meta
                import joblib as _jbl_meta

                _pipe1_m = _ens_results[0][6]
                _pipe2_m = _ens_results[1][6]
                _all_cls_m = list(_pipe1_m.classes_)

                # Воспроизводим тот же split что и в fit_and_evaluate (random_state=42)
                _Xtr_m, _Xva_m, _ytr_m, _yva_m = _tts_meta(
                    X, y,
                    test_size=snap["test_size"],
                    random_state=42,
                    stratify=y,
                )

                def _align_proba_m(pipe, Xval, all_cls):
                    proba = pipe.predict_proba(Xval)
                    pipe_cls = list(pipe.classes_)
                    n, nc = len(Xval), len(all_cls)
                    out = _np_meta.zeros((n, nc))
                    for j, c in enumerate(all_cls):
                        if c in pipe_cls:
                            out[:, j] = proba[:, pipe_cls.index(c)]
                        else:
                            out[:, j] = 1.0 / nc
                    return out

                self.after(0, lambda: self.log_train(
                    "  Мета-классификатор: вычисляю вероятности на валидационной выборке…"
                ))
                _p1_m = _align_proba_m(_pipe1_m, _Xva_m, _all_cls_m)
                _p2_m = _align_proba_m(_pipe2_m, _Xva_m, _all_cls_m)
                _stacked_m = _np_meta.hstack([_p1_m, _p2_m])

                _meta_clf = _LR_meta(C=1.0, max_iter=300, random_state=42)
                _meta_clf.fit(_stacked_m, _yva_m)

                # Добавляем мета-классификатор в бандл М1
                _m1_bundle = _jbl_meta.load(str(_ens_results[0][0]))
                _m1_bundle["meta_learner"] = _meta_clf
                _m1_bundle["meta_learner_classes"] = _all_cls_m
                _jbl_meta.dump(_m1_bundle, str(_ens_results[0][0]), compress=3)

                self.after(0, lambda nc=len(_all_cls_m): self.log_train(
                    f"  ✅ Мета-классификатор сохранён (LogReg, {nc*2} входных признаков)"
                ))
            except Exception as _meta_exc:
                self.after(0, lambda e=str(_meta_exc): self.log_train(
                    f"  ⚠ Мета-классификатор не создан: {e}"
                ))

        # ── Итоговый вывод ──────────────────────────────────────────
        _ens_elapsed = time.time() - t0

        def _done_ensemble():
            ui_prog(100.0, "Ансамбль готов ✅")
            self.train_speed.set("")
            self.train_eta.set(f"Итого: {_ens_elapsed:.0f}с")
            self.model_file.set(str(_ens_results[0][0]))
            self.log_train(f"\n{'═'*44}")
            self.log_train(f"АНСАМБЛЬ ГОТОВ ✅  time={_ens_elapsed:.1f}s")
            _skip_lbl2  = stats.get('rows_skipped_no_label', 0)
            _skip_text2 = stats.get('rows_skipped_empty_text', 0)
            _skip_info2 = ""
            if _skip_lbl2 or _skip_text2:
                _skip_info2 = f"  пропущено: {_skip_lbl2} без метки, {_skip_text2} без текста"
            self.log_train(
                f"  raw={stats.get('rows_raw')}  "
                f"used={stats.get('rows_used')}"
                f"{_skip_info2}  "
                f"roles={stats.get('roles_found_rows')}"
            )
            for _i, (_mp, _mxp, _rep, _vi, _ct, _ext, *_) in enumerate(_ens_results):
                _l = f"М{_i + 1}"
                self.log_train(f"  {_l}: {_mp.name}")
                self.log_train(f"       векторизатор: {_vi}  |  {_ct}")
                if _mxp:
                    self.log_train(f"       метрики Excel: {_mxp}")
                _t75 = _ext.get("thresh_75")
                if _t75 is not None:
                    self.log_train(f"       рек. порог (75%): {float(_t75):.3f}")
            self.log_train(f"{'═'*44}")
            # Пороги из М1 → рекомендация
            _ext0 = _ens_results[0][5]
            if _ext0.get("thresh_75") is not None:
                _t75_0 = float(_ext0["thresh_75"])
                self._rec_threshold_75.set(_t75_0)
                self._rec_thr_label.set(f"← рек. {_t75_0:.3f} (М1 ансамбля)")
            self._processing = False
            self.btn_train.configure(state="normal", text="▶  Обучить / сохранить модель")
            self.btn_train_stop.configure(state="disabled")

        self.after(0, _done_ensemble)
        return True  # не продолжаем обычный путь обучения

    def _run_training_vectorize(
        self, snap: dict,
        use_sbert_hybrid: bool, use_sbert_val: bool, sbert_model_val: str,
        make_tfidf, sbert_log, sbert_prog,
    ):
        """Build and return the features vectorizer for the standard training path."""
        if use_sbert_hybrid:
            self.after(0, lambda: sbert_prog(50.0, "Гибрид SBERT+TF-IDF: векторизация…"))
            tfidf_part = make_tfidf(force_svd=True)
            sbert_part = SBERTVectorizer(
                model_name=sbert_model_val,
                device=snap["sbert_device"],
                batch_size=snap.get("sbert_batch", 32),
                log_cb=sbert_log,
                progress_cb=sbert_prog,
            )
            # TF-IDF (force_svd=True) уже L2-нормирован на выходе SVD; SBERT
            # отдаёт ненормированные эмбеддинги → оборачиваем L2-Normalizer,
            # чтобы магнитуды были сопоставимы. transformer_weights: SBERT 3.0,
            # потому что 384 dense-признака против ~200 SVD-компонент иначе
            # недовешиваются в LinearSVC.
            sbert_pipe = _SklPipeline([
                ("sbert", sbert_part),
                ("l2", _SklNormalizer(norm="l2", copy=False)),
            ])
            features = FeatureUnion(
                [("tfidf", tfidf_part), ("sbert", sbert_pipe)],
                transformer_weights={"tfidf": 1.0, "sbert": 3.0},
            )
            self.after(0, lambda: self.log_train(
                "ℹ Режим: SBERT + TF-IDF гибрид (FeatureUnion). "
                "SVD автоматически включён для TF-IDF части; "
                "SBERT L2-нормирован, вес 3.0 vs TF-IDF 1.0."
            ))

        elif use_sbert_val:
            self.after(0, lambda: sbert_prog(50.0, "Проверка кэша SBERT…"))
            features = SBERTVectorizer(
                model_name=sbert_model_val,
                device=snap["sbert_device"],
                batch_size=snap.get("sbert_batch", 32),
                log_cb=sbert_log,
                progress_cb=sbert_prog,
            )

        else:
            _meta_note = " + мета-признаки" if snap.get("use_meta") else ""
            self.after(0, lambda n=_meta_note: sbert_prog(50.0, f"Сбор признаков / обучение (TF-IDF{n})…"))
            features = make_tfidf()

        return features

    def _run_training_fit(
        self, snap: dict, X: list, y: list, features,
        effective_C: float, workflow, sbert_log, ui_prog,
        use_sbert_val: bool, use_sbert_hybrid: bool, make_tfidf,
    ) -> tuple:
        """Run Optuna tuning, fit_and_evaluate, K-fold ensemble, and hierarchical model.

        Returns:
            (pipe, clf_type, report, labels, cm, train_extras,
             effective_C, kfold_models, kfold_classes, group_model, class_to_group)
        """
        # ── Optuna: автоподбор гиперпараметров ───────────────────────
        _effective_C = effective_C
        if snap.get("use_optuna", False) and not use_sbert_val and not use_sbert_hybrid:
            _n_optuna = int(snap.get("n_optuna_trials", 30))
            self.after(0, lambda n=_n_optuna: ui_prog(48.0,
                f"Optuna: поиск гиперпараметров ({n} испытаний)…"))
            try:
                _opt_feats = make_tfidf()
                _opt_params, _opt_score = optuna_tune(
                    X, y,
                    features=_opt_feats,
                    balanced=snap["balanced"],
                    max_iter=snap["max_iter"],
                    n_trials=_n_optuna,
                    cv=4,
                    progress_cb=lambda p, s: self.after(0, lambda p=p, s=s: ui_prog(
                        48.0 + p * 0.02, s)),
                    cancel_event=self._cancel_event,
                    n_jobs=self._hw.n_jobs_cv,
                )
                if "C" in _opt_params:
                    _effective_C = float(_opt_params["C"])
                self.after(0, lambda p=_opt_params, s=_opt_score: self.log_train(
                    f"  [Optuna] Лучшие параметры: {p} | macro F1={s:.4f}"
                ))
            except Exception as _opt_e:
                self.after(0, lambda e=_opt_e: self.log_train(
                    f"  ⚠ Optuna: ошибка — {e}. Используется C={_effective_C}"
                ))

        # CV включаем только для TF-IDF моделей (SBERT слишком медленный)
        _run_cv = (not use_sbert_val and not use_sbert_hybrid
                   and len(X) >= 100 and len(set(y)) >= 2)
        _train_options = TrainingOptions(
            calib_method=snap.get("calib_method", "sigmoid"),
            use_smote=snap.get("use_smote", True),
            oversample_strategy=snap.get("oversample_strategy", "augment_light"),
            max_dup_per_sample=int(snap.get("max_dup_per_sample", 5)),
            run_cv=_run_cv,
            use_hard_negatives=snap.get("use_hard_negatives", False),
            use_field_dropout=snap.get("use_field_dropout", False),
            field_dropout_prob=float(snap.get("field_dropout_prob", 0.15)),
            field_dropout_copies=int(snap.get("field_dropout_copies", 2)),
            use_label_smoothing=bool(snap.get("use_label_smoothing", False)),
            label_smoothing_eps=float(snap.get("label_smoothing_eps", 0.05)),
            use_fuzzy_dedup=bool(snap.get("use_fuzzy_dedup", False)),
            fuzzy_dedup_threshold=int(snap.get("fuzzy_dedup_threshold", 92)),
        )
        pipe, clf_type, report, labels, cm, train_extras = workflow.fit_and_evaluate(
            X, y, features,
            C=_effective_C,
            max_iter=snap["max_iter"],
            balanced=snap["balanced"],
            test_size=snap["test_size"],
            random_state=42,
            options=_train_options,
            progress_cb=lambda p, s: self.after(0, lambda p=p, s=s: ui_prog(p, s)),
            log_cb=sbert_log,
        )

        if self._cancel_event.is_set():
            raise InterruptedError()

        # ── K-fold ансамбль: K независимых моделей ────────────────
        _kfold_models: list = []
        _kfold_classes: list = []
        if snap.get("use_kfold_ensemble", False) and not use_sbert_val and not use_sbert_hybrid:
            _kf_k = int(snap.get("kfold_k", 5))
            if len(X) >= _kf_k * 10 and len(set(y)) >= 2:
                try:
                    from ml_core import train_kfold_ensemble
                    self.after(0, lambda k=_kf_k: ui_prog(52.0,
                        f"K-fold ансамбль: обучаю {k} моделей…"))
                    _kfold_models, _kfold_classes = train_kfold_ensemble(
                        X, y,
                        features_factory=make_tfidf,
                        balanced=snap["balanced"],
                        C=_effective_C,
                        max_iter=snap["max_iter"],
                        k=_kf_k,
                        calib_method=snap.get("calib_method", "sigmoid"),
                        progress_cb=lambda p, s: self.after(0, lambda p=p, s=s: ui_prog(
                            52.0 + p * 0.1, s)),
                        cancel_event=self._cancel_event,
                        n_jobs=None,
                    )
                    self.after(0, lambda n=len(_kfold_models): self.log_train(
                        f"  [K-fold ансамбль] {n} моделей обучено ✅"
                    ))
                except Exception as _kfe:
                    self.after(0, lambda e=_kfe: self.log_train(
                        f"  ⚠ K-fold ансамбль: ошибка — {e}"
                    ))
            else:
                self.after(0, lambda: self.log_train(
                    "  ⚠ K-fold ансамбль: недостаточно данных (нужно ≥ k×10 примеров)"
                ))

        # ── Иерархическая классификация: group_model ──────────────
        _group_model = None
        _class_to_group: dict = {}
        if snap.get("use_hierarchical", False) and len(set(y)) >= 6:
            try:
                import math as _hmath
                from sklearn.feature_extraction.text import TfidfVectorizer as _TVH
                from sklearn.cluster import KMeans as _KMH
                from sklearn.pipeline import Pipeline as _PH
                from sklearn.calibration import CalibratedClassifierCV as _CCVH
                from sklearn.svm import LinearSVC as _SVCH

                _n_cls = len(set(y))
                _n_groups = max(2, int(_hmath.ceil(_hmath.sqrt(_n_cls))))
                self.after(0, lambda ng=_n_groups, nc=_n_cls: self.log_train(
                    f"  [Иерархия] Кластеризую {nc} классов в {ng} групп…"
                ))
                # Центроиды классов
                _hv = _TVH(analyzer="char_wb", ngram_range=(2, 4),
                           max_features=30_000, sublinear_tf=True)
                _cls_list = sorted(set(y))
                _cls_texts = [" ".join(xi for xi, yi in zip(X, y) if yi == c)[:5000]
                              for c in _cls_list]
                _centroids = _hv.fit_transform(_cls_texts)
                _km = _KMH(n_clusters=_n_groups, random_state=42, n_init=10)
                _km.fit(_centroids)
                _class_to_group = {c: int(_km.labels_[i])
                                   for i, c in enumerate(_cls_list)}
                # Лог групп
                _groups_log: dict = {}
                for c, g in _class_to_group.items():
                    _groups_log.setdefault(g, []).append(c)
                _glog = "\n".join(
                    f"    Группа {g}: {', '.join(cs[:5])}"
                    + ("…" if len(cs) > 5 else "")
                    for g, cs in sorted(_groups_log.items())
                )
                self.after(0, lambda gl=_glog: self.log_train(
                    f"  [Иерархия] Группы:\n{gl}"
                ))
                # Обучаем group_model
                _y_groups = [str(_class_to_group[yi]) for yi in y]
                _g_clf = _CCVH(_SVCH(C=float(snap.get("C", 1.0)),
                                     max_iter=500, class_weight="balanced"),
                               method="sigmoid", cv=3)
                _group_pipe = _PH([
                    ("tfidf", _TVH(analyzer="char_wb", ngram_range=(2, 5),
                                   max_features=80_000, sublinear_tf=True)),
                    ("clf", _g_clf),
                ])
                _group_pipe.fit(X, _y_groups)
                _group_model = _group_pipe
                self.after(0, lambda ng=_n_groups: self.log_train(
                    f"  [Иерархия] Группировочная модель обучена ({ng} групп) ✅"
                ))
            except Exception as _he:
                self.after(0, lambda e=_he: self.log_train(
                    f"  ⚠ Иерархия: ошибка — {e}"
                ))
                _group_model = None
                _class_to_group = {}

        return (
            pipe, clf_type, report, labels, cm, train_extras,
            _effective_C, _kfold_models, _kfold_classes, _group_model, _class_to_group,
        )

    def run_training(self):
        self._right_tabs.select(self._log_tab_indices["train"])
        with self._proc_lock:
            if self._processing:
                return
            self._processing = True

        if not validate_train_preconditions(self):
            return

        snap = build_validated_train_snapshot(self)
        if snap is None:
            return

        # Якорные тексты — читаем из Text-виджета (не tk.Var) до старта потока
        if snap.get("use_anchor_texts", False):
            try:
                _raw_anchors = self.anchor_texts_widget.get("1.0", "end").strip()
                snap["anchor_text_lines"] = _raw_anchors
            except Exception:
                snap["anchor_text_lines"] = ""

        begin_long_task(
            cancel_event=self._cancel_event,
            run_button=self.btn_train,
            run_button_busy_text="⏳ Обучение…",
            stop_button=self.btn_train_stop,
            progress_var=self.train_progress,
            status_var=self.train_status,
            pct_var=self.train_pct,
            phase_var=self.train_phase,
            speed_var=self.train_speed,
            eta_var=self.train_eta,
            start_phase="Подготовка…",
            start_log=self.log_train,
        )
        self.log_train("==== TRAIN START ====")

        # ── Конфигурационный снимок параметров ───────────────────────────────
        self._log_train_config_snapshot(snap)

        xlsx_paths = [Path(p) for p in self.train_files]
        t0 = time.time()
        _t_phase = time.time()   # для фазовых тайминг
        _task_ui = prepare_long_task_ui(
            owner=self,
            progress_var=self.train_progress,
            status_var=self.train_status,
            pct_var=self.train_pct,
            phase_var=self.train_phase,
            speed_var=self.train_speed,
            eta_var=self.train_eta,
            run_button=self.btn_train,
            run_button_idle_text="▶  Обучить / сохранить модель",
            stop_button=self.btn_train_stop,
            log_fn=self.log_train,
        )
        _ctrl = _task_ui.controller
        ui_prog = _task_ui.ui_prog
        _lifecycle = _task_ui.lifecycle

        def worker():
            nonlocal snap
            _t_read = 0.0
            _t_fit  = 0.0
            workflow = TrainingWorkflow()
            try:
                self.after(0, lambda: ui_prog(2.0, "Чтение Excel…"))
                X, y, stats = self._build_training_dataset(
                    xlsx_paths,
                    snap,
                    progress_cb=lambda p, s: self.after(0, lambda p=p, s=s: ui_prog(p, s))
                )
                _t_read = time.time() - _t_phase

                if self._cancel_event.is_set():
                    raise InterruptedError()

                # --- Лог статистики датасета ---
                self.after(0, lambda s=stats, yy=list(y): self._log_dataset_snapshot(s, yy))

                # ── Очистка датасета: дедупликация + исключение редких классов ──
                X, y, _cr = clean_training_data(
                    X, y,
                    min_samples_per_class=2,
                    drop_conflicts=bool(snap.get("drop_conflicts", True)),
                )

                self.after(0, lambda report=dict(_cr): self._log_cleaning_report(report))

                if _cr["n_final"] == 0:
                    raise FeatureBuildError("После очистки датасет пуст — нечего обучать.")
                stats["rows_used"] = _cr["n_final"]

                self._train_step_near_dups(snap, X, y)
                X, y = self._train_step_llm_augment(snap, X, y)
                snap = self._train_step_svd_autocap(snap, _cr["n_final"])
                self._train_step_pseudo_label(snap, X, y)
                self._train_step_anchor_texts(snap, X, y)

                fatal, warn = dataset_health_checks(stats, y)
                if fatal:
                    raise FeatureBuildError(";\n".join(fatal))
                if warn:
                    self.after(0, lambda w="\n".join(warn): self.set_warnings(w))

                use_sbert_val    = snap["use_sbert"]
                use_sbert_hybrid = snap.get("use_sbert_hybrid", False)
                sbert_model_val  = snap["sbert_model"]

                _sbert_log  = lambda msg: self.after(0, lambda m=msg: self.log_train(m))
                _sbert_prog = lambda pct, s: self.after(0, lambda p=pct, s=s: ui_prog(p, s))

                # 6Б: рекомендация SVD при большом словаре
                if snap["max_features"] > 100_000 and not snap.get("use_svd", False):
                    self.after(0, lambda mf=snap["max_features"]: self.log_train(
                        f"💡 Рекомендация: max_features={mf:,} > 100k — "
                        "включите SVD (LSA) для улучшения обобщения. "
                        "SVD сжимает разреженную TF-IDF матрицу до плотного "
                        "низкоразмерного представления."
                    ))

                def _make_tfidf(force_svd: bool = False):
                    """Создаёт TF-IDF пайплайн с учётом snap-параметров."""
                    tfidf = make_hybrid_vectorizer(
                        char_ng=snap["char_ng"],
                        word_ng=snap["word_ng"],
                        min_df=snap["min_df"],
                        max_features=snap["max_features"],
                        sublinear_tf=snap["sublinear_tf"],
                        use_stop_words=snap["use_stop_words"],
                        extra_stop_words=snap["extra_stop_words"] or None,
                        extra_noise_tokens=snap.get("extra_noise_tokens") or None,
                        extra_noise_phrases=snap.get("extra_noise_phrases") or None,
                        use_noise_tokens=snap["use_noise_tokens"],
                        use_noise_phrases=snap["use_noise_phrases"],
                        use_per_field=snap.get("use_per_field", True),
                        base_weights=snap.get("base_w"),
                        use_svd=(True if force_svd else snap.get("use_svd", False)),
                        svd_components=snap.get("svd_components", 300),
                        use_lemma=snap.get("use_lemma", False),
                        use_pos_tags=snap.get("use_pos_tags", False),
                        use_meta=snap.get("use_meta", False),
                    )
                    if snap.get("use_lemma"):
                        from ml_core import Lemmatizer as _Lem
                        _lem_probe = _Lem()
                        _lem_probe.fit(["probe"])
                        if not _lem_probe.is_active_:
                            self.after(0, lambda: self.log_train(
                                "⚠ ПРЕДУПРЕЖДЕНИЕ: «Лемматизация» включена, но ни pymorphy3, "
                                "ни pymorphy2 не установлены — лемматизация НЕ выполняется.\n"
                                "   Рекомендуется: pip install pymorphy3\n"
                                "   Fallback:      pip install pymorphy2"
                            ))
                        else:
                            _backend = getattr(_lem_probe, "backend_", "")
                            if _backend:
                                self.after(0, lambda b=_backend: self.log_train(
                                    f"[Лемматизация] backend: {b}"
                                ))
                    return tfidf

                # ── Режим SetFit: нейросетевой классификатор ─────────────────
                if self._run_training_setfit(snap, X, y, workflow, ui_prog, _sbert_log, t0, _t_phase, _t_read):
                    return



                # ── Режим ансамбля: обучить две модели последовательно ────────
                if self._run_training_ensemble(
                    snap, X, y, workflow, ui_prog, _sbert_log, t0,
                    use_sbert_val, use_sbert_hybrid, sbert_model_val, _make_tfidf,
                ):
                    return

                features = self._run_training_vectorize(
                    snap,
                    use_sbert_hybrid=use_sbert_hybrid,
                    use_sbert_val=use_sbert_val,
                    sbert_model_val=sbert_model_val,
                    make_tfidf=_make_tfidf,
                    sbert_log=_sbert_log,
                    sbert_prog=_sbert_prog,
                )

                (
                    pipe, clf_type, report, labels, cm, train_extras,
                    _effective_C, _kfold_models, _kfold_classes, _group_model, _class_to_group,
                ) = self._run_training_fit(
                    snap, X, y, features,
                    effective_C=float(snap["C"]),
                    workflow=workflow,
                    sbert_log=_sbert_log,
                    ui_prog=ui_prog,
                    use_sbert_val=use_sbert_val,
                    use_sbert_hybrid=use_sbert_hybrid,
                    make_tfidf=_make_tfidf,
                )

                self.after(0, lambda: ui_prog(95.0, "Сохранение модели…"))
                cfg = ModelConfig(
                    version=11,
                    created_at=datetime.now().isoformat(timespec="seconds"),
                    language="ru",
                    classifier_type=clf_type,
                    desc_col=snap["desc_col"],
                    call_col=snap["call_col"],
                    chat_col=snap["chat_col"],
                    summary_col=snap["summary_col"],
                    answer_short_col=snap["ans_short_col"],
                    answer_full_col=snap["ans_full_col"],
                    label_col=snap["label_col"],
                    use_summary=snap["use_summary"],
                    ignore_chatbot=snap["ignore_chatbot"],
                    auto_profile=snap["auto_profile"],
                    use_stop_words=snap["use_stop_words"],
                    w_desc=snap["w_desc"],
                    w_client=snap["w_client"],
                    w_operator=snap["w_operator"],
                    w_summary=snap["w_summary"],
                    w_answer_short=snap["w_ans_short"],
                    w_answer_full=snap["w_ans_full"],
                    char_ng_min=snap["char_ng"][0],
                    char_ng_max=snap["char_ng"][1],
                    word_ng_min=snap["word_ng"][0],
                    word_ng_max=snap["word_ng"][1],
                    min_df=snap["min_df"],
                    max_features=snap["max_features"],
                    sublinear_tf=snap["sublinear_tf"],
                    C=snap["C"],
                    max_iter=snap["max_iter"],
                    class_weight_balanced=snap["balanced"],
                    test_size=snap["test_size"],
                    parent_model=snap["base_model_file"] if snap["train_mode"] == "finetune" else "",
                    use_noise_tokens=snap["use_noise_tokens"],
                    use_noise_phrases=snap["use_noise_phrases"],
                    use_per_field=snap.get("use_per_field", True),
                    use_svd=snap.get("use_svd", False),
                    svd_components=snap.get("svd_components", 300),
                    use_lemma=snap.get("use_lemma", False),
                    use_sbert=use_sbert_val,
                    sbert_model=sbert_model_val if (use_sbert_val or use_sbert_hybrid) else SBERT_DEFAULT,
                    use_sbert_hybrid=use_sbert_hybrid,
                    use_meta=snap.get("use_meta", False),
                    calib_method=snap.get("calib_method", "sigmoid"),
                    use_smote=snap.get("use_smote", True),
                    classifier_backend="linearsvc",
                )

                model_path = MODEL_DIR / f"marker1_model_{now_stamp()}.joblib"
                # class_examples — несколько типичных текстов на класс (для LLM-ре-ранка).
                try:
                    from llm_reranker import build_class_examples_from_training as _bce
                    _class_examples = _bce(X, y, n_per_class=3, max_chars=260)
                except Exception:
                    _class_examples = {}
                _rep_dict_main = train_extras.get("report_dict") or {}
                _eval_metrics = {
                    "macro_f1":   round(float((_rep_dict_main.get("macro avg") or {}).get("f1-score", 0.0)), 4),
                    "accuracy":   round(float(_rep_dict_main.get("accuracy", 0.0)), 4),
                    "roc_auc":    train_extras.get("roc_auc_macro"),
                    "n_train":    int(train_extras.get("n_train", 0)),
                    "n_test":     int(train_extras.get("n_test", 0)),
                    "training_duration_sec": train_extras.get("training_duration_sec"),
                    "model_size_bytes":      train_extras.get("model_size_bytes"),
                    "ece":        train_extras.get("ece"),
                    "mce":        train_extras.get("mce"),
                    "temperature": train_extras.get("temperature", 1.0),
                    "per_class_f1": {
                        cls: round(float(vals.get("f1-score", 0.0)), 4)
                        for cls, vals in _rep_dict_main.items()
                        if isinstance(vals, dict) and "f1-score" in vals
                        and cls not in ("macro avg", "weighted avg", "micro avg")
                    },
                    "per_class_precision": {
                        cls: round(float(vals.get("precision", 0.0)), 4)
                        for cls, vals in _rep_dict_main.items()
                        if isinstance(vals, dict) and "precision" in vals
                        and cls not in ("macro avg", "weighted avg", "micro avg")
                    },
                    "per_class_recall": {
                        cls: round(float(vals.get("recall", 0.0)), 4)
                        for cls, vals in _rep_dict_main.items()
                        if isinstance(vals, dict) and "recall" in vals
                        and cls not in ("macro avg", "weighted avg", "micro avg")
                    },
                    "per_class_thresholds": train_extras.get("per_class_thresholds", {}),
                    "trained_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
                }
                _bundle = {
                    "artifact_type": TRAIN_MODEL_ARTIFACT_TYPE,
                    "pipeline": pipe,
                    "config": asdict(cfg),
                    "schema_version": 1,
                    "per_class_thresholds": train_extras.get("per_class_thresholds", {}),
                    "temperature": train_extras.get("temperature", 1.0),
                    "class_examples": _class_examples,
                    "eval_metrics": _eval_metrics,
                }
                if _group_model is not None:
                    _bundle["group_model"] = _group_model
                    _bundle["class_to_group"] = _class_to_group
                if _kfold_models:
                    _bundle["kfold_models"] = _kfold_models
                    _bundle["kfold_classes"] = _kfold_classes
                workflow.persist_artifact(_bundle, str(model_path))
                try:
                    from experiment_log import log_experiment as _log_exp
                    _log_exp(str(model_path), snap, _eval_metrics)
                except Exception as _logexp_exc:
                    _log.warning("experiment_log failed: %s", _logexp_exc)

                elapsed = time.time() - t0

                thresh_hints = ""
                if train_extras.get("thresh_90") is not None:
                    _roc_line = ""
                    if train_extras.get("roc_auc_macro") is not None:
                        _roc_line = f"\n  ROC-AUC macro (OvR):     {train_extras['roc_auc_macro']:.4f}"
                    thresh_hints = (
                        f"Рекомендуемые пороги review_threshold (на валидационной выборке):\n"
                        f"  90% предсказаний уверены: {train_extras['thresh_90']:.3f}  ← строгий\n"
                        f"  75% предсказаний уверены: {train_extras['thresh_75']:.3f}  ← стандартный\n"
                        f"  50% предсказаний уверены: {train_extras['thresh_50']:.3f}  ← мягкий (медиана)"
                        + _roc_line
                    )

                ascii_cm_text = ""
                if labels is not None and cm is not None and len(labels) <= 12:
                    _cw = max(14, max((len(str(l)) for l in labels), default=0) + 2)
                    _hdr = " " * _cw + "".join(str(l)[:_cw].ljust(_cw) for l in labels)
                    _cm_rows = [_hdr]
                    for _i, _lbl in enumerate(labels):
                        _row = str(_lbl)[:_cw].ljust(_cw)
                        for _j in range(len(labels)):
                            _v = int(cm[_i][_j])
                            _cell = ("*" + str(_v) if _i == _j else str(_v))
                            _row += _cell.ljust(_cw)
                        _cm_rows.append(_row)
                    ascii_cm_text = "\n".join(_cm_rows)

                metrics_path: Optional[Path] = None
                _rep_dict = train_extras.get("report_dict")
                if _rep_dict and labels is not None and cm is not None:
                    try:
                        from openpyxl import Workbook as _WB
                        _mwb = _WB()
                        _mwb.remove(_mwb.active)
                        _ms1 = _mwb.create_sheet("Метрики по классам")
                        _per_thresh = train_extras.get("per_class_thresholds", {})
                        _ms1.append(["Класс", "Precision", "Recall", "F1", "Support", "Порог (P≥0.85)"])
                        for _lbl in labels:
                            _r = _rep_dict.get(str(_lbl), {})
                            _ms1.append([
                                _lbl,
                                round(_r.get("precision", 0), 4),
                                round(_r.get("recall", 0), 4),
                                round(_r.get("f1-score", 0), 4),
                                int(_r.get("support", 0)),
                                round(_per_thresh.get(str(_lbl), 0), 4) or "",
                            ])
                        for _key in ("macro avg", "weighted avg"):
                            _r = _rep_dict.get(_key, {})
                            if _r:
                                _ms1.append([
                                    _key,
                                    round(_r.get("precision", 0), 4),
                                    round(_r.get("recall", 0), 4),
                                    round(_r.get("f1-score", 0), 4),
                                    int(_r.get("support", 0)),
                                    "",
                                ])
                        # ROC-AUC macro (если вычислен)
                        _roc = train_extras.get("roc_auc_macro")
                        if _roc is not None:
                            _ms1.append([])
                            _ms1.append(["ROC-AUC macro (OvR)", round(_roc, 4)])
                        _ms2 = _mwb.create_sheet("Матрица ошибок")
                        _ms2.append(["TRUE / PRED"] + [str(l) for l in labels])
                        for _i, _lbl in enumerate(labels):
                            _ms2.append([_lbl] + [int(cm[_i][_j]) for _j in range(len(labels))])
                        # ── Лист «Анализ ошибок» с примерами текстов ──
                        _conf_ex = train_extras.get("confusion_examples", {})
                        if _conf_ex:
                            _ms3 = _mwb.create_sheet("Анализ ошибок")
                            _ms3.append(["Кол-во", "TRUE (правильный)", "PRED (предсказан)", "Примеры текстов"])
                            for _pair_key, _pair_data in sorted(
                                _conf_ex.items(), key=lambda kv: kv[1]["count"], reverse=True
                            ):
                                _exs_text = "\n---\n".join(_pair_data.get("examples", []))
                                _ms3.append([
                                    _pair_data["count"],
                                    _pair_data["true"],
                                    _pair_data["pred"],
                                    _exs_text,
                                ])
                        # ── Лист «Кросс-валидация» ──
                        if train_extras.get("cv_f1_mean") is not None:
                            _ms4 = _mwb.create_sheet("Кросс-валидация")
                            _cv_n = train_extras.get("cv_n_splits", 5)
                            _ms4.append([f"Стратифицированная {_cv_n}-fold CV (macro F1)"])
                            _ms4.append(["Mean", round(train_extras["cv_f1_mean"], 4)])
                            _ms4.append(["Std",  round(train_extras.get("cv_f1_std", 0.0), 4)])
                            for _ki, _ks in enumerate(train_extras.get("cv_scores", []), 1):
                                _ms4.append([f"Fold {_ki}", _ks])
                            _T = train_extras.get("temperature", 1.0)
                            _ms4.append([])
                            _ms4.append(["Температура калибровки", round(_T, 4)])
                        # ── Лист «Подозрительные метки» ──
                        if snap.get("detect_mislabeled", False):
                            _ml_thr = float(snap.get("mislabeled_threshold", 0.30))
                            _suspicious = detect_mislabeled_examples(
                                pipe, X, y, threshold=_ml_thr, max_results=200,
                                log_cb=lambda m: self.after(0, lambda msg=m: self.log_train(msg)),
                            )
                            if _suspicious:
                                _ms5 = _mwb.create_sheet("Подозрительные метки")
                                _ms5.append([
                                    "Вер-ть истин. класса", "Вер-ть предсказанного",
                                    "Истинная метка", "Предсказанная метка", "Текст",
                                ])
                                for _sr in _suspicious:
                                    _ms5.append([
                                        _sr["prob_true"], _sr["prob_pred"],
                                        _sr["true_label"], _sr["pred_label"],
                                        _sr["text"],
                                    ])
                                self.after(0, lambda n=len(_suspicious): self.log_train(
                                    f"  ⚠ Подозрительных примеров: {n} → лист «Подозрительные метки»"
                                ))
                        # ── Confident Learning (продвинутый поиск ошибок разметки) ──
                        if snap.get("use_confident_learning", False) and not use_sbert_val and not use_sbert_hybrid:
                            try:
                                from ml_core import confident_learning_detect
                                _cl_thr = float(snap.get("confident_learning_threshold", 1.0))
                                self.after(0, lambda: ui_prog(92.0, "Confident Learning: поиск ошибок разметки…"))
                                _cl_feats = _make_tfidf()
                                _cl_suspicious = confident_learning_detect(
                                    X, y,
                                    features=_cl_feats,
                                    balanced=snap["balanced"],
                                    max_iter=snap["max_iter"],
                                    cv=5,
                                    threshold_factor=_cl_thr,
                                    progress_cb=lambda p, s: self.after(0, lambda p=p, s=s: ui_prog(
                                        92.0 + p * 0.04, s)),
                                    cancel_event=self._cancel_event,
                                    n_jobs=self._hw.n_jobs_cv,
                                )
                                if _cl_suspicious:
                                    _cl_sheet = _mwb.create_sheet("Confident Learning")
                                    _cl_sheet.append([
                                        "Вер-ть данной метки", "Вер-ть вероятной метки",
                                        "Данная метка", "Вероятная метка", "Текст",
                                    ])
                                    for _cr in _cl_suspicious[:500]:
                                        _cl_sheet.append([
                                            _cr["p_given"], _cr["p_likely"],
                                            _cr["given_label"], _cr["likely_label"],
                                            _cr["text"],
                                        ])
                                    self.after(0, lambda n=len(_cl_suspicious): self.log_train(
                                        f"  [Confident Learning] {n} подозрительных примеров → лист «Confident Learning»"
                                    ))
                            except Exception as _cl_e:
                                self.after(0, lambda e=_cl_e: self.log_train(
                                    f"  ⚠ Confident Learning: ошибка — {e}"
                                ))
                        metrics_path = MODEL_DIR / (
                            model_path.stem.replace("marker1_model_", "marker1_metrics_") + ".xlsx"
                        )
                        _mwb.save(metrics_path)
                    except OSError as _e:
                        _log.debug("metrics xlsx save: %s", _e)
                        metrics_path = None

                # ── Топ путаниц для лога (с примерами если есть) ──
                top_conf = "(валидация пропущена)"
                if labels is not None and cm is not None:
                    _conf_ex_log = train_extras.get("confusion_examples", {})
                    if _conf_ex_log:
                        _lines = []
                        for _key, _data in sorted(
                            _conf_ex_log.items(), key=lambda kv: kv[1]["count"], reverse=True
                        )[:8]:
                            _exs = _data.get("examples", [])
                            _ex_str = f"  → «{_exs[0][:80]}»" if _exs else ""
                            _lines.append(
                                f"  {_data['count']}× TRUE=«{_data['true']}» → PRED=«{_data['pred']}»{_ex_str}"
                            )
                        top_conf = "\n".join(_lines) if _lines else "(нет путаниц)"
                    else:
                        pairs = []
                        for i in range(len(labels)):
                            for j in range(len(labels)):
                                if i == j: continue
                                v = int(cm[i][j])
                                if v > 0:
                                    pairs.append((v, labels[i], labels[j]))
                        pairs.sort(reverse=True)
                        top_conf = "\n".join([f"  {v} TRUE='{t}' → PRED='{p}'" for v, t, p in pairs[:12]]) or "(нет путаниц)"

                def done():
                    ui_prog(100.0, "Готово ✅")
                    self.train_speed.set("")
                    self.train_eta.set(f"Итого: {elapsed:.0f}с")
                    self.model_file.set(str(model_path))
                    if use_sbert_hybrid:
                        vec_info = f"SBERT+TF-IDF гибрид [{sbert_model_val}]"
                    elif use_sbert_val:
                        vec_info = f"SBERT [{sbert_model_val}]"
                    else:
                        meta_suffix = " + Мета-признаки" if snap.get("use_meta") else ""
                        vec_info = f"TF-IDF (hybrid char+word){meta_suffix}"
                    _skip_lbl  = stats.get('rows_skipped_no_label', 0)
                    _skip_text = stats.get('rows_skipped_empty_text', 0)
                    _skip_info = ""
                    if _skip_lbl or _skip_text:
                        _skip_info = f" | пропущено: {_skip_lbl} без метки, {_skip_text} без текста"
                    self.log_train(f"Готово ✅ | time={elapsed:.1f}s | raw={stats.get('rows_raw')} | used={stats.get('rows_used')}{_skip_info} | roles={stats.get('roles_found_rows')}")
                    self.log_train(f"Векторайзер: {vec_info}")
                    self.log_train(f"Классификатор: {clf_type}")
                    self.log_train(f"Модель сохранена: {model_path}")
                    try:
                        import os as _os
                        _sz = _os.path.getsize(str(model_path)) / (1024 * 1024)
                        self.log_train(f"[Сохранение] Размер файла: {_sz:.1f} МБ")
                    except OSError as _e:
                        _log.debug("setfit model file size stat: %s", _e)
                    _t_fit_std = elapsed - _t_read
                    self.log_train(
                        f"[Время] Чтение: {_t_read:.1f}с"
                        f" | Обучение+Вектор.: {_t_fit_std:.1f}с"
                        f" | Итого: {elapsed:.1f}с"
                    )
                    if metrics_path:
                        self.log_train(f"Метрики Excel: {metrics_path}")
                    self.log_train("\n--- МЕТРИКИ ---\n" + report)
                    if thresh_hints:
                        self.log_train("\n--- ПОРОГИ УВЕРЕННОСТИ ---\n" + thresh_hints)
                        self.log_train("→ Перейдите на вкладку «Классификация» и нажмите «↓ Из обучения»")
                    if ascii_cm_text:
                        self.log_train("\n--- МАТРИЦА ОШИБОК ---\n" + ascii_cm_text)
                    self.log_train("\n--- ЧАСТЫЕ ОШИБКИ (TRUE→PRED) ---\n" + top_conf)
                    # ── CV результаты ──
                    if train_extras.get("cv_f1_mean") is not None:
                        _cv_m = train_extras["cv_f1_mean"]
                        _cv_s = train_extras.get("cv_f1_std", 0.0)
                        _cv_n = train_extras.get("cv_n_splits", 5)
                        self.log_train(
                            f"\n[CV {_cv_n}-fold] macro F1 = {_cv_m:.3f} ± {_cv_s:.3f}"
                            "  ← стабильная оценка (holdout: см. метрики выше)"
                        )
                    # ── Температура калибровки ──
                    _T_log = train_extras.get("temperature", 1.0)
                    if abs(_T_log - 1.0) > 0.05:
                        _dir_log = "вероятности занижены → T > 1" if _T_log > 1.0 else "вероятности завышены → T < 1"
                        self.log_train(
                            f"[Калибровка] T={_T_log:.3f}  ({_dir_log})"
                            "  — температура сохранена в модель, применяется при классификации"
                        )
                    if train_extras.get("thresh_75") is not None:
                        _t75 = float(train_extras["thresh_75"])
                        self._rec_threshold_75.set(_t75)
                        self._rec_thr_label.set(f"← рек. {_t75:.3f} (из обучения)")
                    _lifecycle.complete()

                self.after(0, done)

            except InterruptedError:
                def cancelled_train():
                    _lifecycle.cancelled(
                        ui_prog=ui_prog,
                        log_message="⏹ Обучение отменено пользователем",
                    )
                self.after(0, cancelled_train)

            except FeatureBuildError as e:
                _tb = _traceback.format_exc()
                def err(e=e, tb=_tb):
                    env = ErrorEnvelope.from_exception(
                        e,
                        error_code="TRAIN_FEATURE_BUILD",
                        stage="run_training",
                        hint="Проверьте колонки, качество данных и минимальные объёмы классов.",
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка данных",
                        envelope=env,
                        traceback_text=tb,
                    )
                    messagebox.showerror("Ошибка данных / признаков", str(e))
                self.after(0, err)

            except Exception as e:
                if isinstance(e, MemoryError):
                    import gc as _gc_train
                    _gc_train.collect()
                _tb = _traceback.format_exc()
                def err(e=e, tb=_tb):
                    env = ErrorEnvelope.from_exception(
                        e,
                        error_code="TRAIN_UNEXPECTED",
                        stage="run_training",
                        hint="Повторите запуск на меньшем наборе и приложите traceback из лога.",
                    )
                    _lifecycle.failed(
                        ui_prog=ui_prog,
                        status="Ошибка",
                        envelope=env,
                        traceback_text=tb,
                    )
                    _brief = str(e)[:300] + ("…" if len(str(e)) > 300 else "")
                    messagebox.showerror(
                        "Ошибка обучения",
                        f"{_brief}\n\nПодробности смотри в логе обучения.",
                    )
                self.after(0, err)

            finally:
                # Страховка: сбрасываем флаг в потоке на случай если self.after()
                # не вызовет callback (например, при завершении приложения).
                with self._proc_lock:
                    self._processing = False

        threading.Thread(target=worker, daemon=True).start()

    # ------------------------------------------------- Advanced ML dialogs

    def _run_mlm_pretrain_dialog(self) -> None:
        """Диалог запуска MLM-дообучения SBERT-модели на загруженных текстах."""
        if not self.train_files:
            messagebox.showwarning("MLM Pretrain", "Сначала загрузите файлы обучения.", parent=self)
            return

        dlg = tk.Toplevel(self)
        dlg.title("MLM Pretrain — дообучение SBERT")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.geometry("480x340")

        body = ttk.Frame(dlg, padding=16)
        body.pack(fill="both", expand=True)

        ttk.Label(body, text="Дообучение BERT-encoder на доменных текстах через Masked LM.\n"
                             "Улучшает качество SBERT-векторизации для узкоспециализированных текстов.",
                  wraplength=440, justify="left").pack(anchor="w", pady=(0, 12))

        frm = ttk.LabelFrame(body, text="Параметры", padding=(10, 6))
        frm.pack(fill="x")

        v_model = tk.StringVar(value="ai-forever/ru-en-RoSBERTa")
        v_out   = tk.StringVar(value="mlm_pretrained")
        v_epoch = tk.IntVar(value=3)
        v_batch = tk.IntVar(value=8)

        for r, (lbl, var, wtype, kw) in enumerate([
            ("Модель (HF repo/путь):", v_model, "entry",   {"width": 30}),
            ("Выходная директория:",   v_out,   "entry",   {"width": 30}),
            ("Эпохи:",                 v_epoch, "spinbox", {"from_": 1, "to": 20}),
            ("Batch size:",            v_batch, "spinbox", {"from_": 1, "to": 64}),
        ]):
            ttk.Label(frm, text=lbl, anchor="w").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
            if wtype == "entry":
                w = ttk.Entry(frm, textvariable=var, **kw)
            else:
                w = ttk.Spinbox(frm, textvariable=var, width=8, **kw)
            w.grid(row=r, column=1, sticky="w", pady=3)

        log_var = tk.StringVar(value="Нажмите «Запустить» для начала дообучения.")
        ttk.Label(body, textvariable=log_var, foreground="#555", wraplength=440,
                  justify="left").pack(anchor="w", pady=(10, 0))

        btn_frm = ttk.Frame(dlg)
        btn_frm.pack(fill="x", padx=16, pady=(0, 12))

        def _start_mlm():
            btn_run.config(state="disabled")
            log_var.set("Загрузка датасета…")

            def worker():
                try:
                    from ml_mlm_pretrain import pretrain_mlm, is_available
                    if not is_available():
                        self.after(0, lambda: log_var.set(
                            "Установите: pip install transformers datasets accelerate"))
                        return

                    snap = self._snap_params()
                    X, _, _ = self._build_training_dataset(
                        [Path(p) for p in self.train_files], snap)
                    self.after(0, lambda: log_var.set(f"Загружено {len(X)} текстов. Запуск MLM…"))

                    def _cb(msg: str) -> None:
                        self.after(0, lambda m=msg: log_var.set(m[-120:]))

                    out = pretrain_mlm(
                        X,
                        model_name=v_model.get().strip(),
                        output_dir=v_out.get().strip(),
                        num_train_epochs=v_epoch.get(),
                        per_device_batch_size=v_batch.get(),
                        log_cb=_cb,
                    )
                    self.after(0, lambda: log_var.set(f"Готово! Модель сохранена в: {out}"))
                    self.after(0, lambda: messagebox.showinfo(
                        "MLM Pretrain", f"Дообучение завершено.\nМодель: {out}", parent=dlg))
                except Exception as exc:
                    _msg = str(exc)[:300]
                    self.after(0, lambda m=_msg: log_var.set(f"Ошибка: {m}"))
                    self.after(0, lambda m=_msg: messagebox.showerror(
                        "MLM Pretrain", m, parent=dlg))
                finally:
                    self.after(0, lambda: btn_run.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        btn_run = ttk.Button(btn_frm, text="Запустить", command=_start_mlm)
        btn_run.pack(side="left", padx=(0, 8))
        ttk.Button(btn_frm, text="Закрыть", command=dlg.destroy).pack(side="right")

    def _run_distillation_dialog(self) -> None:
        """Диалог запуска дистилляции знаний (teacher → student)."""
        dlg = tk.Toplevel(self)
        dlg.title("Дистилляция знаний (Teacher → Student)")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.geometry("520x380")

        body = ttk.Frame(dlg, padding=16)
        body.pack(fill="both", expand=True)

        ttk.Label(body, text="Передача знаний от большой модели-учителя к быстрой модели-студенту.\n"
                             "Студент обучается на мягких вероятностях учителя — сохраняя качество при "
                             "меньшем размере.",
                  wraplength=480, justify="left").pack(anchor="w", pady=(0, 12))

        frm = ttk.LabelFrame(body, text="Параметры", padding=(10, 6))
        frm.pack(fill="x")

        v_teacher = tk.StringVar()
        v_temp    = tk.DoubleVar(value=3.0)
        v_alpha   = tk.DoubleVar(value=0.5)
        v_out     = tk.StringVar()

        for r, (lbl, var, wtype, kw) in enumerate([
            ("Модель-учитель (.joblib):", v_teacher, "entry",   {"width": 35}),
            ("Выходной файл (.joblib):",  v_out,     "entry",   {"width": 35}),
            ("Температура (T):",          v_temp,    "spinbox", {"from_": 1.0, "to": 10.0, "increment": 0.5}),
            ("Alpha (вес учителя):",      v_alpha,   "spinbox", {"from_": 0.0, "to": 1.0, "increment": 0.1}),
        ]):
            ttk.Label(frm, text=lbl, anchor="w").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
            if wtype == "entry":
                w = ttk.Entry(frm, textvariable=var, **kw)
            else:
                w = ttk.Spinbox(frm, textvariable=var, width=8, **kw)
            w.grid(row=r, column=1, sticky="w", pady=3)

        def _browse_teacher():
            p = filedialog.askopenfilename(title="Выберите модель-учитель",
                                           filetypes=[("Joblib", "*.joblib"), ("Все", "*.*")])
            if p:
                v_teacher.set(p)
                if not v_out.get():
                    v_out.set(str(Path(p).parent / (Path(p).stem + "_student.joblib")))

        ttk.Button(frm, text="…", width=3, command=_browse_teacher).grid(row=0, column=2, padx=4)

        log_var = tk.StringVar(value="Укажите модель-учитель и нажмите «Запустить».")
        ttk.Label(body, textvariable=log_var, foreground="#555", wraplength=480,
                  justify="left").pack(anchor="w", pady=(10, 0))

        btn_frm = ttk.Frame(dlg)
        btn_frm.pack(fill="x", padx=16, pady=(0, 12))

        def _start_distill():
            teacher_path = v_teacher.get().strip()
            out_path = v_out.get().strip()
            if not teacher_path or not out_path:
                messagebox.showwarning("Дистилляция", "Укажите файлы учителя и выхода.", parent=dlg)
                return
            if not self.train_files:
                messagebox.showwarning("Дистилляция", "Загрузите файлы обучения.", parent=dlg)
                return

            btn_run.config(state="disabled")
            log_var.set("Загрузка…")

            def worker():
                try:
                    import joblib
                    from copy import deepcopy
                    from ml_distillation import distill_soft_labels
                    from model_loader import safe_load_bundle

                    pkg = safe_load_bundle(teacher_path)
                    teacher = pkg["pipeline"]

                    snap = self._snap_params()
                    X, y, _ = self._build_training_dataset(
                        [Path(p) for p in self.train_files], snap)

                    self.after(0, lambda: log_var.set(
                        f"Учитель загружен. Датасет: {len(X)} текстов. Запуск дистилляции…"))

                    student_pipe = deepcopy(teacher)
                    student = distill_soft_labels(
                        teacher, student_pipe, X, y,
                        temperature=v_temp.get(),
                        alpha=v_alpha.get(),
                        log_cb=lambda m: self.after(0, lambda msg=m: log_var.set(msg[-120:])),
                    )

                    bundle = dict(pkg)
                    bundle["pipeline"] = student
                    bundle["config"] = snap
                    joblib.dump(bundle, out_path, compress=3)

                    self.after(0, lambda: log_var.set(f"Готово! Сохранено: {out_path}"))
                    self.after(0, lambda: messagebox.showinfo(
                        "Дистилляция", f"Студент сохранён:\n{out_path}", parent=dlg))
                except Exception as exc:
                    _msg = str(exc)[:300]
                    self.after(0, lambda m=_msg: log_var.set(f"Ошибка: {m}"))
                    self.after(0, lambda m=_msg: messagebox.showerror(
                        "Дистилляция", m, parent=dlg))
                finally:
                    self.after(0, lambda: btn_run.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        btn_run = ttk.Button(btn_frm, text="Запустить", command=_start_distill)
        btn_run.pack(side="left", padx=(0, 8))
        ttk.Button(btn_frm, text="Закрыть", command=dlg.destroy).pack(side="right")

    def _run_active_learning_dialog(self) -> None:
        """Диалог активного обучения: находит строки с наименьшей уверенностью модели."""
        dlg = tk.Toplevel(self)
        dlg.title("Active Learning — приоритеты разметки")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.geometry("520x380")

        body = ttk.Frame(dlg, padding=16)
        body.pack(fill="both", expand=True)

        ttk.Label(body, text="Находит строки, которые модель классифицирует с наименьшей уверенностью.\n"
                             "Разметка именно этих строк даёт максимальный прирост качества модели.",
                  wraplength=480, justify="left").pack(anchor="w", pady=(0, 12))

        frm = ttk.LabelFrame(body, text="Параметры", padding=(10, 6))
        frm.pack(fill="x")

        v_model     = tk.StringVar()
        v_unlabeled = tk.StringVar()
        v_out       = tk.StringVar()
        v_topn      = tk.IntVar(value=50)
        v_strategy  = tk.StringVar(value="entropy")

        _rows = [
            ("Модель (.joblib):",   v_model,     "entry",   {"width": 32}),
            ("Неразмеченный файл:", v_unlabeled, "entry",   {"width": 32}),
            ("Выходной Excel:",     v_out,       "entry",   {"width": 32}),
            ("Топ-N примеров:",     v_topn,      "spinbox", {"from_": 10, "to": 500}),
            ("Стратегия:",          v_strategy,  "combo",
             {"values": ["entropy", "margin", "least_confident"], "width": 18}),
        ]
        for r, (lbl, var, wtype, kw) in enumerate(_rows):
            ttk.Label(frm, text=lbl, anchor="w").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
            if wtype == "entry":
                w = ttk.Entry(frm, textvariable=var, **kw)
            elif wtype == "spinbox":
                w = ttk.Spinbox(frm, textvariable=var, width=8, **kw)
            else:
                w = ttk.Combobox(frm, textvariable=var, state="readonly", **kw)
            w.grid(row=r, column=1, sticky="w", pady=3)

        def _browse_file(var, title, ftypes):
            p = filedialog.askopenfilename(title=title, filetypes=ftypes)
            if p:
                var.set(p)

        def _browse_out():
            p = filedialog.asksaveasfilename(title="Сохранить результат",
                                             defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")])
            if p:
                v_out.set(p)

        ttk.Button(frm, text="…", width=3,
                   command=lambda: _browse_file(v_model, "Модель",
                       [("Joblib", "*.joblib"), ("Все", "*.*")])).grid(row=0, column=2, padx=4)
        ttk.Button(frm, text="…", width=3,
                   command=lambda: _browse_file(v_unlabeled, "Неразмеченный файл",
                       [("Excel/CSV", "*.xlsx *.csv"), ("Все", "*.*")])).grid(row=1, column=2, padx=4)
        ttk.Button(frm, text="…", width=3, command=_browse_out).grid(row=2, column=2, padx=4)

        log_var = tk.StringVar(value="Укажите файлы и нажмите «Запустить».")
        ttk.Label(body, textvariable=log_var, foreground="#555", wraplength=480,
                  justify="left").pack(anchor="w", pady=(10, 0))

        btn_frm = ttk.Frame(dlg)
        btn_frm.pack(fill="x", padx=16, pady=(0, 12))

        def _start_al():
            if not v_model.get().strip() or not v_unlabeled.get().strip() or not v_out.get().strip():
                messagebox.showwarning("Active Learning", "Заполните все поля.", parent=dlg)
                return

            btn_run.config(state="disabled")
            log_var.set("Загрузка модели…")

            def worker():
                try:
                    from model_loader import safe_load_bundle
                    from ml_diagnostics import rank_for_active_learning
                    from excel_utils import open_tabular
                    import openpyxl

                    pkg = safe_load_bundle(v_model.get().strip())
                    pipe = pkg["pipeline"]
                    clf = pipe[-1] if hasattr(pipe, "__len__") else pipe
                    classes = list(getattr(clf, "classes_", []))

                    self.after(0, lambda: log_var.set("Чтение неразмеченных текстов…"))
                    snap = self._snap_params()
                    desc_col = snap.get("desc_col", "")

                    texts: List[str] = []
                    with open_tabular(Path(v_unlabeled.get().strip())) as it:
                        header = [str(h or "") for h in next(it)]
                        col_idx = header.index(desc_col) if desc_col in header else 0
                        for row in it:
                            cell = row[col_idx] if col_idx < len(row) else ""
                            texts.append(str(cell or "").strip())
                    texts = [t for t in texts if t]

                    self.after(0, lambda: log_var.set(
                        f"{len(texts)} текстов. Вычисляем вероятности…"))

                    proba = pipe.predict_proba(texts)
                    ranked = rank_for_active_learning(
                        texts, proba, classes,
                        top_n=v_topn.get(),
                        strategy=v_strategy.get(),
                    )

                    out_path = v_out.get().strip()
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Active Learning"
                    ws.append(["#", "Текст", "Лучший класс", "P(лучш.)", "Score", "Стратегия"])
                    for i, rec in enumerate(ranked, 1):
                        ws.append([
                            i,
                            rec["text"],
                            rec["best_label"],
                            round(rec["best_prob"], 4),
                            round(rec["score"], 4),
                            rec["strategy"],
                        ])
                    wb.save(out_path)

                    self.after(0, lambda: log_var.set(
                        f"Готово! {len(ranked)} строк → {out_path}"))
                    self.after(0, lambda: messagebox.showinfo(
                        "Active Learning",
                        f"Сохранено {len(ranked)} строк для разметки:\n{out_path}",
                        parent=dlg))
                except Exception as exc:
                    _msg = str(exc)[:300]
                    self.after(0, lambda m=_msg: log_var.set(f"Ошибка: {m}"))
                    self.after(0, lambda m=_msg: messagebox.showerror(
                        "Active Learning", m, parent=dlg))
                finally:
                    self.after(0, lambda: btn_run.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        btn_run = ttk.Button(btn_frm, text="Запустить", command=_start_al)
        btn_run.pack(side="left", padx=(0, 8))
        ttk.Button(btn_frm, text="Закрыть", command=dlg.destroy).pack(side="right")
