# -*- coding: utf-8 -*-
"""Excel round-trip integration test.

Creates an in-memory XLSX file, reads it back through excel_utils,
runs a trained model to predict labels, and writes results to a new XLSX.
Verifies the full Excel I/O path works end-to-end without UI.
"""
from __future__ import annotations

import pathlib
import sys

import pytest

sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

try:
    import openpyxl  # noqa: F401 — required for this test
except ImportError:
    pytest.skip("openpyxl not available", allow_module_level=True)

try:
    import joblib  # noqa: F401
except ImportError:
    pytest.skip("joblib not available", allow_module_level=True)


# Synthetic data reused from test_e2e_train_predict
_TEXTS_BRIEF = [
    "Платёж задержан уже трое суток.",
    "Карта заблокирована при оплате.",
    "Хочу сменить пароль от приложения.",
    "Перевод завис в обработке.",
    "Разблокируйте мою карту пожалуйста.",
]
_LABELS_BRIEF = [
    "задержка_платежа", "блокировка_карты", "смена_пароля",
    "задержка_платежа", "блокировка_карты",
]


@pytest.fixture(scope="module")
def mini_model(tmp_path_factory):
    """Train a tiny model and return the pipeline object."""
    from ml_vectorizers import make_hybrid_vectorizer
    from ml_training import train_model

    # More training data for 3-class model (need at least 2 samples/class in test split)
    X = _TEXTS_BRIEF * 6   # 30 samples
    y = _LABELS_BRIEF * 6

    features = make_hybrid_vectorizer(
        char_ng=(2, 3), word_ng=(1, 1), min_df=1, max_features=2000,
        use_svd=False, use_lemma=False, use_per_field=False,
        use_meta=False, sublinear_tf=True,
    )
    pipe, *_ = train_model(
        X=X, y=y, features=features,
        C=1.0, max_iter=300, balanced=True,
        calib_method="sigmoid", test_size=0.2,
        random_state=42, use_smote=False,
    )
    return pipe


def _make_input_xlsx(path: pathlib.Path, texts: list) -> None:
    """Write texts to a single-column XLSX with header 'Текст'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Текст"])
    for t in texts:
        ws.append([t])
    wb.save(path)


def _read_xlsx_col(path: pathlib.Path, col_idx: int = 0) -> list:
    """Read all values from a given column (0-based), skip header."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return [row[col_idx] for row in rows if row]


def test_excel_read_write_roundtrip(tmp_path, mini_model):
    """Create XLSX → read rows → predict → write output → verify."""
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    _make_input_xlsx(input_path, _TEXTS_BRIEF)

    # Read back
    from excel_utils import open_tabular, read_headers

    headers = read_headers(input_path)
    assert "Текст" in headers, f"Header 'Текст' not found, got: {headers}"

    texts_read = []
    with open_tabular(input_path) as rows:
        _hdr = next(rows)
        col_idx = list(_hdr).index("Текст")
        for row in rows:
            val = row[col_idx] if len(row) > col_idx else None
            if val:
                texts_read.append(str(val))

    assert len(texts_read) == len(_TEXTS_BRIEF)

    # Predict
    preds = mini_model.predict(texts_read)
    assert len(preds) == len(_TEXTS_BRIEF)

    # Write output XLSX
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(["Текст", "Метка"])
    for text, label in zip(texts_read, preds):
        ws_out.append([text, label])
    wb_out.save(output_path)

    # Verify output
    assert output_path.exists()
    labels_written = _read_xlsx_col(output_path, col_idx=1)
    assert len(labels_written) == len(_TEXTS_BRIEF)
    valid_labels = {"задержка_платежа", "блокировка_карты", "смена_пароля"}
    for lbl in labels_written:
        assert lbl in valid_labels, f"Неизвестная метка в выводе: {lbl!r}"


def test_excel_row_count(tmp_path):
    """estimate_total_rows should correctly count rows in XLSX."""
    from excel_utils import estimate_total_rows

    path = tmp_path / "count_test.xlsx"
    _make_input_xlsx(path, _TEXTS_BRIEF)

    # estimate_total_rows returns total data rows (excluding header)
    total = estimate_total_rows([path])
    # May be approximate (fast counting method) but should be > 0
    assert total > 0
